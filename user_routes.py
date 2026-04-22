"""
user_routes.py — User Management + Profile + Permission Admin
"""
import json
from flask import send_file, Blueprint, render_template, redirect, url_for, request, flash, jsonify
from flask_login import login_required, current_user
from audit_helper import audit, snapshot
from functools import wraps
from datetime import datetime
from models import db, User, LoginLog, Module, UserPermission
from models.employee import Employee
from permissions import seed_permissions, MODULE_SUB_PERMS

users_bp = Blueprint('users_bp', __name__, url_prefix='/admin')


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════
# USER CRUD
# ══════════════════════════════════════

@users_bp.route('/users')
@login_required
@admin_required
def users():
    from flask import send_file, request
    search   = request.args.get('search', '').strip()
    role_f   = request.args.get('role_f', '')
    status_f = request.args.get('status_f', '')
    sort_by  = request.args.get('sort_by', 'full_name')
    sort_dir = request.args.get('sort_dir', 'asc')

    q = User.query
    if search:
        s = f'%{search}%'
        q = q.filter(User.full_name.ilike(s) | User.email.ilike(s) | User.username.ilike(s))
    if role_f:
        q = q.filter_by(role=role_f)
    if status_f:
        q = q.filter_by(is_active=(status_f == 'active'))

    sort_col = getattr(User, sort_by, User.full_name)
    q = q.order_by(sort_col.asc() if sort_dir == 'asc' else sort_col.desc())
    all_users = q.all()

    return render_template('admin/users/index.html',
        users=all_users, active_page='user_mgmt',
        search=search, role_f=role_f, status_f=status_f,
        sort_by=sort_by, sort_dir=sort_dir)


@users_bp.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def user_add():
    if request.method == 'POST':
        username  = request.form.get('username','').strip().lower()
        email     = request.form.get('email','').strip().lower()
        full_name = request.form.get('full_name','').strip()
        role      = request.form.get('role','user')
        password  = request.form.get('password','').strip() or 'HCP@123'
        is_active = 'is_active' in request.form

        if not username or not email:
            flash('Username and email required.','error')
            return redirect(url_for('users_bp.user_add'))
        if User.query.filter_by(username=username).first():
            flash(f'Username "{username}" exists.','error')
            return redirect(url_for('users_bp.user_add'))
        if User.query.filter_by(email=email).first():
            flash(f'Email "{email}" exists.','error')
            return redirect(url_for('users_bp.user_add'))

        u = User(username=username, email=email, full_name=full_name,
                 role=role, is_active=is_active)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        audit('users','INSERT', u.id, username, f'User created by {current_user.username}: {username} | Role: {u.role} | Email: {u.email}')
        flash(f'User "{full_name or username}" created! Default password: {password}','success')
        return redirect(url_for('users_bp.users'))

    return render_template('admin/users/form.html', user=None, active_page='user_mgmt')


@users_bp.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def user_edit(id):
    u = User.query.get_or_404(id)
    if request.method == 'POST':
        u.full_name = request.form.get('full_name','').strip()
        u.email     = request.form.get('email','').strip().lower()
        u.role      = request.form.get('role','user')
        u.is_active = 'is_active' in request.form
        new_pw = request.form.get('password','').strip()
        if new_pw: u.set_password(new_pw)
        db.session.commit()
        audit('users','UPDATE', u.id, u.username, f'User updated by {current_user.username}: {u.username} | Role: {u.role}')
        flash('User updated!','success')
        return redirect(url_for('users_bp.users'))
    return render_template('admin/users/form.html', user=u, active_page='user_mgmt')


@users_bp.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def user_delete(id):
    u = User.query.get_or_404(id)
    if u.id == current_user.id:
        flash("Can't delete yourself.",'error')
        return redirect(url_for('users_bp.users'))
    db.session.delete(u)
    db.session.commit()
    audit('users','DELETE', id, '', f'User deleted by {current_user.username}')
    flash(f'User deleted.','success')
    return redirect(url_for('users_bp.users'))


@users_bp.route('/users/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def user_toggle(id):
    u = User.query.get_or_404(id)
    if u.id == current_user.id:
        return jsonify(success=False, error="Can't deactivate yourself")
    u.is_active = not u.is_active
    action_lbl = 'ENABLE' if u.is_active else 'DISABLE'
    audit('users', action_lbl, id, f'{u.username}', f'User {action_lbl.lower()}d by {current_user.username}')
    db.session.commit()
    return jsonify(success=True, is_active=u.is_active)


# ══════════════════════════════════════
# PERMISSION MATRIX
# ══════════════════════════════════════

@users_bp.route('/permissions')
@login_required
@admin_required
def permissions():
    """DEPRECATED: Role-based permission matrix removed.
    System is now user-only — har user ko individually permission do.
    Redirect to Access Control Panel.
    """
    flash('Role-based permissions hata diye gaye hain. User-wise permissions yahan manage karo.', 'info')
    return redirect(url_for('users_bp.acp_panel'))


@users_bp.route('/permissions/save', methods=['POST'])
@login_required
@admin_required
def perm_save():
    """DEPRECATED: No-op. Role-based permission save removed."""
    return jsonify(success=False, error='Role-based permissions removed. Use /admin/user-permissions/<user_id>/toggle')


@users_bp.route('/permissions/seed', methods=['POST'])
@login_required
@admin_required
def perm_seed():
    """Seed default modules only (no role permissions)."""
    seed_permissions()
    flash('Default modules seeded!', 'success')
    return redirect(url_for('users_bp.acp_panel'))


# ══════════════════════════════════════
# PROFILE (any logged-in user)
# ══════════════════════════════════════

@users_bp.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    from models.employee import (Employee, Contractor, EmployeeTypeMaster,
                                 EmployeeLocationMaster, DepartmentMaster,
                                 DesignationMaster)
    from datetime import date as date_type, datetime as _dt
    # Find linked employee
    emp = Employee.query.filter_by(user_id=current_user.id).first()

    # ─────────────────────────────────────────────────────────────────
    # Auto-link: if the current user has no employee record yet, try to
    # (a) find one by matching email, and link it; else
    # (b) create a minimal employee record so profile tabs are usable.
    # This fixes cases like "admin" user who was never added via HR form.
    # ─────────────────────────────────────────────────────────────────
    if not emp:
        # (a) Match by email first (case-insensitive)
        if current_user.email:
            emp = Employee.query.filter(
                Employee.email.ilike(current_user.email),
                Employee.user_id.is_(None)
            ).first()
            if emp:
                emp.user_id = current_user.id
                db.session.commit()

        # (b) Still nothing — auto-create minimal employee record
        if not emp:
            # Derive code from username (uppercase, no special chars)
            base_code = ''.join(ch for ch in (current_user.username or 'USER')
                                if ch.isalnum()).upper() or 'USER'
            emp_code = base_code
            i = 1
            while Employee.query.filter(Employee.employee_code.ilike(emp_code)).first():
                i += 1
                emp_code = f'{base_code}{i}'

            # Split full_name into first/last
            fn = (current_user.full_name or current_user.username or 'User').strip()
            parts = fn.split(maxsplit=1)
            first = parts[0]
            last  = parts[1] if len(parts) > 1 else '—'

            emp = Employee(
                employee_code = emp_code,
                first_name    = first,
                last_name     = last,
                mobile        = '0000000000',    # placeholder — user can edit
                email         = current_user.email or f'{current_user.username}@hcp.com',
                user_id       = current_user.id,
                status        = 'active',
                employee_type = 'Full Time',
                country       = 'India',
                is_probation  = False,
                created_by    = current_user.id,
            )
            db.session.add(emp)
            try:
                db.session.commit()
                flash('Profile record auto-created. Please fill in your details.', 'info')
            except Exception as ex:
                db.session.rollback()
                emp = None   # fall through to "no employee" UI
                flash(f'Could not auto-create profile record: {ex}', 'warning')

    def _pd(v):
        try: return date_type.fromisoformat(v) if v else None
        except: return None

    def _dec(v):
        try: return float(v) if v else None
        except: return None

    if request.method == 'POST':
        action = request.form.get('action', 'profile')
        if action == 'password':
            old_pw = request.form.get('old_password', '')
            new_pw = request.form.get('new_password', '')
            conf   = request.form.get('confirm_password', '')
            if not current_user.check_password(old_pw):
                flash('Current password is incorrect.', 'error')
            elif len(new_pw) < 6:
                flash('New password must be at least 6 characters.', 'error')
            elif new_pw != conf:
                flash('Passwords do not match.', 'error')
            else:
                current_user.set_password(new_pw)
                db.session.commit()
                audit('users','PASSWORD_CHANGE', current_user.id, current_user.username, f'Password changed by {current_user.username}')
                flash('Password changed successfully!', 'success')
            return redirect(url_for('users_bp.profile'))

        # ═══════════════════════════════════════════════════════════════════
        # Full profile save — mirrors hr.emp_edit field-for-field
        # ═══════════════════════════════════════════════════════════════════
        # Update user basic info
        current_user.full_name = request.form.get('full_name', '').strip() or current_user.full_name
        new_email = request.form.get('email', '').strip()
        if new_email:
            current_user.email = new_email

        if emp:
            # Photo & QR
            photo  = request.form.get('photo_base64', '').strip()
            if photo: emp.profile_photo = photo
            qr_b64 = request.form.get('qr_base64', '').strip()
            if qr_b64: emp.qr_code_base64 = qr_b64

            # Emp code is read-only on profile
            emp.employee_id   = request.form.get('employee_id', '').strip() or None
            emp.first_name    = request.form.get('first_name', emp.first_name).strip() or emp.first_name
            emp.middle_name   = request.form.get('middle_name', emp.middle_name or '').strip()
            emp.last_name     = request.form.get('last_name', emp.last_name).strip() or emp.last_name
            emp.mobile        = request.form.get('mobile', emp.mobile).strip() or emp.mobile
            emp.email         = new_email or emp.email
            emp.gender        = request.form.get('gender', '')
            emp.linkedin      = request.form.get('linkedin', '').strip()
            emp.facebook      = request.form.get('facebook', '').strip()
            emp.department    = request.form.get('department', '').strip()
            emp.designation   = request.form.get('designation', '').strip()
            emp.employee_type = request.form.get('employee_type', '')
            emp.location      = request.form.get('location', '').strip()
            emp.is_contractor = request.form.get('is_contractor') == 'yes'
            cid = request.form.get('contractor_id') or None
            emp.contractor_id = int(cid) if cid and emp.is_contractor else None
            emp.date_of_joining = _pd(request.form.get('date_of_joining'))
            emp.date_of_birth   = _pd(request.form.get('date_of_birth'))
            emp.blood_group     = request.form.get('blood_group', '').strip()
            emp.marital_status  = request.form.get('marital_status', '')
            emp.is_block        = request.form.get('is_block') == 'yes'
            emp.is_late         = request.form.get('is_late') == 'yes'
            emp.is_probation    = request.form.get('is_probation') == 'yes'
            emp.status          = request.form.get('status', emp.status or 'active')
            emp.remark          = request.form.get('remark', '').strip()
            rto = request.form.get('reports_to', '').strip()
            try:
                emp.reports_to = int(rto) if rto else None
            except (ValueError, TypeError):
                emp.reports_to = None

            # Address
            emp.address  = request.form.get('address', '').strip()
            emp.city     = request.form.get('city', '').strip()
            emp.state    = request.form.get('state', '').strip()
            emp.country  = request.form.get('country', '').strip() or 'India'
            emp.zip_code = request.form.get('zip_code', '').strip()

            # Professional
            emp.pay_grade        = request.form.get('pay_grade', '').strip()
            emp.shift            = request.form.get('shift', '').strip()
            emp.weekly_off       = request.form.get('weekly_off', '').strip()
            emp.notice_period_days = int(request.form.get('notice_period_days') or 30)
            emp.work_hours_per_day = float(request.form.get('work_hours_per_day') or 8)
            emp.rehire_eligible  = request.form.get('rehire_eligible') == 'yes'
            emp.confirmation_date = _pd(request.form.get('confirmation_date'))
            emp.resignation_date  = _pd(request.form.get('resignation_date'))
            emp.last_working_date = _pd(request.form.get('last_working_date'))

            # KYC
            emp.nationality     = request.form.get('nationality', 'Indian').strip()
            emp.religion        = request.form.get('religion', '').strip()
            emp.caste           = request.form.get('caste', '').strip()
            emp.physically_handicapped = request.form.get('physically_handicapped') == 'yes'
            ma_raw = request.form.get('marriage_anniversary', '').strip()
            if emp.marital_status == 'Married' and ma_raw:
                emp.marriage_anniversary = _pd(ma_raw)
            elif emp.marital_status != 'Married':
                emp.marriage_anniversary = None
            emp.aadhar_number   = request.form.get('aadhar_number', '').strip()
            emp.pan_number      = request.form.get('pan_number', '').strip().upper()
            emp.uan_number      = request.form.get('uan_number', '').strip()
            emp.esic_number     = request.form.get('esic_number', '').strip()
            emp.passport_number = request.form.get('passport_number', '').strip()
            emp.passport_expiry = _pd(request.form.get('passport_expiry'))
            emp.driving_license = request.form.get('driving_license', '').strip()
            emp.dl_expiry       = _pd(request.form.get('dl_expiry'))

            # Emergency
            emp.emergency_name     = request.form.get('emergency_name', '').strip()
            emp.emergency_relation = request.form.get('emergency_relation', '').strip()
            emp.emergency_phone    = request.form.get('emergency_phone', '').strip()
            emp.emergency_address  = request.form.get('emergency_address', '').strip()

            # Bank
            emp.bank_account_holder = request.form.get('bank_account_holder', '').strip()
            emp.bank_name           = request.form.get('bank_name', '').strip()
            emp.bank_account_number = request.form.get('bank_account_number', '').strip()
            emp.bank_ifsc           = request.form.get('bank_ifsc', '').strip().upper()
            emp.bank_branch         = request.form.get('bank_branch', '').strip()
            emp.bank_account_type   = request.form.get('bank_account_type', '').strip()

            # Salary — READ-ONLY from Profile page (security).
            # Users cannot modify their own salary. HR must use Employee Edit form.
            # The fields below are intentionally NOT updated, even if they appear in the POST body.
            # ─────────────────────────────────────────────────────────────
            # emp.salary_ctc, emp.salary_basic, emp.salary_hra, emp.salary_da,
            # emp.salary_ta, emp.salary_medical_allow, emp.salary_special_allow,
            # emp.salary_pf_employee, emp.salary_pf_employer,
            # emp.salary_esic_employee, emp.salary_esic_employer,
            # emp.salary_professional_tax, emp.salary_tds, emp.salary_net,
            # emp.salary_mode, emp.salary_effective_date,
            # emp.salary_conveyance, emp.salary_bonus, emp.salary_incentive, emp.salary_gross
            # ─────────────────────────────────────────────────────────────

            # Education
            emp.highest_qualification = request.form.get('highest_qualification', '').strip()
            emp.university            = request.form.get('university', '').strip()
            emp.passing_year          = int(request.form.get('passing_year') or 0) or None
            emp.specialization        = request.form.get('specialization', '').strip()
            emp.prev_company          = request.form.get('prev_company', '').strip()
            emp.prev_designation      = request.form.get('prev_designation', '').strip()
            emp.total_experience_yrs  = _dec(request.form.get('total_experience_yrs'))
            emp.prev_from_date        = _pd(request.form.get('prev_from_date'))
            emp.prev_to_date          = _pd(request.form.get('prev_to_date'))
            emp.prev_leaving_reason   = request.form.get('prev_leaving_reason', '').strip()

            # Documents
            emp.documents_json = request.form.get('documents_json', emp.documents_json or '[]')

            # ── Phase-1: Family / Contact ───────────────────────
            emp.father_name       = request.form.get('father_name', '').strip() or None
            emp.mother_name       = request.form.get('mother_name', '').strip() or None
            emp.alternate_mobile  = request.form.get('alternate_mobile', '').strip() or None
            emp.personal_email    = request.form.get('personal_email', '').strip() or None

            # ── Phase-1: Permanent Address ──────────────────────
            emp.permanent_address    = request.form.get('permanent_address', '').strip() or None
            emp.permanent_city       = request.form.get('permanent_city', '').strip() or None
            emp.permanent_state      = request.form.get('permanent_state', '').strip() or None
            emp.permanent_country    = request.form.get('permanent_country', 'India').strip() or None
            emp.permanent_zip        = request.form.get('permanent_zip', '').strip() or None
            emp.same_as_current_addr = request.form.get('same_as_current_addr') == 'yes'

            # ── Phase-1: Grade / Probation ──────────────────────
            emp.grade_level = request.form.get('grade_level', '').strip() or None
            try:
                emp.probation_period_months = int(request.form.get('probation_period_months') or 6)
            except (ValueError, TypeError):
                emp.probation_period_months = 6
            emp.probation_end_date = _pd(request.form.get('probation_end_date'))

            # ── Phase-1: Salary extras (READ-ONLY — HR only) ────────────
            # emp.salary_conveyance, emp.salary_bonus, emp.salary_incentive, emp.salary_gross
            # NOT updated from profile page. See Salary block above for rationale.

            # ── Phase-1: PF ─────────────────────────────────────
            emp.pf_applicable        = request.form.get('pf_applicable') == 'yes'
            emp.pf_number            = request.form.get('pf_number', '').strip() or None
            emp.eps_applicable       = request.form.get('eps_applicable') == 'yes'
            emp.previous_pf_transfer = request.form.get('previous_pf_transfer') == 'yes'
            emp.previous_pf_number   = request.form.get('previous_pf_number', '').strip() or None

            # ── Phase-1: ESIC ───────────────────────────────────
            emp.esic_applicable       = request.form.get('esic_applicable') == 'yes'
            emp.esic_nominee_name     = request.form.get('esic_nominee_name', '').strip() or None
            emp.esic_nominee_relation = request.form.get('esic_nominee_relation', '').strip() or None
            emp.esic_family_details   = request.form.get('esic_family_details', '').strip() or None
            emp.esic_dispensary       = request.form.get('esic_dispensary', '').strip() or None

            # ── Phase-1: TDS / Tax ──────────────────────────────
            emp.aadhaar_pan_linked      = request.form.get('aadhaar_pan_linked') == 'yes'
            emp.tax_regime              = request.form.get('tax_regime', 'New').strip() or 'New'
            emp.prev_employer_income    = _dec(request.form.get('prev_employer_income'))
            emp.monthly_tds             = _dec(request.form.get('monthly_tds'))
            emp.investment_declaration  = request.form.get('investment_declaration', '').strip() or None
            emp.proof_submission_status = request.form.get('proof_submission_status', 'Pending').strip() or 'Pending'

            # ── Phase-1: Statutory ──────────────────────────────
            emp.professional_tax_applicable = request.form.get('professional_tax_applicable', 'yes') == 'yes'
            emp.labour_welfare_fund         = request.form.get('labour_welfare_fund') == 'yes'
            emp.gratuity_eligible           = request.form.get('gratuity_eligible') == 'yes'
            emp.bonus_eligible              = request.form.get('bonus_eligible', 'yes') == 'yes'

            # ── Phase-1: Attendance / Leave ─────────────────────
            emp.attendance_code       = request.form.get('attendance_code', '').strip() or None
            emp.overtime_eligible     = request.form.get('overtime_eligible') == 'yes'
            emp.casual_leave_balance  = _dec(request.form.get('casual_leave_balance')) or 0
            emp.sick_leave_balance    = _dec(request.form.get('sick_leave_balance')) or 0
            emp.paid_leave_balance    = _dec(request.form.get('paid_leave_balance')) or 0
            emp.leave_policy          = request.form.get('leave_policy', '').strip() or None

            # ── Phase-1: System Access ──────────────────────────
            emp.official_email = request.form.get('official_email', '').strip() or None
            emp.role_access    = request.form.get('role_access', '').strip() or None

            # ── Phase-1: Exit extras ────────────────────────────
            emp.exit_interview_done  = request.form.get('exit_interview_done') == 'yes'
            emp.exit_interview_notes = request.form.get('exit_interview_notes', '').strip() or None
            emp.ff_settlement_status = request.form.get('ff_settlement_status', 'Pending').strip() or 'Pending'
            emp.ff_settlement_amount = _dec(request.form.get('ff_settlement_amount'))
            emp.ff_settlement_date   = _pd(request.form.get('ff_settlement_date'))

            emp.updated_at = _dt.utcnow()

        db.session.commit()
        audit('users', 'PROFILE_UPDATE', current_user.id, current_user.username,
              f'Profile updated by {current_user.username}')
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('users_bp.profile'))

    # GET — load dropdown data same as emp_edit
    contractors   = Contractor.query.filter_by(status=1, is_deleted=0).order_by(Contractor.company_name).all() if emp else []
    all_employees = Employee.query.filter_by(status='active').order_by(Employee.first_name).all() if emp else []
    emp_types     = EmployeeTypeMaster.query.order_by(EmployeeTypeMaster.name).all() if emp else []
    departments   = DepartmentMaster.query.order_by(DepartmentMaster.name).all() if emp else []
    designations  = DesignationMaster.query.order_by(DesignationMaster.name).all() if emp else []
    locations     = EmployeeLocationMaster.query.order_by(EmployeeLocationMaster.name).all() if emp else []

    logs = LoginLog.query.filter_by(user_id=current_user.id)\
               .order_by(LoginLog.timestamp.desc()).limit(20).all()
    return render_template('admin/profile.html',
        employee=emp, logs=logs, active_page='profile',
        contractors=contractors, all_employees=all_employees,
        emp_types=emp_types, departments=departments,
        designations=designations, locations=locations)


# ══════════════════════════════════════
# AUDIT LOG ROUTES
# ══════════════════════════════════════

@users_bp.route('/audit-logs')
@login_required
def audit_logs():
    from models import AuditLog, User
    from datetime import datetime, timedelta
    from sqlalchemy import func

    # Filters
    module    = request.args.get('module', '')
    action    = request.args.get('action', '')
    user_id   = request.args.get('user_id', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    search    = request.args.get('search', '')
    page      = int(request.args.get('page', 1))
    per_page  = int(request.args.get('per_page', 50))

    q = AuditLog.query
    if module:    q = q.filter(AuditLog.module == module)
    if action:    q = q.filter(AuditLog.action == action)
    if user_id:   q = q.filter(AuditLog.user_id == int(user_id))
    if date_from: q = q.filter(AuditLog.created_at >= datetime.strptime(date_from,'%Y-%m-%d'))
    if date_to:   q = q.filter(AuditLog.created_at <= datetime.strptime(date_to+' 23:59:59','%Y-%m-%d %H:%M:%S'))
    if search:    q = q.filter(
        AuditLog.record_label.ilike(f'%{search}%') |
        AuditLog.detail.ilike(f'%{search}%') |
        AuditLog.username.ilike(f'%{search}%')
    )

    total  = q.count()
    logs   = q.order_by(AuditLog.created_at.desc()).offset((page-1)*per_page).limit(per_page).all()
    total_pages = max(1, (total + per_page - 1) // per_page)

    # Distinct filter options
    modules = [r[0] for r in db.session.query(AuditLog.module).distinct().all() if r[0]]
    actions = [r[0] for r in db.session.query(AuditLog.action).distinct().all() if r[0]]
    all_users = User.query.filter_by(is_active=True).order_by(User.full_name).all()

    # Stats — last 30 days action counts
    since = datetime.now() - timedelta(days=30)
    stats = db.session.query(AuditLog.action, func.count(AuditLog.id).label('cnt'))\
                      .filter(AuditLog.created_at >= since)\
                      .group_by(AuditLog.action)\
                      .order_by(func.count(AuditLog.id).desc()).all()

    return render_template('admin/audit/index.html',
        logs=logs, total=total, page=page, per_page=per_page, total_pages=total_pages,
        module=module, action=action, user_id=user_id,
        date_from=date_from, date_to=date_to, search=search,
        modules=modules, actions=actions, all_users=all_users,
        stats=[{'action': s.action, 'cnt': s.cnt} for s in stats],
        active_page='audit_logs'
    )


@users_bp.route('/audit-logs/export')
@login_required
def audit_export():
    from models import AuditLog
    import csv
    from io import StringIO
    from flask import send_file, Response

    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(5000).all()
    si   = StringIO()
    w    = csv.writer(si)
    w.writerow(['ID','Datetime','Username','Role','Module','Action','Record ID','Record Label','Detail','IP'])
    for l in logs:
        w.writerow([l.id, l.created_at.strftime('%d-%m-%Y %H:%M:%S'),
                    l.username, l.user_role, l.module, l.action,
                    l.record_id or '', l.record_label or '', l.detail or '', l.ip_address or ''])
    output = si.getvalue()
    return Response(output, mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=audit_logs.csv'})


@users_bp.route('/users/export')
@login_required
@admin_required
def users_export():
    import io, sys, subprocess
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'], check=True)
        import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    search   = request.args.get('search', '')
    role_f   = request.args.get('role_f', '')
    status_f = request.args.get('status_f', '')
    sort_by  = request.args.get('sort_by', 'full_name')
    sort_dir = request.args.get('sort_dir', 'asc')

    q = User.query
    if search:
        s = f'%{search}%'
        q = q.filter(User.full_name.ilike(s)|User.email.ilike(s)|User.username.ilike(s))
    if role_f:   q = q.filter_by(role=role_f)
    if status_f: q = q.filter_by(is_active=(status_f=='active'))
    sort_col = getattr(User, sort_by, User.full_name)
    users_data = q.order_by(sort_col.asc() if sort_dir=='asc' else sort_col.desc()).all()

    headers = ["#","Full Name","Username","Email","Role","Status",
               "Last Login","Login Attempts","Created At"]
    rows = []
    for i, u in enumerate(users_data, 1):
        rows.append([
            i, u.full_name or '', u.username or '', u.email or '',
            (u.role or '').title(),
            'Active' if u.is_active else 'Inactive',
            u.last_login.strftime('%d-%m-%Y %H:%M') if u.last_login else '',
            u.login_attempts or 0,
            u.created_at.strftime('%d-%m-%Y %H:%M') if u.created_at else '',
        ])

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Users"
    hdr_fill=PatternFill("solid",fgColor="1E3A5F"); hdr_font=Font(bold=True,color="FFFFFF",size=10)
    hdr_align=Alignment(horizontal="center",vertical="center")
    thin=Side(style="thin",color="D0D7E2"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    alt_fill=PatternFill("solid",fgColor="F0F4FA"); d_font=Font(size=9); d_align=Alignment(vertical="center")
    ws.row_dimensions[1].height=28
    for ci,h in enumerate(headers,1):
        cell=ws.cell(1,ci,h); cell.font=hdr_font; cell.fill=hdr_fill
        cell.alignment=hdr_align; cell.border=bdr
    for ri,row in enumerate(rows,2):
        ws.row_dimensions[ri].height=17; fill=alt_fill if ri%2==0 else None
        for ci,val in enumerate(row,1):
            cell=ws.cell(ri,ci,val); cell.font=d_font; cell.alignment=d_align; cell.border=bdr
            if fill: cell.fill=fill
    for ci in range(1,len(headers)+1):
        col=get_column_letter(ci)
        mx=max((len(str(ws.cell(r,ci).value or '')) for r in range(1,ws.max_row+1)),default=8)
        ws.column_dimensions[col].width=min(mx+2,40)
    ws.freeze_panes="A2"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"users_export_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx")


# ═══════════════════════════════════════════════════════════════════
# USER-WISE PERMISSIONS
# ═══════════════════════════════════════════════════════════════════

@users_bp.route('/user-permissions')
@login_required
@admin_required
def user_permissions_list():
    """Redirect to ACP panel."""
    return redirect(url_for('users_bp.acp_panel'))


@users_bp.route('/acp')
@users_bp.route('/acp/<int:user_id>')
@login_required
@admin_required
def acp_panel(user_id=None):
    """Access Control Panel — split panel design.
    
    User <-> Employee link via Employee.user_id. Har user ke saath uska
    employee_code aur employee_id (biometric/device ID) bhi fetch karte hain
    taki permission assign karte waqt HR team sahi employee identify kar sake.
    """
    all_users = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    modules   = Module.query.filter_by(is_active=True).order_by(Module.sort_order).all()

    # ── Build user_id → Employee map (single query, efficient) ──
    user_ids = [u.id for u in all_users]
    emp_map = {}
    if user_ids:
        emps = Employee.query.filter(Employee.user_id.in_(user_ids)).all()
        emp_map = {e.user_id: e for e in emps if e.user_id}

    selected_user = None
    selected_employee = None
    perm_map = {}
    sub_perm_map = {}

    if user_id:
        selected_user = User.query.get_or_404(user_id)
        selected_employee = emp_map.get(user_id)
        for up in UserPermission.query.filter_by(user_id=user_id).all():
            perm_map[up.module_id] = up
        sub_perm_map = {mid: up.get_sub_permissions() for mid, up in perm_map.items()}

    return render_template('admin/permissions/acp_panel.html',
                           all_users=all_users,
                           emp_map=emp_map,
                           selected_user=selected_user,
                           selected_employee=selected_employee,
                           modules=modules,
                           perm_map=perm_map,
                           sub_perm_map=sub_perm_map,
                           module_sub_perms=MODULE_SUB_PERMS,
                           active_page='user_permissions')


@users_bp.route('/user-permissions/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def user_permissions(user_id):
    """User-wise permission set/edit karo."""
    u = User.query.get_or_404(user_id)
    modules = Module.query.filter_by(is_active=True).order_by(Module.sort_order).all()

    if request.method == 'POST':
        action = request.form.get('action', 'save')

        if action == 'reset':
            # User ke saare permissions delete karo (no access)
            UserPermission.query.filter_by(user_id=user_id).delete()
            db.session.commit()
            flash(f'{u.full_name} ke sab permissions reset kar diye gaye — ab koi access nahi.', 'success')
            return redirect(url_for('users_bp.user_permissions', user_id=user_id))

        else:
            # Save individual module permissions
            _enabled_module_ids = set()
            for mod in modules:
                prefix = f'mod_{mod.id}_'
                can_view   = request.form.get(f'{prefix}view') == 'on'
                can_add    = request.form.get(f'{prefix}add') == 'on'
                can_edit   = request.form.get(f'{prefix}edit') == 'on'
                can_delete = request.form.get(f'{prefix}delete') == 'on'
                can_export = request.form.get(f'{prefix}export') == 'on'

                # Sub-permissions
                sub_keys = [k for k, _ in MODULE_SUB_PERMS.get(mod.name, [])]
                sub_dict = {k: (request.form.get(f'{prefix}sub_{k}') == 'on') for k in sub_keys}

                up = UserPermission.query.filter_by(user_id=user_id, module_id=mod.id).first()
                if not up:
                    up = UserPermission(user_id=user_id, module_id=mod.id)
                    db.session.add(up)
                up.can_view   = can_view
                up.can_add    = can_add
                up.can_edit   = can_edit
                up.can_delete = can_delete
                up.can_export = can_export
                up.set_sub_permissions(sub_dict)
                up.updated_by = current_user.id

                if can_view:
                    _enabled_module_ids.add(mod.id)

            # ── Auto-enable parents of every enabled child ──
            # Agar user ke pass child module ka can_view=True hai lekin parent
            # disabled hai, to sidebar me parent nahi dikhega → child bhi nahi
            # milega. Isliye saare enabled children ke parents ko cascade karke
            # can_view=True kar do.
            db.session.flush()
            for mod in modules:
                if mod.id in _enabled_module_ids and mod.parent_id:
                    parent = Module.query.get(mod.parent_id)
                    while parent:
                        p_up = UserPermission.query.filter_by(
                            user_id=user_id, module_id=parent.id
                        ).first()
                        if not p_up:
                            p_up = UserPermission(
                                user_id=user_id, module_id=parent.id, can_view=True
                            )
                            db.session.add(p_up)
                        elif not p_up.can_view:
                            p_up.can_view = True
                        p_up.updated_by = current_user.id
                        parent = Module.query.get(parent.parent_id) if parent.parent_id else None

            db.session.commit()
            audit('users', 'USER_PERM_SAVE', user_id, u.username,
                  f'User permissions saved for {u.full_name} by {current_user.username}')
            flash(f'{u.full_name} ke permissions save ho gaye!', 'success')
            return redirect(url_for('users_bp.user_permissions', user_id=user_id))

    # GET — load existing user permissions
    perm_map = {}  # module_id → UserPermission
    for up in UserPermission.query.filter_by(user_id=user_id).all():
        perm_map[up.module_id] = up

    # Build sub_perm_map: module_id → {key: bool}
    sub_perm_map = {}
    for mod_id, up in perm_map.items():
        sub_perm_map[mod_id] = up.get_sub_permissions()

    return render_template('admin/permissions/user_permissions.html',
                           target_user=u,
                           modules=modules,
                           perm_map=perm_map,
                           sub_perm_map=sub_perm_map,
                           module_sub_perms=MODULE_SUB_PERMS,
                           active_page='user_permissions')


@users_bp.route('/user-permissions/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def user_perm_toggle(user_id):
    """AJAX toggle — single permission on/off."""
    data      = request.json
    module_id = int(data.get('module_id'))
    action    = data.get('action')   # can_view, can_add, can_edit, can_delete, can_export
    value     = bool(data.get('value'))

    up = UserPermission.query.filter_by(user_id=user_id, module_id=module_id).first()
    if not up:
        up = UserPermission(user_id=user_id, module_id=module_id)
        # Naya record — sub_perms sab True set karo by default
        mod = Module.query.get(module_id)
        if mod:
            from permissions import MODULE_SUB_PERMS
            sub_keys = [k for k, _ in MODULE_SUB_PERMS.get(mod.name, [])]
            up.set_sub_permissions({k: True for k in sub_keys})
        db.session.add(up)

    if action in ('can_view', 'can_add', 'can_edit', 'can_delete', 'can_export', 'can_import'):
        # Admin ka can_view kabhi False nahi hoga — warna apna hi module band ho jaata hai
        target_user = User.query.get(user_id)
        if action == 'can_view' and not value and target_user and target_user.role == 'admin':
            return jsonify({'ok': False, 'error': 'Admin ka View permission disable nahi ho sakta'})
        setattr(up, action, value)
        up.updated_by = current_user.id

        # Agar child ka can_view = True kiya → parent ko bhi auto-enable karo
        if action == 'can_view' and value:
            mod = Module.query.get(module_id)
            if mod and mod.parent_id:
                parent_up = UserPermission.query.filter_by(
                    user_id=user_id, module_id=mod.parent_id
                ).first()
                if not parent_up:
                    # Parent record exist nahi karti → create karke can_view=True set karo
                    parent_up = UserPermission(
                        user_id=user_id, module_id=mod.parent_id, can_view=True
                    )
                    db.session.add(parent_up)
                elif not parent_up.can_view:
                    parent_up.can_view = True
                parent_up.updated_by = current_user.id

        db.session.commit()
        return jsonify({'ok': True, 'value': value})

    # Module Enable/Disable ALL — sare permissions ek saath on/off
    if action == 'disable_all':
        # Admin ka can_view kabhi False nahi
        target_user = User.query.get(user_id)
        _is_target_admin = target_user and target_user.role == 'admin'
        up.can_view   = True if _is_target_admin else value
        up.can_add    = value
        up.can_edit   = value
        up.can_delete = value
        up.can_export = value
        up.can_import = value
        # Sub-permissions bhi sab on/off karo
        mod = Module.query.get(module_id)
        if mod:
            from permissions import MODULE_SUB_PERMS
            sub_keys = [k for k, _ in MODULE_SUB_PERMS.get(mod.name, [])]
            subs = {k: value for k in sub_keys}
            up.set_sub_permissions(subs)

            # ── Child modules bhi cascade karo (e.g. CRM → crm_leads, crm_clients) ──
            child_modules = Module.query.filter_by(parent_id=module_id, is_active=True).all()
            for child in child_modules:
                child_up = UserPermission.query.filter_by(user_id=user_id, module_id=child.id).first()
                if not child_up:
                    child_up = UserPermission(user_id=user_id, module_id=child.id)
                    db.session.add(child_up)
                child_up.can_view   = value
                child_up.can_add    = value
                child_up.can_edit   = value
                child_up.can_delete = value
                child_up.can_export = value
                child_up.can_import = value
                child_sub_keys = [k for k, _ in MODULE_SUB_PERMS.get(child.name, [])]
                child_up.set_sub_permissions({k: value for k in child_sub_keys})
                child_up.updated_by = current_user.id

            # Enable karte waqt parent bhi enable karo
            if value and mod.parent_id:
                parent_up = UserPermission.query.filter_by(
                    user_id=user_id, module_id=mod.parent_id
                ).first()
                if not parent_up:
                    # Parent record exist nahi karti → create karke can_view=True set karo
                    parent_up = UserPermission(
                        user_id=user_id, module_id=mod.parent_id, can_view=True
                    )
                    db.session.add(parent_up)
                elif not parent_up.can_view:
                    parent_up.can_view = True
                parent_up.updated_by = current_user.id

        up.updated_by = current_user.id
        db.session.commit()
        return jsonify({'ok': True, 'value': value})

    # Sub-permission toggle
    if action.startswith('sub_'):
        sub_key = action[4:]
        subs = up.get_sub_permissions()
        subs[sub_key] = value
        up.set_sub_permissions(subs)
        up.updated_by = current_user.id
        db.session.commit()
        return jsonify({'ok': True, 'value': value})

    return jsonify({'ok': False, 'error': 'Unknown action'}), 400

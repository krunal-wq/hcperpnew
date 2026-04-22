"""
hr_routes.py — HR Module: Employee & Contractor CRUD
"""
import base64, io, json
from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify
from flask_login import login_required, current_user
from audit_helper import audit, snapshot
from datetime import datetime
from models import db, User, Employee, Contractor, WishLog, SalaryConfig, SalaryComponent, EmployeeTypeMaster, EmployeeLocationMaster, DepartmentMaster, DesignationMaster
from permissions import get_perm, get_grid_columns, save_grid_columns

hr = Blueprint('hr', __name__, url_prefix='/hr')


def _parse_date(val):
    if not val:
        return None
    val = str(val).strip()
    # Try DD-MM-YYYY (flatpickr format)
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d %b %Y', '%d-%b-%Y'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


# Default grid columns
EMP_COLS_DEFAULT = ['employee_code','full_name','mobile','email','department',
                    'designation','employee_type','date_of_joining','status','qr']
EMP_COLS_ALL = {
    'employee_code':  'Emp Code',
    'full_name':      'Name',
    'mobile':         'Mobile',
    'email':          'Email',
    'department':     'Department',
    'designation':    'Designation',
    'employee_type':  'Type',
    'date_of_joining':'Date of Joining',
    'location':       'Location',
    'is_contractor':  'Contractor?',
    'marital_status': 'Marital Status',
    'blood_group':    'Blood Group',
    'status':         'Status',
    'qr':             'QR Code',
    # ── Phase-1 additions: Family / Contact ──────────────────
    'father_name':       'Father Name',
    'mother_name':       'Mother Name',
    'alternate_mobile':  'Alt Mobile',
    'personal_email':    'Personal Email',
    # ── Phase-1: Permanent Address ───────────────────────────
    'permanent_city':    'Perm. City',
    'permanent_state':   'Perm. State',
    # ── Phase-1: Professional extras ─────────────────────────
    'pay_grade':              'Pay Grade',
    'grade_level':            'Grade / Level',
    'shift':                  'Shift',
    'weekly_off':             'Weekly Off',
    'notice_period_days':     'Notice (Days)',
    'probation_period_months':'Probation (Mo)',
    'probation_end_date':     'Probation End',
    'confirmation_date':      'Confirmation',
    'resignation_date':       'Resignation',
    'last_working_date':      'Last Working',
    # ── Phase-1: KYC identifiers ─────────────────────────────
    'aadhar_number':     'Aadhaar',
    'pan_number':        'PAN',
    'uan_number':        'UAN',
    'esic_number':       'ESIC No',
    # ── Phase-1: PF ──────────────────────────────────────────
    'pf_applicable':        'PF Applicable',
    'pf_number':            'PF Number',
    'eps_applicable':       'EPS Applicable',
    'previous_pf_transfer': 'Prev PF Transfer',
    # ── Phase-1: ESIC ────────────────────────────────────────
    'esic_applicable':       'ESIC Applicable',
    'esic_nominee_name':     'ESIC Nominee',
    'esic_dispensary':       'Dispensary',
    # ── Phase-1: TDS / Tax ───────────────────────────────────
    'aadhaar_pan_linked':      'Aadhaar-PAN Linked',
    'tax_regime':              'Tax Regime',
    'monthly_tds':             'Monthly TDS',
    'proof_submission_status': 'Proof Status',
    # ── Phase-1: Statutory flags ─────────────────────────────
    'professional_tax_applicable': 'PT Applicable',
    'labour_welfare_fund':         'LWF',
    'gratuity_eligible':           'Gratuity',
    'bonus_eligible':              'Bonus Eligible',
    # ── Phase-1: Attendance / Leave ──────────────────────────
    'attendance_code':      'Att. Code',
    'overtime_eligible':    'OT Eligible',
    'casual_leave_balance': 'CL Balance',
    'sick_leave_balance':   'SL Balance',
    'paid_leave_balance':   'PL Balance',
    'leave_policy':         'Leave Policy',
    # ── Phase-1: Salary extras ───────────────────────────────
    'salary_ctc':       'CTC',
    'salary_gross':     'Gross',
    'salary_net':       'Net',
    'salary_basic':     'Basic',
    'salary_hra':       'HRA',
    'salary_conveyance':'Conveyance',
    'salary_bonus':     'Bonus',
    'salary_incentive': 'Incentive',
    'salary_mode':      'Salary Mode',
    # ── Phase-1: System Access ───────────────────────────────
    'official_email':  'Official Email',
    'role_access':     'Role Access',
    # ── Phase-1: Exit extras ─────────────────────────────────
    'exit_interview_done':  'Exit Interview',
    'ff_settlement_status': 'F&F Status',
    'ff_settlement_amount': 'F&F Amount',
    'ff_settlement_date':   'F&F Date',
    # ── Previously hidden but common ─────────────────────────
    'date_of_birth':  'Date of Birth',
    'gender':         'Gender',
    'city':           'City',
    'state':          'State',
    'nationality':    'Nationality',
    'reports_to':     'Reports To',
}

CTR_COLS_DEFAULT = ['contract_id','company_name','supply','contact_person','contact_no','email_address','pancard','gstno','status']
CTR_COLS_ALL = {
    'contract_id':    'Contract ID',
    'company_name':   'Company',
    'supply':         'Supply',
    'contact_person': 'Contact Person',
    'contact_no':     'Mobile',
    'email_address':  'Email',
    'pancard':        'PAN',
    'gstno':          'GST No',
    'remarks':        'Remarks',
    'status':         'Status',
}


# ══════════════════════════════════════
# EMPLOYEE
# ══════════════════════════════════════

@hr.route('/employees')
@login_required
def employees():
    perm = get_perm('hr_employees')
    if not perm or not perm.can_view:
        flash('Access denied.', 'error'); return redirect(url_for('dashboard'))

    search     = request.args.get('search', '')
    dept       = request.args.get('dept', '')
    status     = request.args.get('status', '')
    emptype    = request.args.get('emptype', '')
    em_status_f= request.args.get('status_f', '')
    show_trash = request.args.get('trash', '') == '1'

    # Trash view: sirf deleted, Normal view: sirf non-deleted
    q = Employee.query.filter_by(is_deleted=True) if show_trash         else Employee.query.filter_by(is_deleted=False)

    if not show_trash:
        if search:
            q = q.filter(
                Employee.first_name.ilike(f'%{search}%') |
                Employee.last_name.ilike(f'%{search}%') |
                Employee.employee_code.ilike(f'%{search}%') |
                Employee.mobile.ilike(f'%{search}%') |
                Employee.email.ilike(f'%{search}%')
            )
        if dept:    q = q.filter_by(department=dept)
        if status:  q = q.filter_by(status=status)
        if emptype: q = q.filter_by(employee_type=emptype)

    sort_by  = request.args.get('sort_by', 'created_at')
    sort_dir = request.args.get('sort_dir', 'desc')
    sort_col = getattr(Employee, sort_by, Employee.created_at)
    emps = q.order_by(sort_col.asc() if sort_dir == 'asc' else sort_col.desc()).all()

    all_depts     = [r[0] for r in db.session.query(Employee.department).distinct().all() if r[0]]
    grid_cols     = get_grid_columns('employees', EMP_COLS_DEFAULT, list(EMP_COLS_ALL.keys()))
    deleted_count = Employee.query.filter_by(is_deleted=True).count()

    return render_template('hr/employees/index.html',
        employees=emps, perm=perm, active_page='hr_employees',
        search=search, dept=dept, status=status, emptype=emptype,
        em_status_f=em_status_f,
        sort_by=sort_by, sort_dir=sort_dir,
        show_trash=show_trash, deleted_count=deleted_count,
        all_depts=all_depts, grid_cols=grid_cols, all_cols=EMP_COLS_ALL)


@hr.route('/employees/dashboard')
@login_required
def emp_dashboard():
    from sqlalchemy import func, extract
    from datetime import date, timedelta
    perm = get_perm('hr_employees')
    if not perm or not perm.can_view:
        flash('Access denied.', 'error'); return redirect(url_for('dashboard'))

    today = date.today()
    this_month_start = today.replace(day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)

    all_emps = Employee.query.filter(Employee.status != 'terminated').all()
    total       = len(all_emps)
    active      = sum(1 for e in all_emps if e.status == 'active')
    inactive    = sum(1 for e in all_emps if e.status == 'inactive')
    on_leave    = sum(1 for e in all_emps if e.status == 'on_leave')
    terminated  = Employee.query.filter_by(status='terminated').count()
    probation   = sum(1 for e in all_emps if e.is_probation)
    on_block    = sum(1 for e in all_emps if e.is_block)
    contractors = sum(1 for e in all_emps if e.is_contractor)

    # New joinings this month
    new_this_month = sum(1 for e in all_emps if e.date_of_joining and e.date_of_joining >= this_month_start)
    new_last_month = sum(1 for e in all_emps if e.date_of_joining and last_month_start <= e.date_of_joining < this_month_start)

    # By department
    dept_counts = {}
    for e in all_emps:
        d = e.department or 'Unassigned'
        dept_counts[d] = dept_counts.get(d, 0) + 1
    dept_data = sorted(dept_counts.items(), key=lambda x: -x[1])

    # By employee type
    type_counts = {}
    for e in all_emps:
        t = e.employee_type or 'Unknown'
        type_counts[t] = type_counts.get(t, 0) + 1

    # By gender
    gender_counts = {}
    for e in all_emps:
        g = e.gender or 'Not Specified'
        gender_counts[g] = gender_counts.get(g, 0) + 1

    # Monthly joining trend (last 6 months)
    months_trend = []
    for i in range(5, -1, -1):
        m_start = (today.replace(day=1) - timedelta(days=i*28)).replace(day=1)
        if i > 0:
            m_end = (today.replace(day=1) - timedelta(days=(i-1)*28)).replace(day=1)
        else:
            m_end = today
        cnt = sum(1 for e in all_emps if e.date_of_joining and m_start <= e.date_of_joining < m_end)
        months_trend.append({'label': m_start.strftime('%b %Y'), 'count': cnt})

    # Salary stats
    sal_emps = [e for e in all_emps if e.salary_ctc]
    total_ctc = sum(float(e.salary_ctc) for e in sal_emps)
    avg_ctc   = total_ctc / len(sal_emps) if sal_emps else 0

    # Upcoming birthdays (next 30 days)
    bday_emps = []
    for e in all_emps:
        if e.date_of_birth:
            try:
                bday_this_year = e.date_of_birth.replace(year=today.year)
                if 0 <= (bday_this_year - today).days <= 30:
                    bday_emps.append({'emp': e, 'bday': bday_this_year, 'days': (bday_this_year - today).days})
            except ValueError:
                pass
    bday_emps.sort(key=lambda x: x['days'])

    # Work anniversaries (joined exactly N years ago this month)
    anniv_emps = []
    for e in all_emps:
        if e.date_of_joining:
            years = today.year - e.date_of_joining.year
            if years > 0:
                try:
                    anniv = e.date_of_joining.replace(year=today.year)
                    if 0 <= (anniv - today).days <= 30:
                        anniv_emps.append({'emp': e, 'years': years, 'days': (anniv - today).days})
                except ValueError:
                    pass
    anniv_emps.sort(key=lambda x: x['days'])

    # KYC completeness
    kyc_complete   = sum(1 for e in all_emps if e.aadhar_number and e.pan_number)
    bank_complete  = sum(1 for e in all_emps if e.bank_account_number and e.bank_ifsc)
    photo_complete = sum(1 for e in all_emps if e.profile_photo)
    sal_filled     = len(sal_emps)

    # Recent joiners (last 5)
    recent = sorted([e for e in all_emps if e.date_of_joining], key=lambda e: e.date_of_joining, reverse=True)[:5]

    # Expiring documents (passport/DL in next 60 days)
    expiring_docs = []
    for e in all_emps:
        for doc_name, expiry in [('Passport', e.passport_expiry), ('Driving License', e.dl_expiry)]:
            if expiry and 0 <= (expiry - today).days <= 60:
                expiring_docs.append({'emp': e, 'doc': doc_name, 'expiry': expiry, 'days': (expiry - today).days})
    expiring_docs.sort(key=lambda x: x['days'])

    import json
    return render_template('hr/employees/dashboard.html',
        perm=perm, active_page='hr_emp_dashboard',
        total=total, active=active, inactive=inactive, on_leave=on_leave,
        terminated=terminated, probation=probation, on_block=on_block, contractors=contractors,
        new_this_month=new_this_month, new_last_month=new_last_month,
        dept_data=dept_data[:10],
        type_counts=type_counts,
        gender_counts=gender_counts,
        months_trend=months_trend,
        total_ctc=round(total_ctc/12),  # monthly payroll
        avg_ctc=round(avg_ctc),
        sal_filled=sal_filled,
        bday_emps=bday_emps[:8],
        anniv_emps=anniv_emps[:5],
        kyc_complete=kyc_complete, bank_complete=bank_complete,
        photo_complete=photo_complete,
        recent=recent,
        expiring_docs=expiring_docs[:5],
        dept_json=json.dumps(dict(dept_data[:8])),
        type_json=json.dumps(type_counts),
        gender_json=json.dumps(gender_counts),
        trend_json=json.dumps(months_trend),
    )


@hr.route('/api/celebrations')
@login_required
def api_celebrations():
    """Return today/upcoming birthdays, work anniversaries, marriage anniversaries."""
    from datetime import date
    try:
        today = date.today()
        all_emps = Employee.query.filter(Employee.status == 'active').all()

        my_emp = Employee.query.filter_by(user_id=current_user.id).first()
        my_emp_id = my_emp.id if my_emp else None

        # Load wishes safely (wish_logs table may not exist yet)
        wished_today = set()
        try:
            for w in WishLog.query.filter_by(sender_id=current_user.id, wish_date=today).all():
                wished_today.add(f"{w.target_emp_id}_{w.wish_type}")
        except Exception:
            pass

        celebrations = []

        def _ord(n):
            return str(n) + ('st' if n==1 else 'nd' if n==2 else 'rd' if n==3 else 'th')

        for e in all_emps:
            emp_data = {
                'id': e.id,
                'name': e.full_name,
                'first_name': e.first_name or e.full_name,
                'code': e.employee_code or '',
                'dept': e.department or '',
                'designation': e.designation or '',
                'photo': e.profile_photo or '',
                'view_url': f'/hr/employees/{e.id}/view',
                'is_self': e.id == my_emp_id,
            }

            # ── Birthday ──
            if e.date_of_birth:
                try:
                    bday = e.date_of_birth.replace(year=today.year)
                    days = (bday - today).days
                    if -1 <= days <= 7:
                        age = today.year - e.date_of_birth.year
                        fn = e.first_name or e.full_name
                        self_wish = f"🎂 Aaj aapka birthday hai, {fn}ji! Team ki taraf se aapko bahut bahut Happy Birthday! 🎉🥳 Aapki zindagi mein khushiyan aur safalta barhti rahe. Have a wonderful day! 🌟"
                        other_wish = f"Aaj {e.full_name} ka birthday hai! 🎂 Unhe wish karo:\n\nDear {fn},\nWishing you a very Happy Birthday! 🎂🎉 May this year bring you lots of joy, health and success. Team wishes you all the best! 🥳"
                        key = f"{e.id}_birthday"
                        celebrations.append({**emp_data,
                            'type': 'birthday',
                            'type_label': '🎂 Birthday',
                            'days': days,
                            'extra': f'Turning {age}' if days >= 0 else f'Was yesterday ({age} yrs)',
                            'today': days == 0,
                            'self_wish': self_wish,
                            'wish_template': other_wish,
                            'already_wished': key in wished_today,
                        })
                except ValueError:
                    pass

            # ── Work Anniversary ──
            if e.date_of_joining:
                try:
                    anniv = e.date_of_joining.replace(year=today.year)
                    days = (anniv - today).days
                    if -1 <= days <= 7:
                        years = today.year - e.date_of_joining.year
                        if years > 0:
                            fn = e.first_name or e.full_name
                            self_wish = f"🎖️ Aaj aapki {_ord(years)} Work Anniversary hai, {fn}ji! Hamare saath {years} saal poore karne par bahut bahut badhai! Aapka yogdan hamare liye anmol hai. 💪"
                            other_wish = f"Aaj {e.full_name} ki {_ord(years)} Work Anniversary hai! 🎖️ Unhe congratulate karo:\n\nDear {fn},\nHappy {_ord(years)} Work Anniversary! 🎉 Thank you for your {years} wonderful year{'s' if years!=1 else ''} of dedication. We're proud to have you! 💪"
                            key = f"{e.id}_work_anniversary"
                            celebrations.append({**emp_data,
                                'type': 'work_anniversary',
                                'type_label': '🎖️ Work Anniversary',
                                'days': days,
                                'extra': f'{years} year{"s" if years != 1 else ""} with us',
                                'today': days == 0,
                                'self_wish': self_wish,
                                'wish_template': other_wish,
                                'already_wished': key in wished_today,
                            })
                except ValueError:
                    pass

            # ── Marriage Anniversary ──
            if getattr(e, 'marital_status', None) == 'Married' and getattr(e, 'marriage_anniversary', None):
                try:
                    manniv = e.marriage_anniversary.replace(year=today.year)
                    days = (manniv - today).days
                    if -1 <= days <= 7:
                        years = today.year - e.marriage_anniversary.year
                        yrs_label = f'{_ord(years)} Wedding' if years > 0 else 'Wedding'
                        fn = e.first_name or e.full_name
                        self_wish = f"💍 Aaj aapki {yrs_label} Anniversary hai, {fn}ji! Dil se mubarak ho! ❤️ Aapki zindagi mein pyaar aur khushiyaan hamesha bani rahe. Team ki taraf se dher saari shubhkamnayein! 🌹"
                        other_wish = f"Aaj {e.full_name} ki {yrs_label} Anniversary hai! 💍 Unhe wish karo:\n\nDear {fn},\nHappy {yrs_label} Anniversary! 💍❤️ Wishing you and your family a lifetime of love and happiness! 🌹"
                        key = f"{e.id}_marriage_anniversary"
                        celebrations.append({**emp_data,
                            'type': 'marriage_anniversary',
                            'type_label': '💍 Marriage Anniversary',
                            'days': days,
                            'extra': f'{years} year{"s" if years != 1 else ""} of togetherness' if years > 0 else 'First Anniversary!',
                            'today': days == 0,
                            'self_wish': self_wish,
                            'wish_template': other_wish,
                            'already_wished': key in wished_today,
                        })
                except ValueError:
                    pass

        celebrations.sort(key=lambda x: (0 if x['today'] else 1, x['days']))
        return jsonify(celebrations=celebrations, count=len(celebrations),
                       today_count=sum(1 for c in celebrations if c['today']))
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(celebrations=[], count=0, today_count=0, error=str(e))


@hr.route('/api/send-wish', methods=['POST'])
@login_required
def api_send_wish():
    """Log that current user sent a wish. Prevents duplicate notifications."""
    from datetime import date
    data = request.get_json() or {}
    target_emp_id = data.get('emp_id')
    wish_type     = data.get('wish_type')
    wish_text     = data.get('wish_text', '')
    today         = date.today()

    if not target_emp_id or not wish_type:
        return jsonify(success=False, error='Missing params')

    if wish_type not in ('birthday','work_anniversary','marriage_anniversary'):
        return jsonify(success=False, error='Invalid type')

    # Check already wished
    existing = WishLog.query.filter_by(
        sender_id=current_user.id,
        target_emp_id=target_emp_id,
        wish_type=wish_type,
        wish_date=today
    ).first()

    if existing:
        return jsonify(success=True, already=True, msg='Already wished today!')

    try:
        w = WishLog(
            sender_id=current_user.id,
            target_emp_id=target_emp_id,
            wish_type=wish_type,
            wish_date=today,
            wish_text=wish_text[:500] if wish_text else ''
        )
        db.session.add(w)
        db.session.commit()
        return jsonify(success=True, already=False, msg='Wish logged!')
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, error=str(ex))


@hr.route('/employees/grid-config', methods=['POST'])
@login_required
def emp_grid_config():
    cols = request.json.get('cols', [])
    save_grid_columns('employees', cols)
    return jsonify(success=True)


@hr.route('/contractors/grid-config', methods=['POST'])
@login_required
def contractor_grid_config():
    cols = request.json.get('cols', [])
    save_grid_columns('contractors', cols)
    return jsonify(success=True)


@hr.route('/employees/save-qr', methods=['POST'])
@login_required
def emp_save_qr():
    """Save QR code generated by browser JS"""
    data   = request.json
    emp_id = data.get('emp_id')
    qr_b64 = data.get('qr_base64', '').strip()
    if not emp_id or not qr_b64:
        return jsonify(success=False, error='Missing data')
    e = Employee.query.get_or_404(emp_id)
    e.qr_code_base64 = qr_b64
    db.session.commit()
    return jsonify(success=True)


@hr.route('/employees/add', methods=['GET', 'POST'])
@login_required
def emp_add():
    perm = get_perm('hr_employees')
    if not perm or not perm.can_add:
        flash('Access denied.', 'error'); return redirect(url_for('hr.employees'))

    contractors = Contractor.query.filter_by(status=1, is_deleted=0).order_by(Contractor.company_name).all()

    if request.method == 'POST':
        emp_code = request.form.get('employee_code', '').strip()
        if not emp_code:
            flash('Employee Code is required.', 'error')
            return redirect(url_for('hr.emp_add'))

        if Employee.query.filter(Employee.employee_code.ilike(emp_code)).first():
            flash(f'Employee code "{emp_code}" already exists.', 'error')
            return redirect(url_for('hr.emp_add'))

        photo   = request.form.get('photo_base64', '').strip() or None
        qr_b64  = request.form.get('qr_base64', '').strip() or None

        rto = request.form.get('reports_to', '').strip()

        def _dec_add(v):
            try: return float(v) if v else None
            except: return None

        e = Employee(
            employee_code   = emp_code,
            employee_id     = request.form.get('employee_id', '').strip() or None,
            first_name      = request.form.get('first_name', '').strip(),
            middle_name     = request.form.get('middle_name', '').strip(),
            last_name       = request.form.get('last_name', '').strip(),
            mobile          = request.form.get('mobile', '').strip(),
            email           = request.form.get('email', '').strip(),
            gender          = request.form.get('gender', ''),
            profile_photo   = photo,
            qr_code_base64  = qr_b64,
            linkedin        = request.form.get('linkedin', '').strip(),
            facebook        = request.form.get('facebook', '').strip(),
            department      = request.form.get('department', '').strip(),
            designation     = request.form.get('designation', '').strip(),
            employee_type   = request.form.get('employee_type', ''),
            date_of_joining = _parse_date(request.form.get('date_of_joining')),
            location        = request.form.get('location', '').strip(),
            is_contractor   = request.form.get('is_contractor') == 'yes',
            contractor_id   = int(request.form['contractor_id']) if request.form.get('contractor_id') and request.form.get('is_contractor') == 'yes' else None,
            date_of_birth   = _parse_date(request.form.get('date_of_birth')),
            blood_group     = request.form.get('blood_group', '').strip(),
            marital_status  = request.form.get('marital_status', ''),
            is_block        = request.form.get('is_block') == 'yes',
            is_late         = request.form.get('is_late') == 'yes',
            is_probation    = request.form.get('is_probation') == 'yes',
            status          = request.form.get('status', 'active'),
            remark          = request.form.get('remark', '').strip(),
            reports_to      = int(rto) if rto else None,
            # KYC
            nationality     = request.form.get('nationality','Indian').strip(),
            aadhar_number   = request.form.get('aadhar_number','').strip(),
            pan_number      = request.form.get('pan_number','').strip().upper(),
            uan_number      = request.form.get('uan_number','').strip(),
            esic_number     = request.form.get('esic_number','').strip(),
            emergency_name  = request.form.get('emergency_name','').strip(),
            emergency_relation = request.form.get('emergency_relation','').strip(),
            emergency_phone = request.form.get('emergency_phone','').strip(),
            # Bank
            bank_name       = request.form.get('bank_name','').strip(),
            bank_account_number = request.form.get('bank_account_number','').strip(),
            bank_ifsc       = request.form.get('bank_ifsc','').strip().upper(),
            bank_account_type = request.form.get('bank_account_type','').strip(),
            bank_account_holder = request.form.get('bank_account_holder','').strip(),
            # Salary
            salary_ctc      = float(request.form.get('salary_ctc') or 0) or None,
            salary_net      = float(request.form.get('salary_net') or 0) or None,
            salary_mode     = request.form.get('salary_mode','').strip(),
            # Work
            pay_grade       = request.form.get('pay_grade','').strip(),
            shift           = request.form.get('shift','').strip(),
            weekly_off      = request.form.get('weekly_off','').strip(),
            # Education
            highest_qualification = request.form.get('highest_qualification','').strip(),
            # Documents
            documents_json  = request.form.get('documents_json','[]'),
            marriage_anniversary = _parse_date(request.form.get('marriage_anniversary')) if request.form.get('marital_status')=='Married' else None,
            # ── Phase-1 additions: Family / Contact ─────────────────
            father_name       = request.form.get('father_name','').strip() or None,
            mother_name       = request.form.get('mother_name','').strip() or None,
            alternate_mobile  = request.form.get('alternate_mobile','').strip() or None,
            personal_email    = request.form.get('personal_email','').strip() or None,
            # ── Phase-1: Permanent Address ──────────────────────────
            permanent_address    = request.form.get('permanent_address','').strip() or None,
            permanent_city       = request.form.get('permanent_city','').strip() or None,
            permanent_state      = request.form.get('permanent_state','').strip() or None,
            permanent_country    = request.form.get('permanent_country','India').strip() or None,
            permanent_zip        = request.form.get('permanent_zip','').strip() or None,
            same_as_current_addr = request.form.get('same_as_current_addr') == 'yes',
            # ── Phase-1: Grade / Probation ──────────────────────────
            grade_level             = request.form.get('grade_level','').strip() or None,
            probation_period_months = int(request.form.get('probation_period_months') or 6),
            probation_end_date      = _parse_date(request.form.get('probation_end_date')),
            # ── Phase-1: Salary extras ──────────────────────────────
            salary_conveyance = _dec_add(request.form.get('salary_conveyance')),
            salary_bonus      = _dec_add(request.form.get('salary_bonus')),
            salary_incentive  = _dec_add(request.form.get('salary_incentive')),
            salary_gross      = _dec_add(request.form.get('salary_gross')),
            # ── Phase-1: PF ─────────────────────────────────────────
            pf_applicable        = request.form.get('pf_applicable') == 'yes',
            pf_number            = request.form.get('pf_number','').strip() or None,
            eps_applicable       = request.form.get('eps_applicable') == 'yes',
            previous_pf_transfer = request.form.get('previous_pf_transfer') == 'yes',
            previous_pf_number   = request.form.get('previous_pf_number','').strip() or None,
            # ── Phase-1: ESIC ───────────────────────────────────────
            esic_applicable       = request.form.get('esic_applicable') == 'yes',
            esic_nominee_name     = request.form.get('esic_nominee_name','').strip() or None,
            esic_nominee_relation = request.form.get('esic_nominee_relation','').strip() or None,
            esic_family_details   = request.form.get('esic_family_details','').strip() or None,
            esic_dispensary       = request.form.get('esic_dispensary','').strip() or None,
            # ── Phase-1: TDS / Tax ──────────────────────────────────
            aadhaar_pan_linked      = request.form.get('aadhaar_pan_linked') == 'yes',
            tax_regime              = request.form.get('tax_regime','New').strip() or 'New',
            prev_employer_income    = _dec_add(request.form.get('prev_employer_income')),
            monthly_tds             = _dec_add(request.form.get('monthly_tds')),
            investment_declaration  = request.form.get('investment_declaration','').strip() or None,
            proof_submission_status = request.form.get('proof_submission_status','Pending').strip() or 'Pending',
            # ── Phase-1: Statutory ──────────────────────────────────
            professional_tax_applicable = request.form.get('professional_tax_applicable','yes') == 'yes',
            labour_welfare_fund         = request.form.get('labour_welfare_fund') == 'yes',
            gratuity_eligible           = request.form.get('gratuity_eligible') == 'yes',
            bonus_eligible              = request.form.get('bonus_eligible','yes') == 'yes',
            # ── Phase-1: Attendance / Leave ─────────────────────────
            attendance_code       = request.form.get('attendance_code','').strip() or None,
            overtime_eligible     = request.form.get('overtime_eligible') == 'yes',
            casual_leave_balance  = _dec_add(request.form.get('casual_leave_balance')) or 0,
            sick_leave_balance    = _dec_add(request.form.get('sick_leave_balance')) or 0,
            paid_leave_balance    = _dec_add(request.form.get('paid_leave_balance')) or 0,
            leave_policy          = request.form.get('leave_policy','').strip() or None,
            # ── Phase-1: System Access ──────────────────────────────
            official_email = request.form.get('official_email','').strip() or None,
            role_access    = request.form.get('role_access','').strip() or None,
            # ── Phase-1: Exit extras ────────────────────────────────
            exit_interview_done  = request.form.get('exit_interview_done') == 'yes',
            exit_interview_notes = request.form.get('exit_interview_notes','').strip() or None,
            ff_settlement_status = request.form.get('ff_settlement_status','Pending').strip() or 'Pending',
            ff_settlement_amount = _dec_add(request.form.get('ff_settlement_amount')),
            ff_settlement_date   = _parse_date(request.form.get('ff_settlement_date')),
            created_by      = current_user.id,
        )
        db.session.add(e)
        db.session.flush()

        # Auto-create login user for every employee
        uname = emp_code.lower().replace('-', '').replace(' ', '')
        if not User.query.filter_by(username=uname).first():
            u = User(
                username  = uname,
                email     = e.email or f'{uname}@hcp.com',
                full_name = e.full_name,
                role      = 'user',
                is_active = True
            )
            u.set_password('HCP@123')
            db.session.add(u)
            db.session.flush()
            e.user_id = u.id

        db.session.commit()
        flash(f'Employee {emp_code} added! Login: {uname} / HCP@123', 'success')
        return redirect(url_for('hr.employees'))

    all_employees = Employee.query.filter_by(status='active').order_by(Employee.first_name).all()
    emp_types     = EmployeeTypeMaster.query.order_by(EmployeeTypeMaster.name).all()
    departments   = DepartmentMaster.query.order_by(DepartmentMaster.name).all()
    designations  = DesignationMaster.query.order_by(DesignationMaster.name).all()
    locations     = EmployeeLocationMaster.query.order_by(EmployeeLocationMaster.name).all()
    return render_template('hr/employees/form.html',
        employee=None, contractors=contractors, perm=perm, active_page='hr_employees',
        all_employees=all_employees, emp_types=emp_types,
        departments=departments, designations=designations, locations=locations)


@hr.route('/employees/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def emp_edit(id):
    perm = get_perm('hr_employees')
    if not perm or not perm.can_edit:
        flash('Access denied.', 'error'); return redirect(url_for('hr.employees'))

    e = Employee.query.get_or_404(id)
    contractors = Contractor.query.filter_by(status=1, is_deleted=0).order_by(Contractor.company_name).all()

    if request.method == 'POST':
        photo  = request.form.get('photo_base64', '').strip()
        if photo: e.profile_photo = photo

        qr_b64 = request.form.get('qr_base64', '').strip()
        if qr_b64: e.qr_code_base64 = qr_b64

        # Employee code: editable only if currently blank
        new_code = request.form.get('employee_code', '').strip()
        if not e.employee_code and new_code:
            if Employee.query.filter(Employee.employee_code.ilike(new_code), Employee.id != e.id).first():
                flash(f'Code "{new_code}" already in use.', 'error')
                return redirect(url_for('hr.emp_edit', id=id))
            e.employee_code = new_code

        e.employee_id    = request.form.get('employee_id', '').strip() or None
        e.first_name     = request.form.get('first_name', e.first_name).strip()
        e.middle_name    = request.form.get('middle_name', e.middle_name or '').strip()
        e.last_name      = request.form.get('last_name', e.last_name).strip()
        e.mobile         = request.form.get('mobile', e.mobile).strip()
        e.email          = request.form.get('email', '').strip()
        e.gender         = request.form.get('gender', '')
        e.linkedin       = request.form.get('linkedin', '').strip()
        e.facebook       = request.form.get('facebook', '').strip()
        e.department     = request.form.get('department', '').strip()
        e.designation    = request.form.get('designation', '').strip()
        e.employee_type  = request.form.get('employee_type', '')
        e.location       = request.form.get('location', '').strip()
        e.is_contractor  = request.form.get('is_contractor') == 'yes'
        cid              = request.form.get('contractor_id') or None
        e.contractor_id  = int(cid) if cid and e.is_contractor else None
        e.date_of_joining= _parse_date(request.form.get('date_of_joining'))
        e.date_of_birth  = _parse_date(request.form.get('date_of_birth'))
        e.blood_group    = request.form.get('blood_group', '').strip()
        e.marital_status = request.form.get('marital_status', '')
        e.is_block       = request.form.get('is_block') == 'yes'
        e.is_late        = request.form.get('is_late') == 'yes'
        e.is_probation   = request.form.get('is_probation') == 'yes'
        e.status         = request.form.get('status', 'active')
        e.remark         = request.form.get('remark', '').strip()
        rto = request.form.get('reports_to', '').strip()
        try:
            e.reports_to = int(rto) if rto else None
        except (ValueError, TypeError):
            e.reports_to = None

        # Professional
        e.pay_grade      = request.form.get('pay_grade','').strip()
        e.shift          = request.form.get('shift','').strip()
        e.weekly_off     = request.form.get('weekly_off','').strip()
        e.notice_period_days = int(request.form.get('notice_period_days') or 30)
        e.work_hours_per_day = float(request.form.get('work_hours_per_day') or 8)
        e.rehire_eligible = request.form.get('rehire_eligible') == 'yes'
        e.confirmation_date  = _parse_date(request.form.get('confirmation_date'))
        e.resignation_date   = _parse_date(request.form.get('resignation_date'))
        e.last_working_date  = _parse_date(request.form.get('last_working_date'))

        # KYC
        e.nationality        = request.form.get('nationality','Indian').strip()
        e.religion           = request.form.get('religion','').strip()
        e.caste              = request.form.get('caste','').strip()
        e.physically_handicapped = request.form.get('physically_handicapped') == 'yes'
        ma_raw = request.form.get('marriage_anniversary','').strip()
        if e.marital_status == 'Married' and ma_raw:
            e.marriage_anniversary = _parse_date(ma_raw)
        elif e.marital_status != 'Married':
            e.marriage_anniversary = None
        e.aadhar_number      = request.form.get('aadhar_number','').strip()
        e.pan_number         = request.form.get('pan_number','').strip().upper()
        e.uan_number         = request.form.get('uan_number','').strip()
        e.esic_number        = request.form.get('esic_number','').strip()
        e.passport_number    = request.form.get('passport_number','').strip()
        e.passport_expiry    = _parse_date(request.form.get('passport_expiry'))
        e.driving_license    = request.form.get('driving_license','').strip()
        e.dl_expiry          = _parse_date(request.form.get('dl_expiry'))

        # Emergency
        e.emergency_name     = request.form.get('emergency_name','').strip()
        e.emergency_relation = request.form.get('emergency_relation','').strip()
        e.emergency_phone    = request.form.get('emergency_phone','').strip()
        e.emergency_address  = request.form.get('emergency_address','').strip()

        # Bank
        e.bank_account_holder= request.form.get('bank_account_holder','').strip()
        e.bank_name          = request.form.get('bank_name','').strip()
        e.bank_account_number= request.form.get('bank_account_number','').strip()
        e.bank_ifsc          = request.form.get('bank_ifsc','').strip().upper()
        e.bank_branch        = request.form.get('bank_branch','').strip()
        e.bank_account_type  = request.form.get('bank_account_type','').strip()

        # Salary
        def _dec(v): 
            try: return float(v) if v else None
            except: return None
        e.salary_ctc           = _dec(request.form.get('salary_ctc'))
        e.salary_basic         = _dec(request.form.get('salary_basic'))
        e.salary_hra           = _dec(request.form.get('salary_hra'))
        e.salary_da            = _dec(request.form.get('salary_da'))
        e.salary_ta            = _dec(request.form.get('salary_ta'))
        e.salary_medical_allow = _dec(request.form.get('salary_medical_allow'))
        e.salary_special_allow = _dec(request.form.get('salary_special_allow'))
        e.salary_pf_employee   = _dec(request.form.get('salary_pf_employee'))
        e.salary_pf_employer   = _dec(request.form.get('salary_pf_employer'))
        e.salary_esic_employee = _dec(request.form.get('salary_esic_employee'))
        e.salary_esic_employer = _dec(request.form.get('salary_esic_employer'))
        e.salary_professional_tax = _dec(request.form.get('salary_professional_tax'))
        e.salary_tds           = _dec(request.form.get('salary_tds'))
        e.salary_net           = _dec(request.form.get('salary_net'))
        e.salary_mode          = request.form.get('salary_mode','').strip()
        e.salary_effective_date= _parse_date(request.form.get('salary_effective_date'))
        e.pay_grade            = request.form.get('pay_grade','').strip()

        # Education
        e.highest_qualification= request.form.get('highest_qualification','').strip()
        e.university           = request.form.get('university','').strip()
        e.passing_year         = int(request.form.get('passing_year') or 0) or None
        e.specialization       = request.form.get('specialization','').strip()
        e.prev_company         = request.form.get('prev_company','').strip()
        e.prev_designation     = request.form.get('prev_designation','').strip()
        e.total_experience_yrs = _dec(request.form.get('total_experience_yrs'))
        e.prev_from_date       = _parse_date(request.form.get('prev_from_date'))
        e.prev_to_date         = _parse_date(request.form.get('prev_to_date'))
        e.prev_leaving_reason  = request.form.get('prev_leaving_reason','').strip()

        # Documents
        e.documents_json       = request.form.get('documents_json','[]')

        # ── Phase-1 additions: Family / Contact ─────────────────
        e.father_name       = request.form.get('father_name','').strip() or None
        e.mother_name       = request.form.get('mother_name','').strip() or None
        e.alternate_mobile  = request.form.get('alternate_mobile','').strip() or None
        e.personal_email    = request.form.get('personal_email','').strip() or None
        # ── Phase-1: Permanent Address ──────────────────────────
        e.permanent_address    = request.form.get('permanent_address','').strip() or None
        e.permanent_city       = request.form.get('permanent_city','').strip() or None
        e.permanent_state      = request.form.get('permanent_state','').strip() or None
        e.permanent_country    = request.form.get('permanent_country','India').strip() or None
        e.permanent_zip        = request.form.get('permanent_zip','').strip() or None
        e.same_as_current_addr = request.form.get('same_as_current_addr') == 'yes'
        # ── Phase-1: Grade / Probation ──────────────────────────
        e.grade_level             = request.form.get('grade_level','').strip() or None
        try:
            e.probation_period_months = int(request.form.get('probation_period_months') or 6)
        except (ValueError, TypeError):
            e.probation_period_months = 6
        e.probation_end_date      = _parse_date(request.form.get('probation_end_date'))
        # ── Phase-1: Salary extras ──────────────────────────────
        e.salary_conveyance = _dec(request.form.get('salary_conveyance'))
        e.salary_bonus      = _dec(request.form.get('salary_bonus'))
        e.salary_incentive  = _dec(request.form.get('salary_incentive'))
        e.salary_gross      = _dec(request.form.get('salary_gross'))
        # ── Phase-1: PF ─────────────────────────────────────────
        e.pf_applicable        = request.form.get('pf_applicable') == 'yes'
        e.pf_number            = request.form.get('pf_number','').strip() or None
        e.eps_applicable       = request.form.get('eps_applicable') == 'yes'
        e.previous_pf_transfer = request.form.get('previous_pf_transfer') == 'yes'
        e.previous_pf_number   = request.form.get('previous_pf_number','').strip() or None
        # ── Phase-1: ESIC ───────────────────────────────────────
        e.esic_applicable       = request.form.get('esic_applicable') == 'yes'
        e.esic_nominee_name     = request.form.get('esic_nominee_name','').strip() or None
        e.esic_nominee_relation = request.form.get('esic_nominee_relation','').strip() or None
        e.esic_family_details   = request.form.get('esic_family_details','').strip() or None
        e.esic_dispensary       = request.form.get('esic_dispensary','').strip() or None
        # ── Phase-1: TDS / Tax ──────────────────────────────────
        e.aadhaar_pan_linked      = request.form.get('aadhaar_pan_linked') == 'yes'
        e.tax_regime              = request.form.get('tax_regime','New').strip() or 'New'
        e.prev_employer_income    = _dec(request.form.get('prev_employer_income'))
        e.monthly_tds             = _dec(request.form.get('monthly_tds'))
        e.investment_declaration  = request.form.get('investment_declaration','').strip() or None
        e.proof_submission_status = request.form.get('proof_submission_status','Pending').strip() or 'Pending'
        # ── Phase-1: Statutory ──────────────────────────────────
        e.professional_tax_applicable = request.form.get('professional_tax_applicable','yes') == 'yes'
        e.labour_welfare_fund         = request.form.get('labour_welfare_fund') == 'yes'
        e.gratuity_eligible           = request.form.get('gratuity_eligible') == 'yes'
        e.bonus_eligible              = request.form.get('bonus_eligible','yes') == 'yes'
        # ── Phase-1: Attendance / Leave ─────────────────────────
        e.attendance_code       = request.form.get('attendance_code','').strip() or None
        e.overtime_eligible     = request.form.get('overtime_eligible') == 'yes'
        e.casual_leave_balance  = _dec(request.form.get('casual_leave_balance')) or 0
        e.sick_leave_balance    = _dec(request.form.get('sick_leave_balance')) or 0
        e.paid_leave_balance    = _dec(request.form.get('paid_leave_balance')) or 0
        e.leave_policy          = request.form.get('leave_policy','').strip() or None
        # ── Phase-1: System Access ──────────────────────────────
        e.official_email = request.form.get('official_email','').strip() or None
        e.role_access    = request.form.get('role_access','').strip() or None
        # ── Phase-1: Exit extras ────────────────────────────────
        e.exit_interview_done  = request.form.get('exit_interview_done') == 'yes'
        e.exit_interview_notes = request.form.get('exit_interview_notes','').strip() or None
        e.ff_settlement_status = request.form.get('ff_settlement_status','Pending').strip() or 'Pending'
        e.ff_settlement_amount = _dec(request.form.get('ff_settlement_amount'))
        e.ff_settlement_date   = _parse_date(request.form.get('ff_settlement_date'))

        e.updated_at     = datetime.utcnow()
        db.session.commit()
        flash('Employee updated!', 'success')
        return redirect(url_for('hr.employees'))

    all_employees = Employee.query.filter_by(status='active').order_by(Employee.first_name).all()
    emp_types     = EmployeeTypeMaster.query.order_by(EmployeeTypeMaster.name).all()
    departments   = DepartmentMaster.query.order_by(DepartmentMaster.name).all()
    designations  = DesignationMaster.query.order_by(DesignationMaster.name).all()
    locations     = EmployeeLocationMaster.query.order_by(EmployeeLocationMaster.name).all()
    return render_template('hr/employees/form.html',
        employee=e, contractors=contractors, perm=perm, active_page='hr_employees',
        all_employees=all_employees, emp_types=emp_types,
        departments=departments, designations=designations, locations=locations)



# ─────────────────────────────────────────────
# AJAX TAB-WISE SAVE ROUTES
# ─────────────────────────────────────────────

@hr.route('/employees/ajax-init', methods=['POST'])
@login_required
def emp_ajax_init():
    """Step 1: Create employee with basic info only. Returns emp_id."""
    perm = get_perm('hr_employees')
    if not perm or not perm.can_add:
        return {'ok': False, 'error': 'Access denied'}, 403

    from flask import request as req
    data = req.get_json(force=True) or {}

    emp_code = (data.get('employee_code') or '').strip()
    if not emp_code:
        return {'ok': False, 'error': 'Employee Code is required'}, 400

    if Employee.query.filter(Employee.employee_code.ilike(emp_code)).first():
        return {'ok': False, 'error': f'Employee code "{emp_code}" already exists'}, 400

    # Check email duplicate
    email = (data.get('email') or '').strip()
    if email and Employee.query.filter_by(email=email).first():
        return {'ok': False, 'error': f'Email "{email}" already in use'}, 400

    rto = data.get('reports_to', '')
    try:
        rto_id = int(rto) if rto and str(rto).strip() else None
    except (ValueError, TypeError):
        rto_id = None
    e = Employee(
        employee_code  = emp_code,
        employee_id    = (data.get('employee_id') or '').strip() or None,
        first_name     = (data.get('first_name') or '').strip(),
        middle_name    = (data.get('middle_name') or '').strip(),
        last_name      = (data.get('last_name') or '').strip(),
        mobile         = (data.get('mobile') or '').strip(),
        email          = email,
        gender         = data.get('gender', ''),
        profile_photo  = data.get('photo_base64') or None,
        qr_code_base64 = data.get('qr_base64') or None,
        date_of_birth  = _parse_date(data.get('date_of_birth')),
        blood_group    = (data.get('blood_group') or '').strip(),
        marital_status = data.get('marital_status', ''),
        marriage_anniversary = _parse_date(data.get('marriage_anniversary')) if data.get('marital_status') == 'Married' else None,
        address        = (data.get('address') or '').strip(),
        city           = (data.get('city') or '').strip(),
        state          = (data.get('state') or '').strip(),
        country        = (data.get('country') or '').strip(),
        zip_code       = (data.get('zip_code') or '').strip(),
        linkedin       = (data.get('linkedin') or '').strip(),
        facebook       = (data.get('facebook') or '').strip(),
        reports_to     = rto_id,
        department     = (data.get('department') or '').strip(),
        designation    = (data.get('designation') or '').strip(),
        employee_type  = data.get('employee_type', ''),
        date_of_joining      = _parse_date(data.get('date_of_joining')),
        confirmation_date    = _parse_date(data.get('confirmation_date')),
        resignation_date     = _parse_date(data.get('resignation_date')),
        last_working_date    = _parse_date(data.get('last_working_date')),
        location       = (data.get('location') or '').strip(),
        pay_grade      = (data.get('pay_grade') or '').strip(),
        shift          = (data.get('shift') or '').strip(),
        weekly_off     = (data.get('weekly_off') or '').strip(),
        status         = 'active',
        created_by     = current_user.id,
        documents_json = data.get('documents_json', '[]'),
    )
    db.session.add(e)
    db.session.flush()

    # Auto-create login
    uname = emp_code.lower().replace('-','').replace(' ','')
    if not User.query.filter_by(username=uname).first():
        email_for_user = email or f'{uname}@hcp.com'
        # Avoid duplicate email in users table
        if User.query.filter_by(email=email_for_user).first():
            email_for_user = f'{uname}@hcp.com'
        u = User(username=uname, email=email_for_user,
                 full_name=e.full_name, role='user', is_active=True)
        u.set_password('HCP@123')
        db.session.add(u)
        db.session.flush()
        e.user_id = u.id

    db.session.commit()
    return {'ok': True, 'emp_id': e.id, 'emp_code': emp_code,
            'msg': f'Employee {emp_code} created! Login: {uname} / HCP@123'}


@hr.route('/employees/<int:id>/ajax-save-tab', methods=['POST'])
@login_required
def emp_ajax_save_tab(id):
    """Save a specific tab's data for existing employee."""
    perm = get_perm('hr_employees')
    if not perm or not (perm.can_add or perm.can_edit):
        return {'ok': False, 'error': 'Access denied'}, 403

    e = Employee.query.get_or_404(id)
    data = request.get_json(force=True) or {}
    tab  = data.get('tab', '')

    def _dec(v):
        try: return float(v) if v else None
        except: return None

    if tab == 'basic':
        photo = (data.get('photo_base64') or '').strip()
        if photo: e.profile_photo = photo
        qr = (data.get('qr_base64') or '').strip()
        if qr: e.qr_code_base64 = qr
        e.employee_id    = (data.get('employee_id') or '').strip() or None
        e.first_name     = (data.get('first_name') or '').strip()
        e.middle_name    = (data.get('middle_name') or '').strip()
        e.last_name      = (data.get('last_name') or '').strip()
        e.mobile         = (data.get('mobile') or '').strip()
        new_email        = (data.get('email') or '').strip()
        if new_email and new_email != e.email:
            dup = Employee.query.filter(Employee.email == new_email, Employee.id != id).first()
            if dup:
                return {'ok': False, 'error': f'Email "{new_email}" already in use by {dup.employee_code}'}, 400
        e.email          = new_email
        e.gender         = data.get('gender', '')
        e.date_of_birth  = _parse_date(data.get('date_of_birth'))
        e.blood_group    = (data.get('blood_group') or '').strip()
        e.marital_status = data.get('marital_status', '')
        e.marriage_anniversary = _parse_date(data.get('marriage_anniversary')) if data.get('marital_status') == 'Married' else None
        e.address        = (data.get('address') or '').strip()
        e.city           = (data.get('city') or '').strip()
        e.state          = (data.get('state') or '').strip()
        e.country        = (data.get('country') or '').strip()
        e.zip_code       = (data.get('zip_code') or '').strip()
        # ── Phase-1: Family / Contact ──────────────────────────
        e.father_name       = (data.get('father_name') or '').strip() or None
        e.mother_name       = (data.get('mother_name') or '').strip() or None
        e.alternate_mobile  = (data.get('alternate_mobile') or '').strip() or None
        e.personal_email    = (data.get('personal_email') or '').strip() or None
        # ── Phase-1: Permanent Address ─────────────────────────
        e.permanent_address    = (data.get('permanent_address') or '').strip() or None
        e.permanent_city       = (data.get('permanent_city') or '').strip() or None
        e.permanent_state      = (data.get('permanent_state') or '').strip() or None
        e.permanent_country    = (data.get('permanent_country') or 'India').strip() or None
        e.permanent_zip        = (data.get('permanent_zip') or '').strip() or None
        e.same_as_current_addr = data.get('same_as_current_addr') == 'yes'
        

    elif tab == 'professional':
        e.department     = (data.get('department') or '').strip()
        e.designation    = (data.get('designation') or '').strip()
        e.employee_type  = data.get('employee_type', '')
        e.date_of_joining= _parse_date(data.get('date_of_joining'))
        e.location       = (data.get('location') or '').strip()
        e.pay_grade      = (data.get('pay_grade') or '').strip()
        e.shift          = (data.get('shift') or '').strip()
        e.weekly_off     = (data.get('weekly_off') or '').strip()
        e.work_hours_per_day = _dec(data.get('work_hours_per_day')) or 8
        e.notice_period_days = int(data.get('notice_period_days') or 30)
        e.is_contractor  = data.get('is_contractor') == 'yes'
        cid = data.get('contractor_id') or None
        e.contractor_id  = int(cid) if cid and e.is_contractor else None
        e.is_block       = data.get('is_block') == 'yes'
        e.is_late        = data.get('is_late') == 'yes'
        e.is_probation   = data.get('is_probation') == 'yes'
        e.status         = data.get('status', 'active')
        e.remark         = (data.get('remark') or '').strip()
        e.confirmation_date  = _parse_date(data.get('confirmation_date'))
        e.resignation_date   = _parse_date(data.get('resignation_date'))
        e.last_working_date  = _parse_date(data.get('last_working_date'))
        e.rehire_eligible    = data.get('rehire_eligible') == 'yes'
        e.linkedin       = (data.get('linkedin') or '').strip()
        e.facebook       = (data.get('facebook') or '').strip()
        rto = data.get('reports_to', '')
        try:
            e.reports_to = int(rto) if rto and str(rto).strip() else None
        except (ValueError, TypeError):
            e.reports_to = None
        # ── Phase-1: Grade / Probation ─────────────────────────
        e.grade_level = (data.get('grade_level') or '').strip() or None
        try:
            e.probation_period_months = int(data.get('probation_period_months') or 6)
        except (ValueError, TypeError):
            e.probation_period_months = 6
        e.probation_end_date = _parse_date(data.get('probation_end_date'))
        # ── Phase-1: Attendance ────────────────────────────────
        e.attendance_code    = (data.get('attendance_code') or '').strip() or None
        e.overtime_eligible  = data.get('overtime_eligible') == 'yes'
        # ── Phase-1: Leave balances ────────────────────────────
        e.casual_leave_balance = _dec(data.get('casual_leave_balance')) or 0
        e.sick_leave_balance   = _dec(data.get('sick_leave_balance')) or 0
        e.paid_leave_balance   = _dec(data.get('paid_leave_balance')) or 0
        e.leave_policy         = (data.get('leave_policy') or '').strip() or None
        # ── Phase-1: System Access ─────────────────────────────
        e.official_email = (data.get('official_email') or '').strip() or None
        e.role_access    = (data.get('role_access') or '').strip() or None
        # ── Phase-1: Exit extras ───────────────────────────────
        e.exit_interview_done  = data.get('exit_interview_done') == 'yes'
        e.exit_interview_notes = (data.get('exit_interview_notes') or '').strip() or None
        e.ff_settlement_status = (data.get('ff_settlement_status') or 'Pending').strip() or 'Pending'
        e.ff_settlement_amount = _dec(data.get('ff_settlement_amount'))
        e.ff_settlement_date   = _parse_date(data.get('ff_settlement_date'))

    elif tab == 'kyc':
        e.nationality        = (data.get('nationality') or 'Indian').strip()
        e.religion           = (data.get('religion') or '').strip()
        e.caste              = (data.get('caste') or '').strip()
        e.physically_handicapped = data.get('physically_handicapped') == 'yes'
        e.aadhar_number      = (data.get('aadhar_number') or '').strip()
        e.pan_number         = (data.get('pan_number') or '').strip().upper()
        e.uan_number         = (data.get('uan_number') or '').strip()
        e.esic_number        = (data.get('esic_number') or '').strip()
        e.passport_number    = (data.get('passport_number') or '').strip()
        e.passport_expiry    = _parse_date(data.get('passport_expiry'))
        e.driving_license    = (data.get('driving_license') or '').strip()
        e.dl_expiry          = _parse_date(data.get('dl_expiry'))
        e.emergency_name     = (data.get('emergency_name') or '').strip()
        e.emergency_relation = (data.get('emergency_relation') or '').strip()
        e.emergency_phone    = (data.get('emergency_phone') or '').strip()
        e.emergency_address  = (data.get('emergency_address') or '').strip()
        # ── Phase-1: PF ────────────────────────────────────────
        e.pf_applicable        = data.get('pf_applicable') == 'yes'
        e.pf_number            = (data.get('pf_number') or '').strip() or None
        e.eps_applicable       = data.get('eps_applicable') == 'yes'
        e.previous_pf_transfer = data.get('previous_pf_transfer') == 'yes'
        e.previous_pf_number   = (data.get('previous_pf_number') or '').strip() or None
        # ── Phase-1: ESIC ──────────────────────────────────────
        e.esic_applicable       = data.get('esic_applicable') == 'yes'
        e.esic_nominee_name     = (data.get('esic_nominee_name') or '').strip() or None
        e.esic_nominee_relation = (data.get('esic_nominee_relation') or '').strip() or None
        e.esic_family_details   = (data.get('esic_family_details') or '').strip() or None
        e.esic_dispensary       = (data.get('esic_dispensary') or '').strip() or None
        # ── Phase-1: TDS / Tax ─────────────────────────────────
        e.aadhaar_pan_linked      = data.get('aadhaar_pan_linked') == 'yes'
        e.tax_regime              = (data.get('tax_regime') or 'New').strip() or 'New'
        e.prev_employer_income    = _dec(data.get('prev_employer_income'))
        e.monthly_tds             = _dec(data.get('monthly_tds'))
        e.investment_declaration  = (data.get('investment_declaration') or '').strip() or None
        e.proof_submission_status = (data.get('proof_submission_status') or 'Pending').strip() or 'Pending'
        # ── Phase-1: Statutory ─────────────────────────────────
        e.professional_tax_applicable = data.get('professional_tax_applicable', 'yes') == 'yes'
        e.labour_welfare_fund         = data.get('labour_welfare_fund') == 'yes'
        e.gratuity_eligible           = data.get('gratuity_eligible') == 'yes'
        e.bonus_eligible              = data.get('bonus_eligible', 'yes') == 'yes'

    elif tab == 'bank':
        e.bank_account_holder= (data.get('bank_account_holder') or '').strip()
        e.bank_name          = (data.get('bank_name') or '').strip()
        e.bank_account_number= (data.get('bank_account_number') or '').strip()
        e.bank_ifsc          = (data.get('bank_ifsc') or '').strip().upper()
        e.bank_branch        = (data.get('bank_branch') or '').strip()
        e.bank_account_type  = (data.get('bank_account_type') or '').strip()

    elif tab == 'salary':
        e.salary_ctc           = _dec(data.get('salary_ctc'))
        e.salary_basic         = _dec(data.get('salary_basic'))
        e.salary_hra           = _dec(data.get('salary_hra'))
        e.salary_da            = _dec(data.get('salary_da'))
        e.salary_ta            = _dec(data.get('salary_ta'))
        e.salary_medical_allow = _dec(data.get('salary_medical_allow'))
        e.salary_special_allow = _dec(data.get('salary_special_allow'))
        e.salary_pf_employee   = _dec(data.get('salary_pf_employee'))
        e.salary_pf_employer   = _dec(data.get('salary_pf_employer'))
        e.salary_esic_employee = _dec(data.get('salary_esic_employee'))
        e.salary_esic_employer = _dec(data.get('salary_esic_employer'))
        e.salary_professional_tax = _dec(data.get('salary_professional_tax'))
        e.salary_tds           = _dec(data.get('salary_tds'))
        e.salary_net           = _dec(data.get('salary_net'))
        e.salary_mode          = (data.get('salary_mode') or '').strip()
        e.salary_effective_date= _parse_date(data.get('salary_effective_date'))
        # ── Phase-1: Salary extras ──────────────────────────────
        e.salary_conveyance = _dec(data.get('salary_conveyance'))
        e.salary_bonus      = _dec(data.get('salary_bonus'))
        e.salary_incentive  = _dec(data.get('salary_incentive'))
        e.salary_gross      = _dec(data.get('salary_gross'))

    elif tab == 'education':
        e.highest_qualification= (data.get('highest_qualification') or '').strip()
        e.university           = (data.get('university') or '').strip()
        e.passing_year         = int(data.get('passing_year') or 0) or None
        e.specialization       = (data.get('specialization') or '').strip()
        e.prev_company         = (data.get('prev_company') or '').strip()
        e.prev_designation     = (data.get('prev_designation') or '').strip()
        e.total_experience_yrs = _dec(data.get('total_experience_yrs'))
        e.prev_from_date       = _parse_date(data.get('prev_from_date'))
        e.prev_to_date         = _parse_date(data.get('prev_to_date'))
        e.prev_leaving_reason  = (data.get('prev_leaving_reason') or '').strip()

    elif tab == 'documents':
        e.documents_json = data.get('documents_json', '[]')

    else:
        return {'ok': False, 'error': f'Unknown tab: {tab}'}, 400

    e.updated_at = datetime.utcnow()
    db.session.commit()
    return {'ok': True, 'msg': f'{tab.title()} saved successfully!'}

@hr.route('/employees/<int:id>/view')
@login_required
def emp_view(id):
    perm = get_perm('hr_employees')
    if not perm or not perm.can_view:
        flash('Access denied.', 'error'); return redirect(url_for('hr.employees'))
    e = Employee.query.get_or_404(id)
    audit('employees','VIEW', id, f'{e.employee_code or id} / {e.full_name}', obj=e)
    from datetime import date
    sub_perms = {
        'kyc_details'    : True,
        'bank_details'   : True,
        'salary_details' : perm.can_view if perm else False,
        'documents'      : True,
    }
    return render_template('hr/employees/view.html',
        employee=e, perm=perm, sub_perms=sub_perms,
        active_page='hr_employees', today=date.today())



@hr.route('/employees/<int:id>/id-card')
@login_required
def emp_id_card(id):
    perm = get_perm('hr_employees')
    if not perm or not perm.can_view:
        flash('Access denied.', 'error'); return redirect(url_for('hr.employees'))
    e = Employee.query.get_or_404(id)
    from datetime import date
    LOGO_SRC = 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAA0gAAAM6CAYAAAC7DWMRAAAACXBIWXMAAA7DAAAOwwHHb6hkAAAAGXRFWHRTb2Z0d2FyZQB3d3cuaW5rc2NhcGUub3Jnm+48GgAAIABJREFUeJzs3XuYZlddJ/rf2lXVSciluwlKUtVyOYHxEXFGR8ZHB/XIqOMFR5xxkpNUdSITlaAOKreAXLoSFCWIAWQ8jiAqpKoDxHFELjri7fGoc27OOSpzdEBBSV+iCAQIJunu2uv80W+gk3R31e6933fvvd7P53n4g67d7/tLv937Xd+91vqtlHOOebG8uvHNKaWXRcRTt7s25/jlraj+/d8evOZvZ1AaAAAwAGkeAtLKDW98RL7nwp9KETdERGrwWz9Wp7j26Mbar0+rNgAAYDiKD0grq7c/KlL9GxHx5ef4ElsR6XmHN1d/usu6AACA4Sk6IH3+Vb902dLS0m9GxJe0fKmcIj/z0Ob+t3ZRFwAAMEzFBqTHXbtx+fE6/reIdEVHL3nfVqQvvmtz9UMdvR4AADAwVd8FTMOl+zcvOV6n93YYjiIizl+I+nUdvh4AADAwxc0gPfmqO3Z9YunYeyLiG6bx+keO71rM77hyaxqvDQAA9KuoGaR01R0LH188thlTCkcREY8+/77zp/XaAABAv4oKSMu7jr02pfi303yPxZQFJAAAKFQxAWll/+bNkeM5036f6tiu86b9HgAAQD8W+y6gC/vWDn5fRByYxXsdWzouIAEAQKFGP4O0vLrxnTnyG2b1fruO7Rr9nxkAAHB6ox7sL68d/NqU0kZELPRdCwAAMH6jDUiX7998Uor8qxEx06YJeWFrtH9mAADA2Y1ysP+Ya962XOV4b0Ts7bsWAACgHKMLSJfu37xkq9p6T0Q8tu9aAACAsowqID35qjt2nZ/jVyLiS/uuBQAAKM9oAlK66o6FTyzd/5aI+Po+68hbJ1Kf7w8AAEzPKAJSSpGWF4/9bES6uu9aAACAcg0+IKUUaXl142cixff2XUtERK6WBv9nBgAAnJvBD/aXr9l8VUT6vr7rAAAAyjfogLS8f/PHIsWNfdcBAADMh8EGpH2rmy9NOV7adx0PlWtNGgAAoFSLfRdwOiv7N2+MFD/Wdx0AAMB8GdwM0sr+zRsjxy191wEAAMyfQQUk4QgAAOjTYALSWMJRXqzsQQIAgEINIiCNJRwBAABl671Jw8rawRdEjCgc1WaQAACgVL3OIC2vHbw2Ir+6zxoAAAAe0FtAWrl249tT5F+IiFHNyORF5yABAECpellit2/19q+KlG7v6/0BAABOZ+YB5Qv2b35JTvGeiHjErN8bAADgbGa6xO4xV7/9ijriNyNi7yzft0tLmjQAAECxZjaDdNl1t33+wkL6L5HTZbN6z2nIKQ+iNToAANC9mQz2V2544yMWtqp3RqQrZvF+AAAA52LqASlddcdC3HPhwYj4ymm/FwAAQBtTDUgpRVpeOv6miHjGNN9nlvKJRUvsAACgUFPdg7S8uvm6iPh303yPmXMOEgAAFGtqsyH71g6+MiJ+cFqvDwAA0LWpBKTltY2X5MgvmcZr9y1r8w0AAMXqPCDt27/5/BTplV2/LgAAwLR1GpD2rW48L+d4TZevCQAAMCudBaR9qxvPyyn9VFevN1j1liV2AABQqE4C0r61g8+di3AEAAAUrXVA2rd28Lk58q1dFDMGeUGTBgAAKFWrc5CW92/+cIqYm3AEAACU7ZxnkJb3b/5wyvHaLosZg8WczCABAEChzikgLa9t/FDKZo4AAICyNA5Iy6sbN6RIr42I+ZxJ0cUOAACK1SggLa8dfFZK6WdjXsMRAABQtB0HpJX9mz+QIv/HmPNwlNNip4frAgAAw7Gjwf7y6ub+yPHTMefhCAAAKNu2AWnl2o1vTyl+cSfXAgAAjNlZz0FaXjv4tSnS27e7bp7krEkDAACU6oyzQstX3/ZFKfKvRsT5M6xn+BacgwQAAKU6bUC67LrbPj8tVO+NiL0zrgcAAKA3DwtI6eabq4Wt6q0R8bjZlwMAANCfhwWkyz/4hBdFxDf1UMso5JQ1qwAAgEI9aLB/+ertX5xy3NxXMQAAAH36bEBKKVKV6jdExFKP9QAAAPTmswFpeW3zqoh4Wo+1jENd6WIHAACFqiJOzh5Fzi/quxgAAIA+VRERK2ubT49IX9Z3MWOwUNWaNAAAQKGqiIhcx/f2XQgAAEDfqseuHdwbKWvrDQAAzL1qK+Vvj0jn9V3IWGRNGgAAoFhVzul/7rsIAACAIagi4ql9FzEqecsMEgAAFKqKyE/su4hRqSyxAwCAUlURYcAPAAAQkzbf7FzOSaAEAIBCCUhN5VpAAgCAQglIAAAAEwJSU5UldgAAUCoBqSF7kAAAoFwCEgAAwISABAAAMCEgNZRT9mcGAACFMtgHAACYEJAaWtSkAQAAiiUgAQAATAhIDeW6NoMEAACFEpAAAAAmBKSmFuxBAgCAUglIDeW6EpAAAKBQAhIAAMCEgNTUwpYZJAAAKJSA1JQldgAAUCwBqamszTcAAJRKQAIAAJgQkBrKlTbfAABQKgEJAABgQkBqKptBAgCAUglITWnSAAAAxRKQAAAAJgSkhipNGgAAoFgCUlP2IAEAQLEEJAAAgAkBqaGcsj8zAAAolME+AADAhIAEAAAwISA1pUkDAAAUS0BqykGxAABQLAEJAABgQkBqykGxAABQLAGpoWwPEgAAFEtAAgAAmBCQmsrZDBIAABRKQGoqWWIHAAClEpAAAAAmBKSGqnAOEgAAlEpAasoSOwAAKJaA1JQ23wAAUCwBCQAAYEJAaihne5AAAKBUAlJTlSV2AABQKgGpKXuQAACgWAISAADAhIDUlD1IAABQLAGpoWwPEgAAFEtAAgAAmBCQmtKkAQAAiiUgAQAATAhIAAAAEwJSU7rYAQBAsQQkAACACQGpoUqbbwAAKJaABAAAMCEgNaXNNwAAFEtAaijnLCABAEChBCQAAIAJAampKswgAQBAoQQkAACACQGpKXuQAACgWAISAADAhIDUlDbfAABQLAEJAABgQkBqKplBAgCAUglIDeWoBSQAACiUgAQAADAhIDVliR0AABRLQAIAAJgQkBqqtPkGAIBiCUhNZU0aAACgVAISAADAhIAEAAAwISA1VdmDBAAApRKQGsqaNAAAQLEEJAAAgAkBqSld7AAAoFgCEgAAwISA1JQmDQAAUCwBCQAAYEJAakoXOwAAKJaABAAAMCEgNZRT9mcGAACFMtgHAACYEJAAAAAmBKSmctakAQAACiUgAQAATAhIDVXafAMAQLEEJAAAgAkBCQAAYEJAaipZYgcAAKUSkBqrBSQAACiUgAQAADAhIDWUdbEDAIBiCUhN2YMEAADFEpAAAAAmBKTGNGkAAIBSCUgAAAATAlJT9iABAECxBKSmdLEDAIBiCUgAAAATAlJDObIZJAAAKJSA1FQKAQkAAAolIAEAAEwISA1VmjQAAECxBCQAAIAJAamprEkDAACUSkBqKFeaNAAAQKkEJAAAgAkBqSlL7AAAoFgCUlO62AEAQLEEJAAAgAkBqSlNGgAAoFgCUlP2IAEAQLEEJAAAgAkBqSlNGgAAoFgCUkM5CUgAAFAqAQkAAGBCQGqsNoMEAACFEpAaquxBAgCAYglIAAAAEwJSU5o0AABAsQSkhnI4KBYAAEolIAEAAEwISE2lMIMEAACFEpCa0sUOAACKJSABAABMCEiNadIAAAClEpAAAAAmBKSmNGkAAIBiCUgAAAATAlJT2R4kAAAolYDUUA5tvgEAoFQCEgAAwISA1JQmDQAAUCwBCQAAYEJAaqjSpAEAAIolIAEAAEwISA3lrIsdAACUSkBqKglIAABQKgEJAABgQkBqrDaDBAAAhRKQAAAAJgSkpjRpAACAYglIAAAAEwJSU7rYAQBAsQSkxrKABAAAhRKQAAAAJgSkpnKYQQIAgEIJSAAAABMCUkNZkwYAACiWgAQAADAhIDVU6WIHAADFEpCaSpo0AABAqQQkAACACQGpoZwtsQMAgFIJSI3pYgcAAKUSkAAAACYEpKY0aQAAgGIJSE3ZgwQAAMUSkBqzBwkAAEolIAEAAEwISE3ZgwQAAMUSkJrKAhIAAJRKQAIAAJgQkBrKZpAAAKBYAhIAAMCEgNRQlbT5BgCAUglIjTkoFgAASiUgAQAATAhIDWnSAAAA5RKQAAAAJgSkpjRpAACAYglIAAAAEwJSY7rYAQBAqQQkAACACQGpKV3sAACgWAISAADAhIDUVLIHCQAASiUgAQAATAhIjTkHCQAASiUgNZU0aQAAgFIJSAAAABMCUkM5a9IAAAClEpAaquxBAgCAYglIAAAAEwJSQ1mTBgAAKJaA1FQWkAAAoFQCEgAAwISA1JwZJAAAKJSA1FzuuwAAAGA6BKSmkoAEAAClEpAAAAAmBCQAAIAJAQkAAGBCQAIAAJgQkBrKdZzfdw0AAMB0CEgNpSoe3XcNAADAdAhITeW4rO8SAACA6RCQmtv7mKvffkXfRQAAAN0TkM5BvXjiG/uuAQAA6J6AdA5yjqf3XQMAANA9AencfPOjV2/XrAEAAAojIJ2bxYWqvqbvIgAAgG4JSOco5fiuvmsAAAC6lZZXN3LfRYzYb1W5ev6dB6/5074LAQCgP0++6o5dn146fuED//94xCV1xEJExOLWicV6YfHiB36Wc7ovp3xvREQsbOX6vvM/+nfvuPKemRfNaQlI7Z3IkQ5WKd+xtHT8tz78i8+8r++CAADOJt18c/WYDzxx95l+fu/xY+dVS7secbqfLdTV7lzVD1qFVKW8lHO66GEX5/qilKqlB/1SylXO6TTvXV+Qojr/4cXmPZFzethLR+yqIi582PUnf3ZJRFo4zQ8WI8XFD/v1FI+IHOc9+G3jvEjxoD+D/PBf2x3drcj6hxRxV05xV9TxFxH5DyIW/uvhg9f8RUevzw4JSN26LyI+miOOpohPTfet8r0RaUZhLG9N/7/nIe948v22pvkeKaqtPOP/rlNM7b2riGN1xIdOHD/2gb97xzPvmsZ7AMOQUqTHrB7cs911Jxa3Lt46sbB4up8tVXW1VVcPG6wuVPXFdV2d9veklPdETikioj55r/6b+1P+4Mc21vq6p87ME7/1Defds+dRT1io6sfnOh4fEXsj4pKU8sURETmqpYh8UaQ4kXL+9MNeIKVH5AcPxE83wL4oIpYe8msXpIgHhYd8cnbikrOUe0lMZjAYs/ynkdLrdy0dP9j1g/hDr/2CCy6+/zShtIGt+7cuXIgTu9rWshXp/IUqXdD2dT77enW9tBALF+VUby3khcMXP/KCv4nnfPD+nfxeAQnK9omI+P0c+XfTYv7Ph99y7Uf6LqhLT/zWN5x3395LH1Pn+nEpVXtyrnfnVO34SV6KeldEddqnj02kiKUc9cOfnJ7bq12QIrf6sjqdHGlvpy+Y4uLIcdrBc2OneUp7zlI8Ij3kKfC5yt0+Gd7O6QbEY3RnjvS7kfP7zjvv+Ls//IvPvLvvgrqwfM1tT6lS9W9yyl8TEf8sInXydwwaurOK9Ow7N1ffu9PfcPfNyydCSH5AHRGHUsQf5BS/sZDTr1+8fvjvT3ehgATzo46I34qcf/bI7fvfmXOM8t/+5fs3n1TVcX2k+OaIeFJEPGzZBTAI9+ZId1Qp/+yhjbX/ve9imnrKDW9cOvqZC58VOW6IiC/pux54QIq47d4U/34nM7Z337z81xHx2OlXNUr3pki/GAtbr9n9srs+fOoPBCSYQyni/8qRXnJ4c/W3+q5lp77g2o0n1Fvp1ZHiX/ddC9DYb+dIrziyufr7fReyEytrB78hIr8uIr6471rgDD6Yqnzlodv2/8nZLvrkzSvfkyO/aVZFjVM6HhE/sSf2vjLW338sQkCC+Vbnpx6+ff8f9V3GdvatbVyXI/2vcYbNuMA4VJGe3mR5UB+Wr934mlSnUQQ55t69OdL3Htlc3TzbRZ+8efkncsSLZ1XUiP33rYW48tKXHflz5yDBPKuqF/ZdwnaW1w5+U470lhCOYPTqyK9/8lV3tN7MPU2pTi/tuwbYoQtS5NtW1g5+99ku2r1+5Eci4rUzqmnMvnhhKw7GzWlRQIK5lp+xfPVtX9R3FWeSbr65SpF/vO86gM484e6l4z/QdxFnsnLN274sIv5l33VAAyki/9zK2sGrz3bRnvUjz0sRr5xVUSP2pXfHZasCEsy3FAvVD/ZdxJlc/oEnrkXEP+27DqA7OfKLnnLDG4fZsa/a+sHQ+IXxWYjIbz25d+7Mdq8feVlE3DqjmkYsfbeABHMuRf7OdPPNg7wXpMg/1HcNQOcefeQzF35b30U81OP/3S+dH6EJDKO1FJH/0+Wrt5+1sciem46+ICL90oxqGqn0VYMcFAGzlD5v5YNP+Iq+q3iofddu/JOI+PK+6wC6l+p4Rt81PNTxY7v+VZw8+wrG6pIq1e+6/Ko7Pu+MV+Sc98SR743IvzbDukYmLwlIQNQRg3uam7fStX3XAExJiq/su4SHyhHf0ncN0IHHV0vHbjvrypD1fOKeSxavjog/nF1Z4yIgAZHq+Nq+a3iolIZXE9CZK574rW84r+8iTpVy/md91wAd+abl//GEG892wb7n3nnvwq7qGTnif8yqqDERkICIKp7YdwmneuK3vuG8HPkf910HMDWL9zzykVf0XcQDLrvutgtzisF29ITGUvzoyurtX322Sy7+kUMfS5G/OUXcNauyxkJAAiJyXHbFVXcMZu39vXsv/aKINKiny0C3qjqv9F3DAxa3Fp4YEQt91wEdWoxUv/Xzr7rjorNdtGf96F/XVf0vI9InZ1XYGAhIQERE3Lt0/Al91/CAHPGovmsApmwhLui7hM/KW4/suwSYgscv7jr2mu0u2vvyu/4sRVwVkY7PoqgxEJCAiIhIEXv7ruEBQ6oFmJKtuLDvEh5Qp4U9fdcA05ByPGtl/+a2hx/vXj/8mzny9RFRz6CswROQgJNyGszT3BxxSd81AFNWDeeek1IWkChVioi3PO7ajcu3u3Dv+pGNlOLZEZGnX9awCUhARESklM/vu4bPyrWT7KFweUh7fupsPES5clx2vE7/6Sk3vHFpu0t3HzjyphTpWTHnIckNAZioB/M0NyK2+i4AmLIBhZKUqhN91wBT9lV33XPRT+7kwt3rh38+Ujx32gUN2WBuTkC/ch7Q09wqmUGCwlWpGswYJJu1Zg7kyD+0srrxwp1cu+fAkddHxPOnXNJgDebmBPRMKAFmKOfsngOzltIt+9Y2rtvJpXvWj9wakW6edklDJCABJw1ouQtQvpTSXO9xgJ6kHOkX9q0dfOZOLt6zfvimiPixqVY0QAZEwEmpGs5gJZvNgtLVuR5OO2Ez6MyXhRz5F1ZWN79/JxfvWT/y8pTSi6dd1JAISADA7A3poQzMnxQp/sPK6uZzdnLx7gOHb5mnkCQgASelbLACzCez1synFClev7x/84d3cvHuA4dvySleNu2ihkBAAoZHRykAmIWUcty6b+3gjtp67z1w5JUp8o464Y2ZgAQAAPMr5ci3Lq9tvGQnF+9eP/qa0pfbCUgAwOyZKYZBSZFeueOQdODwLRHpFdOuqS8CEjA8Nm9D8aohdY6zBxMi4mRIWtm/eeNOrt2zfng9R371tGvqg4AEDI8ny1C8uhZKYJBy3LJvbePlO7l07013vThF/MdplzRrAhIAMHtmimGwcqRXrKwdXN/+wpx333T0+yPizdOvanYEJOCkIbW5NXCC8g1ppthsFpxGvmmnIWnPk47ekCO/fQZFzYSABAzPkAZOwHQM6UHIkPZDwaDkm1bWNl617WVX5q29cel1EfG+6dc0fQIScJJQAsySxggwEulFK6ubt2x72fr7j5244Ph3RsQfT7+m6RKQgIiIyJ6gAgCnk+LGnYSkR9340U8v7srfEhEfmEFVUyMgARERkazBB2bIPQdGJsWNK/s3t23rfdGPHP1oivpbIuJvZ1DVVAhIAADA9nK8cCchaff6XR9KVTw9Iu6ZQVWdE5AAAICdyfHClbWNn9zust0vP/LHKepnRMT9M6iqUwISMDj2QwHAkKUXrKxuvma7q3av3/U7OeJ7ImJUS2oFJAAAoJkUz99JSNq7fmQjRXpWjCgkCUgAAEBzKZ6/vHbwp7a7bPf64Z9PKW6IkYQkAQkAADgnKfLzdhSSDhx5U6R47ixqaktAAoYn24MEpbPXEMqRIj9vJ40b9hw48vpI8bxZ1NSGgAQAALSUXrBvbePl212158CR10bKL5hFRedKQAIAAFrLkV6xb+3gtsvo9hw4+lMp5RtnUdO5EJAAAIBO5Miv2be6cdV21+0+cPQnI+WbZlBSYwISADB79hpCqaqc0ltWrtn459tduOfA0ZtTpBfNoqgmBCQAAKBL50cVv/qYq99+xXYX7l4//OqhLbcTkACA2cu1GSQoWvq8rYXj/+XRq7c/ersrJ8vtBtO4QUACAACmIF2xmPK7P/+qOy7a7so9B47+VEQ8fwZFbUtAAgAApiQ/ZWnp2K8++ao7dm135Z71I7emlF48i6rORkACAACm6es/sXT/L6QU2y6t3X3g8C0p4idmUdSZCEgAAMCUpbXltc0f38mVu9ePvCQibp1yQWckIAEAANOX48X7VjdfupNL99x09AUR8eYpV3RaAhIAADATOcWPrawd3P7so5zznjj67Ij4z9Ov6sEEJCAiInLl0EYAYBbyq/bt39y+Y916PrEnHnl1RLxv+jV9joAEAADMVM7xk/v2bz572wvX338sn3/flRH5T2dQVkQISADAvKtz7rsEmEMp5/iZlf2b12934d4XffyTJ2LpWyLir6dfloAEAMw7S4yhL1Xk+PmV/Zs/sN2Fj1r/yJEqFr4xIj46/aIAAAD6kSLHG1ZWN5+z3YWXrN/5l1Xkb4uIz0yzIAEJAJi5yqwN8DkpUrx+Ze3gD2534SXrR//PHPXVEXFiWsUISMBJ2WAFmCH3HODBUkR+3fL+zR/e7sK963e9O0X6voiYyv5BAQkAmL00oMYImjTAUKSU47U7aQG+e/3wz0dKPzSNIgQkYHhy7ckyFK7Oqe67BmCYco7X7Fs7+LLtrttz4PAbIuK5Xb+/gAQAzN6QZpCAwcmRf3RldfPAdtftWT/yukh52+uaEJCAk8zaADOULGsDtpPi5pW1jVdtd9meA0d/NEW8squ3FZAAgNlLlYAE7EB60U5C0u71Iy/LkW/p4h0FJGB4DJwAgM/aWUjae9NdPxKR/0PbdxOQgOGx3A/KZw8S0MgOQlLOec9Nd/1gRPx0m3cSkAAAgBHYaUg6+sMR8fpzfRcBCQAAGIkdh6TnRsTrzuUdBCQAoA+H+i4AGKv0opXVzbM3ZMg571k/8twU8RNNX11AAgBmLaeFrT/uuwhgxFLcuG1Iiojd60dekiJ+vMlLC0gAwIzlDx16y3Uf67sKYOR2HpJe2uScJAEJGJyqSrrYQcFSVG/vuwagEDsPSS/LkV+9k5cUkACAWdraqqs39V0EUJAdhqS960dfFBE/tt11AhIwOHXtfBQoVY60efT2q/+67zqAwqS4cWX/5rYzRHvWj7w8Iv3o2a4RkIDBscQOinVk4fjSc/suAihUjhfuLCQdPhCRXnGmnwtIAMAsHI+cn3nnO678eN+FAAXbeUhaP1NIEpAAgGmrI+fvOnxw//v6LgSYAzleuLx/c9u9RpOQdPNDf11AAgCm6ZMpxfWHD+6/ve9CgPmRcrx0Ze3gC7a7bs/64ZtSSi8+9dcEJABgGv6/FPnArl3HH3doY+0tfRcDzKP86pX9m9dvd9XuA4dviZRveuD/L061JgDG4GMR8TcRMYW9IfneiHRf9697xvf7VIrYmsYr1ynllPPd03jtB6s+Mf33iMhR3xdR3dvV66WU78kR96a6+uhCTh/6yO1XH+nqtSnSPRHxlxHxsRTxyRzxqRT5WNMXyRGXRKSF7svrWI7FSHHxOfzOvRFxaUQ8MiIu6baouZAix8/tu2bjbw/dvv89Z7twz4GjN3/yFSv35ZxfJSBNU47DkeL9keIDkfM/RFTHIuUcOZ/X7IXSBSny+e1KSbuj2Yzhroi4sM17PsTenV2WLk6RFyMi8sn33zWFWjgNnePmyl9Ejl+LlN537Pz7/o+Pvvn6T/ddEFC8T6cU78x1fm/Ewv9z5As/8IG8vl73XdSYPOWGNy7d9Q+PeFLU8U9yTv84Ir4iUvzziBh+QOzXYq7S2/Zdu/HVh27b/ydnu3D3gcO33H3zygUCUvdyitjIKX7uyMG1P8o5nOfSgX3f9dZLc109Nm+lJ6WIp0WKp0XE4/uuC8Yl/1VU8bwjG/vf5d4EzEb+g5TS69OxXe+58x1XdjZjOY/+75971vGI+JPJ/yIi4guuuuORW7uOfWfK6VkR+Sn9VTd4F+U6vfOy6277irveeu3fne3CPeuHbxKQuvXBHOl7Dm+u/n5ERGz0XE1BDr3luo/FyWVA/y0mf7Ira5vvj4gv7rMuGIsU8Sv3n3/smR998/Wfjtv6rgaYA38Yubrl8MFr3tV3ISWbtM1/U0S8aWVt8y8i4gt7LmnIHruwVf3KE7/1DV//wfc+5/6zXSggdSbfn7fyM4687do/77uS+ZHuiMgCUkfqOptRKNfBw8d3XZc316ayNwfgId55eHPtO/ouYu6k+LXI8cK+yxi4p35m76VvTCmeebaVFLrYdSSn9BrhaLbqnH657xpKMqg9SHlAtYzf359/fNf353dcKRwBs7C1sLX4/L6LmEc5ZbN1O5AiX7d8zearznaNgNSNf1g4tuvWvouYN0cPXvPfI0KXpBIls1ndSa/5q3dc+cm+qwDmRXrXR972v/xV31XMo6P3n/dHEfH3fdcxCiluXF7dvGNl9fZHne7HAlIXUrxzsgaUGUspfqfvGpgCM0hdyVFXB/suApgfOeX39l3DvJqsFPjNvusYi5Ti30ZV/9nK/s2nPvRnAlIHck6/3ncN8yrX8V/7rqEU9iCVKH3w8O1X39l3FcAcyckAvU8l2JURAAAgAElEQVQ5/rDvEkYlx2WR82/vW9u47tRfFpA6sJDy/9t3DfOr+tO+KyjFoPYg0Y2UP9R3CcBc+cyRzdW/6buIuZYXPDhuLJ2XI71l3/7Nz+6dE5A6UOV0qO8a5tX5Jxb/rO8aYMA+0XcBwDzJd/Vdwbw7ctdlfxYRx/uuY4xyjp/cd83G0yMEpC585m82Vw1CejLZfH5333XAEKUcJ/quAZgn6aN9VzDv8u9+3YlI8bG+6xiplKv0psuuu+1CAam9T/VdwLxLOczgdcAepPLk7CkiMEPuOcOQBaQWLl+sq2cLSO05W6RntZaWnbAHqUAp1X2XAMyPlLJ7zhDk0Fm5hRzx3QJSewJSz1I680nI7JwZpAJlnykwOzk8lBmEZHl1Kzm+SEBqz82gd8kgsANmkMqTHLgLzJTvY8ogILWUwuwFMEw5Kvd42IGUKt/lnfBQZiA8vG/Jl2dL2c2AQlhiV6Cc3eNhB3KuzaB3w/cIRfDl2Zrp5P4Z2HdhUEvsLA3rRnKPh50wg9QVY6KB8Dm05MuTErgRdGBQM0h5QGFt1AxWYCfMIHUl+3McBp9DSwISMDxmkLqhix3siBmkzhiYD4KHjG0JSBTAU3IA6J+BOWUQkFozndw/B9N1YVB7kOiGmTgAaExAAgCgAx7KUAYBqTXLuyjDoJo0ADBGvkcogoDUnpsBAAAD4YFnWwISEBH2IBVJu3TYGfv1OmJVzUD4HFoSkFpzMwCAUfMwoSOCJmUQkFpzM6AMQ9qDNKRaAID5IiAxfslUMgAA3RCQAID5Zg8SRbH9oy0BCaBUyb4KYJbcc4ZB4G9LQGL0UrbEDk4r+5IEgKYEJAAAOuChzDCYyWtLQGrPzaBntRsyAAAdEZAAAAAmBCRGLyXdWgAA6IaABAxPqoReYHY0NAFOISAxfrrYlSfXNpgCs6MlPnAKAYkCePIHQAtmkIBTCEgAAHRB0ByEbEa0JQGJ8cuaNAAA0A0Bqa3kaQkAAJRCQAIAAJgQkNrSQQ0AwKoaiiEgMX5uyAAAdERAAgAAmBCQ2jN7AQAAhRCQKIA233BaKTkLA3agjsr3CPBZAhJAqXI26ANmJmlcRSEEJMbPIBAAgI4ISC15WgIAAOUQkBi9pM13J6rKfpXi2IMEAI0JSAAAABMCUkvZ7AWFqGt7uQA4d3X4HqEMAhIF0OYbAIBuCEgAAAATAlJrppP7VmvzDUAbvkeAUwhIAMB80/GxEylZ8j4Q/j63JCAxetp8A9CKGSTgFAISAADAhIDUVjad3DszSAAAdERAAoYnVUJvJ7J16ADQkIDE+GUzSMXJtYE9wNj4PqYQAlJblncBAEAxBCQKoPsQAADdEJAAAAAmBKTWzF4AABgTUQoBifFzcjcAAB0RkICTtNYGABCQKIC2onB6ObRLh52wEgE4hYDUmpsqAACUQkCiADaFFsdyP4DxyR4aUwYBCQAAYEJAas/Tkp4la8cBAOiIgMT4adLQjVzb0A8AzD0BCQCYaynZywp8joDUmuVdFEJjBGBO5ZzMoHchWdFBGQQkRq8WUuG0cjLoA4CmBCQAAIAJAaml5AweAICw7YBSCEiMXspCKgDnTpMG4FQCEgAAwISA1FLtoNj+6ZoDAEBHBCQAANqz5J1CCEgtJbMXAAAMh7FpSwISBdA1pzg2THeiiuwcJABoSEACThpSKHGqPQDQEwGprWwas3fWPANA72w7oBQCEuM3pJmPMTNrAwAgIAETgiYAgIDUAYNKAAAohIAEUKoUlk3CTtRm0CmJ7r5tCUht2ZA4AJXPAAB6Z2A+DI54aEtAAgAAmBCQGD9tvsujYQQA0BMBqS3nIAEARO2BJYUQkAAAACYEpNY8LelbshwLAICOCEgAhcpZJyNgdpLOvhRCQGL06qStKAAtJMdFdCFnZ69RBgGprWxwDgCjlmsD+w6YQaIUAhLjp5MgAAAdEZCAiIhItWYXwJyyxK4bZpAohIDE6CWdBDuRq2SJCQAw9wSktjwtoRBmkEok9AJAUwJSS2YvAACgHAIS46fNdycssQOgFU2TKISABESEJXYAABECEiXwxKoTZpAAAASk1moHxVIIM0gAAAISQLlSmBWEnUgeEFESD+/bEpBaStp8D4AbAQAtZEuMuyFoDkM2vm/JHyDj58kfAPRPV9mBEPjbEpBaczMAgFHzoK0bORuYUwQBCQCA9swgUQgBifHT5hsA+uf7eCDMiLYlILXmLyEwUJa7AMwjY9OWBCRGL9kHVhxnMgHAuTIuaktAas9fQgAYMQ9lKIvVA20JSIxf9sUGwLnLlbbI3fB9PAgpHtF3CWMnIAEAQCmygNSWgNSadZ59q51fAQC9S9p8D8UFfRcwdgISQLEsG4KdsAeJouS4qO8Sxk5Aas1NtX+VzwCAc2YPEqVIT/u9xUjx6L7rGDsBCQCYa2aQKMXly0dWImKx7zrGTkBi9JIudgDQv+zok75VKX9J3zWUQEBqz80AAID+5fxlfZdQAgGprSQg9U4XOwDoXa2zb+/qnP5F3zWUQEBi/LQVhdNLYeM5wJy44qo7dqcUT+27jhIISC0l620BAOjZvbuOXRURS33XUQIBCQCA1jRN6lfKcUPfNZRCG0AAABipldXbvzpSvR4RX953LaUQkFqqHRTbvzqynRYAnLNUZU1pGZuV1Y1vjEgvjxRf03ctpRGQAABgBB5z9duvqBeOr+ZIa5HSF/ZdT6kEpJaSDmr9SzmHKSQA6NdIjj557NrBvVsRj42cLsgpX/jAr9cpLy3kdNHZfm+d8p7I6YyDjhRxfkRccMYXSHlP5LzNoCVdlCIvRUTkSHsjYlfkeFyk+J9iIS425pk+AQkAgGKlp/3e4srK4e/IOX9nRPrGiLj05A8enOdOLrQ8e8Y7+Vta5MAcsZOA87C9AzLRTAlIjF/SbB0A+je8VTUrawe/YXk535pzfImUwU4JSG0ZmgNDpX0JMMeW1w6upcgbfdfB+DgHidFLtZAKAHzOZdfd9vgU+Wf6roNxEpBa0+YbACAGdFDs4lb1mojY3XcdjJOAxPjpJAhAG2k4A3vaW7l245/miH/ddx2Ml4AEAMy3s7RtpoGhBM06XR86MtCCgNRWNnvRt3ooN2QAxsn3SDcGEDRTihRmj2hJQAIA5tsABvZFGEDQXL76bfsiYrnvOhg3AamtkZwaXTSt1gFgAKrev4/rha0n910D4ycgARQqOwcJdmYAMx90o6rzZX3XwPgJSIxeGuDJ3QBAD1JlbEtr/hK1ZnAOADCkc5CgDQGJ8bM0AgCAjghIbXlaAgAwDLm295LWBCTGT0gFgN6lAazoyJWW7bQnIAEAUAZnWtEBAaml5BwkAIBhsMSODghIAIWqkiepwOzUqf/OvpUldnRAQKIA/Z/cDcB41XX/e2foiCV2dEBAaq3/pyUAwLkz6wCcSkBi/HSxA6AFM0gdyQPYlz2ATnqMn4DUUm1wDgAAxRCQGL0hnLsAAPTPbCBdEJAAAAAmBKSWnIM0BLrYAUDfUpi9oQwCEkCxss5cwOwM4Bwk6IKAxPhplAEAQEcEpLYssQOAcUuWagOfIyAxerUudgC0kWvLUbswhHOQoAMCUltuBgAAUAwBqTWzFwAwapbYAacQkBg/XXMAAOiIgNSWwXn/slbGANA/YyLKICC1ZXDePyEVAICOCEhtGZwDA5VzeIADAA0JSG3pYte7VPsMAIDQcINOCEgAAAATAhLjZ5kjAPTPwe0UQkACAACYEJBa87QEAGAQzGLRAQGJ8XMzBACgIwISADDfPGgrRqp9lrQnILWUNAjon5shAPRvAEef5Co5/43WBCQAANpL/R9ObQaJLghIAAC0ZwaJQghIbQ3gZjD3nJpdHF9wAEBfBKS2BjCdPO+SzbUAQERE9oCN9gSktswgAQBECo2rKIOABAAAMCEgtVR7WgIAMAy5tsSO1gQkxk9LTwAAOiIgAQBQhEoXVDogILWUstmL3mnzDQBARwQkAADa89CYQghIAAAAEwJSW8k5SH2rHRQLAERErXETHRCQAEqVbFYGgKYEJMbPmmcAIELjJjohIAEAUAYHxdIBAam15ElFz1J4WgQAfRvCnuAqVca2tOYvEQAARcg5m0GiNQEJAABgQkBqS4OA/vkMypN1XwOguZRsfaA9AQkAgA70vye4znXddw2Mn4DE+A1gUygAMABVEpBoTUBqy+AcACAi+m+xnSK2+q6B8ROQAADoQP9L7HJO9/ddA+MnIDF+NmQCABERub6v7xIYPwEJAIAiLKTq3r5rYPwEpNb6n06ee3X4DACgZ2kAx24cX9j6SN81MH4CEhAREbly9hAA43bx39/9wdCogZYEJAAAivDB9z7n/oj4y77rYNwEJEbPqdkAwANSjl/vuwbGTUBqawDrbQEAejeQsyFPLNZvjrA/mXOV/0BAYvwGckMGgLmWh7GX9a63Xvv+iPTOvutgnFJKtwhILSWDcwCAQT2w3FrYel5EfKbvOhiXnOOXD22svVtAAgDmWqqHM7CnG3e99doPpxQ/0HcdjMqhhRO7boiwB4kC1L7YAGjBMQcdGVjTpEMba29Jka6KiI/3XQsDl+I9KcVX3fmOKz8eISABAHPODFJHch5c0Dy0uXpH1AtfGhG/3XctDNKHI8XVhzfWvu3QxtqhB35xsc+KoBOpyprVwOkMb7ACFGxgM0gPOHz71XdGxDfs27/5bTnH90TE10fERT2XRT/ujog/zyn+tNrK7zq8dd5v5Hdc+bCDhQWkluqB3gxg1HKdwooXADp0aGPt3RHx7pQiLV932xdUJxYuqyMuSRF7T70up7yUczoZoFLOVU5391HvmeSUPxk51X29f0r5njqn4329/6lyyvfmnO7b7rrzIj5R7TqWP/yLz9zRZykgAQAwN3KOHHHtRyLiI33XwjDZg8T4Dait6KgN5PwKgFnTpAE4lYDUVrb5hUIImgAAAhLjp/tQR4Y0g5Qqn2kXcgznM4UB8z0CnEpAaimFmyqFGNIMUq4N7IGZscSuI7VVNZRBQGL8zDaUx2cKAPREQAKGxwwSANATAakt5yD1b0hLwwBgXvk+phACEgAAwISA1JY23/3TfQgAgI4ISAAAtGfbAYUQkBi9pOMZAAAdEZBa87SEQgzpoFgARic5B4lCCEgAAAATAhLjp60oAAAdEZDaMjiHzlWV5X4Ao6NJA4UQkBi/LKQCANANAakt5yABAEAxBCRGrw5tvgGgb7VtBxRCQGL8LLEDAKAjAlJLyTlIAABQDAEJAID27MumEAJSW5Z39U9b0eLUtX9XAEA/BCRgcJyDBAD0RUBi9JKuOQDQO/uyKYWA1JKWlgAAUA4BifGzXwVOL2VLFQGgIQEJAID2rKqhEAJSa5WbQd+SzwAAgG4ISAAAABMCUkvJOUj9M6UPZ6BdOuxEspcVOIWABABAex4aUwgBidHz5A8AgK4ISG1Z3gUAAMUQkBg/Xezg9FLYgwTMkO9jyiAgMX5m8bqRa4NpAGDuCUhtpWRwDgAAhRCQAABoTxc7CiEgtVWHm0HPal3siuMz7UbO2bJJ2Al7WbuRnL1GGQQkAADaM4NEIQQkxs+TPwDoXdI0iUIISG25GQDAuFW173LgswQkxk9IhdOqwn4AAGhKQGpLm29KMaClilVlYA8wPsP5HoE2BCRGL+l4BgBARwQkAACACQGJ8RvQ0jAAmFvafFMIAaml5KBYAAAohoDUliYN/dOeFU4rp9DsAgAaEpAAAGitduwGhRCQGL+tBTdkAAA6ISC15GkJAACUQ0Bi9JI9SADQP/uyKYSA1FbWxQ66Vjv8txtZkwYAaEpAYvy2DKZLU1XJwB4A6IWABABAa86GpBQCEqOXqsoNuQu5NmtTHp8pADQkILWUwoZEABg1x0UApxCQGL0TWq0D0MbCltnWLuhiRyEEpLYMzoHB0uwCAJoSkAAAACYEJEYvOTMHziCbQQKAhgSktrLBOQCAbQeUQkBi/JI23wAAdENAAiIioqps6C+QzxQAGhKQGL1kSh/OxL8NAGhIQAIoljNJAKApAYnx08UOzsC/DdiRrQX/VoDPEpBa0yAAGCr7yoAZ8sCSQghIAMVyDhIANCUgMX7afMMZJPd4AGjIl2dbDooFBssMEgA0JSC1lazx75s23x3J/i4D0IIVHRRCQGrLDBKlEDRLtNB3AQAwNgJSS2YvBqCufQZweo/puwAAGBsBCaBcj7l0/+YlfRcBAGMiIDF+yQF/cAbV+XX+F30XAQBjIiABFCylan/fNQDzwbYDSiEgMXqpsgcJziRH/o59V7/1H/VdBwyZ7xHgVAJSa1paAoO2kBcXbu27CAAYCwEJoHQ5nr6ytvm+y6677cl9lwIAQ7fYdwGj5xyk3p3YWshVqvsuA4buGxa2qv+2srp5W0R+25Gj+343/+7Xnei7KAAYGgEJiIiIuhb258BSpLg+Il2/vHzo/pW1zSMR8bcRcc803ixF+mSOPOWnF/lTKWJruu8RkSM+FVN7n+pYRHxmOq99ejnXd0fqcIl4yjnV+a+PnzjxZ3/3jmfe1dnrMi6+RyiEgMTo6ZoD5yKdFxGPn/xvKnLM4p9mmsm7TNfs/wtSSt2+b46IlGJpaSlWVjcP5xTvqlL88uHNtd/Js/mLANAZe5CAwTGbBSOWYiVFPDvn+K3l1c33r+zfvD497fc8kAVGQ0BqqTZ70T/tWYtTVSn1XQPQiSdFjjcvLx9+/8rq7f+q72IAdkJAAgCm7Qsj1b+2vLrxzX0XwhR1ua8NeiQgMXppa8ENGWAEUpVutdwOGDoBqa2UDM4BYCdyfNHK8pHv7bsMgLMRkACAmcmRX5RSDGuf4Zb9xMDnCEiMnyYNAGPy2JVrbv/KvosAOBMBifHz5K842nxD2XJVX9N3DQBnIiC1lGoH4EHXtPmGwuX4ur5LoHuOPqEUAhJwkvaswOx8oW52wFAJSG3pYgcATe1a3nfoir6LoGPZDBJlEJAYvZScgwQwNmkrr/RdA8DpCEhtWW8LAI3lWLyw7xoATkdAYvy0+QYYn5QFJGCQBKS2tCOGzmnzDXMg5Uf0XQLdSqHZD2UQkIDB0eYbypcjFvqu4QGpMrAHPkdAYvTSliV2pTGDBHOgzsYgwCC5ObXl7BgAaKxKlTFIabT5phBuTgDAzOWcLaUFBklAYvwq5yABjE1y0DowUAJSS8k5SADQWJ3ruu8a6JgxEYUQkNqymbx3J9yQy2NvH5RvQP/OfY8ApxKQAAAAJgSktgb0BAwAAGhHQGL0nIMEAAOg8QaFEJAAgNnLtTbfwCAJSC3VNnb2T5vv8vh3BcWrqjSYgJQ0XAJOISABADNXCyXlqcNnShEEpLayGzwANKbJETBQAhKjlypNGgAA6IaAxPidWBSQADh3ZrOAUwhILaVwUwWAxjRjAQZKQAIGR0cpgPFJzkGiEAISo2cPUnnygNr/AtPhQQgwVAISAADAhIDUljbf0DlPlgGAvghIjJ8ldt0Y0IZpS+wAgL4ISAAAtDegB23QhoDE6KWtBTdkAAA6ISC15WkJAAAUQ0Bi/OxB6oTGCMAs2WsIDJWABAAAMCEgMXrHzSB1wtNcAAABCZiwxA4AQECiAOnEooE9wNjk4cxaJw2XgFMISG2l5Kbas7x4YjBfsqOWKn+XgbmUBxTWxqy2EoFCCEht1eFm0DMzSAAjlGuhBBgkAQkAmGuW2HXESgQKISC1lCyxA4BxszQMOIWAxPhp810cHfUAgL4ISADAfLM0rBuWKlIIAaktNwPonsEKANATAaklLS37l05YYgfAudOkATiVgAQMj8EKMEPOQeqG/aOUQkBqy1Ig6J7BCgDQEwGJ0UuVg2I7MaRDG80gAQA9EZAYv4Utg+nSmEECZsgeJOBUAhJwkuWiwJyyBwk4lYDE6KXjuth1YkhPUIe03A+YiqoSSoBhEpDaMpDrXVqwB6k4ZrOgfGZtgIESkIDh8eABilfnuu67hgfYg9SNbFaQQghIbVVpMDd4aGVI51eYQQJmqbZUuwvOQaIUAhIwPJ7mQvkG9IAxLQynFqB/AlJLnpZA91LEVt81APOj3lo40XcNRRhQ6IU2BKS2LAWiFKkazP0g1/lY3zUA05XycAbTVT5xvO8agOEYzIAIzlll7XgnBrSsLafq/r5rAOZHvZgEJOCzBKS2BjSonFfp+JLPoAt1/njfJTygSmaQoHgD+v7cSvkTfddQBNsOKISA1JabAaXI8Xd9l/CAtFV/tO8agPlxwcLW0b5rKEGy7YD/v703D5Ojuu7+v+dWz4ZWYgRmMQiBCDZC0nRV9wxCOB5kE0O8EOdFNottcAIktvN6SfKz/SZekngheWM7Mdg/2/GCbSSI7BhsA3bwIhJgmOmu7tGIAdsIxCbMIkDLSJqZ7q573j+mR2jpulXdXdXL6HyeZ55H6nvrnjM9VbfuufcsswQxkIS2p0gsAf1RkMCLzVZhBqWsh5utgyAIRw6PffvKSQC7mq2HIAitgRhIQtujJONZFHjFqa6nm63EDE+su2wH0DoGmyAIsbCt2QocBOPxZqsgCEJrIAZSnShSzzVbhyMdSwykuiHGr5/fcMmeZutxCA82WwFBEGKDVaHTbbYSh5BrtgLtjmb9bLN1EIQoEAOpTqaKhXyzdTjSKXQUpX5FvRD9stkqHA79qtkaCIIQG488teGSlkkMAwCkkG22Du1OqVTa1GwdBCEKxECqj8ee33Cl7JY0mR7L2w05RaoLUvq2ZutwKAT8qNk6CIIQF3RLszU4DKahZqvQzhCwVdZEwmxBAXis2Uq0LYT/aLYKwkxwLT3abD3amN9uO/3R/2m2Eoeybd1lmwDIbqQgzD5KRPz1ZitxKOU555Fm69G2sKyJhNmDAlrRtaYtKFHJ+3azlRD2c1ezFWhXiPA5/uQnW6ai/YEw6AvN1kEQhKjh72y76fLWStBQhkAbmq1Dm1IiIlkTCbMGpVtwF6ctYHx72y3vklTErYJW32q2Cm1K7umlj3yv2Ur4ccLcPbcAJM+ZIMwenklA/U2zlfBDe95NzdahHSHwt55ad9mWZushCFGhnrnp8iyBrgKwr9nKtBEvlKA+3mwlhJd5+uZ3jMh9XDWTinBVq54eAYD7tWuKCvhQs/UQBCESPAX6s3Ia/5bkd7e889cE/kSz9WgzXiiyJd+ZMKtQALBt3WU3alZpgDc3W6F2gJjf99z6SyW9d4sh93G10PufuunyB5qtRRBPrbvsThD+EIzRZusiCELtMPivnlp32Z3N1iOIbeuu+EeAPgqAm61LOyBrImE2QswvP/80cHfixBOf/gAzPgVgbtO0al2eAPMHn15/Rctl/BJe5oD7+JMA5jVbnxaEGfjM79Zd3lanoESgEy9f90fM+EsAAwA6mq2TIAihKAD0T0+vu6ytThlOuOymNxLRjQCOa7YuLYqsiYRZy0EG0gynXnXjwmKh8yomvhqMVzdBr1biGQJ+wYQ7jp+z94fu164pNlshIRynXnXjwqlix5UEXCP38XQKVjB+zBZ/9envXfHbZutTD6dcvv7oEvGFYH4zQL0ATgXQ2Wy9BEHYTwHAAwT+EYi+3apJGYIozzVXg/EXABY3W58W4CUA9zDo+yfM3bNB1kTCbKWigXQgi9950/ElrVYzsCQ2JYDdupY6Nsz7FNFU/QrwBDNN7h8WKMDztpNKbH96/aUv1D2+0HQWv/Om40tM52pgJTEnqrqYaR+iuM8iQ+9lqEJQL2KtAbzAFl70POvh2ewCQQN3J1513HOnaMs7mYAeBuYwMJ8Bq9m6CcKRAgG7GShYTI8cN2/817Nt8fzKy9cvUcyriOjEGMVMMDAZ3O1wFBBJbBcrPQWt9h30/yK/wLr7hWc2XLI9ChmC0OoEGkiCIAiCIAiCIAhHCqrZCgiCIAiCIAiCILQKYiAJgiAIgiAIgiCUEQNJEARBEARBEAShjBhIgiAIgiAIgiAIZcRAEgRBEARBEARBKCMGkiAIgiAIgiAIQhkxkARBEARBEARBEMqIgSQIgiAIgiAIglBGDCRBEARBEARBEIQyYiAJgiAIgiAIgiCUEQNJEARBEARBEAShjBhIgiAIgiAIgiAIZcRAEgRBEARBEARBKCMGkiAIgiAIgiAIQhkxkARBEARBEARBEMqIgSQIgiAIgiAIglBGDCRBEARBEARBEIQyYiAJgiAIgiAIgiCUEQNJEARBEARBEAShjBhIgiAIgiAIgiAIZcRAEgRBEARBEARBKCMGkiAIgiAIgiAIQhkxkARBEARBEARBEMqIgSQIgiAIgiAIglBGDCRBEARBEARBEIQyYiAJgiAIgiAIgiCUEQNJEARBEARBEAShjBhIgiAIgiAIgiAIZcRAEgRBEARBEARBKCMGkiAIgiAIgiAIQhkxkARBEARBEARBEMqIgSQIgiAIgiAIglBGDCRBEARBEARBEIQyYiAJgiAIgiAIgiCUSTRK0NKlS7sWLFjwe7t27Xppy5YtU42QOTAw0L13795FWutFAIpW0dpDPfTi0NDQ7qhknNPbu7hECZsUn8TMJxDUwv2NxDuYebciep6Zt7FSj5166qmPbNiwwYtKvlAdRES2bb8qwbyoUrv2rEl0YuLQz7smJ3cc+tkOrfeOjY0VqtWh0c9CM+QdffTRpyutTywBL3R2dv56cHDwsO+0XohIrT777AUz/7/3gQd2MbOuZazzli8/2ps7l0ql0sLg3v4opV6IYn5Zu3at9cQTTyxk5gWe5x098zkRFQA86brurnplVGLFihXHdlnW2Rp4DYCjiKmbibQi3gbgmYTnPTC4adPTcciuQsc5nZ2dqxTz/r+VJtpb/m52Kc/TnEiMM3NJKbV3aGjouSaqGztERL29vcckEol5M59NTU3tYeZiPeOeeeaZu1v9XRXF3EZEyrbtM5RWJyIBArDL0trzlNoZsboz7Fu8eD9ybMEAACAASURBVPH2sN9tf3///I59+6yp7u6jtdaUYF5YBHYlSgmvkyd27tC6ODY2ticmXQXhiIXStsOVGobdLEUhoC+V+hCYPgrwsQeIfR7E1w1ns1+MQsYM6d50ihQPMLAa4HMAHOPT9RlijDCQU8R3Dufzw8xc8XuohOM4SaXxZyD8MYBXVqnmBMBjAG0CYZMGbndd98kqx6iJvpV9Z7Dl/bZSmyac7rruo3WNn0yew6QGK7Wxor5sNpupZ/wZUrb9NwT650M/p5K1ZHh0+LFK15xj26/2mL4CQj+A7ij0OJAS6xPz+fzvTH36U6mLmPFVAK96+dN4ngUASNvpNxPxVwCc1Ah5AJBKpT6kGP/AwNwDPtYARsF0fTafvbGaZ81Xjp36IBEO+x2Y8KfZbPZbYcY4p/ecxdryvgnwaxHpZhHtAPSnh133C2GvGBgY6N6za9dagP4XAX0MHBtwyTMA/gfgu0rMPxwZGal5MXdOb+/iklKXg+lSgM8K6s/A02Dcp4Afa4tuj8tYq0Q6nV4FjVsPfp8EUiTgdwAeZPAoE23cuXPn/4RdUNu2/WrFeKhSmwKvHM7nR6vQJTTpXuevmfj/Vmh6MZvPHWPb9vkK+BqA0wBE8r6uBkvrM4ZGRrY0Wi4Q3VyattNXEulPA3RiDGqaYADPEei6ITfzb4c2Llu2rLOnq+sTBHongJNDjjkBwlPMeJoYj0Ihr4E8gLzrunUZy4JwJBKrix0RKTC+cPjLjI8F4wtr16616pWRTCZP6Es6H0nbzq+hdIbB/wTwm+FvHAHA8Uy4CISPa9D9qaT9RNpOfdFxnDPDyFSMHAh/geqNIwDoASgF4GowvqwYW/ps+3rHcRYEXlkn2ipd5Ndmabyh3vGZ1If82pTW7693fABwHOd4An22YmOHt8bvOg/4BgivQwzGEQAkiOygPppxHQ56oQMzzwIRRf4sEvHncJBxFK88ACCmjx5iHAHT80wviL+VcpzvLlu2rLNuOYS/qywfnwn7u2mr9EWAz0fkJ+l8NECfD9vbtu037dm1+zcAfQfAm0MYRwBwPIC3A/TNBKlnU0lnveM4fdVo2dfX94p0MvlvJVIPg/HpMMYRABBwIhHWMuEm0vx8utf5WDVy60LzV6s0jgCgg4FTGLgIoI8R466jFyx8sc9OfdNxnGTQxZami/3VoXdXqUso0un0Eib+jE/zwwCgmP8BwOlognEEACXLurAZcoHo5lIi/mwTjCNg+m/2Sgb/a6XGo7q7P0qgv0V44wgAesA4g4ABEP4MjK8oxpBibE/bzoY+x7ls6dKlXZFoLwhHALEaSGeddZZx4bF169aaH1bHcY5JO851CVKPMuE6AKGMGx9eBfAHFeOhtO38JJ1Mv76Osaqlk0HvV4wH073pVbFKIkr7NTHhgnqGXrVy5YkA/th3fNDac1esqHZhcxjE/Hb4LGhZo+Jix3GcowA6p17ZRr2gjAbS2rVrLZp2W6rI6173ukifxYGBgQQMz0TU8oDpBXfg4pVxxVE9PTfXY6D19/fPB/AKn+ZXptPpU8KNRH9Qqw5RkU46n1CMHwMIqXNFugC+lDQPhb3AWelcoIvFhxn0vwF01CG7k4k/m7btq+oYIxTnLV9+NIBlEQ03hwnvUaBcykndmk6nl/h1ZIKv4UmEyx3Hqef7q4guld4NoOJGAhN+Wxb++1HLrQZifm0z5EY7l3IuCp2ihhnviHC4BQAuYca6o+cveDJl259OJpMVXcwFQXiZWA2k7u5uo4HUUyjUZCClHeejivE4GB9BtCcCBOBNIP3zCMcMy4lQ+ldp235zXAKIabGh+fzyoromvETiQph34rtKic6rax1/BgK93dD4qkofWyXrJMS8y8pgo4H0xBNPLARgOjGN9BRjfHy8ofIAwPO8OaE6Mt6Wsu3Q7mcV5Bh3VbnIvovdGaaNZj46qF+cOMnkZxn892jwCUDKtt9Hiu8A6PciG5RxXVynkjNMdHbOQwzfFQEXQ/ODqVTK55Sbfd9TDByrmN8YvU7Kd4FMGjNubT1Ry60GAhY3Q26kcylRyxlIZQMwLuP3WAL9bYLUw31J58NxGPeCMFuI9YWWSCSME9VUd3fVBlJfMvWnYHwOQLjFWHvRBdCGVCrVG8/w/i96AAv27tzre8IUODIHu+gx8Z/XY4Q5jnMMAF8diX3uZ1VsxELCMTUys/FFND4+HumzqJQyfs9Ry6se+kDaTte6sDS6oxJ5gYkzat2ciYp0r301gRrnmrZfbu9KMG5AxAYyA8c6y52lUY55KEqpumPXDHQT4/q0k15HRAcbYYzfmC5kRqRudqmVqTTAZ/i1k8ID5X82LMlSJTTM30tcRDqXau3WrVDElA3AuOfnhUz4vAJG+pNJ39M4QTiSifUh7Ni3LyjGqKrdi3Qy/Qc8HXQ+m+lWmjeU3Ygihp43tiquyc2uvHPsG/+zH8ZJe3btqTnWiTS9AYZ7Viu84NMU5y7ZXgCPM/jTpk5BBkt3d3fd8XiHYPydY5BX/QKW9OdrMZiV1kbjhi1rMmiMWjZnoqK/t3cpE5ozjxHF5grHxPW4CdZLYTo5BrYe8rO9umH4sj7bPniOIt4YcNGby5s30UD6UkMra6L7y/9u6u4/KfX/N0NulHMpW1bLGUjMfFTIruMAdpR/9tUmDGdpUpm0bV9Z0/WCMIuJdQdqbyKRiGoGdxzneEW4FT5+2RWYxMuTxwSmA8dPADDPdFG9ZHKu0QUklUy9h4hvgME9goHTuVD6PIC6XdIOgngUDN9de4Z+A4BPVTtsamXKBrRfTMjBKhCvBfDTamUAAIH/0NyOzZVlUsK0cj/wb0ZEtPrss0Olek684hXjGzduLIXpS1OUYMNre+/evZEaLESUmE6U1Bh5tUGv2bN7z7UAvlzNVVqpbjKbYoGLBa11p+XjrRVVBk8/PLI+DHDQ3LsTjG8z6Jes+FmLeQ4DJ4IpTYR+BnoBVGXkrV271mLQZaE6E0aY+YfQKqOU1h5RNzGvImANQDYquDhZSr9UjT5RknGzvt/FqlWrevSkPpUt7xzWeAsIF8JgXDDoagB3zfy/4Hm/7LQSe3B48pEZOsuxkVXdx5UgIuX0Jtf664ZHXNd9AQCyuVzVz3DKtn2fnGwu15RkD9US5VyayWSeRZVum31Oyvc7jGLuoClKmGYH0xqj7DJ3smJeQcAfMehPEHDiDmAOQN9O2/a+TC63oSalBWEWEquB1ON5VFLRrMMU0z+HihlgvATi92Zyuf+o1Lxq1aqe0kRpGSteTeDzALwW/gHfkZPNZ7/Vl0zmmNRtMPlwE67qt+3rh3K5iov+WiDmOxj0EUOPtOM4C6pO22vpNYa1+KFcvHTp0j+vtmYFEVEqaRtPuDRwb6XPuYqd1nIK6sNqHtVL0SpafgtyALAsK9Jn0SqVLM8QEhK1vFohwqfOW758/T2bN4f+zpVWXUz+5Y48zwtzbzVl9336xEyvNa3JmLEBFr1vZiF8CDcD0/Vf5s+f32sBy5lpGYHPYuBsk+zHt2xJglTQSceTzPTXuZHcDyqkY78dmE6SoaemlkGpZQfKniiV6ioTEBflGlwPlX++2ZfsW8FK3wOfzTIGLly1alXPTO2u0dHRvamkfSuAd/oKmXazq9tASiaT54Jxgl+7AlcsoxAFRERRpOCPm0bPpY2mlCgZfz8T5XTej5Z/frh69eoPTk1M/RmB/xaB6xz68rkrVtx93+io0dNEEI4UYp1Ipjo6tDGSspAIVdSxz7bPA+jy4J70M634Pa6be8avR/mlly3/fNFxnA4LuIAZlwK4GA2IbRrO50fT6fQaePoewPdlaGnQFxHGdS0kp5x22uDjj2x9CQS/4OyExXw+gFurGZcZA1VM5wsXzlv4hwB+XI2M1MrUCkAfb+iyY968eRUDbrVSFunWfu93FYuRursWiEybrJHLA6Zd7LSuuq7kMYWOro8D+HDoK8jrMRkYlmXVVCy2EYyPjy8nY3IEutkdcQNPecobDEPln1CwUgOmjQwCr9dKvTdog6RcCHew/NMwary/DmM4PzyaSqWuIZ42NiswRxcK5+GAUyTWdBMp9jeQgFR/MvmaoXy+Yr2ksBDwJ8YOTHfXM36w+Cq2ulqUOOa2RkJEHNVf4d577x0H8MXzli+/caqj8wsArjR0P6ZodXwRQIi1liDMfmKdSDo7O40LFaUmQ04D9H8RdAzO+Mbi0059k+u6vsZRJVzXLQ677h2ZnHvFVKl4HBGuBih2v+RMJrNVgS8EYNrtPn/FihWRGWwbNmzwQHyXqQ8zVRUjtGzZsk4Czq3mmrKbXXVYOiCgn+7yc3cjL8AhqwEExedMdXRE6vLWaHkAoCZrC6Jnwvv7+vp8g9IPRRMZM1cSNf/v7YdiNgVEFz3ov45LNjGvMMnO5POXN7Loa7UUCoXIDN9sNnsLgE1+7RoHp/ZecsaSX2K6OK8vXp01kYiIiMlkIGlV6qjJPTkMcaT+j4NmzG3tzj2bN+/I5nPvAeM7xo6ES+NLEiUI7UW8aVknJowvtMkQQd39tm0z/OtQlLk9O5K7dsOGDXVtL46Oju4ddt1vZHLZVD3jhGUol9tMoC+Z+hylVNT1Cu4wtlZZD2lOV1cfqj91e1O16UWZzXoR406/Nq0i2HaOmagzdGmtjRsKcWQEKyRqXsB2sNafC9uZ2JiNEcwceKAZc0Y0XxgwPM/0XD6f/118smmxoTlULF0zif5vRrf7NjEdZMhOv1vI78SpPByuqKf4eWrlynMAPsm/B7tDDww9V+v4RwrNerajIi79mZn3FSavAcNUxoRI4+/jkC8I7UasBlJXV1fACVLwRMDAewO6jJdYX8vMLetWY4JK1pdgcGvw0BGpgZTo6roT5sXQaaaiiYeigfNrUGOB0jp04dYVK1bMIcBURNfzFPsaSK18ohAXzVgk1CWT8ba+ZF+oe4IiqP8SZEDGBWlD8BQ4VHKQOjjR0NbjOI7JhbXpRH1PM7HJU+DwWluabgoY8oTHtjxWy3w4rQ/U/zK1E5R5c6tOxsfH2yJJw2wnzrlpbGysQB3WpQAMcUb8pmrWAIIwW4nVQOrZs6cuo2XFihVzGGSuKE34hzh3XeNmaHRoGwyuHqBSpOm+BwcHXwJwv6kPlXToUyQCBmpSRKmLwnbttDrXwJSxizHkE9DeMgS99DqmpiI18BstLwQlBozpkrXSnwkzEEMZT5CaZfyEQRNMNZrmOknnL+KTzmbXRGbzXDvL0Fo/6NtIdNg9lt2UHQFhzDQmKx0uS+Bh4ohAAfFHuvSftYwdlsnJyZZ9bg6kBee2tmJ4ePhFInzI0IXI43c1TCFBaFFiNZCm5s+vy7WpO5EYAGCqCfDijl27rq9HRivA8D/9ACcmIhdI5gQJHNLNbtWqVT0A9deiAjMuDNuXyJzeGyrAbbANmLCshp74xCHPtMNPwCQIHzddT8BAn22b/9YAAG1c6LcyBG3M1kfgf0v1pmquFRaA0a016oKnURP1CVKX12WYWyu7aTLTOtOYxHjb9LxYHene3n5UOrV6md9kNm3yN+iE/TR6Lm1Hhl13PflkfQUABr+9kfoIQisSt4udcaIKsdNjXiwx1lWbLroVIaKdfm2c4MgNJI/5toAuA2EKeJYmS+eiylosMxBw9qqVK00uPwdivA90Of1wO9Nol7g45FkTE75jMjCZzWbvA+i/TGMwWZ8hIuMOMbM5SUMrxyAQc1Aq7A6Qvi2dTL4+BulG2QSsSCeTN0aZGCZKTPdXLRDvLRqaK74bO3XpewBM7635hYnCm6rVRQe41zFofbVjVsvU1FRbnCAF0crPfyvB4BsMzWf2r+g3xMMJwuwnVgNp+/btRgMoeKeHzJnLWH27aqVaEGJ+tV+b53mRF1/M5XKPgPCwocvCvTv3poPGIWJ/9zrCYwB+a7q8qDoCTwts2z4dwGmGLk+5rvtA0DitjjX7dz0nAUCx97cwphJm23Eco6sRkTlJQyvHnCWKPWFqBR3FoNtTyeR7opRNTDcG9WHQuzutRDa1MhX4/Lc7pe5uQ8wXjVf6dHDTpqdB+G/TuKS4Kje76Q0BNt/zHsVewPOkk06aFQbSETCXRsK+qalbwfBN+sGWV3M8nSDMBmI1kBYtWlRzkoZly5Z1MnC64fKJzEjGP3anraDX+jRsz+fzT8QikgPc7JQX6ObDYN8JlJg3MvhHpusJHFjjSQWdIgI/CRzD81reJ73RL/WgFPy1YNrwoLKBNJTL5UDmEz9ifMZ0gkkBJ0itzODY4EsAfhmiaxdA30wlnW/09/dHEoeoLdwI4NkQXV8NpQdTvfYXopIdBVG7TmmtTzG0PunbxGxO1sC4aNWyVYZaVwdj23YagK8uBOSyo1nTZlMkzJYTJDGQwjE2NlYgYl/3dCY2Fp4WhNlOrAbS3XffbYxBMk1k8+fPnxswfOQnK80glUxd6GcIMpCJTTBro2FBMMchrV69eh4Ax3d4YCNpy2ggAVgT5E5FUEYDicg/vXc7Ye3Z09CXehzyTBseXDaQAABEn4C5IOUZe/bs8Q0S5oATpFaHFQWcoh3U+0+9QvGBdK99ORHVNV+7rruPmYxxYAdggfAhr1B4OGXb19aTvjoqonadIqY/9m0j8t2Ysjo7fwBgn2HozkJH4W2hFWE2utdpcOzudQBQKBRmh4HU4Lm0rWFlOA2l32+cIoLQevju0vY5qbonmbTtu34OpFQqBfnBG4Od24Hzli8/mjo6v+Lfg38Rl+zFp59+3+OPbn0BwDGVe1DacZwFfoUjp6amziPD/ZPwvI33j+afSSXtZwG8srIIHOc4zllA5cxQy5Yt6zyqq/t1hl9jwgvIjAYAnlLatLRK207N93om50ayqNjT4F3PRssDvWwgZTKZTWk79UNT1i5ifHJgYGD9xo0bJw9rZOoOKBtdF/XMfR74hKBi1a7rDqeS9tcBXBty2JOZcJPTm/xwOpn8SCafr3leyG3KfdNemXwrEULGydBxYHz1sUe2/qVt23+Vy+WMMWTtguM4xyvQpX7trOlxv7ahoaHdqaT9EwC+gexE+nIA3wjSg4jI6bX/xFTpQQO3BI0jvEzD57aIUUoxGuT0UIT33wm/fXJiSfUtHNE0tXK26QTJ8rygE6S2NpDWrl1rTXV0fhfAYp8uEx1dXd+NS/6GDRs8ApuqsicsNrjQaUP8EfCbwU2bnmZmDQ5IoKC1bzD63O7ucwHM87+Yf+66rmknF0BrxKTMhiQMQZieZ2IcbOhY9EmYg91P3jc+XrEGGhNa1sXOYrbD9GNFf0nVnwwkGfTzVK99T6q39621nCgxM3dP9FwGIF/llWcpxs9Stv1ftm0vr1ZuFETpOqWYrgfgF4M03lWc/IHpemb6nlkCvbZ/xYrAIPd0b68N8KmGLnc3qoxFsVhsixOk2Z6EoZFlCowu/IzQbqKCMBtproFkOArXzMYTJKL2dbHr6+t7xeOPbr0D8N/FZcL6cs2i+CAyutlpIpObnb+BxAec6hAbY50A8jWQtDa7+TFUKPc6KjXfQAoiqKhyO8gzLWD5EAMpk8k8yMB/mMZj0McqxcBQBAZSbIsspUIdm7uuW8zk85cT4VoA1WWqJKwGqducZPKhVDL5p9W6v937m3vHWdH5BPyqKrkAwLhAMUZSvfa3whgAUVKvgUREqs9xLk7bqfvMNYfoX+/ZvNm8AWfhLjIW24TSqiOwrlRgcVhGUHHayPA8ry0MpCAaPZdGTSMNwAAX97gLVwtCS9NUA8l0FK4ty3iCxIy2TO+dSqXewCUvD3Pygd0M/EvcuuydnPwp4P89EqNioobe3t6FAFb6Xceg/QuvHbt33wWgYkYoACDgtY7jVK7PomDKYshWyQpV/0ip1k/SMBsw+v4TH+YqZ3nWJwGUDEMeo0ulvzrsU27dEySEPEGaIZPLfZ20lyJgqHpZ+H2AvvHYlkfvS69ceVY1l7quu2vx6addAManABzuxmhGgXCVZyV+6ySTHwlTEiAKao0tSaVSv5920n+dsp2HGHQrCKsM3R/T0NcFjem6blETbjb14RDZ7BjG+KNJbdGtQWNExdy5c5u6HhCagslAMtZNE4TZTkNebLVAWnebn11qSd3TtrMHwHMAP0tE25nxPEAM4CgG2wT4pvQuM8mK3upms7+JW9exsbE96aTzP6DKhhCA09Lp9JJMJrP1wA87Let1zPDbtWZWfPfMf7Zs2TKVtp1fAPALiJ5HRGkA9x34YX9//3FgrDCoPzI0OrTN0L6fEhG3ulNGIpFoqIaNlkegwxbhQyNDW9JOeh3AvgVKCfTh/v7+rwwNDR2Qjpa6Q+c4aDhUdeBlZtOmB4noXCeZfC8YnwFQXeY4Qh/Duj+VTL49m8+b3GYPYsOGDR6Av3cc5ybS+DzAb61S9aMIdN2eXbvXOo5zueu6sc5ZeyyL/VZsaSc1AWDn9A/tBDEzYw4BxxBwQsj7ZRKK3u5msoFuuwDAwPcI+IChQ69t26/O5XK/rtTsOE6SDCUMiPCTrE8MqOBPo+e2duaSSy6hxx/d6tdsTLIlCLOdlo1BIqBgvpo7o9YnIuYAWALQKma8FcDVAF8D8BUhjKMSwJdns9m741dzGgpIk00lfZibG2uDex2w2XXdFw4ag2DOZlchDskreBfAZCFXkb2uFWKQgujYubOhOsYhz3QizFT5lMLyrE/B8KwzMFeXSh875FNjFjtmDnQViuGemATwFMCfreViZtbZXO4GVnQmgID4lgoQ5gH0I2elY3o2K+K67qPZvHuxJqxBuBTkh5JUmv87tSK1rIZro6Ib08lgzgS4H4xzCFgO4ISQ1xeY8I5MJpMNKzA3nbK+YoKZGZTG2lraAICD0olHTKlUmhUudo2eS6OGphr3vtq2bZtpHWU63ReEWY/vKcywm617siQiStuOr3uTcaeHeQLmDNA+2dfaEwL2gPCOYTcXym0sKorQP05AfcmvnQkXAPjqQR+Sv4FE4MPiGjzgDjW9G1Xx1IlAawD8/UGfEZvTe2sd2kBSnqfZEM8eVSY6E9ML8sa9txstDwiIEeHDT5AA4P6R+x9PO853ALraf2T6i3Q6/aWZk0wGuuP8gx049w0MDCRKL75oSBQCeHPnTg4ODlYXR+RDOQPeu1LJ5M0AfR7BmyoH0kGKN6TT6bMzmUyYekcHkcvlfgXgV33JpKOh/g7gt8B8jL8fBo6Fpe/s7e1dPjIysrNa2WGIsb7NNih6ezaTGaz2QmZaR+DP+XYgugSHzG37rzUWh+WXdo6PNzRjYLvEIDVjbmsk3MXUsLObfft+D5bPMpDaOxGWINRLrCdIl1xyiXF8k4HkKbUnYPhja1KqNXncI6wadt2GGkfAdBYbBjYbugwcGGNw7ooVxwIw7RQfZiBNnyjRfZU6l+lbtmzZ/pizcnYuU6HaF045/fT4akQ1gZciXvwFnaKMd3dHHpdlep6JD49BmkEDn4YhFg5AJ5j/bv9YMMcgRZkFauPGjaV7Nm/eYfqJyjg6kGw+/9O5C+YvLydx2F7FpcdwyftaPbKH83k3m3cvZkXnVJnI4VUJsgxlC+ojBtepIghf1uBlmRqMIwBIeMWbYMzGyGf1J5OvOfTT1MpUL4yF0NWGLVu2NDTOtl0MpCCinksbTSNPkErUucivjbl9E2EJQhTEaiBt3bq15vEtywpKbXry0qVL27pgJAF7wLhuqlRc5rruA83ThE1udgv37tybnvlPMZF4Hfx3lUse0T2VGojZ5GbX2dPZc97Mf2zbXgmDAUzgn5bjJ9qGIIOlo6OjrV/qgHkByxWSNMzguu6TAH/TODjjXf3J/pmFptFACpMFKowbXrPZuHFjKZPLfb2j0HUmgK8h/Lb5W5LJZFjXMl9c1x3O5HNrmGkthS6rwJceuNkRJREaSEVifE951lmZbPb9frXewjA0OroNAUakx3TJoZ+x8g777OB2rKtVp9nOkTCXNgom9o2BA1D1KbQgzCZiNZB27dplHN/kKzw8PPwiAaZTpMQr5s9vx0rPzwD8Y4Cv7SwWTs7k3Y+Njo7ubapGShlTcbPy9p/mMJHBvQ45v8UGJ9RtJhkEXrP/38ym7HVgoOEnbXHT2dnZ0Jd6o+UBypgpLVHq/CzM2dQsbXn/WP63cWOkkXVEGsHg2OBL2XzuzzXhDwA8EuYaizkVlXx3xP2+VrQSoE1h+s9JJOo2zuqgANCOg3/wOwD3gfBtMF0LRacN57LvGhoZ2hKNSDbHjNHhBhIRvc1wxZO5XM504h4LPT09syKLXePntmjhrsZt3ijCmX5txDVk1hSEWUSsmeBOOukk2jce5CnnDxOeAMM3fS1PZz8zuYc1nGw+p5i5rSZo13XdVNJ+BsDxldoJuABlP3pi//gjJtzt15bJZLambWcMfu55hNe//E9j/aNSV7F4l6G9JdFak2UI5+jq6or0nmm0PADYuXMnH9Xd46eR0V1ocNPg02nH+RpApqxgb+u37X6QqvsEKej7aUVyudw9q89cnZzqmfgXEK4x9SWyHCAgOUoVuK77ZG9v70CC1J0AzjH1ZaWOB/BwVLJnMN1fmQhiZmul4Hn/2WklvgzA5+Rs2s1uKJ9/CADSvb0rQcp3c4/B69rtHdJImjG3NRKtdfUVoGuEgdX+raqiN4ggHCnE+hyOj48bxw/0FdZwTc3M5kKizSAo7qoVYWZNMBV0pbTjOAtWrVx5IgDfF7sKcDVhGN3slieTyUWrV6+eB5ChTgnfF1jEsQ3Ztm1bQ1/qccgzB9GbT5Cmu6jrABhTLGuofwRgzGA5206QDuTe39w7nh3JXQvgF6Z+DD4s7qVeRkZGdlqdHW9E0CkW0Sujlg3EmqShLsoeAMZ6RQe62QUWh/Ws9RGpVhWzJQap0XNp1DSqUOx0iAKf69Nc2lfYJydIwhFNrIv5eZOT9Y2vMBzQ4/WtFof00EMPKIBDCwAADSNJREFUVVXVvoUwveATFvP5JcsypRAuTBSLRrcQ0pbJQKIE0flTe6fOh6FAHTEZ3QErXtMGab5nAybff1MM0gyZTOZZZtxg7DR90mg8+W5kJfpmoQkfD+hS8TS4XoaGhnYzBbq4xlJgsqVjS1gFuNnhHQf82zd7HQOj2dGsMXV4XMzmjYV2olF/h99bsOAtAPwydLpjY2O1u/8IwiwgVgOpuHCh8UGfM2eOMZOWBu4NEHH00fPnX1y1YjEyf/78tjSQPKJfwRCIzUxvIBjrHw0HxVJlN2VdEAzFXWlNUHpvaOt2Y3sFSm1gIC1YsCDyrHKtJE8Zstgd1C+h/hnA7pjVaXtyudwQA4/7dmAcF5dsBdxvateaGnpvtQLuJveXAJkKV5/Zn0y+xnGcswFj3EfTkjPMlhOkRs9t7Qpr+JdWIPyggaoIQksSq4E0GXCCFOQrXM7s9luzFGWon9J4CoVCW75kXNctgvFT3w6EC0wJGsDYGCSDpx3rTQbOGpDBbZLw8PCm4apjG9rhBGnevHmR6tiMUxRTcDRzCBc7TCdnAfhfo9Nq9kImQ5KMCW7qQnv0vKmdSMeyQG3l4Htm1iD+rqmPx3QJtDa512lLl26OWLUjjqjn0kbTiLm7z7bPB/mW0iipYuI/4tZBEFqdWA2kKHakCHSLuQev6UsmjUHDjSTIKGxlGGRyszsNjFP9m1WoeikB6b6XAPBPO6phSkfe0jTaYFFFZTzJXLJkSeSLWNOGB6twJ0gAoIEvAHgxEqV8mA1ueAS8ytAcVCahdrkBGw5EFEtWzlYPvmeib8GUip3oEoD8DSTCf5fThjeFnp6ettjcmw3PbjNZtWpVD4P+zdDl1qHRoabdh4LQKsS6mO/s7DQu0sIEU5agbwLMdaWZ1KerVO0wli1bNjflON9J2w6nbYf7VvadUcs4c+fObVsDaaIw8TMAtRS9nNgxviNUQOeO3bs3okYXKraoavc6AFCe13S3R6toGQ2Su+++O1KDpWSVjL/z97///cgXGabnWWkd2kByXXcXEf45Gq0qE/T3aHVSvak3MHC0oUuodOA1QZ6pdgoU86NxiG314HvXdR+FMVENn0WAf/IM5qbWPiqVSm3x7mr0XNpoSqVSrDX+vKmpG+Bf7F1rwnVxyheEdiHWNN/FYlF1GDayw+wI5nK5R1KOs4EYlxq6nZ+ynduVZ324FhesPts+76iu7n8HH5ChTXlvwvROdlVMTU21xUumEmNjY3vStv1zgN5S3ZV8X9iq71u2bJlK287PAKytUr2dzFxTbRBtWYp0c9dW3MUllMxdopRHRImAESP/QownSJYV2kACAKuj4/pSofC/ATqxWj08zwt8BkP8PSLHcZyjyOMHAPwCxD/vKHT/anBssOpq9Y7jLCDF1xv/gqwOcpd1HOc00vp1BLWYiT3WtKmzp/O/BgcHa9gQUY7h9tlXUiqi+kLtCH8DoDXB/Q5jssT8n5GrExO9vb1LLeAiZrIVsIiJniVGTlv4WS6Xi884R+Pn0kbTrXWpZEW/jEjb9hsB+j8Aneffi77jutl8mPGIiFYtX74IAAY3b94uqemF2UasBlKP51HJ7OkTDqLPgvntMJx4EfBHbHlvTNnOTxT4ux09Pb+49957x03DppPpPwDpTwB0/qFtTBhADQbSLCi2dyuA6gyk6QQPoWGm24i4SgOJf+a6uWJ110xjaa10k2veaK1LJg2ifrlMG0j+Qzb6ZaarOEECgMHBwYm0nf4oKKAIZwU6dWeg6RP094gDrfUii2gJgGsAuqbYOVVMJZM/J6b1+4pTPwqTNSqVTF5IoOthckUF9s5dOPegZ5I05wGaz2CAASJGcXJqd6rX/k+26AbXdUMtiqZhv9gFABh0Xbem5zSIVnexA4CJQuGHPZ1d2wEsqvLSO0ZGRnbGoVNYOkOeIKV6e8+1QPcCANGMNcJgwpXl9ByxPlqNnksbTamz04MXzSEYEVEqmXwTmP4O03UjDZ2xjSz1N0Fj9jvpDzD4I2nbOa5UXpOlbUf3OannCPRPQ27G5L4nCG1DvCdIXV3GnftFixaFmgWy2exY2na+C+DKgK4WARcz6OLCxGQpbTuPAfQkoJmYCkxcAtQigI8DcBwIcwxjnRJGt0OZnJxsujtXPVAi8RMueSVUcW/ogPpHh8KK7yRGAQH1bA66hlXV6b1n8JSymn2CVCqVPNNpatQoz7O4YeUGp+nu7vb9ki1d3QkSAGTz2XUpO3UNwIYdz8MpWsXAeaXRfw8A6AAOnfA6ALqICRf1dHbtSyXt+wl0v2ZsJtB2Qmmvp9RxCjgFjD4ArwUocF5iwvqNGzce8n3TUwAfWnR7PghXkearUr32vUz8K0U0qokeUkoVtNaktD4RRKdp5qUEWgmgF4BvnSMifD/ct1E9pvurVRgbGyukeu2bQPhQNdcpal72umphUFMTIzXj2W0kWuu6fB3Xrl1rPfnooys10R+lkva7ASwJYbJOMNE7MsPDxtjPgYGBBFdOoqMAHM/gfx0YGPjyxo0bG3w+LwjRE6uBFGW1ek34oGK8AUBYl5sEgKUALwUITMD0xlbod2xNBlJ3d3es/sNxMzw8/GLKdu4JSOl9IOPz5s3LVSPDdd1dadvZCMCc0vtlPNWh7qpGRjWkbSeShVcm5/re7HNKpVLBEJLX56Tq0mHYzR4kmzlBMGRbjloeMJ09au945UMQbVV3ggRM7wT3Jfs+zIqHAIReESUSiUADyfT3qPe7maHCd2Q6WTkKwBoGryECAAZDofpwdNrmae//O/xj3gLGoQbSAe1YTaDVzAAxg7VXni2njwgozDzOGC+y3lCtxmEx3V+tBFv0ddL8QYQ8SSFgx0u7d98Zs1qB7GUOe/LnF7/SEBo9lzaankLBm+rw3ztM204J03G8+wBM4YDyHAQsYOBVAHVV4WhYJPDaTNYNdmHfvr0b3T3BfRBfFk1BaBTxJmkodRrHX7RoUehH2HXdXcz03vq1CgcTvlPLdVrrtt85UWyuCn8wvLGW3SJi3FZF7/uGA3a2TCRKJWNa4kbA8+c39L6gTnq2kfIA8/NcrYvdDMP5YZcI/1TNNYVCITAertF/DwDgRKLqeKMq8cD0Th9XrV/ELBtMfEOz3cRaAdd1f4OAWlEHogk/CBvDGScc0kAiIJYshWFpxrPbSEL8fhamE7SciOnMr/bMDwOnA+gKLwwvEfiNw7lcqARI493d3VH0EYR2IFYDSXebK0JXm0krm8/+GOArEH8hydtPXbKkKheJGUqlUttP3pxQtyHkURsR1ZR6uwj9YwChXCwZugqDrYIsy3oaAZkQ42ZycrKh8k8++eRngEanIfCno6OjJgMJADzmT4Hghu2vlApcpDf67wEAiUTihVgFED6YHcneXamJiX4Wq2zglqLnfSZOAdVsqDUbInwjbF/V5Ox1M2itQxlIDHoqbl1MNOPZbSQN2mRlAN8nz3KGc7mqXOQF4UghVgOpY9++yHeaMrncOipZK1Fl3EtoGOv3TU3+yYYNG2qahCcnJ1tmUVormUzmKYDDLEi1B9xRi4x8Pv87ArIhupasjo66iie6rlsk4LF6xqiXRt8X0/cvN/R3Nm14lEqlmg0k13WLILoSgDHpSpnnwiQJaMZzOjg4OEGEdyL6DR4m0CezudwNfh2mU1DzZQCiPuHZQ0zvy+Zzl46OjsZ6shBHavq4SHR13YJw3/VT2ZGRe+LWJwxTU1PhXOyIx2JWxchseMea2L17d5wGoAZwKyvqz+TctcOjw1W9I+ZNTgbO42H6CEI7EKuBdM/mzTtgfknU9MIbHh1+LJNz14D16wFE9XJ5HuBrsiO5K8bGxgq1DvLggw/OjsmbjEVjy/Dtrus+U6sIpjBudnzn0NDQc7XKmEFT/C5GJppzX6hG/86+z/NknS/NTCbzIBPegaCaaCE3Tpr1nGZyuZsoYfVi+t6PYMFP25jpbZm8+w9BPbP5/M2UsJYDdDNCnt4a8AB8L8H67MyI+5U6xwpL2xhI0+nTg4qcAwRaz8wtUbfnkUceCXeCRLQpbl1MzJp3rA9x/H4M/BqMf9SEMzI5923ZbDZT00CLFgXP42H6CEIbEHuaKwb7FnGtNx1nJp//ZSbnvhaWOg2EjwH0S1S/Q5oH4X2qI7E0k8v9ez06MfhTrfKyqxcqWYE1OVipuop5Ks+LXcb+cYB/j2KcmuUza4D/CkC8blYHoKBDu/lEgenZWRTBSzObzd5JYKPrKynyPUU5kAP+Htvr1ataMpnM1mwu98eakCLwl8DYWv0onCHGFROFydPcETd0PF8mk3kqm3cvg6dWMOFfAFRT1HUPCBuJ6f9Qwjo1m8+96/6RkcerVr1G2i19M6vgOUerlsleF/rdZVmWn4H0OxA+EKFOFWnGXNpImFkz6C/BqGVjcBzAFgbuYsL1xPRuKllLsjn3NZm8+4lyMeOa2bhxY4nBHwDwDA7eZNEAnmHwBySDnTBb+H82UjvVZD7otAAAAABJRU5ErkJggg=='
    if e.profile_photo and e.profile_photo.startswith('data:image'):
        photo_html = '<img src="' + e.profile_photo + '" style="width:100%;height:100%;object-fit:cover;">'
    else:
        initial = (e.first_name or e.full_name or '?')[:1].upper()
        photo_html = '<div style="font-size:60px;font-weight:900;color:#fff;line-height:1;">' + initial + '</div>'
    if e.qr_code_base64 and e.qr_code_base64.startswith('data:image'):
        qr_html = '<img src="' + e.qr_code_base64 + '" style="width:100%;height:100%;">'
    else:
        qr_html = ''
    issue_date = date.today().strftime('%d-%m-%Y')
    dept = (e.department or '').upper()
    html = (
        '<!DOCTYPE html><html><head><meta charset="UTF-8">'
        '<title>ID Card - ' + (e.employee_code or '') + '</title>'
        '<style>'
        '@page { size: 100mm 70mm; margin: 0; }'
        '* { margin:0; padding:0; box-sizing:border-box; }'
        'html, body { width:100mm; height:70mm; background:#fff; }'
        'body { display:flex; flex-direction:column; align-items:center; justify-content:flex-start; }'
        '.card { width:100mm; height:70mm; background:#fff; border:1.5px solid #111; overflow:hidden; display:flex; flex-direction:column; }'
        '.hdr { display:flex; align-items:center; padding:2.5mm 3mm; border-bottom:1.5px solid #111; position:relative; overflow:hidden; }'
        '.hdr-logo { width:14mm; flex-shrink:0; margin-right:3mm; }'
        '.hdr-logo img { width:100%; display:block; }'
        '.hdr-text h1 { font-size:8.5pt; font-weight:900; color:#111; line-height:1.2; white-space:nowrap; }'
        '.hdr-text p { font-size:4pt; color:#555; line-height:1.5; margin-top:1mm; }'
        '.corner { position:absolute; top:0; right:0; width:0; height:0; border-top:12mm solid #111; border-left:12mm solid transparent; }'
        '.id-bar { background:#111; display:flex; flex-shrink:0; }'
        '.id-cell { flex:1; padding:1.5mm 3mm; }'
        '.id-cell+.id-cell { border-left:1px solid #555; }'
        '.id-label { font-size:4.5pt; font-weight:700; color:#aaa; letter-spacing:0.5pt; text-transform:uppercase; margin-bottom:0.5mm; }'
        '.id-value { font-size:11pt; font-weight:900; color:#fff; letter-spacing:0.5pt; }'
        '.body { display:flex; align-items:center; padding:2.5mm 3mm; gap:3mm; flex:1; }'
        '.photo { width:22mm; height:22mm; background:#F5A800; flex-shrink:0; display:flex; align-items:center; justify-content:center; overflow:hidden; }'
        '.info { flex:1; }'
        '.emp-name { font-size:9pt; font-weight:900; color:#111; margin-bottom:1mm; }'
        '.emp-gender { font-size:7pt; font-weight:700; color:#333; margin-bottom:2mm; }'
        '.emp-meta { font-size:6pt; color:#333; line-height:1.8; }'
        '.emp-meta strong { font-weight:700; }'
        '.qr { width:18mm; height:18mm; flex-shrink:0; }'
        '.note { padding:1.5mm 3mm; font-size:4pt; color:#888; border-top:0.5px solid #ddd; flex-shrink:0; }'
        '.footer { background:#111; text-align:center; padding:2mm; flex-shrink:0; }'
        '.footer span { font-size:7pt; font-weight:900; color:#fff; letter-spacing:2pt; }'
        '@media screen { body { background:#e0e0e0; min-height:100vh; justify-content:center; } .card { box-shadow:0 4px 20px rgba(0,0,0,0.3); } }'
        '@media print { html,body { width:100mm; height:70mm; } .no-print { display:none!important; } }'
        '</style></head><body>'
        '<div class="card">'
          '<div class="hdr">'
            '<div class="hdr-logo"><img src="' + LOGO_SRC + '"></div>'
            '<div class="hdr-text">'
              '<h1>HCP WELLNESS PVT. LTD.</h1>'
              '<p>#8, Ozone Industrial Park, Nr. Kerala GIDC, Bhayla, Bavla, Ahmedabad<br>382220, Gujarat, India. &nbsp;www.hcpwellness.in &nbsp;|&nbsp; Email: info@hcpwellness.in</p>'
            '</div>'
            '<div class="corner"></div>'
          '</div>'
          '<div class="id-bar">'
            '<div class="id-cell"><div class="id-label">Employee ID</div><div class="id-value">' + (e.employee_code or '—') + '</div></div>'
            '<div class="id-cell"><div class="id-label">Department</div><div class="id-value">' + (dept or '—') + '</div></div>'
          '</div>'
          '<div class="body">'
            '<div class="photo">' + photo_html + '</div>'
            '<div class="info">'
              '<div class="emp-name">' + e.full_name + '</div>'
              '<div class="emp-gender">' + (e.gender or '') + '</div>'
              '<div class="emp-meta">Issue Date : <strong>' + issue_date + '</strong><br>Issue By : <strong>HR HCP</strong></div>'
            '</div>'
            '<div class="qr">' + qr_html + '</div>'
          '</div>'
          '<div class="note">This Card is System Generated, Doesn\'t Require Signature</div>'
          '<div class="footer"><span>HCP WELLNESS PVT. LTD.</span></div>'
        '</div>'
        '<div class="no-print" style="margin-top:8mm;display:flex;gap:8px;">'
          '<button onclick="window.print()" style="background:#111;color:#fff;border:none;padding:8px 20px;border-radius:6px;font-size:13px;cursor:pointer;font-weight:700;">&#128424; Print / Save PDF</button>'
          '<button onclick="window.close()" style="background:#f1f5f9;color:#111;border:1px solid #ccc;padding:8px 20px;border-radius:6px;font-size:13px;cursor:pointer;">Close</button>'
        '</div>'
        '</body></html>'
    )
    from flask import Response
    return Response(html, mimetype='text/html')


@hr.route('/employees/<int:id>/delete', methods=['POST'])
@login_required
def emp_delete(id):
    perm = get_perm('hr_employees')
    if not perm or not perm.can_delete:
        flash('Access denied.', 'error'); return redirect(url_for('hr.employees'))
    e = Employee.query.get_or_404(id)
    name = e.full_name
    e.is_deleted = True
    e.deleted_at = datetime.utcnow()
    db.session.commit()
    flash(f'Employee "{name}" moved to trash.', 'warning')
    return redirect(url_for('hr.employees'))


@hr.route('/employees/<int:id>/restore', methods=['POST'])
@login_required
def emp_restore(id):
    perm = get_perm('hr_employees')
    if not perm or not perm.can_delete:
        flash('Access denied.', 'error'); return redirect(url_for('hr.employees'))
    e = Employee.query.get_or_404(id)
    e.is_deleted = False
    e.deleted_at = None
    db.session.commit()
    flash(f'Employee "{e.full_name}" restored successfully!', 'success')
    return redirect(url_for('hr.employees', trash=1))


@hr.route('/employees/<int:id>/regenerate-qr', methods=['POST'])
@login_required
def regen_qr(id):
    """QR is now generated client-side; this just clears the stored QR so form re-generates"""
    e = Employee.query.get_or_404(id)
    return jsonify(success=True, employee_code=e.employee_code or '')




@hr.route('/employees/<int:id>/create-login', methods=['POST'])
@login_required
def emp_create_login(id):
    e = Employee.query.get_or_404(id)
    if e.user_id and User.query.get(e.user_id):
        flash('Login already exists for this employee.', 'info')
        return redirect(url_for('hr.employees'))
    uname = (e.employee_code or str(e.id)).lower().replace('-','').replace(' ','')
    if User.query.filter_by(username=uname).first():
        flash(f'Username "{uname}" already taken.', 'error')
        return redirect(url_for('hr.employees'))
    u = User(
        username  = uname,
        email     = e.email or f'{uname}@hcp.com',
        full_name = e.full_name,
        role      = 'user',
        is_active = True
    )
    u.set_password('HCP@123')
    db.session.add(u)
    db.session.flush()
    e.user_id = u.id
    db.session.commit()
    flash(f'Login created! Username: {uname}  Password: HCP@123', 'success')
    return redirect(url_for('hr.employees'))

# ══════════════════════════════════════
# CONTRACTOR
# ══════════════════════════════════════

@hr.route('/contractors')
@login_required
def contractors():
    perm = get_perm('hr_contractors')
    if not perm or not perm.can_view:
        flash('Access denied.', 'error'); return redirect(url_for('dashboard'))

    search      = request.args.get('search', '')
    ct_status_f = request.args.get('status_f', '')
    ct_supply_f = request.args.get('supply_f', '')
    show_trash  = request.args.get('trash', '') == '1'

    # Trash view: sirf deleted, Normal view: sirf non-deleted
    q = Contractor.query.filter_by(is_deleted=True) if show_trash         else Contractor.query.filter_by(is_deleted=False)

    if not show_trash:
        if search:
            q = q.filter(
                Contractor.company_name.ilike(f'%{search}%') |
                Contractor.contact_person.ilike(f'%{search}%') |
                Contractor.contract_id.ilike(f'%{search}%') |
                Contractor.contact_no.ilike(f'%{search}%')
            )
        if ct_status_f != '':
            q = q.filter_by(status=int(ct_status_f))
        if ct_supply_f:
            q = q.filter_by(supply=ct_supply_f)

    sort_by  = request.args.get('sort_by', 'created_date')
    sort_dir = request.args.get('sort_dir', 'desc')
    sort_col = getattr(Contractor, sort_by, Contractor.created_date)
    ctrs = q.order_by(sort_col.asc() if sort_dir == 'asc' else sort_col.desc()).all()

    all_supplies  = [r[0] for r in db.session.query(Contractor.supply).distinct().all() if r[0]]
    deleted_count = Contractor.query.filter_by(is_deleted=True).count()
    grid_cols     = get_grid_columns('contractors', CTR_COLS_DEFAULT, list(CTR_COLS_ALL.keys()))

    return render_template('hr/contractors/index.html',
        contractors=ctrs, perm=perm, active_page='hr_contractors',
        search=search, ct_status_f=ct_status_f, ct_supply_f=ct_supply_f,
        sort_by=sort_by, sort_dir=sort_dir,
        show_trash=show_trash, deleted_count=deleted_count,
        all_supplies=all_supplies,
        grid_cols=grid_cols, all_cols=CTR_COLS_ALL)


@hr.route('/contractors/add', methods=['GET', 'POST'])
@login_required
def contractor_add():
    perm = get_perm('hr_contractors')
    if not perm or not perm.can_add:
        flash('Access denied.', 'error'); return redirect(url_for('hr.contractors'))

    if request.method == 'POST':
        last = Contractor.query.order_by(Contractor.id.desc()).first()
        num  = (last.id + 1) if last else 1
        c = Contractor(
            company_name   = request.form.get('company_name', '').strip(),
            supply         = request.form.get('supply', '').strip(),
            pancard        = request.form.get('pancard', '').strip(),
            gstno          = request.form.get('gstno', '').strip(),
            remarks        = request.form.get('remarks', '').strip(),
            contract_id    = f"CTR-{num:04d}",
            contact_person = request.form.get('contact_person', '').strip(),
            contact_no     = request.form.get('contact_no', '').strip(),
            email_address  = request.form.get('email_address', '').strip(),
            address        = request.form.get('address', '').strip(),
            status         = 1,
            created_by     = current_user.full_name or current_user.username,
        )
        db.session.add(c)
        db.session.commit()
        flash(f'Contractor {c.contract_id} added!', 'success')
        return redirect(url_for('hr.contractors'))

    return render_template('hr/contractors/form.html',
        contractor=None, perm=perm, active_page='hr_contractors')


@hr.route('/contractors/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def contractor_edit(id):
    perm = get_perm('hr_contractors')
    if not perm or not perm.can_edit:
        flash('Access denied.', 'error'); return redirect(url_for('hr.contractors'))

    c = Contractor.query.get_or_404(id)
    if request.method == 'POST':
        c.company_name   = request.form.get('company_name', '').strip()
        c.supply         = request.form.get('supply', '').strip()
        c.pancard        = request.form.get('pancard', '').strip()
        c.gstno          = request.form.get('gstno', '').strip()
        c.remarks        = request.form.get('remarks', '').strip()
        c.contact_person = request.form.get('contact_person', '').strip()
        c.contact_no     = request.form.get('contact_no', '').strip()
        c.email_address  = request.form.get('email_address', '').strip()
        c.address        = request.form.get('address', '').strip()
        c.status         = int(request.form.get('status', 1))
        c.modified_by    = current_user.full_name or current_user.username
        c.modified_date  = datetime.utcnow()
        db.session.commit()
        flash('Contractor updated!', 'success')
        return redirect(url_for('hr.contractors'))

    return render_template('hr/contractors/form.html',
        contractor=c, perm=perm, active_page='hr_contractors')


@hr.route('/contractors/<int:id>/delete', methods=['POST'])
@login_required
def contractor_delete(id):
    perm = get_perm('hr_contractors')
    if not perm or not perm.can_delete:
        flash('Access denied.', 'error'); return redirect(url_for('hr.contractors'))
    c = Contractor.query.get_or_404(id)
    c.is_deleted = True
    c.deleted_at = datetime.utcnow()
    c.modified_by = current_user.full_name or current_user.username
    db.session.commit()
    flash(f'Contractor "{c.company_name}" moved to trash.', 'warning')
    return redirect(url_for('hr.contractors'))


@hr.route('/contractors/<int:id>/restore', methods=['POST'])
@login_required
def contractor_restore(id):
    perm = get_perm('hr_contractors')
    if not perm or not perm.can_delete:
        flash('Access denied.', 'error'); return redirect(url_for('hr.contractors'))
    c = Contractor.query.get_or_404(id)
    c.is_deleted = False
    c.deleted_at = None
    c.modified_by = current_user.full_name or current_user.username
    db.session.commit()
    flash(f'Contractor "{c.company_name}" restored successfully!', 'success')
    return redirect(url_for('hr.contractors', trash=1))



# ─────────────────────────────────────────────
# EMPLOYEE IMPORT ROUTES
# ─────────────────────────────────────────────

@hr.route('/employees/import', methods=['GET','POST'])
@login_required
def emp_import():
    perm = get_perm('hr_employees')
    if not perm or not perm.can_add:
        flash('Access denied.', 'error'); return redirect(url_for('hr.employees'))

    if request.method == 'POST':
        import openpyxl, json as _json
        from datetime import datetime as _dt

        f = request.files.get('import_file')
        if not f or not f.filename:
            flash('Please select a file!', 'warning')
            return redirect(url_for('hr.emp_import'))

        ext = f.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls'):
            flash('Only Excel (.xlsx) files allowed!', 'danger')
            return redirect(url_for('hr.emp_import'))

        added = skipped = 0
        errors = []

        def _gv(row, *keys):
            for k in keys:
                v = row.get(k)
                if v and str(v).strip() not in ('', 'None', 'nan'):
                    return str(v).strip()
            return ''

        def _pd(val):
            if not val or str(val).strip() in ('', 'None', 'nan'): return None
            for fmt in ('%d-%m-%Y','%Y-%m-%d','%d/%m/%Y','%d %b %Y','%d-%b-%Y'):
                try: return _dt.strptime(str(val).strip(), fmt).date()
                except: pass
            return None

        def _dec(val):
            try: return float(str(val).replace(',','').replace('₹','').strip()) if val else None
            except: return None

        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

            def sheet_rows(sheet_name):
                """Read a sheet by name, return list of dicts. Returns [] if sheet not found."""
                if sheet_name not in wb.sheetnames:
                    return []
                ws = wb[sheet_name]
                hdrs = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None and str(v).strip() not in ('','None') for v in row):
                        rows.append(dict(zip(hdrs, [str(v).strip() if v is not None else '' for v in row])))
                return rows

            # Try multi-sheet format first
            has_sheets = any(s in wb.sheetnames for s in ['1 - Basic Info','Basic Info','Sheet1'])

            # Read all sheets
            basic_rows   = sheet_rows('1 - Basic Info') or sheet_rows('Basic Info') or []
            prof_rows    = sheet_rows('2 - Professional') or sheet_rows('Professional') or []
            kyc_rows     = sheet_rows('3 - KYC') or sheet_rows('KYC') or []
            bank_rows    = sheet_rows('4 - Bank Details') or sheet_rows('Bank Details') or []
            sal_rows     = sheet_rows('5 - Salary') or sheet_rows('Salary') or []
            edu_rows     = sheet_rows('6 - Education') or sheet_rows('Education') or []

            # If no multi-sheet, try first sheet as basic
            if not basic_rows:
                ws = wb.active
                hdrs = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None and str(v).strip() not in ('','None') for v in row):
                        basic_rows.append(dict(zip(hdrs, [str(v).strip() if v is not None else '' for v in row])))

            # Index supplementary sheets by employee code
            def idx_by_code(rows):
                d = {}
                for r in rows:
                    code = _gv(r,'Code','Employee Code','employee_code')
                    if code: d[code.upper()] = r
                return d

            prof_idx = idx_by_code(prof_rows)
            kyc_idx  = idx_by_code(kyc_rows)
            bank_idx = idx_by_code(bank_rows)
            sal_idx  = idx_by_code(sal_rows)
            edu_idx  = idx_by_code(edu_rows)

            for i, row in enumerate(basic_rows, 2):
                emp_code = _gv(row, 'Employee Code', 'Code', 'employee_code')
                if not emp_code:
                    errors.append(f'Row {i}: Employee Code missing — skipped')
                    continue

                existing = Employee.query.filter(Employee.employee_code.ilike(emp_code)).first()

                try:
                    p  = prof_idx.get(emp_code.upper(), {})
                    k  = kyc_idx.get(emp_code.upper(), {})
                    bk = bank_idx.get(emp_code.upper(), {})
                    sl = sal_idx.get(emp_code.upper(), {})
                    ed = edu_idx.get(emp_code.upper(), {})

                    email = _gv(row,'Email','email')
                    if email:
                        email_owner = Employee.query.filter_by(email=email).first()
                        if email_owner and (not existing or email_owner.id != existing.id):
                            email = ''

                    marital = _gv(row,'Marital Status','marital_status') or 'Single'

                    if existing:
                        # ── UPDATE existing employee ──
                        e = existing
                        eid = _gv(row,'Employee ID','employee_id')
                        if eid: e.employee_id = eid                           
                        fn = _gv(row,'First Name','first_name')
                        mn = _gv(row,'Middle Name','middle_name')
                        ln = _gv(row,'Last Name','last_name')
                        if fn: e.first_name   = fn
                        if mn: e.middle_name  = mn
                        if ln: e.last_name    = ln
                        if email: e.email     = email
                        mob = _gv(row,'Mobile','mobile')
                        if mob: e.mobile      = mob
                        gen = _gv(row,'Gender','gender')
                        if gen: e.gender      = gen
                        dob = _pd(_gv(row,'DOB','Date of Birth','date_of_birth'))
                        if dob: e.date_of_birth = dob
                        bg = _gv(row,'Blood Group','blood_group')
                        if bg: e.blood_group  = bg
                        e.marital_status = marital
                        addr = _gv(row,'Address','address')
                        if addr: e.address    = addr
                        city = _gv(row,'City','city')
                        if city: e.city       = city
                        st = _gv(row,'State','state')
                        if st: e.state        = st
                        cntry = _gv(row,'Country','country')
                        if cntry: e.country   = cntry
                        zp = _gv(row,'ZIP','Zip Code','zip_code')
                        if zp: e.zip_code     = zp
                        sts = (_gv(row,'Status','status') or '').lower().replace(' ','_')
                        if sts: e.status      = sts
                        # Professional
                        dept = _gv(p,'Department','department') or _gv(row,'Department')
                        if dept: e.department = dept
                        desig = _gv(p,'Designation','designation') or _gv(row,'Designation')
                        if desig: e.designation = desig
                        et = _gv(p,'Employee Type','employee_type') or _gv(row,'Employee Type')
                        if et: e.employee_type = et
                        loc = _gv(p,'Location','location') or _gv(row,'Location')
                        if loc: e.location    = loc
                        doj = _pd(_gv(p,'DOJ','Date of Joining','date_of_joining') or _gv(row,'DOJ','Date of Joining'))
                        if doj: e.date_of_joining = doj
                        conf = _pd(_gv(p,'Confirmation','Confirmation Date'))
                        if conf: e.confirmation_date = conf
                        resign = _pd(_gv(p,'Resignation Date','resignation_date'))
                        if resign: e.resignation_date = resign
                        lwd = _pd(_gv(p,'Last Working Date','last_working_date'))
                        if lwd: e.last_working_date = lwd
                        # KYC
                        pe = _pd(_gv(k,'Passport Expiry','passport_expiry'))
                        if pe: e.passport_expiry = pe
                        dle = _pd(_gv(k,'DL Expiry','dl_expiry'))
                        if dle: e.dl_expiry = dle
                        # Bank
                        bn = _gv(bk,'Bank Name','bank_name')
                        if bn: e.bank_name   = bn
                        ban = _gv(bk,'Account Number','bank_account_number')
                        if ban: e.bank_account_number = ban
                        ifsc = (_gv(bk,'IFSC Code','IFSC','bank_ifsc') or '').upper()
                        if ifsc: e.bank_ifsc = ifsc
                        bat = _gv(bk,'Account Type','bank_account_type')
                        if bat: e.bank_account_type = bat
                        # Salary
                        ctc = _dec(_gv(sl,'CTC Annual','salary_ctc'))
                        if ctc: e.salary_ctc = ctc
                        net = _dec(_gv(sl,'Net Salary','salary_net'))
                        if net: e.salary_net = net
                        # Education - prev dates
                        pfd = _pd(_gv(ed,'Prev From','prev_from_date'))
                        if pfd: e.prev_from_date = pfd
                        ptd = _pd(_gv(ed,'Prev To','prev_to_date'))
                        if ptd: e.prev_to_date = ptd

                        # ── Phase-1: Basic — Family / Contact / Permanent Addr ────────
                        fth = _gv(row,'Father Name','father_name')
                        if fth: e.father_name = fth
                        mth = _gv(row,'Mother Name','mother_name')
                        if mth: e.mother_name = mth
                        altm = _gv(row,'Alternate Mobile','alternate_mobile')
                        if altm: e.alternate_mobile = altm
                        pem = _gv(row,'Personal Email','personal_email')
                        if pem: e.personal_email = pem
                        pa = _gv(row,'Permanent Address','permanent_address')
                        if pa: e.permanent_address = pa
                        pc = _gv(row,'Permanent City','permanent_city')
                        if pc: e.permanent_city = pc
                        ps = _gv(row,'Permanent State','permanent_state')
                        if ps: e.permanent_state = ps
                        pcn = _gv(row,'Permanent Country','permanent_country')
                        if pcn: e.permanent_country = pcn
                        pz = _gv(row,'Permanent ZIP','Permanent Zip','permanent_zip')
                        if pz: e.permanent_zip = pz
                        sac = _gv(row,'Same as Current','same_as_current_addr')
                        if sac: e.same_as_current_addr = sac.lower() == 'yes'

                        # ── Phase-1: Professional — Grade / Probation / Attendance / Leave / System / Exit extras
                        gl = _gv(p,'Grade Level','grade_level')
                        if gl: e.grade_level = gl
                        pm = _gv(p,'Probation Months','probation_period_months')
                        if pm:
                            try: e.probation_period_months = int(pm)
                            except: pass
                        ped = _pd(_gv(p,'Probation End','probation_end_date'))
                        if ped: e.probation_end_date = ped
                        ac = _gv(p,'Attendance Code','attendance_code')
                        if ac: e.attendance_code = ac
                        ot = _gv(p,'Overtime Eligible','overtime_eligible')
                        if ot: e.overtime_eligible = ot.lower() == 'yes'
                        clv = _dec(_gv(p,'CL Balance','casual_leave_balance'))
                        if clv is not None: e.casual_leave_balance = clv
                        slv = _dec(_gv(p,'SL Balance','sick_leave_balance'))
                        if slv is not None: e.sick_leave_balance = slv
                        plv = _dec(_gv(p,'PL Balance','paid_leave_balance'))
                        if plv is not None: e.paid_leave_balance = plv
                        lp = _gv(p,'Leave Policy','leave_policy')
                        if lp: e.leave_policy = lp
                        oe = _gv(p,'Official Email','official_email')
                        if oe: e.official_email = oe
                        ra = _gv(p,'Role Access','role_access')
                        if ra: e.role_access = ra
                        eid_ = _gv(p,'Exit Interview','exit_interview_done')
                        if eid_: e.exit_interview_done = eid_.lower() == 'yes'
                        ein = _gv(p,'Exit Notes','exit_interview_notes')
                        if ein: e.exit_interview_notes = ein
                        ffs = _gv(p,'FF Status','ff_settlement_status')
                        if ffs: e.ff_settlement_status = ffs
                        ffa = _dec(_gv(p,'FF Amount','ff_settlement_amount'))
                        if ffa is not None: e.ff_settlement_amount = ffa
                        ffd = _pd(_gv(p,'FF Date','ff_settlement_date'))
                        if ffd: e.ff_settlement_date = ffd

                        # ── Phase-1: KYC — PF / ESIC / TDS / Statutory
                        pfa = _gv(k,'PF Applicable','pf_applicable')
                        if pfa: e.pf_applicable = pfa.lower() == 'yes'
                        pfn = _gv(k,'PF Number','pf_number')
                        if pfn: e.pf_number = pfn
                        eps = _gv(k,'EPS Applicable','eps_applicable')
                        if eps: e.eps_applicable = eps.lower() == 'yes'
                        ppt = _gv(k,'Previous PF Transfer','previous_pf_transfer')
                        if ppt: e.previous_pf_transfer = ppt.lower() == 'yes'
                        ppn = _gv(k,'Previous PF Number','previous_pf_number')
                        if ppn: e.previous_pf_number = ppn
                        esa = _gv(k,'ESIC Applicable','esic_applicable')
                        if esa: e.esic_applicable = esa.lower() == 'yes'
                        enn = _gv(k,'ESIC Nominee','esic_nominee_name')
                        if enn: e.esic_nominee_name = enn
                        enr = _gv(k,'Nominee Relation','esic_nominee_relation')
                        if enr: e.esic_nominee_relation = enr
                        efd = _gv(k,'ESIC Family','esic_family_details')
                        if efd: e.esic_family_details = efd
                        edp = _gv(k,'Dispensary','esic_dispensary')
                        if edp: e.esic_dispensary = edp
                        apl = _gv(k,'Aadhaar PAN Linked','aadhaar_pan_linked')
                        if apl: e.aadhaar_pan_linked = apl.lower() == 'yes'
                        tr = _gv(k,'Tax Regime','tax_regime')
                        if tr: e.tax_regime = tr
                        pei = _dec(_gv(k,'Prev Employer Income','prev_employer_income'))
                        if pei is not None: e.prev_employer_income = pei
                        mtd = _dec(_gv(k,'Monthly TDS','monthly_tds'))
                        if mtd is not None: e.monthly_tds = mtd
                        idec = _gv(k,'Investment Declaration','investment_declaration')
                        if idec: e.investment_declaration = idec
                        pss = _gv(k,'Proof Status','proof_submission_status')
                        if pss: e.proof_submission_status = pss
                        pta = _gv(k,'PT Applicable','professional_tax_applicable')
                        if pta: e.professional_tax_applicable = pta.lower() == 'yes'
                        lwf = _gv(k,'LWF','labour_welfare_fund')
                        if lwf: e.labour_welfare_fund = lwf.lower() == 'yes'
                        ge = _gv(k,'Gratuity Eligible','gratuity_eligible')
                        if ge: e.gratuity_eligible = ge.lower() == 'yes'
                        be = _gv(k,'Bonus Eligible','bonus_eligible')
                        if be: e.bonus_eligible = be.lower() == 'yes'

                        # ── Phase-1: Salary — Conveyance / Bonus / Incentive / Gross
                        cv = _dec(_gv(sl,'Conveyance','salary_conveyance'))
                        if cv is not None: e.salary_conveyance = cv
                        bn = _dec(_gv(sl,'Bonus','salary_bonus'))
                        if bn is not None: e.salary_bonus = bn
                        inc = _dec(_gv(sl,'Incentive','salary_incentive'))
                        if inc is not None: e.salary_incentive = inc
                        gr = _dec(_gv(sl,'Gross','salary_gross'))
                        if gr is not None: e.salary_gross = gr

                        skipped += 1
                        errors.append(f'Row {i}: Code "{emp_code}" updated \u2705')

                    else:
                        # ── CREATE new employee ──
                        e = Employee(
                        employee_code   = emp_code,
                        employee_id     = _gv(row,'Employee ID','employee_id') or None,                                                               
                        first_name      = _gv(row,'First Name','first_name'),
                        middle_name     = _gv(row,'Middle Name','middle_name'),
                        last_name       = _gv(row,'Last Name','last_name'),
                        mobile          = _gv(row,'Mobile','mobile'),
                        email           = email,
                        gender          = _gv(row,'Gender','gender'),
                        date_of_birth   = _pd(_gv(row,'DOB','Date of Birth','date_of_birth')),
                        blood_group     = _gv(row,'Blood Group','blood_group'),
                        marital_status  = marital,
                        marriage_anniversary = _pd(_gv(row,'Marriage Anniversary','marriage_anniversary')) if marital=='Married' else None,
                        address         = _gv(row,'Address','address'),
                        city            = _gv(row,'City','city'),
                        state           = _gv(row,'State','state'),
                        country         = _gv(row,'Country','country') or 'India',
                        zip_code        = _gv(row,'ZIP','Zip Code','zip_code'),
                        linkedin        = _gv(row,'LinkedIn','linkedin'),
                        facebook        = _gv(row,'Facebook','facebook'),
                        status          = (_gv(row,'Status','status') or 'active').lower().replace(' ','_'),
                        # Professional
                        department      = _gv(p,'Department','department') or _gv(row,'Department'),
                        designation     = _gv(p,'Designation','designation') or _gv(row,'Designation'),
                        employee_type   = _gv(p,'Employee Type','employee_type') or _gv(row,'Employee Type') or 'Full Time',
                        location        = _gv(p,'Location','location') or _gv(row,'Location'),
                        pay_grade       = _gv(p,'Pay Grade','pay_grade'),
                        date_of_joining = _pd(_gv(p,'DOJ','Date of Joining','date_of_joining') or _gv(row,'DOJ','Date of Joining')),
                        confirmation_date = _pd(_gv(p,'Confirmation','Confirmation Date')),
                        shift           = _gv(p,'Shift','shift'),
                        work_hours_per_day = _dec(_gv(p,'Work Hrs','Work Hours Per Day')) or 8,
                        weekly_off      = _gv(p,'Weekly Off','weekly_off'),
                        notice_period_days = int(_gv(p,'Notice Days','Notice Period (Days)') or 30),
                        is_contractor   = _gv(p,'Contractor','Is Contractor','is_contractor').lower() == 'yes',
                        is_probation    = _gv(p,'Probation','Is Probation','is_probation').lower() == 'yes',
                        is_block        = _gv(p,'Block','Is Block','is_block').lower() == 'yes',
                        is_late         = _gv(p,'Late Mark','Is Late','is_late').lower() == 'yes',
                        rehire_eligible = _gv(p,'Rehire Eligible','rehire_eligible').lower() == 'yes',
                        remark          = _gv(p,'Remark','remark'),
                        # KYC
                        nationality     = _gv(k,'Nationality','nationality') or 'Indian',
                        religion        = _gv(k,'Religion','religion'),
                        aadhar_number   = _gv(k,'Aadhaar','Aadhaar Number','aadhar_number'),
                        pan_number      = (_gv(k,'PAN','PAN Number','pan_number') or '').upper(),
                        uan_number      = _gv(k,'UAN','UAN Number','uan_number'),
                        esic_number     = _gv(k,'ESIC','ESIC Number','esic_number'),
                        passport_number = _gv(k,'Passport No.','Passport Number','passport_number'),
                        passport_expiry = _pd(_gv(k,'Passport Expiry','passport_expiry')),
                        driving_license = _gv(k,'DL No','Driving License','driving_license'),
                        dl_expiry       = _pd(_gv(k,'DL Expiry','dl_expiry')),
                        emergency_name  = _gv(k,'Emergency Name','emergency_name'),
                        emergency_relation = _gv(k,'Emergency Relation','emergency_relation'),
                        emergency_phone = _gv(k,'Emergency Phone','emergency_phone'),
                        emergency_address = _gv(k,'Emergency Address','emergency_address'),
                        # Bank
                        bank_account_holder = _gv(bk,'Account Holder','bank_account_holder'),
                        bank_name       = _gv(bk,'Bank Name','bank_name'),
                        bank_account_number = _gv(bk,'Account Number','bank_account_number'),
                        bank_ifsc       = (_gv(bk,'IFSC Code','IFSC','bank_ifsc') or '').upper(),
                        bank_branch     = _gv(bk,'Branch','bank_branch'),
                        bank_account_type = _gv(bk,'Account Type','bank_account_type'),
                        # Salary
                        salary_ctc      = _dec(_gv(sl,'CTC Annual','salary_ctc')),
                        salary_basic    = _dec(_gv(sl,'Basic','salary_basic')),
                        salary_hra      = _dec(_gv(sl,'HRA','salary_hra')),
                        salary_da       = _dec(_gv(sl,'DA','salary_da')),
                        salary_ta       = _dec(_gv(sl,'TA','salary_ta')),
                        salary_medical_allow = _dec(_gv(sl,'Medical','salary_medical_allow')),
                        salary_special_allow = _dec(_gv(sl,'Special','salary_special_allow')),
                        salary_pf_employee   = _dec(_gv(sl,'PF Emp','salary_pf_employee')),
                        salary_pf_employer   = _dec(_gv(sl,'PF Er','salary_pf_employer')),
                        salary_esic_employee = _dec(_gv(sl,'ESIC Emp','salary_esic_employee')),
                        salary_esic_employer = _dec(_gv(sl,'ESIC Er','salary_esic_employer')),
                        salary_professional_tax = _dec(_gv(sl,'Prof Tax','salary_professional_tax')),
                        salary_tds      = _dec(_gv(sl,'TDS','salary_tds')),
                        salary_net      = _dec(_gv(sl,'Net Salary','salary_net')),
                        salary_mode     = _gv(sl,'Mode','salary_mode'),
                        salary_effective_date = _pd(_gv(sl,'Effective Date','salary_effective_date')),
                        # Education
                        highest_qualification = _gv(ed,'Qualification','highest_qualification'),
                        university      = _gv(ed,'University / Board','university'),
                        passing_year    = int(_gv(ed,'Year','passing_year') or 0) or None,
                        specialization  = _gv(ed,'Specialization','specialization'),
                        prev_company    = _gv(ed,'Prev Company','prev_company'),
                        prev_designation = _gv(ed,'Prev Designation','prev_designation'),
                        prev_from_date  = _pd(_gv(ed,'Prev From','prev_from_date')),
                        prev_to_date    = _pd(_gv(ed,'Prev To','prev_to_date')),
                        prev_leaving_reason = _gv(ed,'Leaving Reason','prev_leaving_reason'),
                        total_experience_yrs = _dec(_gv(ed,'Experience (Yrs)','total_experience_yrs')),
                        documents_json  = '[]',
                        # ── Phase-1: Basic — Family / Contact / Permanent Addr ──
                        father_name       = _gv(row,'Father Name','father_name') or None,
                        mother_name       = _gv(row,'Mother Name','mother_name') or None,
                        alternate_mobile  = _gv(row,'Alternate Mobile','alternate_mobile') or None,
                        personal_email    = _gv(row,'Personal Email','personal_email') or None,
                        permanent_address = _gv(row,'Permanent Address','permanent_address') or None,
                        permanent_city    = _gv(row,'Permanent City','permanent_city') or None,
                        permanent_state   = _gv(row,'Permanent State','permanent_state') or None,
                        permanent_country = _gv(row,'Permanent Country','permanent_country') or 'India',
                        permanent_zip     = _gv(row,'Permanent ZIP','Permanent Zip','permanent_zip') or None,
                        same_as_current_addr = _gv(row,'Same as Current','same_as_current_addr').lower() == 'yes',
                        # ── Phase-1: Professional ───────────────────
                        grade_level             = _gv(p,'Grade Level','grade_level') or None,
                        probation_period_months = int(_gv(p,'Probation Months','probation_period_months') or 6),
                        probation_end_date      = _pd(_gv(p,'Probation End','probation_end_date')),
                        attendance_code         = _gv(p,'Attendance Code','attendance_code') or None,
                        overtime_eligible       = _gv(p,'Overtime Eligible','overtime_eligible').lower() == 'yes',
                        casual_leave_balance    = _dec(_gv(p,'CL Balance','casual_leave_balance')) or 0,
                        sick_leave_balance      = _dec(_gv(p,'SL Balance','sick_leave_balance')) or 0,
                        paid_leave_balance      = _dec(_gv(p,'PL Balance','paid_leave_balance')) or 0,
                        leave_policy            = _gv(p,'Leave Policy','leave_policy') or None,
                        official_email          = _gv(p,'Official Email','official_email') or None,
                        role_access             = _gv(p,'Role Access','role_access') or None,
                        exit_interview_done     = _gv(p,'Exit Interview','exit_interview_done').lower() == 'yes',
                        exit_interview_notes    = _gv(p,'Exit Notes','exit_interview_notes') or None,
                        ff_settlement_status    = _gv(p,'FF Status','ff_settlement_status') or 'Pending',
                        ff_settlement_amount    = _dec(_gv(p,'FF Amount','ff_settlement_amount')),
                        ff_settlement_date      = _pd(_gv(p,'FF Date','ff_settlement_date')),
                        # ── Phase-1: KYC — PF / ESIC / TDS / Statutory
                        pf_applicable        = _gv(k,'PF Applicable','pf_applicable').lower() == 'yes',
                        pf_number            = _gv(k,'PF Number','pf_number') or None,
                        eps_applicable       = _gv(k,'EPS Applicable','eps_applicable').lower() == 'yes',
                        previous_pf_transfer = _gv(k,'Previous PF Transfer','previous_pf_transfer').lower() == 'yes',
                        previous_pf_number   = _gv(k,'Previous PF Number','previous_pf_number') or None,
                        esic_applicable       = _gv(k,'ESIC Applicable','esic_applicable').lower() == 'yes',
                        esic_nominee_name     = _gv(k,'ESIC Nominee','esic_nominee_name') or None,
                        esic_nominee_relation = _gv(k,'Nominee Relation','esic_nominee_relation') or None,
                        esic_family_details   = _gv(k,'ESIC Family','esic_family_details') or None,
                        esic_dispensary       = _gv(k,'Dispensary','esic_dispensary') or None,
                        aadhaar_pan_linked      = _gv(k,'Aadhaar PAN Linked','aadhaar_pan_linked').lower() == 'yes',
                        tax_regime              = _gv(k,'Tax Regime','tax_regime') or 'New',
                        prev_employer_income    = _dec(_gv(k,'Prev Employer Income','prev_employer_income')),
                        monthly_tds             = _dec(_gv(k,'Monthly TDS','monthly_tds')),
                        investment_declaration  = _gv(k,'Investment Declaration','investment_declaration') or None,
                        proof_submission_status = _gv(k,'Proof Status','proof_submission_status') or 'Pending',
                        professional_tax_applicable = (_gv(k,'PT Applicable','professional_tax_applicable') or 'yes').lower() == 'yes',
                        labour_welfare_fund         = _gv(k,'LWF','labour_welfare_fund').lower() == 'yes',
                        gratuity_eligible           = _gv(k,'Gratuity Eligible','gratuity_eligible').lower() == 'yes',
                        bonus_eligible              = (_gv(k,'Bonus Eligible','bonus_eligible') or 'yes').lower() == 'yes',
                        # ── Phase-1: Salary extras ──────────────────
                        salary_conveyance = _dec(_gv(sl,'Conveyance','salary_conveyance')),
                        salary_bonus      = _dec(_gv(sl,'Bonus','salary_bonus')),
                        salary_incentive  = _dec(_gv(sl,'Incentive','salary_incentive')),
                        salary_gross      = _dec(_gv(sl,'Gross','salary_gross')),
                        created_by      = current_user.id,
                    )

                    # Validate status
                    valid_statuses = ['active','inactive','on_leave','terminated']
                    if e.status not in valid_statuses:
                        e.status = 'active'

                    db.session.add(e)
                    db.session.flush()

                    # Auto-create login
                    uname = emp_code.lower().replace('-','').replace(' ','')
                    if not User.query.filter_by(username=uname).first():
                        ue_email = e.email or f'{uname}@hcp.com'
                        if User.query.filter_by(email=ue_email).first():
                            ue_email = f'{uname}@hcp.com'
                        u = User(username=uname, email=ue_email,
                                 full_name=e.full_name, role='user', is_active=True)
                        u.set_password('HCP@123')
                        db.session.add(u)
                        db.session.flush()
                        e.user_id = u.id

                    added += 1

                except Exception as ex:
                    errors.append(f'Row {i} ({emp_code}): {str(ex)[:100]}')
                    db.session.rollback()

            db.session.commit()
            msg = f'✅ {added} employees imported!'
            if skipped: msg += f' ⚠️ {skipped} skipped (duplicate codes).'
            flash(msg, 'success')
            if errors:
                for err in errors[:8]:
                    flash(err, 'warning')

        except Exception as ex:
            db.session.rollback()
            flash(f'Import failed: {str(ex)}', 'danger')

        return redirect(url_for('hr.employees'))

    return render_template('hr/employees/import.html', perm=perm, active_page='hr_employees')


@hr.route('/employees/import/template')
@login_required
def emp_import_template():
    """Download 7-tab Excel import template."""
    import io, sys, subprocess
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable,'-m','pip','install','openpyxl','--quiet'],check=True)
        import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    def mk_hdr(cell, color):
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D0D7E2")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def mk_note(cell):
        cell.font = Font(size=9, color="64748B", italic=True, name="Arial")
        cell.fill = PatternFill("solid", fgColor="F8FAFC")
        thin = Side(style="thin", color="E2E8F0")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical="center")

    def mk_sample(cell):
        cell.font = Font(size=9, name="Arial")
        thin = Side(style="thin", color="E2E8F0")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical="center")

    def build_tpl(ws, color, cols, notes, samples):
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 18
        for ci, h in enumerate(cols, 1):
            mk_hdr(ws.cell(1, ci, h), color)
        for ci, n in enumerate(notes, 1):
            mk_note(ws.cell(2, ci, n))
        for ci, s in enumerate(samples, 1):
            mk_sample(ws.cell(3, ci, s))
        # Note row style
        ws.cell(2, 1).font = Font(bold=True, size=9, color="B45309", name="Arial")
        ws.cell(2, 1).fill = PatternFill("solid", fgColor="FFFBEB")
        for ci in range(1, len(cols)+1):
            col = get_column_letter(ci)
            mx = max(len(str(ws.cell(r,ci).value or '')) for r in range(1,4))
            ws.column_dimensions[col].width = min(mx+3, 35)
        ws.freeze_panes = "A3"
        ws.cell(2,1,"⚠️ Row 2 = Notes (delete before import). Row 3+ = your data")

    # ── Sheet 1: Basic Info ──
    ws1 = wb.active; ws1.title = "1 - Basic Info"
    build_tpl(ws1, "1E3A5F",
        ["Employee Code","Employee ID","First Name","Middle Name","Last Name","Mobile","Email","Gender","Date of Birth","Blood Group","Marital Status","Marriage Anniversary","Father Name","Mother Name","Alternate Mobile","Personal Email","Address","City","State","Country","ZIP","Permanent Address","Permanent City","Permanent State","Permanent Country","Permanent ZIP","Same as Current","LinkedIn","Facebook","Status"],
        ["Required. Unique code","Biometric/Device ID","Required","Optional","Optional","10 digits","Valid email","Male/Female/Other","DD-MM-YYYY","A+/B+/O+...","Single/Married/Divorced","DD-MM-YYYY if Married","Father's name","Mother's name","Optional","Personal Gmail/Yahoo","Street address","City","State","Default: India","Pincode","Permanent street","City","State","India","Pincode","Yes/No","URL optional","URL optional","active/inactive"],
        ["EMP0001","1001","Krunal","Naresh","Chandi","9876543210","krunal@hcp.com","Male","15-06-1990","A+","Married","20-02-2015","Naresh Chandi","Suman Chandi","9876512345","krunal.p@gmail.com","123 MG Road","Ahmedabad","Gujarat","India","380001","123 MG Road","Ahmedabad","Gujarat","India","380001","Yes","","","active"]
    )

    # ── Sheet 2: Professional ──
    ws2 = wb.create_sheet("2 - Professional")
    build_tpl(ws2, "1D4ED8",
        ["Code","Full Name","Department","Designation","Employee Type","Location","Pay Grade","Grade Level","DOJ","Confirmation","Shift","Work Hrs","Weekly Off","Notice Days","Probation Months","Probation End","Contractor","Probation","Block","Status","Reports To","Attendance Code","Overtime Eligible","CL Balance","SL Balance","PL Balance","Leave Policy","Official Email","Role Access","Exit Interview","Exit Notes","FF Status","FF Amount","FF Date"],
        ["Match Sheet1 Code","For reference","e.g. Sales","e.g. Manager","Full Time/Part Time/Contract/Intern","City/Branch","G1/G2/G3","L1/L2/Senior","DD-MM-YYYY","DD-MM-YYYY","General/Night","8","Sunday","30","6","DD-MM-YYYY","Yes/No","Yes/No","Yes/No","active","Manager full name","ATT1001","Yes/No","0","0","0","Standard","name@co.com","Staff/Manager/HR Admin","Yes/No","Feedback text","Pending/Processed/Paid","Amount ₹","DD-MM-YYYY"],
        ["EMP0001","Krunal Chandi","Sales","Sales Manager","Full Time","Ahmedabad","G2","L2","01-01-2022","01-07-2022","General (9-6)","8","Sunday","30","6","01-07-2022","No","No","No","active","Rajesh Shah","ATT0001","Yes","10","8","12","Standard","krunal@hcp.com","Manager","No","","Pending","","",""]
    )

    # ── Sheet 3: KYC ──
    ws3 = wb.create_sheet("3 - KYC")
    build_tpl(ws3, "7C3AED",
        ["Code","Full Name","Nationality","Religion","Aadhaar","PAN","UAN","ESIC","Passport No.","Passport Expiry","DL No","DL Expiry","Emergency Name","Emergency Relation","Emergency Phone","Emergency Address","PF Applicable","PF Number","EPS Applicable","Previous PF Transfer","Previous PF Number","ESIC Applicable","ESIC Nominee","Nominee Relation","ESIC Family","Dispensary","Aadhaar PAN Linked","Tax Regime","Prev Employer Income","Monthly TDS","Investment Declaration","Proof Status","PT Applicable","LWF","Gratuity Eligible","Bonus Eligible"],
        ["Match Sheet1 Code","For reference","Indian","Hindu/Muslim/..","12 digits","ABCDE1234F","12 digits","17 digits","A1234567","DD-MM-YYYY","GJ01 2024 123456","DD-MM-YYYY","Contact name","Father/Spouse/..","10 digits","Address","Yes/No","PF A/c No.","Yes/No","Yes/No","Old PF No.","Yes/No","Nominee name","Spouse/Father/..","Family details","Dispensary name","Yes/No","New/Old","Amount ₹","Amount ₹","Declaration text","Pending/Submitted/Verified","Yes/No","Yes/No","Yes/No","Yes/No"],
        ["EMP0001","Krunal Chandi","Indian","Hindu","123456789012","ABCDE1234F","100123456789","1234567890123456","A1234567","31-12-2030","GJ01 2024 123","31-12-2030","Ramesh Chandi","Father","9876500000","123 MG Road Ahmedabad","Yes","PF/GJ/12345/67890","Yes","No","","Yes","Priya Chandi","Spouse","Spouse + 1 child","ESIC Ahmedabad","Yes","New","","","","Pending","Yes","No","No","Yes"]
    )

    # ── Sheet 4: Bank ──
    ws4 = wb.create_sheet("4 - Bank Details")
    build_tpl(ws4, "065F46",
        ["Code","Full Name","Account Holder","Bank Name","Account Number","IFSC Code","Branch","Account Type"],
        ["Match Sheet1 Code","For reference","As per bank record","Bank name","Account number","11 char IFSC","Branch name","Savings/Current"],
        ["EMP0001","Krunal Chandi","Krunal N Chandi","HDFC Bank","50100123456789","HDFC0001234","Ahmedabad Main","Savings"]
    )

    # ── Sheet 5: Salary ──
    ws5 = wb.create_sheet("5 - Salary")
    build_tpl(ws5, "B45309",
        ["Code","Full Name","CTC Annual","Basic","HRA","DA","TA","Conveyance","Medical","Special","Bonus","Incentive","Gross","PF Emp","PF Er","ESIC Emp","ESIC Er","Prof Tax","TDS","Net Salary","Mode","Effective Date"],
        ["Match Sheet1 Code","For reference","Annual CTC in ₹","Monthly basic","Monthly HRA","Monthly DA","Transport","Conveyance","Medical allow","Special allow","Monthly bonus","Monthly incentive","Monthly gross","PF deduction","PF employer","ESIC employee","ESIC employer","Prof. tax","TDS monthly","Net take-home","Cash/Bank Transfer/Cheque","DD-MM-YYYY"],
        ["EMP0001","Krunal Chandi","480000","16000","8000","1600","1600","800","1250","0","0","0","28250","1920","1920","0","0","200","0","26130","Bank Transfer","01-01-2022"]
    )

    # ── Sheet 6: Education ──
    ws6 = wb.create_sheet("6 - Education")
    build_tpl(ws6, "0F766E",
        ["Code","Full Name","Qualification","University / Board","Year","Specialization","Prev Company","Prev Designation","Prev From","Prev To","Leaving Reason","Experience (Yrs)"],
        ["Match Sheet1 Code","For reference","12th/Diploma/BCA/MBA..","University name","Passing year","Branch/Subject","Previous employer","Designation there","DD-MM-YYYY","DD-MM-YYYY","Reason","Total yrs"],
        ["EMP0001","Krunal Chandi","MBA","Gujarat University","2015","Marketing","ABC Pvt Ltd","Sales Executive","01-06-2015","31-12-2021","Better opportunity","6.5"]
    )

    # ── Instructions sheet ──
    wsi = wb.create_sheet("📋 Instructions", 0)
    wsi.column_dimensions['A'].width = 8
    wsi.column_dimensions['B'].width = 55
    wsi.column_dimensions['C'].width = 45
    instructions = [
        ("","📋 EMPLOYEE IMPORT TEMPLATE — Instructions",""),
        ("","",""),
        ("","HOW TO USE THIS FILE:",""),
        ("1️⃣","Sheet 1 - Basic Info is REQUIRED. Fill employee code, name, mobile etc.","Employee Code must be unique"),
        ("2️⃣","Sheets 2-6 are OPTIONAL. Fill only what you have.","Match Employee Code exactly in each sheet"),
        ("3️⃣","Delete the Notes row (Row 2) from each sheet before importing.","Or keep it — system will skip non-data rows"),
        ("4️⃣","Date format: DD-MM-YYYY (e.g. 15-06-2024)",""),
        ("5️⃣","Salary fields: Numbers only, no ₹ symbol needed","e.g. 500000 not ₹5,00,000"),
        ("6️⃣","Yes/No fields: type Yes or No exactly",""),
        ("7️⃣","Upload this file at: HR > Employees > Import",""),
        ("","",""),
        ("","SHEETS IN THIS FILE:",""),
        ("📗","1 - Basic Info","Name, Contact, Address, DOB, Gender"),
        ("📘","2 - Professional","Dept, Designation, DOJ, Shift, Status"),
        ("📙","3 - KYC","Aadhaar, PAN, Passport, Emergency Contact"),
        ("📕","4 - Bank Details","Bank Account, IFSC, Branch"),
        ("📒","5 - Salary","CTC, Basic, HRA, PF, ESIC, Net Pay"),
        ("📓","6 - Education","Qualification, Previous Company"),
    ]
    for ri, (a, b, c_) in enumerate(instructions, 1):
        wsi.row_dimensions[ri].height = 20
        ca = wsi.cell(ri, 1, a)
        cb = wsi.cell(ri, 2, b)
        cc = wsi.cell(ri, 3, c_)
        if ri == 1:
            cb.font = Font(bold=True, size=14, color="1E3A5F", name="Arial")
        elif b.startswith(("HOW","SHEETS")):
            cb.font = Font(bold=True, size=10, color="1D4ED8", name="Arial")
        else:
            ca.font = Font(size=10, name="Arial")
            cb.font = Font(size=9, name="Arial")
            cc.font = Font(size=9, color="64748B", italic=True, name="Arial")

    from flask import send_file
    import io as _io
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='employee_import_template.xlsx')


@hr.route('/employees/<int:id>/export')
@login_required
def emp_export_single(id):
    """Export single employee as 7-tab Excel workbook."""
    import io, sys, subprocess, json
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable,'-m','pip','install','openpyxl','--quiet'],check=True)
        import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt2

    e = Employee.query.get_or_404(id)

    wb = openpyxl.Workbook()

    # ── Style helpers ──
    def hdr_style(cell, color="1E3A5F"):
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D0D7E2")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def data_style(cell, bold=False):
        thin = Side(style="thin", color="E2E8F0")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.font = Font(size=9, name="Arial", bold=bold)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    def alt_style(cell, row_idx):
        if row_idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F8FAFC")

    def write_sheet(ws, title_color, pairs):
        """Write label-value pairs to sheet."""
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 42
        # Header row
        ws.row_dimensions[1].height = 26
        h1 = ws.cell(1, 1, "Field"); h2 = ws.cell(1, 2, "Value")
        hdr_style(h1, title_color); hdr_style(h2, title_color)
        for i, (lbl, val) in enumerate(pairs, 2):
            ws.row_dimensions[i].height = 18
            c1 = ws.cell(i, 1, lbl)
            c2 = ws.cell(i, 2, val if val is not None else "—")
            data_style(c1, bold=True); data_style(c2)
            alt_style(c1, i); alt_style(c2, i)
        ws.freeze_panes = "A2"

    def fmt_date(d):
        return d.strftime('%d-%m-%Y') if d else None

    def fmt_cur(v):
        return f"₹{v:,.0f}" if v else None

    # ── Sheet 1: Basic Info ──
    ws1 = wb.active; ws1.title = "1 - Basic Info"
    write_sheet(ws1, "1E3A5F", [
        ("Employee Code", e.employee_code),
        ("First Name", e.first_name),
        ("Last Name", e.last_name),
        ("Full Name", e.full_name),
        ("Mobile", e.mobile),
        ("Email", e.email),
        ("Gender", e.gender),
        ("Date of Birth", fmt_date(e.date_of_birth)),
        ("Blood Group", e.blood_group),
        ("Marital Status", e.marital_status),
        ("Marriage Anniversary", fmt_date(e.marriage_anniversary)),
        ("Address", e.address),
        ("City", e.city),
        ("State", e.state),
        ("Country", e.country),
        ("ZIP / Pin Code", e.zip_code),
        ("LinkedIn", e.linkedin),
        ("Facebook", e.facebook),
    ])

    # ── Sheet 2: Professional ──
    ws2 = wb.create_sheet("2 - Professional")
    write_sheet(ws2, "1D4ED8", [
        ("Department", e.department),
        ("Designation", e.designation),
        ("Employee Type", e.employee_type),
        ("Location", e.location),
        ("Pay Grade", e.pay_grade),
        ("Date of Joining", fmt_date(e.date_of_joining)),
        ("Confirmation Date", fmt_date(e.confirmation_date)),
        ("Shift", e.shift),
        ("Work Hours Per Day", str(e.work_hours_per_day) if e.work_hours_per_day else None),
        ("Weekly Off", e.weekly_off),
        ("Notice Period (Days)", str(e.notice_period_days) if e.notice_period_days else None),
        ("Is Contractor", "Yes" if e.is_contractor else "No"),
        ("Status", (e.status or "").title()),
        ("Is Probation", "Yes" if e.is_probation else "No"),
        ("Is Block", "Yes" if e.is_block else "No"),
        ("Is Late", "Yes" if e.is_late else "No"),
        ("Rehire Eligible", "Yes" if e.rehire_eligible else "No"),
        ("Resignation Date", fmt_date(e.resignation_date)),
        ("Last Working Date", fmt_date(e.last_working_date)),
        ("Remark", e.remark),
        ("Reports To", e.manager_emp.full_name if e.manager_emp else None),
    ])

    # ── Sheet 3: KYC ──
    ws3 = wb.create_sheet("3 - Personal & KYC")
    write_sheet(ws3, "7C3AED", [
        ("Nationality", e.nationality),
        ("Religion", e.religion),
        ("Caste", e.caste),
        ("Physically Handicapped", "Yes" if e.physically_handicapped else "No"),
        ("Aadhaar Number", e.aadhar_number),
        ("PAN Number", e.pan_number),
        ("UAN Number", e.uan_number),
        ("ESIC Number", e.esic_number),
        ("Passport Number", e.passport_number),
        ("Passport Expiry", fmt_date(e.passport_expiry)),
        ("Driving License", e.driving_license),
        ("DL Expiry", fmt_date(e.dl_expiry)),
        ("Emergency Name", e.emergency_name),
        ("Emergency Relation", e.emergency_relation),
        ("Emergency Phone", e.emergency_phone),
        ("Emergency Address", e.emergency_address),
    ])

    # ── Sheet 4: Bank ──
    ws4 = wb.create_sheet("4 - Bank Details")
    write_sheet(ws4, "065F46", [
        ("Account Holder", e.bank_account_holder),
        ("Bank Name", e.bank_name),
        ("Account Number", e.bank_account_number),
        ("IFSC Code", e.bank_ifsc),
        ("Branch", e.bank_branch),
        ("Account Type", e.bank_account_type),
    ])

    # ── Sheet 5: Salary ──
    ws5 = wb.create_sheet("5 - Salary")
    write_sheet(ws5, "B45309", [
        ("CTC (Annual)", fmt_cur(e.salary_ctc)),
        ("Basic Salary", fmt_cur(e.salary_basic)),
        ("HRA", fmt_cur(e.salary_hra)),
        ("DA (Dearness Allow.)", fmt_cur(e.salary_da)),
        ("Transport Allow.", fmt_cur(e.salary_ta)),
        ("Medical Allow.", fmt_cur(e.salary_medical_allow)),
        ("Special Allow.", fmt_cur(e.salary_special_allow)),
        ("PF Employee (12%)", fmt_cur(e.salary_pf_employee)),
        ("PF Employer (12%)", fmt_cur(e.salary_pf_employer)),
        ("ESIC Employee (0.75%)", fmt_cur(e.salary_esic_employee)),
        ("ESIC Employer (3.25%)", fmt_cur(e.salary_esic_employer)),
        ("Professional Tax", fmt_cur(e.salary_professional_tax)),
        ("TDS", fmt_cur(e.salary_tds)),
        ("Net Salary", fmt_cur(e.salary_net)),
        ("Salary Mode", e.salary_mode),
        ("Effective Date", fmt_date(e.salary_effective_date)),
    ])

    # ── Sheet 6: Education ──
    ws6 = wb.create_sheet("6 - Education")
    write_sheet(ws6, "0F766E", [
        ("Highest Qualification", e.highest_qualification),
        ("University / Board", e.university),
        ("Passing Year", str(e.passing_year) if e.passing_year else None),
        ("Specialization", e.specialization),
        ("Previous Company", e.prev_company),
        ("Previous Designation", e.prev_designation),
        ("Prev From Date", fmt_date(e.prev_from_date)),
        ("Prev To Date", fmt_date(e.prev_to_date)),
        ("Leaving Reason", e.prev_leaving_reason),
        ("Total Experience (Yrs)", str(e.total_experience_yrs) if e.total_experience_yrs else None),
    ])

    # ── Sheet 7: Documents ──
    ws7 = wb.create_sheet("7 - Documents")
    ws7.column_dimensions['A'].width = 5
    ws7.column_dimensions['B'].width = 28
    ws7.column_dimensions['C'].width = 22
    ws7.column_dimensions['D'].width = 22
    ws7.column_dimensions['E'].width = 14
    ws7.column_dimensions['F'].width = 18
    ws7.row_dimensions[1].height = 26

    doc_headers = ["#", "Document Name", "Type", "Filename", "Size (KB)", "Uploaded At"]
    for ci, h in enumerate(doc_headers, 1):
        cell = ws7.cell(1, ci, h)
        hdr_style(cell, "BE185D")

    docs = []
    try:
        if e.documents_json and e.documents_json != '[]':
            docs = json.loads(e.documents_json)
    except Exception:
        pass

    if docs:
        for i, doc in enumerate(docs, 2):
            ws7.row_dimensions[i].height = 18
            vals = [
                i - 1,
                doc.get('name', ''),
                doc.get('type', ''),
                doc.get('filename', ''),
                round(doc.get('size', 0) / 1024, 1),
                (doc.get('uploaded_at', '') or '')[:10],
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws7.cell(i, ci, val)
                data_style(cell)
                alt_style(cell, i)
    else:
        ws7.cell(2, 1, "No documents uploaded").font = Font(italic=True, color="94A3B8", name="Arial")

    ws7.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    from flask import send_file
    fname = f"employee_{e.employee_code or e.id}_{dt2.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)

@hr.route('/employees/export')
@login_required
def emp_export():
    import io, sys, subprocess
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'], check=True)
        import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    search  = request.args.get('search', '')
    dept    = request.args.get('dept', '')
    status  = request.args.get('status', '')
    emptype = request.args.get('emptype', '')
    sort_by = request.args.get('sort_by', 'created_at')
    sort_dir= request.args.get('sort_dir', 'desc')

    q = Employee.query
    if search:
        s = f'%{search}%'
        q = q.filter(Employee.first_name.ilike(s)|Employee.last_name.ilike(s)|
                     Employee.employee_code.ilike(s)|Employee.mobile.ilike(s)|Employee.email.ilike(s))
    if dept:    q = q.filter_by(department=dept)
    if status:  q = q.filter_by(status=status)
    if emptype: q = q.filter_by(employee_type=emptype)
    sort_col = getattr(Employee, sort_by, Employee.created_at)
    emps = q.order_by(sort_col.asc() if sort_dir=='asc' else sort_col.desc()).all()

    from models.user import User as UserModel
    users = {u.id: u.full_name for u in UserModel.query.all()}

    headers = ["Employee Code","First Name","Middle Name","Last Name","Full Name","Mobile","Email","Gender",
               "Department","Designation","Employee Type","Date of Joining","Location",
               "Date of Birth","Blood Group","Marital Status","Status",
               "Is Contractor","Is Block","Is Late","Is Probation",
               "LinkedIn","Facebook","Remark","Created At","Updated At"]

    rows = []
    for e in emps:
        full = (e.first_name or '') + ' ' + (e.middle_name or '') + ' ' + (e.last_name or '')
        rows.append([
            e.employee_code or '', e.first_name or '', e.middle_name or '', e.last_name or '', full.strip(),
            e.employee_code or '', e.first_name or '', e.middle_name or '', e.last_name or '', full.strip(),
            e.mobile or '', e.email or '', e.gender or '',
            e.department or '', e.designation or '', e.employee_type or '',
            e.date_of_joining.strftime('%d-%m-%Y') if e.date_of_joining else '',
            e.location or '',
            e.date_of_birth.strftime('%d-%m-%Y') if e.date_of_birth else '',
            e.blood_group or '', e.marital_status or '',
            (e.status or '').title(),
            'Yes' if e.is_contractor else 'No',
            'Yes' if e.is_block else 'No',
            'Yes' if e.is_late else 'No',
            'Yes' if e.is_probation else 'No',
            e.linkedin or '', e.facebook or '', e.remark or '',
            e.created_at.strftime('%d-%m-%Y %H:%M') if e.created_at else '',
            e.updated_at.strftime('%d-%m-%Y %H:%M') if e.updated_at else '',
        ])


    # ─── 7 Tab-Wise Sheets ───
    wb = openpyxl.Workbook()

    def mk_hdr(cell, color="1E3A5F"):
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D0D7E2")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def mk_data(cell, ri):
        thin = Side(style="thin", color="E2E8F0")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.font = Font(size=9, name="Arial")
        cell.alignment = Alignment(vertical="center")
        if ri % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F8FAFC")

    def build_sheet(ws, color, headers, rows_data):
        ws.row_dimensions[1].height = 28
        for ci, h in enumerate(headers, 1):
            mk_hdr(ws.cell(1, ci, h), color)
        for ri, row in enumerate(rows_data, 2):
            ws.row_dimensions[ri].height = 17
            for ci, val in enumerate(row, 1):
                mk_data(ws.cell(ri, ci, val), ri)
        for ci in range(1, len(headers)+1):
            col = get_column_letter(ci)
            mx = max((len(str(ws.cell(r, ci).value or '')) for r in range(1, ws.max_row+1)), default=8)
            ws.column_dimensions[col].width = min(mx + 2, 40)
        ws.freeze_panes = "A2"

    def fd(d): return d.strftime('%d-%m-%Y') if d else ''
    def fc(v):
        try: return round(float(v),2) if v is not None and v != '' else ''
        except: return ''

    # Sheet 1: Basic Info
    ws1 = wb.active; ws1.title = "1 - Basic Info"
    h1 = ["Code","First Name","Last Name","Full Name","Mobile","Email","Gender","DOB","Blood Group","Marital Status","Father Name","Mother Name","Alternate Mobile","Personal Email","Address","City","State","Country","ZIP","Permanent Address","Permanent City","Permanent State","Permanent Country","Permanent ZIP","Same as Current","LinkedIn","Facebook","Status","Created At"]
    r1 = []
    for e in emps:
        r1.append([e.employee_code or '',e.first_name or '',e.last_name or '',e.full_name,
            e.mobile or '',e.email or '',e.gender or '',fd(e.date_of_birth),e.blood_group or '',
            e.marital_status or '',
            e.father_name or '', e.mother_name or '', e.alternate_mobile or '', e.personal_email or '',
            e.address or '',e.city or '',e.state or '',e.country or '',e.zip_code or '',
            e.permanent_address or '', e.permanent_city or '', e.permanent_state or '',
            e.permanent_country or '', e.permanent_zip or '',
            'Yes' if e.same_as_current_addr else 'No',
            e.linkedin or '',e.facebook or'',(e.status or '').title(),
            e.created_at.strftime('%d-%m-%Y') if e.created_at else ''])
    build_sheet(ws1,"1E3A5F",h1,r1)

    # Sheet 2: Professional
    ws2 = wb.create_sheet("2 - Professional")
    h2 = ["Code","Full Name","Department","Designation","Employee Type","Location","Pay Grade","Grade Level","DOJ","Confirmation","Shift","Work Hrs","Weekly Off","Notice Days","Probation Months","Probation End","Contractor","Probation","Block","Status","Reports To","Attendance Code","Overtime Eligible","CL Balance","SL Balance","PL Balance","Leave Policy","Official Email","Role Access","Exit Interview","Exit Notes","FF Status","FF Amount","FF Date"]
    r2 = []
    for e in emps:
        r2.append([e.employee_code or '',e.full_name,e.department or '',e.designation or '',
            e.employee_type or '',e.location or '',e.pay_grade or '', e.grade_level or '',
            fd(e.date_of_joining),
            fd(e.confirmation_date),e.shift or '',str(e.work_hours_per_day or ''),e.weekly_off or '',
            str(e.notice_period_days or ''),
            str(e.probation_period_months or ''), fd(e.probation_end_date),
            'Yes' if e.is_contractor else 'No',
            'Yes' if e.is_probation else 'No','Yes' if e.is_block else 'No',
            (e.status or '').title(),e.manager_emp.full_name if e.manager_emp else '',
            e.attendance_code or '', 'Yes' if e.overtime_eligible else 'No',
            str(e.casual_leave_balance or 0), str(e.sick_leave_balance or 0), str(e.paid_leave_balance or 0),
            e.leave_policy or '', e.official_email or '', e.role_access or '',
            'Yes' if e.exit_interview_done else 'No', e.exit_interview_notes or '',
            e.ff_settlement_status or '', fc(e.ff_settlement_amount), fd(e.ff_settlement_date)])
    build_sheet(ws2,"1D4ED8",h2,r2)

    # Sheet 3: KYC
    ws3 = wb.create_sheet("3 - KYC")
    h3 = ["Code","Full Name","Nationality","Religion","Aadhaar","PAN","UAN","ESIC","Passport No","Passport Expiry","DL No","DL Expiry","Emergency Name","Emergency Phone","PF Applicable","PF Number","EPS Applicable","Previous PF Transfer","Previous PF Number","ESIC Applicable","ESIC Nominee","Nominee Relation","ESIC Family","Dispensary","Aadhaar PAN Linked","Tax Regime","Prev Employer Income","Monthly TDS","Investment Declaration","Proof Status","PT Applicable","LWF","Gratuity Eligible","Bonus Eligible"]
    r3 = []
    for e in emps:
        r3.append([e.employee_code or '',e.full_name,e.nationality or '',e.religion or '',
            e.aadhar_number or '',e.pan_number or '',e.uan_number or '',e.esic_number or '',
            e.passport_number or '',fd(e.passport_expiry),e.driving_license or '',fd(e.dl_expiry),
            e.emergency_name or '',e.emergency_phone or '',
            'Yes' if e.pf_applicable else 'No', e.pf_number or '',
            'Yes' if e.eps_applicable else 'No',
            'Yes' if e.previous_pf_transfer else 'No', e.previous_pf_number or '',
            'Yes' if e.esic_applicable else 'No', e.esic_nominee_name or '',
            e.esic_nominee_relation or '', e.esic_family_details or '', e.esic_dispensary or '',
            'Yes' if e.aadhaar_pan_linked else 'No', e.tax_regime or '',
            fc(e.prev_employer_income), fc(e.monthly_tds), e.investment_declaration or '',
            e.proof_submission_status or '',
            'Yes' if e.professional_tax_applicable else 'No',
            'Yes' if e.labour_welfare_fund else 'No',
            'Yes' if e.gratuity_eligible else 'No',
            'Yes' if e.bonus_eligible else 'No'])
    build_sheet(ws3,"7C3AED",h3,r3)

    # Sheet 4: Bank
    ws4 = wb.create_sheet("4 - Bank Details")
    h4 = ["Code","Full Name","Account Holder","Bank Name","Account Number","IFSC","Branch","Account Type"]
    r4 = []
    for e in emps:
        r4.append([e.employee_code or '',e.full_name,e.bank_account_holder or '',
            e.bank_name or '',e.bank_account_number or '',e.bank_ifsc or '',
            e.bank_branch or '',e.bank_account_type or ''])
    build_sheet(ws4,"065F46",h4,r4)

    # Sheet 5: Salary
    ws5 = wb.create_sheet("5 - Salary")
    h5 = ["Code","Full Name","CTC Annual","Basic","HRA","DA","TA","Conveyance","Medical","Special","Bonus","Incentive","Gross","PF Emp","PF Er","ESIC Emp","ESIC Er","Prof Tax","TDS","Net Salary","Mode","Effective Date"]
    r5 = []
    for e in emps:
        r5.append([e.employee_code or '',e.full_name,fc(e.salary_ctc),fc(e.salary_basic),
            fc(e.salary_hra),fc(e.salary_da),fc(e.salary_ta), fc(e.salary_conveyance),
            fc(e.salary_medical_allow),
            fc(e.salary_special_allow), fc(e.salary_bonus), fc(e.salary_incentive), fc(e.salary_gross),
            fc(e.salary_pf_employee),fc(e.salary_pf_employer),
            fc(e.salary_esic_employee),fc(e.salary_esic_employer),fc(e.salary_professional_tax),
            fc(e.salary_tds),fc(e.salary_net),e.salary_mode or '', fd(e.salary_effective_date)])
    build_sheet(ws5,"B45309",h5,r5)

    # Sheet 6: Education
    ws6 = wb.create_sheet("6 - Education")
    h6 = ["Code","Full Name","Qualification","University","Year","Specialization","Prev Company","Prev Designation","Prev From","Prev To","Experience (Yrs)"]
    r6 = []
    for e in emps:
        r6.append([e.employee_code or '',e.full_name,e.highest_qualification or '',
            e.university or '',str(e.passing_year or ''),e.specialization or '',
            e.prev_company or '',e.prev_designation or '',fd(e.prev_from_date),
            fd(e.prev_to_date),str(e.total_experience_yrs or '')])
    build_sheet(ws6,"0F766E",h6,r6)

    # Sheet 7: Documents summary
    ws7 = wb.create_sheet("7 - Documents")
    h7 = ["Code","Full Name","Doc Name","Doc Type","Filename","Size (KB)","Uploaded At"]
    r7 = []
    import json as _json2
    for e in emps:
        try:
            docs = _json2.loads(e.documents_json or '[]')
        except Exception:
            docs = []
        if docs:
            for doc in docs:
                r7.append([e.employee_code or '',e.full_name,
                    doc.get('name',''),doc.get('type',''),doc.get('filename',''),
                    round(doc.get('size',0)/1024,1),
                    (doc.get('uploaded_at','') or '')[:10]])
        else:
            r7.append([e.employee_code or '',e.full_name,'—','—','—','—','—'])
    build_sheet(ws7,"BE185D",h7,r7)
    from flask import send_file
    from datetime import datetime as dt2
    import io as io2
    buf = io2.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"employees_export_{dt2.now().strftime('%Y%m%d_%H%M')}.xlsx")

# ── Salary Config Routes ──────────────────────────────────────────────────────

@hr.route('/salary-config', methods=['GET'])
@login_required
def salary_config_get():
    """Return current salary config as JSON."""
    try:
        cfg = SalaryConfig.get_config()
        return jsonify(ok=True, config=cfg)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@hr.route('/salary-config', methods=['POST'])
@login_required
def salary_config_save():
    """Save salary config from JSON body."""
    try:
        data = request.get_json() or {}
        allowed_keys = {
            'basic_pct', 'hra_pct', 'da_pct', 'ta_fixed', 'med_fixed',
            'pf_emp_pct', 'pf_er_pct', 'esic_emp_pct', 'esic_er_pct',
            'esic_limit', 'pt_fixed',
            # ── HCP Policy keys (Phase 7) ─────────────────────────
            'hcp_enabled', 'hcp_high_gross_thresh', 'hcp_esic_limit',
            'hcp_low_basic_fixed', 'hcp_high_basic_pct', 'hcp_hra_pct_of_basic',
            'hcp_conv_pct_of_basic', 'hcp_medical_fixed', 'hcp_pt_threshold',
            'hcp_pt_amount', 'hcp_pf_emp_pct', 'hcp_pf_er_pct',
            'hcp_esic_emp_pct', 'hcp_esic_er_pct', 'hcp_bonus_pct',
        }
        clean = {k: v for k, v in data.items() if k in allowed_keys}
        if not clean:
            return jsonify(ok=False, error='No valid keys provided'), 400
        updated_by = current_user.full_name or current_user.username
        SalaryConfig.save_config(clean, updated_by=updated_by)
        return jsonify(ok=True, msg='Salary config saved successfully!')
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@hr.route('/salary-config/page')
@login_required
def salary_config_page():
    """Render salary config admin page."""
    cfg = SalaryConfig.get_config()
    last = SalaryConfig.query.order_by(SalaryConfig.updated_at.desc()).first()
    cfg['updated_by'] = last.updated_by if last else None
    cfg['updated_at'] = last.updated_at.strftime('%d %b %Y, %I:%M %p') if last and last.updated_at else None
    comps_raw = SalaryComponent.query.order_by(SalaryComponent.component_type, SalaryComponent.sort_order).all()
    components_dict = [c.to_dict() for c in comps_raw]
    return render_template('hr/salary_config.html', cfg=cfg, components=comps_raw, components_json=components_dict, active_page='salary_config')


# ── Salary Components CRUD ────────────────────────────────────────────────────

@hr.route('/salary-components', methods=['GET'])
@login_required
def salary_components_list():
    """Return all components as JSON (for employee form)."""
    comps = SalaryComponent.get_all_active()
    return jsonify(ok=True, components=[c.to_dict() for c in comps])


@hr.route('/salary-components/add', methods=['POST'])
@login_required
def salary_component_add():
    try:
        d = request.get_json() or {}
        # Validate required
        for f in ('name', 'code', 'component_type', 'calc_type'):
            if not d.get(f):
                return jsonify(ok=False, error=f'{f} required'), 400
        # Check duplicate code
        if SalaryComponent.query.filter_by(code=d['code'].strip().lower()).first():
            return jsonify(ok=False, error='Code already exists'), 400
        comp = SalaryComponent(
            name               = d['name'].strip(),
            code               = d['code'].strip().lower(),
            component_type     = d['component_type'],
            calc_type          = d['calc_type'],
            value              = float(d.get('value', 0)),
            cap_amount         = float(d['cap_amount']) if d.get('cap_amount') else None,
            apply_if_gross_lte = float(d['apply_if_gross_lte']) if d.get('apply_if_gross_lte') else None,
            sort_order         = int(d.get('sort_order', 0)),
            is_active          = bool(d.get('is_active', True)),
            description        = d.get('description', ''),
            updated_by         = current_user.full_name or current_user.username,
        )
        db.session.add(comp)
        db.session.commit()
        return jsonify(ok=True, component=comp.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify(ok=False, error=str(e)), 500


@hr.route('/salary-components/<int:cid>', methods=['POST'])
@login_required
def salary_component_edit(cid):
    try:
        comp = SalaryComponent.query.get_or_404(cid)
        d = request.get_json() or {}
        comp.name               = d.get('name', comp.name).strip()
        comp.component_type     = d.get('component_type', comp.component_type)
        comp.calc_type          = d.get('calc_type', comp.calc_type)
        comp.value              = float(d['value']) if 'value' in d else comp.value
        comp.cap_amount         = float(d['cap_amount']) if d.get('cap_amount') else None
        comp.apply_if_gross_lte = float(d['apply_if_gross_lte']) if d.get('apply_if_gross_lte') else None
        comp.sort_order         = int(d.get('sort_order', comp.sort_order))
        comp.is_active          = bool(d.get('is_active', comp.is_active))
        comp.description        = d.get('description', comp.description)
        comp.updated_by         = current_user.full_name or current_user.username
        comp.updated_at         = datetime.utcnow()
        db.session.commit()
        return jsonify(ok=True, component=comp.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify(ok=False, error=str(e)), 500


@hr.route('/salary-components/<int:cid>/delete', methods=['POST'])
@login_required
def salary_component_delete(cid):
    try:
        comp = SalaryComponent.query.get_or_404(cid)
        if comp.is_system:
            return jsonify(ok=False, error='System components cannot be deleted'), 400
        db.session.delete(comp)
        db.session.commit()
        return jsonify(ok=True)
    except Exception as e:
        db.session.rollback()
        return jsonify(ok=False, error=str(e)), 500

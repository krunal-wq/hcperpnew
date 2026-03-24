import os
from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, send_file, current_app
from flask_login import login_required, current_user
from datetime import datetime, date
from werkzeug.utils import secure_filename
from models import (db, User, ClientMaster, ClientBrand, ClientAddress,
                    Lead, LeadDiscussion, LeadAttachment,
                    LeadReminder, LeadNote, LeadActivityLog,
                    SampleOrder, Quotation, LeadContribution, ContributionConfig,
                    Customer, CustomerAddress,
                    LeadStatus, LeadSource, LeadCategory, ProductRange)

from permissions import get_perm, get_grid_columns, save_grid_columns
from audit_helper import audit, snapshot, diff

LEAD_COLS_DEFAULT = ['created_at','name','company','email','mobile','product','team','status','lead_type','last_contact','lead_age']
LEAD_COLS_ALL = {
    'created_at':    'Created Date',
    'name':          'Name',
    'company':       'Company',
    'email':         'Email',
    'mobile':        'Mobile',
    'product':       'Product',
    'category':      'Category',
    'source':        'Source',
    'city':          'City',
    'assigned_to':   'Assigned To',
    'team':          'Team',
    'follow_up':     'Follow Up Date',
    'status':        'Status',
    'lead_type':     'Lead Type',
    'last_contact':  'Last Contact',
    'priority':      'Priority',
    'expected_value':'Expected Value',
    'lead_age':      'Days (Age)',
}

CLIENT_COLS_DEFAULT = ['code','created_at','contact_name','company_name','mobile','email','city','brands','status']
CLIENT_COLS_ALL = {
    'code':          'Client Code',
    'created_at':    'Created Date',
    'contact_name':  'Contact Name',
    'company_name':  'Company',
    'mobile':        'Mobile',
    'email':         'Email',
    'city':          'City',
    'state':         'State',
    'gstin':         'GSTIN',
    'website':       'Website',
    'brands':        'Brands',
    'status':        'Status',
}

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx', 'txt'}

crm = Blueprint('crm', __name__, url_prefix='/crm')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def gen_code(model, prefix):
    last = model.query.order_by(model.id.desc()).first()
    num  = (last.id + 1) if last else 1
    return f"{prefix}-{num:04d}"


def log_activity(lead_id, action, user_id=None):
    uid = user_id or (current_user.id if current_user.is_authenticated else None)
    db.session.add(LeadActivityLog(lead_id=lead_id, user_id=uid, action=action))


def add_contribution(lead_id, action_type, user_id=None, note=''):
    """Track contribution points for a user on a lead."""
    uid = user_id or (current_user.id if current_user.is_authenticated else None)
    if not uid: return

    # Read points from DB config, fallback to defaults
    DEFAULT_POINTS = {
        'comment': 1, 'status_change': 2, 'close_fast': 8,
        'close_slow': 0, 'cancel': 0, 'follow_up': 1, 'reminder': 1, 'edit': 1,
    }
    try:
        cfg = ContributionConfig.query.filter_by(action_type=action_type).first()
        pts = cfg.points if cfg else DEFAULT_POINTS.get(action_type, 1)
    except Exception:
        pts = DEFAULT_POINTS.get(action_type, 1)

    db.session.add(LeadContribution(
        lead_id=lead_id, user_id=uid,
        action_type=action_type, points=pts, note=note
    ))


def _handle_close_contribution(lead):
    """
    Smart close contribution with 5 slabs:
    Slab 1: 1-7 days
    Slab 2: 8-14 days
    Slab 3: 15-21 days
    Slab 4: 22-28 days
    Slab 5: 29+ days
    Points divided among active members only.
    """
    try:
        age = lead.lead_age

        # Determine slab
        if   age <=  7: slab = 'close_slab1'
        elif age <= 14: slab = 'close_slab2'
        elif age <= 21: slab = 'close_slab3'
        elif age <= 28: slab = 'close_slab4'
        else:           slab = 'close_slab5'

        # Get points for this slab from DB
        SLAB_DEFAULTS = {
            'close_slab1': 10, 'close_slab2': 8,
            'close_slab3': 6,  'close_slab4': 4, 'close_slab5': 0
        }
        cfg = ContributionConfig.query.filter_by(action_type=slab).first()
        base_pts = cfg.points if cfg else SLAB_DEFAULTS.get(slab, 0)

        if base_pts == 0:
            return  # No points for this slab

        # Find active members — who had prior activity on this lead
        active_users = db.session.query(LeadContribution.user_id).filter(
            LeadContribution.lead_id == lead.id,
            LeadContribution.action_type.in_(['comment', 'edit', 'status_change', 'follow_up', 'reminder'])
        ).distinct().all()
        active_ids = [r[0] for r in active_users]

        # Add current closer if not in active list
        current_uid = current_user.id if current_user.is_authenticated else None
        if current_uid and current_uid not in active_ids:
            active_ids.append(current_uid)

        if not active_ids:
            return

        # Divide equally (minimum 1 per active member)
        pts_each = max(1, round(base_pts / len(active_ids)))

        slab_labels = {
            'close_slab1':'1-7d', 'close_slab2':'8-14d',
            'close_slab3':'15-21d', 'close_slab4':'22-28d', 'close_slab5':'29+d'
        }

        for uid in active_ids:
            db.session.add(LeadContribution(
                lead_id     = lead.id,
                user_id     = uid,
                action_type = slab,
                points      = pts_each,
                note        = f'Closed in {age}d ({slab_labels[slab]}) — {pts_each}pts/{len(active_ids)} members'
            ))

    except Exception as e:
        # Fallback — give default points to closer
        if current_user.is_authenticated:
            db.session.add(LeadContribution(
                lead_id=lead.id, user_id=current_user.id,
                action_type='close_slab5', points=0, note='Lead closed (fallback)'
            ))


# ══════════════════════════════════════
# CLIENT MASTER ROUTES
# ══════════════════════════════════════

@crm.route('/clients')
@login_required
def clients():
    search      = request.args.get('search', '')
    city        = request.args.get('city', '')
    state       = request.args.get('state', '')
    status_f    = request.args.get('status_f', '')
    sort_by     = request.args.get('sort_by', 'created_at')
    sort_dir    = request.args.get('sort_dir', 'desc')

    show_trash = request.args.get('trash', '') == '1'
    query = ClientMaster.query.filter_by(is_deleted=True) if show_trash else ClientMaster.query.filter_by(is_deleted=False)

    if search:
        query = query.filter(
            ClientMaster.company_name.ilike(f'%{search}%') |
            ClientMaster.contact_name.ilike(f'%{search}%') |
            ClientMaster.mobile.ilike(f'%{search}%') |
            ClientMaster.email.ilike(f'%{search}%')
        )
    if city:        query = query.filter(ClientMaster.city.ilike(f'%{city}%'))
    if state:       query = query.filter(ClientMaster.state.ilike(f'%{state}%'))
    if status_f:    query = query.filter_by(status=status_f)

    sort_col = getattr(ClientMaster, sort_by, ClientMaster.created_at)
    if sort_dir == 'asc':
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    all_clients = query.all()

    all_cities = [r[0] for r in db.session.query(ClientMaster.city).distinct().all() if r[0]]
    all_states = [r[0] for r in db.session.query(ClientMaster.state).distinct().all() if r[0]]

    grid_cols = get_grid_columns('clients', CLIENT_COLS_DEFAULT, list(CLIENT_COLS_ALL.keys()))

    deleted_count = ClientMaster.query.filter_by(is_deleted=True).count()
    perm = get_perm('crm_clients')
    from permissions import get_sub_perm
    sub_perms = {
        'create_npd': get_sub_perm('crm_clients', 'create_npd'),
        'create_epd': get_sub_perm('crm_clients', 'create_epd'),
        'npd_quote':  get_sub_perm('crm_clients', 'npd_quote'),
    }
    return render_template('crm/clients/clients.html',
        clients=all_clients, search=search, show_trash=show_trash,
        deleted_count=deleted_count,
        city=city, state=state, status_f=status_f,
        sort_by=sort_by, sort_dir=sort_dir,
        all_cities=all_cities, all_states=all_states,
        grid_cols=grid_cols, all_cols=CLIENT_COLS_ALL,
        perm=perm,
        sub_perms=sub_perms,
        active_page='clients')


@crm.route('/clients/add', methods=['GET', 'POST'])
@login_required
def client_add():
    if request.method == 'POST':
        c = ClientMaster(
            code             = gen_code(ClientMaster, 'CLT'),
            contact_name     = request.form.get('contact_name', '').strip(),
            position         = request.form.get('position', '').strip(),
            company_name     = request.form.get('company_name', '').strip(),
            email            = request.form.get('email', '').strip(),
            website          = request.form.get('website', '').strip(),
            mobile           = request.form.get('mobile', '').strip(),
            alternate_mobile = request.form.get('alternate_mobile', '').strip(),
            gstin            = request.form.get('gstin', '').strip().upper(),
            status           = request.form.get('status', 'active'),
            notes            = request.form.get('notes', '').strip(),
            created_by       = current_user.id
        )
        db.session.add(c)
        db.session.flush()

        # Save addresses (multiple)
        addr_titles  = request.form.getlist('addr_title[]')
        addr_types   = request.form.getlist('addr_type[]')
        addr_streets = request.form.getlist('addr_street[]')
        addr_cities  = request.form.getlist('addr_city[]')
        addr_states  = request.form.getlist('addr_state[]')
        addr_countries = request.form.getlist('addr_country[]')
        addr_zips    = request.form.getlist('addr_zip[]')
        addr_defaults= request.form.getlist('addr_default[]')
        for i, title in enumerate(addr_titles):
            if title.strip() or (addr_streets[i] if i < len(addr_streets) else '').strip():
                db.session.add(ClientAddress(
                    client_id  = c.id,
                    title      = title.strip() or 'Address',
                    addr_type  = addr_types[i]    if i < len(addr_types)    else 'billing',
                    address    = addr_streets[i]  if i < len(addr_streets)  else '',
                    city       = addr_cities[i]   if i < len(addr_cities)   else '',
                    state      = addr_states[i]   if i < len(addr_states)   else '',
                    country    = addr_countries[i] if i < len(addr_countries) else 'India',
                    zip_code   = addr_zips[i]     if i < len(addr_zips)     else '',
                    is_default = (str(i) in addr_defaults),
                ))

        # Save brands
        brand_names = request.form.getlist('brand_name[]')
        brand_cats  = request.form.getlist('brand_category[]')
        brand_descs = request.form.getlist('brand_description[]')
        for i, bname in enumerate(brand_names):
            if bname.strip():
                db.session.add(ClientBrand(
                    client_id   = c.id,
                    brand_name  = bname.strip(),
                    category    = brand_cats[i] if i < len(brand_cats) else '',
                    description = brand_descs[i] if i < len(brand_descs) else '',
                ))

        # ── Link client to Lead & NPD Project if came from convert flow ──
        lead_id_link = request.form.get('lead_id_link') or request.args.get('lead_id')
        proj_id_link = request.form.get('proj_id_link') or request.args.get('proj_id')

        if lead_id_link:
            try:
                _lead = Lead.query.get(int(lead_id_link))
                if _lead:
                    _lead.client_id  = c.id
                    _lead.updated_at = datetime.now()
                    _lead.modified_by= current_user.id
                    db.session.add(LeadActivityLog(
                        lead_id    = _lead.id,
                        user_id    = current_user.id,
                        action     = f'Client created & linked: {c.contact_name} (Code: {c.code})',
                        created_at = datetime.now(),
                    ))
            except: pass

        if proj_id_link:
            try:
                from models import NPDProject
                _proj = NPDProject.query.get(int(proj_id_link))
                if _proj:
                    from models.npd import NPDActivityLog
                    db.session.add(NPDActivityLog(
                        project_id = _proj.id,
                        user_id    = current_user.id,
                        action     = f'Client linked: {c.contact_name} (Code: {c.code}, ID: {c.id})',
                        created_at = datetime.now(),
                    ))
            except: pass

        db.session.commit()
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return jsonify(success=True, message=f'Client {c.contact_name} added! (Code: {c.code})', redirect=url_for('crm.clients'))
        flash(f'Client {c.contact_name} added! (Code: {c.code})', 'success')

        # If came from NPD convert flow → go to client view
        if proj_id_link:
            return redirect(url_for('crm.client_view', id=c.id))
        return redirect(url_for('crm.clients'))

    # Pre-fill from URL params (when coming from NPD convert flow)
    _prefill = {
        'contact_name': request.args.get('contact_name', ''),
        'company_name': request.args.get('company_name', ''),
        'email':        request.args.get('email', ''),
        'mobile':       request.args.get('mobile', ''),
        'city':         request.args.get('city', ''),
        'state':        request.args.get('state', ''),
        'lead_id_link': request.args.get('lead_id', ''),
        'proj_id_link': request.args.get('proj_id', ''),
    }
    _from_npd = bool(request.args.get('proj_id'))
    return render_template('crm/clients/client_form.html',
        client=None, brands=[], active_page='clients',
        prefill=_prefill, from_npd=_from_npd)


@crm.route('/clients/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def client_edit(id):
    c = ClientMaster.query.get_or_404(id)
    if request.method == 'POST':
        c.contact_name     = request.form.get('contact_name', '').strip()
        c.position         = request.form.get('position', '').strip()
        c.company_name     = request.form.get('company_name', '').strip()
        c.email            = request.form.get('email', '').strip()
        c.website          = request.form.get('website', '').strip()
        c.mobile           = request.form.get('mobile', '').strip()
        c.alternate_mobile = request.form.get('alternate_mobile', '').strip()
        c.gstin            = request.form.get('gstin', '').strip().upper()
        c.status           = request.form.get('status', 'active')
        c.notes            = request.form.get('notes', '').strip()
        c.updated_at       = datetime.utcnow()

        # Delete old addresses and re-save
        ClientAddress.query.filter_by(client_id=c.id).delete()
        addr_titles   = request.form.getlist('addr_title[]')
        addr_types    = request.form.getlist('addr_type[]')
        addr_streets  = request.form.getlist('addr_street[]')
        addr_cities   = request.form.getlist('addr_city[]')
        addr_states   = request.form.getlist('addr_state[]')
        addr_countries= request.form.getlist('addr_country[]')
        addr_zips     = request.form.getlist('addr_zip[]')
        addr_defaults = request.form.getlist('addr_default[]')
        for i, title in enumerate(addr_titles):
            if title.strip() or (addr_streets[i] if i < len(addr_streets) else '').strip():
                db.session.add(ClientAddress(
                    client_id  = c.id,
                    title      = title.strip() or 'Address',
                    addr_type  = addr_types[i]     if i < len(addr_types)     else 'billing',
                    address    = addr_streets[i]   if i < len(addr_streets)   else '',
                    city       = addr_cities[i]    if i < len(addr_cities)    else '',
                    state      = addr_states[i]    if i < len(addr_states)    else '',
                    country    = addr_countries[i] if i < len(addr_countries) else 'India',
                    zip_code   = addr_zips[i]      if i < len(addr_zips)      else '',
                    is_default = (str(i) in addr_defaults),
                ))

        # Delete old brands and re-save
        ClientBrand.query.filter_by(client_id=c.id).delete()
        brand_names = request.form.getlist('brand_name[]')
        brand_cats  = request.form.getlist('brand_category[]')
        brand_descs = request.form.getlist('brand_description[]')
        for i, bname in enumerate(brand_names):
            if bname.strip():
                db.session.add(ClientBrand(
                    client_id   = c.id,
                    brand_name  = bname.strip(),
                    category    = brand_cats[i] if i < len(brand_cats) else '',
                    description = brand_descs[i] if i < len(brand_descs) else '',
                ))

        db.session.commit()
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return jsonify(success=True, message='Client updated successfully!', redirect=url_for('crm.clients'))
        flash('Client updated successfully!', 'success')
        return redirect(url_for('crm.clients'))

    brands = ClientBrand.query.filter_by(client_id=c.id).all()
    return render_template('crm/clients/client_form.html', client=c, brands=brands, active_page='clients')


@crm.route('/clients/<int:id>')
@login_required
def client_view(id):
    c = ClientMaster.query.get_or_404(id)
    brands = ClientBrand.query.filter_by(client_id=c.id).all()
    from permissions import get_sub_perm
    perm = get_perm('crm_clients')
    sub_perms = {
        'create_npd': get_sub_perm('crm_clients', 'create_npd'),
        'create_epd': get_sub_perm('crm_clients', 'create_epd'),
        'npd_quote':  get_sub_perm('crm_clients', 'npd_quote'),
    }
    return render_template('crm/clients/client_view.html', client=c, brands=brands,
        perm=perm, sub_perms=sub_perms, active_page='clients')


@crm.route('/clients/<int:id>/delete', methods=['POST'])
@login_required
def client_delete(id):
    c = ClientMaster.query.get_or_404(id)
    name = c.contact_name
    c.is_deleted = True
    c.deleted_at = datetime.utcnow()
    db.session.commit()
    flash(f'Client "{name}" moved to trash.', 'warning')
    return redirect(url_for('crm.clients'))


@crm.route('/clients/<int:id>/restore', methods=['POST'])
@login_required
def client_restore(id):
    c = ClientMaster.query.get_or_404(id)
    c.is_deleted = False
    c.deleted_at = None
    db.session.commit()
    flash(f'Client "{c.contact_name}" restored successfully!', 'success')
    return redirect(url_for('crm.clients', trash=1))


# ══════════════════════════════════════
# LEAD ROUTES
# ══════════════════════════════════════

@crm.route('/leads')
@login_required
def leads():
    # Basic filters
    status   = request.args.get('status', '')
    search   = request.args.get('search', '')
    # Advanced filters
    source   = request.args.get('source', '')
    category = request.args.get('category', '')
    p_range  = request.args.get('product_range', '')
    city     = request.args.get('city', '')
    date_from= request.args.get('date_from', '')
    date_to  = request.args.get('date_to', '')
    # Performance dashboard filters
    assigned_to_filter = request.args.get('assigned_to_filter', '')
    perf_period        = request.args.get('perf_period', '')
    # Sorting
    sort_by  = request.args.get('sort_by', 'created_at')
    sort_dir = request.args.get('sort_dir', 'desc')

    show_trash = request.args.get('trash', '') == '1'
    query = Lead.query.filter_by(is_deleted=True) if show_trash else Lead.query.filter_by(is_deleted=False)

    # ── Performance period filter ──
    if perf_period and not date_from and not date_to:
        from datetime import date as _date
        today = datetime.now().date()
        if perf_period == 'today':
            pf = datetime.combine(today, datetime.min.time())
            pt = datetime.combine(today, datetime.max.time())
        elif perf_period == 'this_week':
            pf = datetime.combine(today - timedelta(days=today.weekday()), datetime.min.time())
            pt = datetime.combine(today, datetime.max.time())
        elif perf_period == 'this_month':
            pf = datetime(today.year, today.month, 1)
            pt = datetime.combine(today, datetime.max.time())
        elif perf_period == 'last_month':
            first = today.replace(day=1)
            last_prev = first - timedelta(days=1)
            pf = datetime(last_prev.year, last_prev.month, 1)
            pt = datetime.combine(last_prev, datetime.max.time())
        elif perf_period == 'last_3_months':
            pf = datetime.combine(today - timedelta(days=90), datetime.min.time())
            pt = datetime.combine(today, datetime.max.time())
        elif perf_period == 'last_6_months':
            pf = datetime.combine(today - timedelta(days=180), datetime.min.time())
            pt = datetime.combine(today, datetime.max.time())
        elif perf_period == 'this_year':
            pf = datetime(today.year, 1, 1)
            pt = datetime.combine(today, datetime.max.time())
        else:
            pf = pt = None
        if pf and pt:
            query = query.filter(Lead.created_at >= pf, Lead.created_at <= pt)

    # ── Assigned to filter (from performance dashboard) ──
    if assigned_to_filter:
        try:
            uid = int(assigned_to_filter)
            query = query.filter(
                db.or_(
                    Lead.assigned_to == uid,
                    Lead.team_members.like(f'%{uid}%')
                )
            )
        except ValueError:
            pass

    if not show_trash:
        if status:   query = query.filter_by(status=status)
    if source:   query = query.filter_by(source=source)
    if category: query = query.filter_by(category=category)
    if p_range:  query = query.filter_by(product_range=p_range)
    if city:     query = query.filter(Lead.city.ilike(f'%{city}%'))
    if date_from:
        query = query.filter(Lead.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(Lead.created_at <= datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    if search:
        s = f'%{search}%'
        query = query.filter(
            Lead.contact_name.ilike(s)    |
            Lead.company_name.ilike(s)    |
            Lead.phone.ilike(s)           |
            Lead.alternate_mobile.ilike(s)|
            Lead.email.ilike(s)           |
            Lead.product_name.ilike(s)    |
            Lead.category.ilike(s)        |
            Lead.product_range.ilike(s)   |
            Lead.source.ilike(s)          |
            Lead.city.ilike(s)            |
            Lead.state.ilike(s)           |
            Lead.country.ilike(s)         |
            Lead.zip_code.ilike(s)        |
            Lead.address.ilike(s)         |
            Lead.position.ilike(s)        |
            Lead.title.ilike(s)           |
            Lead.tags.ilike(s)            |
            Lead.remark.ilike(s)          |
            Lead.notes.ilike(s)           |
            Lead.lost_reason.ilike(s)     |
            Lead.requirement_spec.ilike(s)|
            Lead.order_quantity.ilike(s)
        )

    # Sort
    # Map frontend sort names to actual DB columns
    sort_map = {
        'name':    'contact_name',
        'mobile':  'phone',
        'company': 'company_name',
    }
    actual_sort = sort_map.get(sort_by, sort_by)
    sort_col = getattr(Lead, actual_sort, Lead.created_at)
    if sort_dir == 'asc':
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    all_leads = query.all()

    # Dynamic counts — all statuses from DB
    _all_statuses = LeadStatus.query.filter_by(is_active=True).order_by(LeadStatus.sort_order).all()
    counts = {st.name: Lead.query.filter_by(status=st.name, is_deleted=False).count()
              for st in _all_statuses}
    # Always include core statuses for backward compat
    for _s in ('open','in_process','close','cancel'):
        counts.setdefault(_s, Lead.query.filter_by(status=_s, is_deleted=False).count())
    deleted_count = Lead.query.filter_by(is_deleted=True).count()

    # Filter options
    all_sources   = [r[0] for r in db.session.query(Lead.source).distinct().all() if r[0]]
    all_categories= [r[0] for r in db.session.query(Lead.category).distinct().all() if r[0]]
    all_ranges    = [r[0] for r in db.session.query(Lead.product_range).distinct().all() if r[0]]
    all_cities    = [r[0] for r in db.session.query(Lead.city).distinct().all() if r[0]]
    all_users     = User.query.filter_by(is_active=True).all()
    grid_cols     = get_grid_columns('leads', LEAD_COLS_DEFAULT, list(LEAD_COLS_ALL.keys()))
    lead_statuses  = LeadStatus.query.filter_by(is_active=True).order_by(LeadStatus.sort_order).all()
    lead_sources   = LeadSource.query.filter_by(is_active=True).order_by(LeadSource.sort_order).all()
    lead_categories= LeadCategory.query.filter_by(is_active=True).order_by(LeadCategory.sort_order).all()
    product_ranges = ProductRange.query.filter_by(is_active=True).order_by(ProductRange.sort_order).all()

    perm = get_perm('crm_leads')
    return render_template('crm/leads/leads.html',
        leads=all_leads, counts=counts, deleted_count=deleted_count,
        show_trash=show_trash, all_users=all_users,
        status=status, search=search,
        source=source, category=category, p_range=p_range,
        city=city, date_from=date_from, date_to=date_to,
        sort_by=sort_by, sort_dir=sort_dir,
        all_sources=all_sources, all_categories=all_categories,
        all_ranges=all_ranges, all_cities=all_cities,
        lead_statuses=lead_statuses, lead_sources=lead_sources,
        lead_categories=lead_categories, product_ranges=product_ranges,
        grid_cols=grid_cols, all_cols=LEAD_COLS_ALL,
        perm=perm,
        active_page='leads')




@crm.route('/leads/export')
@login_required
def leads_export():
    """Export filtered leads to Excel with ALL fields."""
    import io, sys, subprocess
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'], check=True)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    # ── Apply same filters as leads list ──
    status   = request.args.get('status', '')
    search   = request.args.get('search', '')
    source   = request.args.get('source', '')
    category = request.args.get('category', '')
    p_range  = request.args.get('product_range', '')
    city     = request.args.get('city', '')
    date_from= request.args.get('date_from', '')
    date_to  = request.args.get('date_to', '')
    sort_by  = request.args.get('sort_by', 'created_at')
    sort_dir = request.args.get('sort_dir', 'desc')

    query = Lead.query
    if status:   query = query.filter_by(status=status)
    if source:   query = query.filter_by(source=source)
    if category: query = query.filter_by(category=category)
    if p_range:  query = query.filter_by(product_range=p_range)
    if city:     query = query.filter(Lead.city.ilike(f'%{city}%'))
    if date_from:
        query = query.filter(Lead.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(Lead.created_at <= datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    if search:
        s = f'%{search}%'
        query = query.filter(
            Lead.contact_name.ilike(s)    | Lead.company_name.ilike(s)    |
            Lead.phone.ilike(s)           | Lead.alternate_mobile.ilike(s)|
            Lead.email.ilike(s)           | Lead.product_name.ilike(s)    |
            Lead.category.ilike(s)        | Lead.product_range.ilike(s)   |
            Lead.source.ilike(s)          | Lead.city.ilike(s)            |
            Lead.state.ilike(s)           | Lead.country.ilike(s)         |
            Lead.zip_code.ilike(s)        | Lead.address.ilike(s)         |
            Lead.position.ilike(s)        | Lead.title.ilike(s)           |
            Lead.tags.ilike(s)            | Lead.remark.ilike(s)          |
            Lead.notes.ilike(s)           | Lead.lost_reason.ilike(s)     |
            Lead.requirement_spec.ilike(s)| Lead.order_quantity.ilike(s)
        )
    sort_map = {'name': 'contact_name', 'mobile': 'phone', 'company': 'company_name'}
    actual_sort = sort_map.get(sort_by, sort_by)
    sort_col = getattr(Lead, actual_sort, Lead.created_at)
    query = query.order_by(sort_col.asc() if sort_dir == 'asc' else sort_col.desc())
    leads_data = query.all()

    # User lookup for assigned_to / created_by
    from models.user import User as UserModel
    users = {u.id: u.full_name for u in UserModel.query.all()}

    # ── Build Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D7E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # All columns — getattr used so missing DB columns don't crash
    def _s(l, f, default=''):
        v = getattr(l, f, None)
        return v if v is not None else default

    def _d(l, f, fmt='%d-%m-%Y'):
        v = getattr(l, f, None)
        return v.strftime(fmt) if v else ''

    def _f(l, f):
        v = getattr(l, f, None)
        return float(v) if v else ''

    COLUMNS = [
        ("Title",            lambda l: _s(l,'title')),
        ("Contact Name",     lambda l: _s(l,'contact_name')),
        ("Company",          lambda l: _s(l,'company_name')),
        ("Position",         lambda l: _s(l,'position')),
        ("Email",            lambda l: _s(l,'email')),
        ("Mobile",           lambda l: _s(l,'phone')),
        ("Alternate Mobile", lambda l: _s(l,'alternate_mobile')),
        ("Website",          lambda l: _s(l,'website')),
        ("Address",          lambda l: _s(l,'address')),
        ("City",             lambda l: _s(l,'city')),
        ("State",            lambda l: _s(l,'state')),
        ("Country",          lambda l: _s(l,'country')),
        ("Zip Code",         lambda l: _s(l,'zip_code')),
        ("Source",           lambda l: _s(l,'source')),
        ("Category",         lambda l: _s(l,'category')),
        ("Product Range",    lambda l: _s(l,'product_range')),
        ("Product Name",     lambda l: _s(l,'product_name')),
        ("Order Quantity",   lambda l: _s(l,'order_quantity')),
        ("Requirement Spec", lambda l: _s(l,'requirement_spec')),
        ("Status",           lambda l: _s(l,'status','').replace('_',' ').title()),
        ("Lead Type",        lambda l: _s(l,'lead_type','Quality')),
        ("Priority",         lambda l: _s(l,'priority','').title()),
        ("Expected Value",   lambda l: _f(l,'expected_value')),
        ("Average Cost",     lambda l: _f(l,'average_cost')),
        ("Tags",             lambda l: _s(l,'tags')),
        ("Remark",           lambda l: _s(l,'remark')),
        ("Notes",            lambda l: _s(l,'notes')),
        ("Lost Reason",      lambda l: _s(l,'lost_reason')),
        ("Assigned To",      lambda l: users.get(l.assigned_to,'') if getattr(l,'assigned_to',None) else ''),
        ("Follow Up Date",   lambda l: _d(l,'follow_up_date')),
        ("Last Contact",     lambda l: _d(l,'last_contact','%d-%m-%Y %H:%M')),
        ("Created By",       lambda l: users.get(l.created_by,'') if getattr(l,'created_by',None) else ''),
        ("Created At",       lambda l: _d(l,'created_at','%d-%m-%Y %H:%M')),
        ("Closed At",        lambda l: _d(l,'closed_at','%d-%m-%Y %H:%M')),
        ("Lead Age (Days)",  lambda l: l.lead_age),
        ("Updated At",       lambda l: _d(l,'updated_at','%d-%m-%Y %H:%M')),
    ]

    # Write header row
    ws.row_dimensions[1].height = 32
    for col_idx, (label, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font    = hdr_font
        cell.fill    = hdr_fill
        cell.alignment = hdr_align
        cell.border  = border

    # Write data rows
    alt_fill = PatternFill("solid", fgColor="F0F4FA")
    data_font = Font(size=9)
    data_align = Alignment(vertical="center", wrap_text=False)

    for row_idx, lead in enumerate(leads_data, 2):
        row_fill = alt_fill if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 18
        for col_idx, (_, getter) in enumerate(COLUMNS, 1):
            try:
                val = getter(lead)
            except Exception:
                val = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = border
            if row_fill:
                cell.fill = row_fill

    # Auto column widths
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 40))
                except:
                    pass
        ws.column_dimensions[col_letter].width = max_len + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    # Filter info sheet
    ws2 = wb.create_sheet("Filter Info")
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 40
    info_rows = [
        ("Exported At",    datetime.now().strftime('%d-%m-%Y %H:%M:%S')),
        ("Total Records",  len(leads_data)),
        ("Status Filter",  status or 'All'),
        ("Search",         search or '—'),
        ("Source",         source or 'All'),
        ("Category",       category or 'All'),
        ("Product Range",  p_range or 'All'),
        ("City",           city or 'All'),
        ("Date From",      date_from or '—'),
        ("Date To",        date_to or '—'),
        ("Sort By",        sort_by),
        ("Sort Direction", 'Ascending' if sort_dir == 'asc' else 'Descending'),
    ]
    for r, (k, v) in enumerate(info_rows, 1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=str(v))

    # Output
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)



# ══════════════════════════════════════════════════════════
# SHARED EXCEL HELPER
# ══════════════════════════════════════════════════════════

def _make_excel(ws, headers, rows):
    """Write styled header + data rows to openpyxl worksheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="D0D7E2")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill("solid", fgColor="F0F4FA")
    d_font    = Font(size=9)
    d_align   = Alignment(vertical="center")

    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = bdr

    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 17
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = d_font; cell.alignment = d_align; cell.border = bdr
            if fill: cell.fill = fill

    for ci in range(1, len(headers)+1):
        col = get_column_letter(ci)
        mx = max((len(str(ws.cell(r, ci).value or '')) for r in range(1, ws.max_row+1)), default=10)
        ws.column_dimensions[col].width = min(mx + 2, 45)
    ws.freeze_panes = "A2"


@crm.route('/clients/export')
@login_required
def clients_export():
    import io, sys, subprocess
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'], check=True)
        import openpyxl

    search      = request.args.get('search', '')
    city        = request.args.get('city', '')
    state       = request.args.get('state', '')
    status_f    = request.args.get('status_f', '')
    sort_by     = request.args.get('sort_by', 'created_at')
    sort_dir    = request.args.get('sort_dir', 'desc')

    q = ClientMaster.query
    if search:
        s = f'%{search}%'
        q = q.filter(ClientMaster.contact_name.ilike(s)|ClientMaster.company_name.ilike(s)|
                     ClientMaster.mobile.ilike(s)|ClientMaster.email.ilike(s)|
                     ClientMaster.city.ilike(s)|ClientMaster.gstin.ilike(s))
    if city:        q = q.filter(ClientMaster.city.ilike(f'%{city}%'))
    if state:       q = q.filter(ClientMaster.state.ilike(f'%{state}%'))
    if status_f:    q = q.filter_by(status=status_f)
    sort_col = getattr(ClientMaster, sort_by, ClientMaster.created_at)
    q = q.order_by(sort_col.asc() if sort_dir=='asc' else sort_col.desc())
    clients = q.all()

    from models.user import User as UserModel
    users = {u.id: u.full_name for u in UserModel.query.all()}

    headers = ["Code","Company","Contact Name","Position","Email","Mobile","Alt Mobile",
               "GSTIN","Status","Address","City","State","Country","Zip Code",
               "Notes","Brands","Created By","Created At","Updated At"]

    rows = []
    for cl in clients:
        brands = ', '.join([b.brand_name for b in cl.brands]) if hasattr(cl,'brands') else ''
        rows.append([
            cl.code or '', cl.company_name or '', cl.contact_name or '', cl.position or '',
            cl.email or '', cl.mobile or '', cl.alternate_mobile or '', cl.gstin or '',
(cl.status or '').title(),
            cl.address or '', cl.city or '', cl.state or '', cl.country or '', cl.zip_code or '',
            cl.notes or '', brands,
            users.get(cl.created_by,'') if cl.created_by else '',
            cl.created_at.strftime('%d-%m-%Y %H:%M') if cl.created_at else '',
            cl.updated_at.strftime('%d-%m-%Y %H:%M') if cl.updated_at else '',
        ])

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Clients"
    _make_excel(ws, headers, rows)

    ws2 = wb.create_sheet("Filter Info")
    for r,(k,v) in enumerate([
        ("Exported At", datetime.now().strftime('%d-%m-%Y %H:%M')),
        ("Total",       len(clients)),
        ("Search",      search or '—'), ("City", city or '—'),
        ("Status",      status_f or 'All'),
    ], 1):
        from openpyxl.styles import Font as F2
        ws2.cell(r,1,k).font = F2(bold=True); ws2.cell(r,2,str(v))

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"clients_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

@crm.route('/leads/grid-config', methods=['POST'])
@login_required
def lead_grid_config():
    cols = request.json.get('cols', [])
    save_grid_columns('leads', cols)
    return jsonify(success=True)


@crm.route('/clients/grid-config', methods=['POST'])
@login_required
def client_grid_config():
    cols = request.json.get('cols', [])
    save_grid_columns('clients', cols)
    return jsonify(success=True)


@crm.route('/leads/<int:id>/update-status', methods=['POST'])
@login_required
def lead_update_status(id):
    """Kanban drag-drop status update"""
    lead = Lead.query.get_or_404(id)
    data = request.get_json()
    new_status = data.get('status', '').strip()
    # Load valid statuses dynamically from DB
    valid = {st.name for st in LeadStatus.query.filter_by(is_active=True).all()}
    valid.update({'open', 'in_process', 'close', 'cancel'})  # always allow core statuses
    if new_status not in valid:
        return jsonify(success=False, error='Invalid status'), 400
    old_status = lead.status
    lead.status = new_status
    lead.updated_at = datetime.now()
    log_activity(id, f'Status changed: {old_status} → {new_status}')
    audit('leads','KANBAN', id, f'{lead.contact_name}', f'Kanban: {old_status} → {new_status}')
    lead.modified_by = current_user.id
    db.session.commit()
    return jsonify(success=True, id=id, status=new_status)



@crm.route('/leads/<int:id>/inline-edit', methods=['POST'])
@login_required
def lead_inline_edit(id):
    perm = get_perm('crm_leads')
    if not perm or not perm.can_edit:
        return jsonify(success=False, error='Edit permission nahi hai'), 403
    lead = Lead.query.get_or_404(id)
    data = request.get_json()
    field = data.get('field','').strip()
    value = data.get('value','')
    allowed = {
        'contact_name','company_name','email','phone','city','state',
        'product_name','category','source','status','priority',
        'product_range','order_quantity','remark','tags'
    }
    if field not in allowed:
        return jsonify(success=False, error='Field not allowed'), 400
    old_val = getattr(lead, field, None)
    setattr(lead, field, value.strip() if isinstance(value, str) else value)
    lead.updated_at = datetime.now()
    lead.modified_by = current_user.id
    log_activity(id, f'Inline edit: {field} changed')
    db.session.commit()
    return jsonify(success=True, field=field, value=value)


@crm.route('/clients/<int:id>/inline-edit', methods=['POST'])
@login_required
def client_inline_edit(id):
    perm = get_perm('crm_clients')
    if not perm or not perm.can_edit:
        return jsonify(success=False, error='Edit permission nahi hai'), 403
    c = ClientMaster.query.get_or_404(id)
    data = request.get_json()
    field = data.get('field','').strip()
    value = data.get('value','')
    allowed = {
        'contact_name','company_name','mobile','alternate_mobile',
        'email','website','city','state','gstin','status','notes'
    }
    if field not in allowed:
        return jsonify(success=False, error='Field not allowed'), 400
    setattr(c, field, value.strip() if isinstance(value, str) else value)
    c.updated_at = datetime.now()
    db.session.commit()
    return jsonify(success=True, field=field, value=value)


@crm.route('/leads/add', methods=['GET', 'POST'])
@login_required
def lead_add():
    if request.method == 'POST':
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        try:
            # Get team members
            team_ids = request.form.getlist('team_members[]')
            team_str = ','.join(team_ids) if team_ids else ''

            pname = request.form.get('name', '').strip()
            if not pname:
                if is_ajax:
                    return jsonify(success=False, message='Lead name is required.')
                flash('Lead name is required.', 'danger')
                return redirect(request.url)

            l = Lead(
                code             = gen_code(Lead, 'LD'),
                contact_name     = pname,
                title            = request.form.get('product_name', pname).strip() or pname,
                position         = request.form.get('position', '').strip(),
                email            = request.form.get('email', '').strip(),
                website          = request.form.get('website', '').strip(),
                phone            = request.form.get('mobile', '').strip(),
                alternate_mobile = request.form.get('alternate_mobile', '').strip(),
                company_name     = request.form.get('company', '').strip(),
                address          = request.form.get('address', '').strip(),
                city             = request.form.get('city', '').strip(),
                state            = request.form.get('state', '').strip(),
                country          = request.form.get('country', 'India').strip(),
                zip_code         = request.form.get('zip_code', '').strip(),
                average_cost     = request.form.get('average_cost') or 0,
                product_name     = request.form.get('product_name', '').strip(),
                category         = request.form.get('category', '').strip(),
                product_range    = request.form.get('product_range', '').strip(),
                order_quantity   = request.form.get('order_quantity', '').strip(),
                requirement_spec = request.form.get('requirement_spec', '').strip(),
                tags             = request.form.get('tags', '').strip(),
                remark           = request.form.get('remark', '').strip(),
                source           = request.form.get('source', '').strip(),
                status           = request.form.get('status', 'open'),
                lead_type        = request.form.get('lead_type', 'Quality'),
                follow_up_date   = datetime.strptime(request.form['follow_up_date'], '%Y-%m-%d').date()
                                   if request.form.get('follow_up_date') else None,
                team_members     = team_str,
                client_id        = request.form.get('client_id') or None,
                created_by       = current_user.id
            )
            db.session.add(l)
            db.session.flush()
            log_activity(l.id, 'New Lead Added')
            db.session.commit()
            if is_ajax:
                return jsonify(success=True, message=f'Lead {l.contact_name} added successfully!', redirect=url_for('crm.lead_view', id=l.id))
            flash(f'Lead {l.contact_name} added!', 'success')
            return redirect(url_for('crm.lead_view', id=l.id))
        except Exception as e:
            db.session.rollback()
            import traceback; traceback.print_exc()
            err_msg = f'Error saving lead: {str(e)}'
            if is_ajax:
                return jsonify(success=False, message=err_msg)
            flash(err_msg, 'danger')
            return redirect(request.url)

    clients      = ClientMaster.query.filter_by(status='active').order_by(ClientMaster.contact_name).all()
    all_users    = User.query.filter_by(is_active=True).all()
    lead_statuses  = LeadStatus.query.filter_by(is_active=True).order_by(LeadStatus.sort_order).all()
    lead_sources   = LeadSource.query.filter_by(is_active=True).order_by(LeadSource.sort_order).all()
    lead_categories= LeadCategory.query.filter_by(is_active=True).order_by(LeadCategory.sort_order).all()
    product_ranges = ProductRange.query.filter_by(is_active=True).order_by(ProductRange.sort_order).all()
    return render_template('crm/leads/lead_form.html', lead=None, clients=clients,
                           all_users=all_users, lead_statuses=lead_statuses,
                           lead_sources=lead_sources, lead_categories=lead_categories,
                           product_ranges=product_ranges, active_page='leads')


@crm.route('/leads/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def lead_edit(id):
    l = Lead.query.get_or_404(id)
    if request.method == 'POST':
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        try:
            team_ids = request.form.getlist('team_members[]')
            team_str = ','.join(team_ids) if team_ids else ''
            from audit_helper import model_to_dict
            _old_snap = model_to_dict(l)

            l.contact_name     = request.form.get('name', '').strip()
            l.title            = request.form.get('product_name', l.contact_name).strip() or l.contact_name
            l.position         = request.form.get('position', '').strip()
            l.email            = request.form.get('email', '').strip()
            l.website          = request.form.get('website', '').strip()
            l.phone            = request.form.get('mobile', '').strip()
            l.alternate_mobile = request.form.get('alternate_mobile', '').strip()
            l.company_name     = request.form.get('company', '').strip()
            l.address          = request.form.get('address', '').strip()
            l.city             = request.form.get('city', '').strip()
            l.state            = request.form.get('state', '').strip()
            l.country          = request.form.get('country', 'India').strip()
            l.zip_code         = request.form.get('zip_code', '').strip()
            l.average_cost     = request.form.get('average_cost') or 0
            l.product_name     = request.form.get('product_name', '').strip()
            l.category         = request.form.get('category', '').strip()
            l.product_range    = request.form.get('product_range', '').strip()
            l.order_quantity   = request.form.get('order_quantity', '').strip()
            l.requirement_spec = request.form.get('requirement_spec', '').strip()
            l.tags             = request.form.get('tags', '').strip()
            l.remark           = request.form.get('remark', '').strip()
            l.source           = request.form.get('source', '').strip()
            l.status           = request.form.get('status', 'open')
            l.lead_type        = request.form.get('lead_type', 'Quality')
            l.follow_up_date   = None  # removed from form
            l.team_members     = team_str
            l.client_id        = request.form.get('client_id') or None
            l.updated_at       = datetime.utcnow()

            log_activity(l.id, f'Lead Record Updated')
            add_contribution(l.id, 'edit', note='Lead record updated')
            db.session.commit()
            if is_ajax:
                return jsonify(success=True, message='Lead updated successfully!', redirect=url_for('crm.lead_view', id=l.id))
            flash('Lead updated!', 'success')
            return redirect(url_for('crm.lead_view', id=l.id))
        except Exception as e:
            db.session.rollback()
            import traceback; traceback.print_exc()
            err_msg = f'Error updating lead: {str(e)}'
            if is_ajax:
                return jsonify(success=False, message=err_msg)
            flash(err_msg, 'danger')
            return redirect(request.url)

    clients        = ClientMaster.query.filter_by(status='active').order_by(ClientMaster.contact_name).all()
    all_users      = User.query.filter_by(is_active=True).all()
    lead_statuses  = LeadStatus.query.filter_by(is_active=True).order_by(LeadStatus.sort_order).all()
    lead_sources   = LeadSource.query.filter_by(is_active=True).order_by(LeadSource.sort_order).all()
    lead_categories= LeadCategory.query.filter_by(is_active=True).order_by(LeadCategory.sort_order).all()
    product_ranges = ProductRange.query.filter_by(is_active=True).order_by(ProductRange.sort_order).all()
    return render_template('crm/leads/lead_form.html', lead=l, clients=clients,
                           all_users=all_users, lead_statuses=lead_statuses,
                           lead_sources=lead_sources, lead_categories=lead_categories,
                           product_ranges=product_ranges, active_page='leads')


@crm.route('/leads/<int:id>/delete', methods=['POST'])
@login_required
def lead_delete(id):
    l = Lead.query.get_or_404(id)
    name = l.contact_name
    l.is_deleted = True
    l.deleted_at = datetime.utcnow()
    db.session.commit()
    flash(f'Lead "{name}" moved to trash.', 'warning')
    return redirect(url_for('crm.leads'))


@crm.route('/leads/<int:id>/restore', methods=['POST'])
@login_required
def lead_restore(id):
    l = Lead.query.get_or_404(id)
    l.is_deleted = False
    l.deleted_at = None
    db.session.commit()
    flash(f'Lead "{l.contact_name}" restored successfully!', 'success')
    return redirect(url_for('crm.leads', trash=1))


@crm.route('/leads/<int:id>')
@login_required
def lead_view(id):
    l           = Lead.query.get_or_404(id)
    tab         = request.args.get('tab', 'overview')
    discussions = LeadDiscussion.query.filter_by(lead_id=id).order_by(LeadDiscussion.created_at.desc()).all()
    reminders   = LeadReminder.query.filter_by(lead_id=id).order_by(LeadReminder.remind_at).all()
    notes_list  = LeadNote.query.filter_by(lead_id=id, user_id=current_user.id).order_by(LeadNote.created_at.desc()).all()
    activity    = LeadActivityLog.query.filter_by(lead_id=id).order_by(LeadActivityLog.created_at.desc()).all()
    attachments = LeadAttachment.query.filter_by(lead_id=id).order_by(LeadAttachment.created_at.desc()).all()
    team_members = l.get_team_member_objects()
    all_users   = User.query.filter_by(is_active=True).all()
    audit('leads','VIEW', id, f'{l.contact_name}', obj=l)
    _view_statuses = LeadStatus.query.filter_by(is_active=True).order_by(LeadStatus.sort_order).all()

    # Permissions
    perm = get_perm('crm_leads')
    from permissions import get_sub_perm
    sub_perms = {
        'discussion_board' : get_sub_perm('crm_leads', 'discussion_board'),
        'activity_log'     : get_sub_perm('crm_leads', 'activity_log'),
        'reminder'         : get_sub_perm('crm_leads', 'reminder'),
        'quotation'        : get_sub_perm('crm_leads', 'quotation'),
        'sample_order'     : get_sub_perm('crm_leads', 'sample_order'),
        'attachments'      : get_sub_perm('crm_leads', 'attachments'),
        'whatsapp'         : get_sub_perm('crm_leads', 'whatsapp'),
        'personal_notes'   : True,
    }

    return render_template('crm/leads/lead_view.html',
        lead=l, tab=tab,
        discussions=discussions, reminders=reminders,
        notes_list=notes_list, activity=activity,
        attachments=attachments, team_members=team_members,
        all_users=all_users,
        lead_statuses=_view_statuses,
        perm=perm,
        sub_perms=sub_perms,
        now=datetime.utcnow(),
        active_page='leads')


# ── Discussion Board ──

@crm.route('/leads/<int:id>/discussion/add', methods=['POST'])
@login_required
def lead_discussion_add(id):
    l       = Lead.query.get_or_404(id)
    comment = request.form.get('comment', '').strip()
    if not comment:
        flash('Comment cannot be empty!', 'warning')
        return redirect(url_for('crm.lead_view', id=id, tab='discussion'))

    disc = LeadDiscussion(
        lead_id    = id,
        user_id    = current_user.id,
        comment    = comment,
        created_at = datetime.utcnow()
    )
    db.session.add(disc)
    db.session.flush()

    # Handle file attachments
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    files = request.files.getlist('attachments[]')
    for f in files:
        if f and f.filename and allowed_file(f.filename):
            fname = secure_filename(f.filename)
            fpath = os.path.join(UPLOAD_FOLDER, fname)
            f.save(fpath)
            att = LeadAttachment(
                lead_id       = id,
                discussion_id = disc.id,
                file_name     = fname,
                file_path     = f'uploads/{fname}',
                file_type     = f.content_type,
                uploaded_by   = current_user.id
            )
            db.session.add(att)

    # Update last contact
    l.last_contact = datetime.utcnow()
    log_activity(id, 'New Comment Added in Lead')
    add_contribution(id, 'comment', note='Added comment')
    db.session.commit()
    audit('leads','DISCUSSION', id, f'Lead #{id}', f'Comment added by {current_user.username}')
    flash('Comment added!', 'success')
    return redirect(url_for('crm.lead_view', id=id, tab='discussion'))


# ── Reminders (Followup) ──

@crm.route('/leads/<int:id>/reminder/add', methods=['POST'])
@login_required
def lead_reminder_add(id):
    Lead.query.get_or_404(id)
    title      = request.form.get('title', '').strip()
    remind_str = request.form.get('remind_at', '')
    if not title or not remind_str:
        flash('Title and reminder date/time required!', 'warning')
        return redirect(url_for('crm.lead_view', id=id, tab='reminder'))

    try:
        remind_dt = datetime.strptime(remind_str, '%Y-%m-%dT%H:%M')
    except:
        remind_dt = datetime.strptime(remind_str, '%Y-%m-%d %H:%M')

    r = LeadReminder(
        lead_id     = id,
        user_id     = current_user.id,
        title       = title,
        description = request.form.get('description', '').strip(),
        remind_at   = remind_dt
    )
    db.session.add(r)
    log_activity(id, f'Reminder set: {title}')
    add_contribution(id, 'reminder', note=f'Reminder: {title}')
    db.session.commit()
    audit('leads','REMINDER', id, f'Lead #{id}', f'Reminder set: {title}')
    flash('Reminder added!', 'success')
    return redirect(url_for('crm.lead_view', id=id, tab='reminder'))


@crm.route('/leads/reminder/<int:rid>/done', methods=['POST'])
@login_required
def lead_reminder_done(rid):
    r = LeadReminder.query.get_or_404(rid)
    r.is_done = not r.is_done
    db.session.commit()
    return redirect(url_for('crm.lead_view', id=r.lead_id, tab='reminder'))


@crm.route('/leads/reminder/<int:rid>/delete', methods=['POST'])
@login_required
def lead_reminder_delete(rid):
    r = LeadReminder.query.get_or_404(rid)
    lid = r.lead_id
    db.session.delete(r)
    db.session.commit()
    flash('Reminder deleted!', 'success')
    return redirect(url_for('crm.lead_view', id=lid, tab='reminder'))


# ── Personal Notes ──

@crm.route('/leads/<int:id>/note/add', methods=['POST'])
@login_required
def lead_note_add(id):
    Lead.query.get_or_404(id)
    note_text = request.form.get('note', '').strip()
    if not note_text:
        flash('Note cannot be empty!', 'warning')
        return redirect(url_for('crm.lead_view', id=id, tab='notes'))

    n = LeadNote(lead_id=id, user_id=current_user.id, note=note_text)
    db.session.add(n)
    db.session.commit()
    audit('leads','NOTE', id, f'Lead #{id}', f'Note added by {current_user.username}')
    flash('Note saved!', 'success')
    return redirect(url_for('crm.lead_view', id=id, tab='notes'))


@crm.route('/leads/note/<int:nid>/delete', methods=['POST'])
@login_required
def lead_note_delete(nid):
    n = LeadNote.query.get_or_404(nid)
    lid = n.lead_id
    db.session.delete(n)
    db.session.commit()
    flash('Note deleted!', 'success')
    return redirect(url_for('crm.lead_view', id=lid, tab='notes'))


# ── Status Change ──

@crm.route('/leads/<int:id>/status', methods=['POST'])
@login_required
def lead_status_change(id):
    l = Lead.query.get_or_404(id)
    new_status = request.form.get('status')
    _valid_statuses = {st.name for st in LeadStatus.query.filter_by(is_active=True).all()}
    _valid_statuses.update({'open', 'in_process', 'close', 'cancel'})
    if new_status in _valid_statuses:
        old = l.status
        l.status = new_status
        if new_status in ('close', 'cancel') and not l.closed_at:
            l.closed_at = datetime.now()
        elif new_status in ('open', 'in_process'):
            l.closed_at = None  # reopen hone pe reset
        log_activity(id, f'Status changed: {old} → {new_status}')

        if new_status == 'close':
            _handle_close_contribution(l)
        elif new_status == 'cancel':
            add_contribution(id, 'cancel', note=f'Status: {old} → cancel')
        else:
            add_contribution(id, 'status_change', note=f'Status: {old} → {new_status}')

        db.session.commit()
    return redirect(url_for('crm.lead_view', id=id))


# ══════════════════════════════════════
# LEGACY CUSTOMER ROUTES (kept)
# ══════════════════════════════════════

@crm.route('/customers')
@login_required
def customers():
    return redirect(url_for('crm.clients'))


@crm.route('/customers/add')
@login_required
def customer_add():
    return redirect(url_for('crm.client_add'))


# ══════════════════════════════════════
# DASHBOARD API (JSON for Charts)
# ══════════════════════════════════════

@crm.route('/api/emp-leads-list')
@login_required
def emp_leads_list():
    """Return filtered leads as JSON for popup modal."""
    from datetime import timedelta as _td2
    emp_id  = request.args.get('emp_id', '')
    period  = request.args.get('period', 'this_month')
    status  = request.args.get('status', '')
    today   = datetime.now().date()

    if period == 'today':
        pf = datetime.combine(today, datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_week':
        pf = datetime.combine(today - _td2(days=today.weekday()), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_month':
        pf = datetime(today.year, today.month, 1)
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_month':
        first = today.replace(day=1); lp = first - _td2(days=1)
        pf = datetime(lp.year, lp.month, 1); pt = datetime.combine(lp, datetime.max.time())
    elif period == 'last_3_months':
        pf = datetime.combine(today - _td2(days=90), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_6_months':
        pf = datetime.combine(today - _td2(days=180), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_year':
        pf = datetime(today.year, 1, 1); pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_year':
        pf = datetime(today.year - 1, 1, 1); pt = datetime(today.year - 1, 12, 31, 23, 59, 59)
    else:
        pf = datetime(today.year, today.month, 1); pt = datetime.combine(today, datetime.max.time())

    q = Lead.query.filter(Lead.is_deleted == False, Lead.created_at >= pf, Lead.created_at <= pt)

    if emp_id:
        try:
            uid = int(emp_id)
            q = q.filter(db.or_(Lead.assigned_to == uid, Lead.team_members.like(f'%{uid}%')))
        except ValueError: pass

    if status:
        q = q.filter(Lead.status == status)

    leads = q.distinct().order_by(Lead.created_at.desc()).all()

    _st_objs = LeadStatus.query.filter_by(is_active=True).all()
    status_labels = {st.name: st.name.replace('_',' ').title() for st in _st_objs}
    status_labels.update({'open':'Open','in_process':'In Process','close':'Close','cancel':'Cancel'})
    result = []
    for l in leads:
        result.append({
            'id':       l.id,
            'name':     l.contact_name or '',
            'company':  l.company_name or '',
            'mobile':   l.phone or '',
            'email':    l.email or '',
            'product':  l.product_name or '',
            'status':   l.status or 'open',
            'status_label': status_labels.get(l.status, l.status),
            'lead_type': l.lead_type or 'Quality',
            'age':      l.lead_age,
            'created':  l.created_at.strftime('%d-%m-%Y') if l.created_at else '',
            'city':     l.city or '',
        })

    return jsonify(leads=result, total=len(result))


@crm.route('/api/emp-dashboard-stats')
@login_required
def emp_dashboard_stats():
    """Dashboard stats filtered by employee + period — for Employee Performance tab."""
    from sqlalchemy import func, extract
    from datetime import timedelta as _td

    emp_id  = request.args.get('emp_id', '')
    period  = request.args.get('period', 'this_month')
    today   = datetime.now().date()

    # ── Period date range ──
    if period == 'today':
        pf = datetime.combine(today, datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_week':
        pf = datetime.combine(today - _td(days=today.weekday()), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_month':
        pf = datetime(today.year, today.month, 1)
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_month':
        first = today.replace(day=1)
        lp = first - _td(days=1)
        pf = datetime(lp.year, lp.month, 1)
        pt = datetime.combine(lp, datetime.max.time())
    elif period == 'last_3_months':
        pf = datetime.combine(today - _td(days=90), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_6_months':
        pf = datetime.combine(today - _td(days=180), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_year':
        pf = datetime(today.year, 1, 1)
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_year':
        pf = datetime(today.year - 1, 1, 1)
        pt = datetime(today.year - 1, 12, 31, 23, 59, 59)
    else:
        pf = datetime(today.year, today.month, 1)
        pt = datetime.combine(today, datetime.max.time())

    # ── Base query with employee + period filter ──
    def base_q():
        q = Lead.query.filter(
            Lead.is_deleted == False,
            Lead.created_at >= pf,
            Lead.created_at <= pt,
        )
        if emp_id:
            try:
                uid = int(emp_id)
                q = q.filter(db.or_(
                    Lead.assigned_to == uid,
                    Lead.team_members.like(f'%{uid}%')
                ))
            except ValueError:
                pass
        return q

    # ── Status counts ──
    all_leads = base_q().distinct().all()
    _all_st = LeadStatus.query.filter_by(is_active=True).all()
    counts = {st.name: 0 for st in _all_st}
    counts.update({'open':0,'in_process':0,'close':0,'cancel':0})
    for l in all_leads:
        s = l.status or 'open'
        if s in counts: counts[s] += 1
        else: counts[s] = 1
    total = len(all_leads)

    # ── Monthly trend (last 6 months within period) ──
    monthly = []
    for i in range(5, -1, -1):
        dt = (today.replace(day=1) - _td(days=i*30))
        m, y = dt.month, dt.year
        q = Lead.query.filter(
            Lead.is_deleted == False,
            extract('month', Lead.created_at) == m,
            extract('year', Lead.created_at) == y,
        )
        if emp_id:
            try:
                uid = int(emp_id)
                q = q.filter(db.or_(Lead.assigned_to == uid, Lead.team_members.like(f'%{uid}%')))
            except: pass
        monthly.append({'month': dt.strftime('%b %Y'), 'count': q.count()})

    # ── Status distribution ──
    status_data = [
        {'label':'Open',       'value': counts['open'],       'color':'#94a3b8'},
        {'label':'In Process', 'value': counts['in_process'], 'color':'#1e2d5e'},
        {'label':'Close',      'value': counts['close'],      'color':'#0d9488'},
        {'label':'Cancel',     'value': counts['cancel'],     'color':'#ef4444'},
    ]

    # ── Source wise ──
    src = {}
    for l in all_leads:
        k = l.source or 'Unknown'
        src[k] = src.get(k, 0) + 1
    source_data = [{'label':k,'value':v} for k,v in sorted(src.items(), key=lambda x:-x[1])[:8]]

    # ── Category wise ──
    cat = {}
    for l in all_leads:
        k = l.category or 'N/A'
        cat[k] = cat.get(k, 0) + 1
    cat_data = [{'label':k,'value':v} for k,v in sorted(cat.items(), key=lambda x:-x[1])[:8]]

    # ── Last 7 days ──
    week_data = []
    for i in range(6, -1, -1):
        d = today - _td(days=i)
        q = Lead.query.filter(
            Lead.is_deleted == False,
            func.date(Lead.created_at) == d
        )
        if emp_id:
            try:
                uid = int(emp_id)
                q = q.filter(db.or_(Lead.assigned_to == uid, Lead.team_members.like(f'%{uid}%')))
            except: pass
        week_data.append({'day': d.strftime('%a'), 'count': q.count()})

    # ── Conversion rate ──
    conv = round((counts['close']/total*100), 1) if total > 0 else 0

    # ── Quality breakdown ──
    quality     = sum(1 for l in all_leads if (l.lead_type or 'Quality') == 'Quality')
    non_quality = total - quality

    return jsonify(
        total=total, counts=counts, conversion=conv,
        quality=quality, non_quality=non_quality,
        monthly=monthly, status=status_data,
        sources=source_data, categories=cat_data, week=week_data,
    )


@crm.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    from sqlalchemy import func, extract
    from datetime import datetime, timedelta

    # Monthly lead counts (last 6 months)
    monthly = []
    for i in range(5, -1, -1):
        dt = datetime.utcnow().replace(day=1) - timedelta(days=i*30)
        m, y = dt.month, dt.year
        count = Lead.query.filter(
            extract('month', Lead.created_at) == m,
            extract('year', Lead.created_at) == y
        ).count()
        monthly.append({'month': dt.strftime('%b %Y'), 'count': count})

    # Status distribution
    status_data = [
        {'label': 'Open',       'value': Lead.query.filter_by(status='open').count(),       'color': '#94a3b8'},
        {'label': 'In Process', 'value': Lead.query.filter_by(status='in_process').count(), 'color': '#1e2d5e'},
        {'label': 'Close',      'value': Lead.query.filter_by(status='close').count(),      'color': '#0d9488'},
        {'label': 'Cancel',     'value': Lead.query.filter_by(status='cancel').count(),     'color': '#ef4444'},
    ]

    # Source wise
    sources = db.session.query(Lead.source, func.count(Lead.id)).group_by(Lead.source).all()
    source_data = [{'label': s or 'Unknown', 'value': c} for s, c in sources]

    # Category wise
    cats = db.session.query(Lead.category, func.count(Lead.id)).group_by(Lead.category).all()
    cat_data = [{'label': c or 'N/A', 'value': cnt} for c, cnt in cats]

    # Product range wise
    ranges = db.session.query(Lead.product_range, func.count(Lead.id)).group_by(Lead.product_range).all()
    range_data = [{'label': r or 'N/A', 'value': c} for r, c in ranges]

    # Recent activity - last 7 days leads
    week_data = []
    for i in range(6, -1, -1):
        d = datetime.utcnow().date() - timedelta(days=i)
        c = Lead.query.filter(func.date(Lead.created_at) == d).count()
        week_data.append({'day': d.strftime('%a'), 'count': c})

    return jsonify({
        'monthly': monthly,
        'status': status_data,
        'sources': source_data,
        'categories': cat_data,
        'ranges': range_data,
        'week': week_data,
    })


# ══════════════════════════════════════
# LEAD IMPORT (CSV / Excel)
# ══════════════════════════════════════

@crm.route('/leads/import', methods=['GET', 'POST'])
@login_required
def lead_import():
    if request.method == 'POST':
        import csv, io
        f = request.files.get('import_file')
        if not f or not f.filename:
            flash('Please select a file!', 'warning')
            return redirect(url_for('crm.lead_import'))

        ext = f.filename.rsplit('.', 1)[-1].lower()
        added = 0
        errors = []

        try:
            if ext == 'csv':
                content = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif ext in ('xls', 'xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
            else:
                flash('Only CSV and Excel (.xlsx) files allowed!', 'danger')
                return redirect(url_for('crm.lead_import'))

            for i, row in enumerate(rows, 2):
                name = (row.get('name') or row.get('Name') or '').strip()
                if not name:
                    errors.append(f'Row {i}: Name is required')
                    continue
                try:
                    l = Lead(
                        code             = gen_code(Lead, 'LD'),
                        contact_name     = name,
                        position         = row.get('position') or row.get('Position') or '',
                        email            = row.get('email') or row.get('Email') or '',
                        phone            = row.get('mobile') or row.get('Mobile') or '',
                        alternate_mobile = row.get('alternate_mobile') or row.get('Alternate Mobile') or '',
                        company_name     = row.get('company') or row.get('Company') or '',
                        website          = row.get('website') or row.get('Website') or '',
                        city             = row.get('city') or row.get('City') or '',
                        state            = row.get('state') or row.get('State') or '',
                        country          = row.get('country') or row.get('Country') or 'India',
                        zip_code         = row.get('zip_code') or row.get('Zip Code') or '',
                        product_name     = row.get('product_name') or row.get('Product Name') or '',
                        category         = row.get('category') or row.get('Category') or '',
                        product_range    = row.get('product_range') or row.get('Product Range') or '',
                        source           = row.get('source') or row.get('Source') or '',
                        status           = row.get('status') or row.get('Status') or 'open',
                        lead_type        = row.get('lead_type') or row.get('Lead Type') or 'Quality',
                        order_quantity   = row.get('order_quantity') or row.get('Order Quantity') or '',
                        requirement_spec = row.get('requirement_spec') or row.get('Requirement Specification') or '',
                        remark           = row.get('remark') or row.get('Remark') or '',
                        tags             = row.get('tags') or row.get('Tags') or '',
                        created_by       = current_user.id
                    )
                    # Validate status
                    _valid_import = {st.name.lower() for st in LeadStatus.query.filter_by(is_active=True).all()}
                    _valid_import.update(['open','in_process','close','cancel'])
                    if l.status.lower() not in _valid_import:
                        l.status = 'open'
                    # Validate lead_type
                    if l.lead_type not in ['Quality', 'Non-Quality']:
                        l.lead_type = 'Quality'
                    db.session.add(l)
                    db.session.flush()
                    log_activity(l.id, 'Lead Imported via CSV/Excel')
                    added += 1
                except Exception as e:
                    errors.append(f'Row {i}: {str(e)}')

            db.session.commit()
            flash(f'✅ {added} leads imported successfully!{(" ⚠️ " + str(len(errors)) + " rows had errors.") if errors else ""}', 'success')
            if errors:
                for e in errors[:5]:
                    flash(e, 'warning')

        except Exception as e:
            db.session.rollback()
            flash(f'Import failed: {str(e)}', 'danger')

        return redirect(url_for('crm.leads'))

    return render_template('crm/leads/lead_import.html', active_page='leads')


@crm.route('/leads/import/template')
@login_required
def lead_import_template():
    import io, sys, subprocess
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'], check=True)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lead Import'

    # ── Styles ──
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', fgColor='1E3A5F')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    req_fill  = PatternFill('solid', fgColor='FFE4E4')  # light red for required
    opt_fill  = PatternFill('solid', fgColor='F0F4FF')  # light blue for optional
    data_align= Alignment(vertical='center')
    thin      = Side(style='thin', color='D1D5DB')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Columns: (header, width, required, example, note) ──
    columns = [
        ('name',             18, True,  'John Doe',              'Contact full name'),
        ('position',         15, False, 'CEO',                   'Job title / designation'),
        ('email',            22, False, 'john@abc.com',          'Email address'),
        ('mobile',           15, False, '9999999999',            'Primary mobile'),
        ('alternate_mobile', 15, False, '8888888888',            'Alternate mobile'),
        ('company',          22, False, 'ABC Corp',              'Company / firm name'),
        ('website',          22, False, 'www.abc.com',           'Website URL'),
        ('city',             14, False, 'Mumbai',                'City'),
        ('state',            14, False, 'Maharashtra',           'State'),
        ('country',          14, False, 'India',                 'Country'),
        ('zip_code',         10, False, '400001',                'PIN / ZIP code'),
        ('product_name',     20, False, 'Face Wash',             'Product name'),
        ('category',         16, False, 'Skin Care',             'Product category'),
        ('product_range',    16, False, 'Premium',               'Product range'),
        ('order_quantity',   15, False, '500 units',             'Required quantity'),
        ('source',           16, False, 'HCP Website',           'Lead source'),
        ('status',           14, False, 'open',                  'open / in_process / close / cancel / NPD Project / Existing Project'),
        ('lead_type',        14, False, 'Quality',               'Quality / Non-Quality'),
        ('requirement_spec', 28, False, 'Vitamin C 500ml',       'Product specification'),
        ('remark',           22, False, 'Urgent requirement',    'Remarks / notes'),
        ('tags',             18, False, 'skincare,premium',      'Comma-separated tags'),
    ]

    # ── Header row (row 1) ──
    ws.row_dimensions[1].height = 36
    for col_idx, (col_name, width, required, example, note) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Comment/note on header
        from openpyxl.comments import Comment
        comment = Comment(f'{"REQUIRED" if required else "Optional"}\n{note}', 'HCP ERP')
        comment.width  = 160
        comment.height = 50
        cell.comment   = comment

    # ── Example row (row 2) ──
    ws.row_dimensions[2].height = 22
    example_data = [col[3] for col in columns]
    for col_idx, value in enumerate(example_data, 1):
        required = columns[col_idx-1][2]
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill      = req_fill if required else opt_fill
        cell.alignment = data_align
        cell.border    = border
        cell.font      = Font(size=9, italic=True, color='6B7280')

    # ── Data rows 3-101 styling ──
    for row in range(3, 102):
        ws.row_dimensions[row].height = 20
        for col_idx in range(1, len(columns)+1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border    = border
            cell.alignment = data_align
            cell.font      = Font(size=9)

    # ── Data Validations ──
    # status dropdown
    dv_status = DataValidation(
        type='list', formula1='"open,in_process,close,cancel"',
        allow_blank=True, showDropDown=False,
        error='Use: open, in_process, close, cancel',
        errorTitle='Invalid Status', showErrorMessage=True
    )
    status_col = get_column_letter([c[0] for c in columns].index('status') + 1)
    dv_status.sqref = f'{status_col}3:{status_col}101'
    ws.add_data_validation(dv_status)

    # lead_type dropdown
    dv_type = DataValidation(
        type='list', formula1='"Quality,Non-Quality"',
        allow_blank=True, showDropDown=False,
        error='Use: Quality or Non-Quality',
        errorTitle='Invalid Lead Type', showErrorMessage=True
    )
    type_col = get_column_letter([c[0] for c in columns].index('lead_type') + 1)
    dv_type.sqref = f'{type_col}3:{type_col}101'
    ws.add_data_validation(dv_type)

    # ── Freeze header row ──
    ws.freeze_panes = 'A3'

    # ── Info sheet ──
    ws2 = wb.create_sheet('Instructions')
    instructions = [
        ('LEAD IMPORT INSTRUCTIONS', True),
        ('', False),
        ('1. Fill data from Row 3 onwards (Row 1 = Headers, Row 2 = Example)', False),
        ('2. "name" column is REQUIRED — all other fields are optional', False),
        ('3. status values: open, in_process, close, cancel, NPD Project, Existing Project', False),
        ('4. lead_type values: Quality OR Non-Quality (default: Quality)', False),
        ('5. Dropdown available for status and lead_type columns', False),
        ('6. Max file size: 5MB', False),
        ('', False),
        ('FIELD DESCRIPTIONS:', True),
    ]
    for col_name, width, required, example, note in columns:
        instructions.append((f'  {col_name} — {"[REQUIRED]" if required else "[Optional]"} — {note} — Example: {example}', False))

    for i, (text, bold) in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=11, color='1E3A5F')
        else:
            cell.font = Font(size=9)
    ws2.column_dimensions['A'].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='lead_import_template.xlsx')


# ══════════════════════════════════════
# CLIENT IMPORT (CSV / Excel)
# ══════════════════════════════════════

@crm.route('/clients/import', methods=['GET', 'POST'])
@login_required
def client_import():
    if request.method == 'POST':
        import csv, io
        f = request.files.get('import_file')
        if not f or not f.filename:
            flash('Please select a file!', 'warning')
            return redirect(url_for('crm.client_import'))

        ext = f.filename.rsplit('.', 1)[-1].lower()
        added = 0
        errors = []

        try:
            if ext == 'csv':
                content = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif ext in ('xls', 'xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
            else:
                flash('Only CSV and Excel (.xlsx) files allowed!', 'danger')
                return redirect(url_for('crm.client_import'))

            for i, row in enumerate(rows, 2):
                name = (row.get('contact_name') or row.get('Contact Name') or '').strip()
                if not name:
                    errors.append(f'Row {i}: Contact Name required')
                    continue
                try:
                    c = ClientMaster(
                        code             = gen_code(ClientMaster, 'CLT'),
                        contact_name     = name,
                        position         = row.get('position') or row.get('Position') or '',
                        company_name     = row.get('company_name') or row.get('Company Name') or '',
                        email            = row.get('email') or row.get('Email') or '',
                        phone            = row.get('mobile') or row.get('Mobile') or '',
                        alternate_mobile = row.get('alternate_mobile') or row.get('Alternate Mobile') or '',
                        website          = row.get('website') or row.get('Website') or '',
                        city             = row.get('city') or row.get('City') or '',
                        state            = row.get('state') or row.get('State') or '',
                        country          = row.get('country') or row.get('Country') or 'India',
                        zip_code         = row.get('zip_code') or row.get('Zip Code') or '',
                        gstin            = row.get('gstin') or row.get('GSTIN') or '',
                        status           = 'active',
                        created_by       = current_user.id
                    )
                    db.session.add(c)
                    db.session.flush()
                    # Brands from import
                    brands_str = row.get('brands') or row.get('Brands') or ''
                    if brands_str:
                        for b in brands_str.split('|'):
                            b = b.strip()
                            if b:
                                db.session.add(ClientBrand(client_id=c.id, brand_name=b))
                    added += 1
                except Exception as e:
                    errors.append(f'Row {i}: {str(e)}')

            db.session.commit()
            flash(f'✅ {added} clients imported!{(" ⚠️ " + str(len(errors)) + " errors.") if errors else ""}', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Import failed: {str(e)}', 'danger')

        return redirect(url_for('crm.clients'))

    return render_template('crm/clients/client_import.html', active_page='clients')


@crm.route('/clients/import/template')
@login_required
def client_import_template():
    import csv, io
    from flask import Response
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['contact_name','position','company_name','email','mobile','alternate_mobile',
                     'website','gstin','city','state','country','zip_code','brands'])
    writer.writerow(['Jane Doe','CEO','ABC Corp','jane@abc.com','9999999999','','www.abc.com',
                     '27ABCDE1234F1Z5','Mumbai','Maharashtra','India','400001','regular','Brand A|Brand B'])
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=client_import_template.csv'}
    )



# ══════════════════════════════════════
# CRM DASHBOARD
# ══════════════════════════════════════

@crm.route('/dashboard')
@login_required
def crm_dashboard():
    from models import LeadReminder
    _dash_statuses = LeadStatus.query.filter_by(is_active=True).order_by(LeadStatus.sort_order).all()
    _dash_lead_statuses = _dash_statuses  # reuse for template
    lead_counts = {st.name: Lead.query.filter_by(status=st.name).count() for st in _dash_statuses}
    lead_counts.update({
        'open':       Lead.query.filter_by(status='open').count(),
        'in_process': Lead.query.filter_by(status='in_process').count(),
        'close':      Lead.query.filter_by(status='close').count(),
        'cancel':     Lead.query.filter_by(status='cancel').count(),
        'total':      Lead.query.count(),
    })
    total_clients     = ClientMaster.query.count()
    recent_leads      = Lead.query.order_by(Lead.created_at.desc()).limit(5).all()
    upcoming_reminders = LeadReminder.query.filter(
        LeadReminder.is_done == False,
        LeadReminder.remind_at >= datetime.utcnow()
    ).order_by(LeadReminder.remind_at).limit(5).all()

    # Users who have leads assigned
    all_users = User.query.filter_by(is_active=True).order_by(User.full_name).all()

    return render_template('crm/dashboard/index.html',
        active_page='crm_dashboard',
        lead_counts=lead_counts,
        lead_statuses=_dash_lead_statuses,
        total_clients=total_clients,
        recent_leads=recent_leads,
        upcoming_reminders=upcoming_reminders,
        all_users=all_users,
        is_admin=(current_user.role == 'admin'),
        now=datetime.utcnow())


# ══════════════════════════════════════
# NOTIFICATION API — Reminder Alerts
# ══════════════════════════════════════

@crm.route('/api/employee-performance')
@login_required
def api_employee_performance():
    """Employee performance stats for CRM dashboard."""
    emp_id  = request.args.get('emp_id', '')
    period  = request.args.get('period', 'this_month')

    today = datetime.now().date()

    # ── Period calculation ──
    if period == 'today':
        date_from = datetime.combine(today, datetime.min.time())
        date_to   = datetime.combine(today, datetime.max.time())
        label     = 'Today'
    elif period == 'this_week':
        start = today - timedelta(days=today.weekday())
        date_from = datetime.combine(start, datetime.min.time())
        date_to   = datetime.combine(today, datetime.max.time())
        label     = 'This Week'
    elif period == 'this_month':
        date_from = datetime(today.year, today.month, 1)
        date_to   = datetime.combine(today, datetime.max.time())
        label     = today.strftime('%B %Y')
    elif period == 'last_month':
        first_this = today.replace(day=1)
        last_prev  = first_this - timedelta(days=1)
        date_from  = datetime(last_prev.year, last_prev.month, 1)
        date_to    = datetime.combine(last_prev, datetime.max.time())
        label      = last_prev.strftime('%B %Y')
    elif period == 'last_3_months':
        date_from = datetime.combine(today - timedelta(days=90), datetime.min.time())
        date_to   = datetime.combine(today, datetime.max.time())
        label     = 'Last 3 Months'
    elif period == 'last_6_months':
        date_from = datetime.combine(today - timedelta(days=180), datetime.min.time())
        date_to   = datetime.combine(today, datetime.max.time())
        label     = 'Last 6 Months'
    elif period == 'this_year':
        date_from = datetime(today.year, 1, 1)
        date_to   = datetime.combine(today, datetime.max.time())
        label     = str(today.year)
    else:
        date_from = datetime(today.year, today.month, 1)
        date_to   = datetime.combine(today, datetime.max.time())
        label     = today.strftime('%B %Y')

    # ── Base query ──
    q = Lead.query.filter(
        Lead.is_deleted == False,
        Lead.created_at >= date_from,
        Lead.created_at <= date_to,
    )

    if emp_id:
        try:
            uid = int(emp_id)
            q = q.filter(
                db.or_(
                    Lead.assigned_to == uid,
                    Lead.team_members.like(f'%{uid}%')
                )
            )
        except ValueError:
            pass

    leads = q.distinct().all()

    # ── Count by status ──
    _all_st2 = LeadStatus.query.filter_by(is_active=True).all()
    counts = {st.name: 0 for st in _all_st2}
    counts.update({'open':0,'in_process':0,'close':0,'cancel':0})
    for l in leads:
        s = l.status or 'open'
        if s in counts: counts[s] += 1
        else: counts[s] = 1

    total = len(leads)

    # ── Lead type breakdown ──
    quality     = sum(1 for l in leads if (l.lead_type or 'Quality') == 'Quality')
    non_quality = total - quality

    # ── Recent leads (last 5) ──
    recent = []
    for l in sorted(leads, key=lambda x: x.created_at or datetime.min, reverse=True)[:5]:
        recent.append({
            'id':      l.id,
            'name':    l.contact_name or '',
            'company': l.company_name or '',
            'status':  l.status or 'open',
            'age':     l.lead_age,
        })

    return jsonify(
        period_label = label,
        total        = total,
        counts       = counts,
        quality      = quality,
        non_quality  = non_quality,
        recent       = recent,
    )


@crm.route('/api/due-reminders')
@login_required
def api_due_reminders():
    """
    Returns reminders due in next 5 min OR overdue (not done yet).
    Uses IST (Asia/Kolkata = UTC+5:30) to match browser local time.
    """
    from datetime import timedelta, timezone
    # IST = UTC + 5:30
    IST = timezone(timedelta(hours=5, minutes=30))
    now    = datetime.now(IST).replace(tzinfo=None)   # naive IST datetime
    window = now + timedelta(minutes=5)

    # Get pending reminders due within next 5 min or already overdue
    reminders = LeadReminder.query.filter(
        LeadReminder.is_done == False,
        LeadReminder.remind_at <= window
    ).all()

    results = []
    for r in reminders:
        lead = r.lead
        if not lead:
            continue

        # Check if current user should be notified
        should_notify = False

        # 1. User who set the reminder
        if r.user_id == current_user.id:
            should_notify = True

        # 2. Lead assigned to current user
        if lead.assigned_to == current_user.id:
            should_notify = True

        # 3. Current user is in team_members
        if lead.team_members:
            try:
                team_ids = [int(x) for x in lead.team_members.split(',') if x.strip()]
                if current_user.id in team_ids:
                    should_notify = True
            except Exception:
                pass

        # Admin sees all
        if current_user.role == 'admin':
            should_notify = True

        if should_notify:
            mins_left = int((r.remind_at - now).total_seconds() / 60)
            results.append({
                'id':        r.id,
                'title':     r.title,
                'lead_name': lead.contact_name or '',
                'company':   lead.company_name or '',
                'remind_at': r.remind_at.strftime('%d %b %Y %I:%M %p'),
                'mins_left': mins_left,
                'overdue':   mins_left < 0,
                'lead_url':  f'/crm/leads/{lead.id}?tab=reminder',
            })

    return jsonify(reminders=results, server_time=now.strftime('%d-%m-%Y %H:%M:%S IST'))


@crm.route('/api/stale-leads')
@login_required
def api_stale_leads():
    """
    Returns open leads created 2+ days ago with no followup taken.
    Used for bell notification — "Yeh lead ka followup nahi liya!"
    """
    from datetime import timedelta, timezone
    from config import Config

    IST = timezone(timedelta(hours=5, minutes=30))
    now = datetime.now(IST).replace(tzinfo=None)

    # Get configurable days threshold (default 2)
    stale_days = 2

    cutoff = now - timedelta(days=stale_days)

    # Find open leads older than cutoff
    q = Lead.query.filter(
        Lead.is_deleted == False,
        Lead.status == 'open',
        Lead.created_at <= cutoff,
    )

    # Non-admin sees only their leads
    if current_user.role != 'admin':
        uid_str = str(current_user.id)
        q = q.filter(
            db.or_(
                Lead.assigned_to == current_user.id,
                Lead.team_members.like(f'%{uid_str}%'),
                Lead.created_by == current_user.id,
            )
        )

    stale = q.order_by(Lead.created_at.asc()).limit(20).all()

    results = []
    for l in stale:
        age = l.lead_age
        results.append({
            'id':       l.id,
            'name':     l.contact_name or '',
            'company':  l.company_name or '',
            'age':      age,
            'created':  l.created_at.strftime('%d-%m-%Y') if l.created_at else '',
            'lead_url': f'/crm/leads/{l.id}',
            'has_followup': bool(l.follow_up_date),
            'followup_date': l.follow_up_date.strftime('%d-%m-%Y') if l.follow_up_date else None,
        })

    return jsonify(stale_leads=results, count=len(results), threshold_days=stale_days)


@crm.route('/api/reminder/<int:rid>/snooze', methods=['POST'])
@login_required
def api_reminder_snooze(rid):
    """Snooze reminder by 5 minutes"""
    from datetime import timedelta
    r = LeadReminder.query.get_or_404(rid)
    r.remind_at = r.remind_at + timedelta(minutes=5)
    db.session.commit()
    return jsonify(success=True)


@crm.route('/api/reminder/<int:rid>/done', methods=['POST'])
@login_required
def api_reminder_mark_done(rid):
    """Mark reminder done via notification"""
    r = LeadReminder.query.get_or_404(rid)
    r.is_done = True
    db.session.commit()
    return jsonify(success=True)


# ══════════════════════════════════════
# SAMPLE ORDER PDF
# ══════════════════════════════════════


# ── Global helper: next unique Sample Order number ──
def _next_so_number():
    """Generate next unique HCPSMPLXXX order number (collision-safe)."""
    last = (SampleOrder.query
            .filter(SampleOrder.order_number.like('HCPSMPL%'))
            .order_by(SampleOrder.id.desc())
            .first())
    if last:
        try:
            num = int(last.order_number.replace('HCPSMPL', '')) + 1
        except ValueError:
            num = SampleOrder.query.count() + 1
    else:
        num = 1
    candidate = f'HCPSMPL{num:03d}'
    while SampleOrder.query.filter_by(order_number=candidate).first():
        num += 1
        candidate = f'HCPSMPL{num:03d}'
    return candidate


@crm.route('/sample-orders/next-number')
@login_required
def sample_order_next_number():
    """API: return the next available sample order number as JSON."""
    return jsonify(order_number=_next_so_number())


@crm.route('/leads/<int:id>/sample-order', methods=['POST'])
@login_required
def lead_sample_order(id):
    """Generate Sample Order PDF for a Lead"""
    import io, base64
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    lead = Lead.query.get_or_404(id)

    # ── Form data ──
    so_number    = request.form.get('so_number', '') or _next_so_number()
    so_date      = request.form.get('so_date', date.today().strftime('%d-%m-%Y'))
    so_category  = request.form.get('so_category', 'Sample Order')
    gst_pct      = float(request.form.get('so_gst_pct', '18') or '0')
    bill_company = request.form.get('bill_company', lead.company_name or '')
    bill_address = request.form.get('bill_address', '')
    bill_phone   = request.form.get('bill_phone', lead.phone or '')
    bill_email   = request.form.get('bill_email', lead.email or '')
    bill_gst     = request.form.get('bill_gst', '')
    items_raw    = request.form.getlist('item_name[]')
    qtys         = request.form.getlist('item_qty[]')
    units        = request.form.getlist('item_unit[]')
    rates        = request.form.getlist('item_rate[]')
    descs        = request.form.getlist('item_desc[]')
    terms        = request.form.get('terms', 'Payment: Advance\nDelivery: 7-10 working days\nSamples are for evaluation purpose only.')

    # ── Backend validation ──
    if not bill_company.strip():
        flash('Company Name required', 'danger')
        return redirect(url_for('crm.lead_view', id=id))
    if not bill_address.strip():
        flash('Address required', 'danger')
        return redirect(url_for('crm.lead_view', id=id))
    if not bill_phone.strip():
        flash('Mobile required', 'danger')
        return redirect(url_for('crm.lead_view', id=id))
    valid_items = [n.strip() for n in items_raw if n.strip()]
    if not valid_items:
        flash('Kam se kam ek product add karo', 'danger')
        return redirect(url_for('crm.lead_view', id=id))

    # ── Build PDF ──
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=12*mm, bottomMargin=15*mm)
    W = A4[0] - 30*mm  # usable width

    styles = getSampleStyleSheet()
    def S(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    title_style   = S('T', fontSize=18, fontName='Helvetica-Bold', textColor=colors.HexColor('#1e3a5f'), spaceAfter=2)
    sub_style     = S('Sub', fontSize=8.5, textColor=colors.HexColor('#6b7280'), spaceAfter=0)
    h2_style      = S('H2', fontSize=10, fontName='Helvetica-Bold', textColor=colors.HexColor('#1e3a5f'), spaceBefore=6, spaceAfter=3)
    normal_style  = S('N', fontSize=9, leading=13)
    small_style   = S('Sm', fontSize=8, textColor=colors.HexColor('#6b7280'), leading=11)
    right_style   = S('R', fontSize=9, alignment=TA_RIGHT)
    center_style  = S('C', fontSize=9, alignment=TA_CENTER)

    story = []

    # ── HEADER: Logo left, Company info right ──
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'icons', 'hcp-logo.png')
    company_info = (
        '<b>HCP Wellness Pvt. Ltd.</b><br/>'
        '403, Maruti Vertex Elanza,<br/>'
        'Opp. Global Hospital, Sindhu Bhavan Road, Bodakdev,<br/>'
        'Ahmedabad-380054, Gujarat, India.<br/>'
        '<b>GST :</b> 24AAFCH7246H1ZK'
    )
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=40*mm, height=18*mm, kind='proportional')
        hdr_left = logo
    else:
        hdr_left = Paragraph('<b style="font-size:16px;color:#1e3a5f;">HCP</b>', title_style)

    hdr_tbl = Table([[hdr_left, Paragraph(company_info, S('CI', fontSize=8.5, alignment=TA_RIGHT, leading=13))]],
                    colWidths=[W*0.35, W*0.65])
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1e3a5f'), spaceAfter=5))

    # ── BILLING ADDRESS + ORDER INFO ──
    bill_lines = []
    if bill_company: bill_lines.append(f'<b>{bill_company}</b>')
    if bill_address: bill_lines.append(bill_address.replace('\n', '<br/>'))
    if bill_phone:   bill_lines.append(bill_phone)
    if bill_email:   bill_lines.append(bill_email)
    if bill_gst:     bill_lines.append(f'GST: {bill_gst}')
    bill_txt = (
        f'<font size="7.5" color="#6b7280"><b>BILLING ADDRESS</b></font><br/>'
        + '<br/>'.join(bill_lines)
    )
    order_info_txt = (
        f'<font size="7.5" color="#6b7280">Date</font><br/>'
        f'<b>{so_date}</b><br/><br/>'
        f'<font size="7.5" color="#6b7280">Order ID</font><br/>'
        f'<b>{so_number}</b><br/><br/>'
        f'<font size="7.5" color="#6b7280">Category</font><br/>'
        f'<b>{so_category}</b>'
    )
    addr_tbl = Table([
        [Paragraph(bill_txt, S('BT', fontSize=9, leading=14)),
         Paragraph(order_info_txt, S('OI', fontSize=9, leading=13, alignment=TA_RIGHT))]
    ], colWidths=[W*0.55, W*0.45])
    addr_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (0,-1), 0),
        ('RIGHTPADDING', (-1,0), (-1,-1), 0),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
    ]))
    story.append(addr_tbl)
    story.append(Spacer(1, 5*mm))

    # ── ITEMS TABLE ──
    tbl_header = [
        Paragraph('<b>Product Name</b>', S('TH', fontSize=9, fontName='Helvetica-Bold')),
        Paragraph('<b>Rate (₹)</b>', S('THR', fontSize=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
        Paragraph('<b>Quantity</b>', S('THC', fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph('<b>Amount (₹)</b>', S('THA', fontSize=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
    ]
    tbl_data    = [tbl_header]
    sub_total   = 0.0
    for i, name in enumerate(items_raw):
        if not name.strip():
            continue
        try: qty  = float(qtys[i])  if i < len(qtys)  else 0
        except: qty  = 0
        try: rate = float(rates[i]) if i < len(rates) else 0
        except: rate = 0
        amount    = qty * rate
        sub_total += amount
        tbl_data.append([
            Paragraph(name, normal_style),
            Paragraph(f'{rate:,.2f}', right_style),
            Paragraph(str(int(qty) if qty == int(qty) else qty), center_style),
            Paragraph(f'{amount:,.2f}', right_style),
        ])

    col_w = [W*0.45, W*0.18, W*0.15, W*0.22]
    items_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#e2e8f0')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('ALIGN', (2,0), (2,-1), 'CENTER'),
        ('ALIGN', (3,0), (3,-1), 'RIGHT'),
    ]))
    story.append(items_tbl)
    story.append(Spacer(1, 4*mm))

    # ── TOTALS ──
    gst_amt   = sub_total * gst_pct / 100
    total_amt = sub_total + gst_amt
    totals_data = [
        [Paragraph('Sub Total', S('TL', fontSize=9, textColor=colors.HexColor('#6b7280'))),
         Paragraph(f'{sub_total:,.2f}', right_style)],
        [Paragraph(f'GST ({int(gst_pct) if gst_pct == int(gst_pct) else gst_pct}%)', S('TL2', fontSize=9, textColor=colors.HexColor('#6b7280'))),
         Paragraph(f'{gst_amt:,.2f}', right_style)],
        [Paragraph('<b>Total Amount</b>', S('TLB', fontSize=10, fontName='Helvetica-Bold', textColor=colors.HexColor('#1e3a5f'))),
         Paragraph(f'<b>{total_amt:,.2f}</b>', S('TRB', fontSize=10, fontName='Helvetica-Bold', alignment=TA_RIGHT, textColor=colors.HexColor('#1e3a5f')))],
    ]
    totals_tbl = Table(totals_data, colWidths=[W*0.6, W*0.4])
    totals_tbl.setStyle(TableStyle([
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,1), 0.5, colors.HexColor('#e5e7eb')),
        ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#e8f0fe')),
        ('BOX', (0,2), (-1,2), 0.5, colors.HexColor('#c7d2fe')),
    ]))
    story.append(totals_tbl)
    story.append(Spacer(1, 5*mm))

    # ── TERMS & SIGNATURE ──
    terms_sign = Table([
        [Paragraph(f'<b>Terms & Conditions:</b><br/>{terms.replace(chr(10),"<br/>")}', small_style),
         Paragraph('<br/><br/><br/>________________________<br/><b>Authorised Signature</b>', S('Sig', fontSize=9, alignment=TA_CENTER))]
    ], colWidths=[W*0.6, W*0.4])
    terms_sign.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
    ]))
    story.append(terms_sign)
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#d1d5db')))
    story.append(Paragraph(f'<i>Generated by ERP Demo · {datetime.now().strftime("%d-%m-%Y %H:%M")} · {current_user.full_name}</i>',
                           S('Footer', fontSize=7.5, textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER, spaceBefore=3)))

    doc.build(story)
    buf.seek(0)
    filename = f'SampleOrder_{lead.id}_{so_number}.pdf'

    # ── Save to DB ──
    import json as _json
    items_list = []
    for i, name in enumerate(items_raw):
        if not name.strip(): continue
        try: qty  = float(qtys[i])  if i < len(qtys)  else 0
        except: qty = 0
        try: rate = float(rates[i]) if i < len(rates) else 0
        except: rate = 0
        items_list.append({'name': name, 'qty': qty, 'rate': rate, 'amount': qty*rate})

    try:
        so_date_obj = datetime.strptime(so_date, '%Y-%m-%d').date()
    except Exception:
        so_date_obj = date.today()

    so_rec = SampleOrder(
        order_number = so_number,
        lead_id      = id,
        order_date   = so_date_obj,
        category     = so_category,
        bill_company = bill_company,
        bill_address = bill_address,
        bill_phone   = bill_phone,
        bill_email   = bill_email,
        bill_gst     = bill_gst,
        gst_pct      = gst_pct,
        sub_total    = sub_total,
        gst_amount   = gst_amt,
        total_amount = total_amt,
        items_json   = _json.dumps(items_list),
        terms        = terms,
        created_by   = current_user.id,
    )
    db.session.add(so_rec)
    log_activity(id, f'Sample Order generated: {so_number}')
    audit('sample_orders','INSERT', None, so_number, f'Sample Order {so_number} generated by {current_user.username} for {lead.contact_name} | Total: {total_amt} | Email: {bill_email}')
    try:
        db.session.commit()
    except Exception as _db_err:
        db.session.rollback()
        flash(f'❌ Order number "{so_number}" already exists. Please refresh and try again.', 'danger')
        return redirect(url_for('crm.lead_view', id=id))

    # ── Auto-send confirmation email with PDF attachment ──
    to_email = (bill_email or lead.email or '').strip()
    if to_email:
        try:
            from mail_routes import _get_or_create_sample_order_template, _render_template_vars, _send_smtp
            sender_name = current_user.full_name if hasattr(current_user, 'full_name') and current_user.full_name else 'Administrator'
            t         = _get_or_create_sample_order_template()
            subject   = _render_template_vars(t.subject, lead, so_rec, sender_name)
            body      = _render_template_vars(t.body,    lead, so_rec, sender_name)
            pdf_buf   = _build_sample_order_pdf(so_rec, lead)
            pdf_bytes = pdf_buf.getvalue()
            success, err = _send_smtp(
                to_email, subject, body, t.from_email, t.from_name,
                attachment_bytes=pdf_bytes,
                attachment_name=f'{so_number}.pdf'
            )
            if success:
                flash(f'✅ Sample Order {so_number} created! Mail sent to: {to_email}', 'success')
            else:
                flash(f'✅ Sample Order {so_number} created! (Mail failed: {err})', 'warning')
        except Exception as mail_err:
            flash(f'✅ Sample Order {so_number} created! (Mail error: {mail_err})', 'warning')
    else:
        flash(f'✅ Sample Order {so_number} created! (No email — mail not sent)', 'warning')

    return redirect(url_for('crm.sample_orders_list'))


# ══════════════════════════════════════
# SAMPLE ORDERS — LIST & VIEW
# ══════════════════════════════════════

@crm.route('/sample-orders')
@login_required
def sample_orders_list():
    search   = request.args.get('search', '').strip()
    tab      = request.args.get('tab', 'active')
    page     = request.args.get('page', 1, type=int)
    per_page = 20

    q = SampleOrder.query.join(Lead, SampleOrder.lead_id == Lead.id)

    if tab == 'deleted':
        q = q.filter(SampleOrder.is_deleted == True)
    else:
        q = q.filter((SampleOrder.is_deleted == False) | (SampleOrder.is_deleted == None))

    if search:
        q = q.filter(
            SampleOrder.order_number.ilike(f'%{search}%') |
            SampleOrder.bill_company.ilike(f'%{search}%') |
            Lead.contact_name.ilike(f'%{search}%') |
            Lead.company_name.ilike(f'%{search}%')
        )
    q = q.order_by(SampleOrder.created_at.desc())
    pagination = q.paginate(page=page, per_page=per_page, error_out=False)

    active_count  = SampleOrder.query.filter((SampleOrder.is_deleted == False) | (SampleOrder.is_deleted == None)).count()
    deleted_count = SampleOrder.query.filter(SampleOrder.is_deleted == True).count()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        rows = []
        for so in pagination.items:
            rows.append({
                'id':           so.id,
                'order_number': so.order_number,
                'order_date':   so.order_date.strftime('%d-%m-%Y') if so.order_date else '—',
                'contact_name': so.lead.contact_name if so.lead else '—',
                'company_name': so.lead.company_name if so.lead else '',
                'lead_id':      so.lead_id,
                'bill_company': so.bill_company or '—',
                'sub_total':    float(so.sub_total or 0),
                'gst_pct':      int(so.gst_pct or 0),
                'gst_amount':   float(so.gst_amount or 0),
                'total_amount': float(so.total_amount or 0),
                'creator':      so.creator.full_name if so.creator else '—',
                'created_at':   so.created_at.strftime('%d-%m-%Y %H:%M') if so.created_at else '—',
                'invoice_file': so.invoice_file or '',
                'bill_email':   so.bill_email or '',
                'lead_email':   so.lead.email if so.lead and hasattr(so.lead, 'email') else '',
                'deleted_at':   so.deleted_at.strftime('%d-%m-%Y %H:%M') if so.deleted_at else '',
                'is_deleted':   bool(so.is_deleted),
            })
        return jsonify(rows=rows, total=pagination.total, page=pagination.page,
                       pages=pagination.pages, active_count=active_count, deleted_count=deleted_count)

    return render_template('crm/sample_orders/list.html',
        orders=pagination.items, pagination=pagination,
        search=search, tab=tab,
        active_count=active_count, deleted_count=deleted_count,
        active_page='sample_orders')


@crm.route('/sample-orders/bulk-delete', methods=['POST'])
@login_required
def sample_orders_bulk_delete():
    """Soft-delete selected sample orders."""
    ids = request.form.getlist('ids[]')
    if not ids:
        return jsonify(success=False, message='No orders selected.'), 400
    count = 0
    for sid in ids:
        so = SampleOrder.query.get(int(sid))
        if so and not so.is_deleted:
            so.is_deleted = True
            so.deleted_at = datetime.utcnow()
            so.deleted_by = current_user.id
            count += 1
    db.session.commit()
    return jsonify(success=True, message=f'{count} order(s) moved to Deleted tab.')


@crm.route('/sample-orders/bulk-restore', methods=['POST'])
@login_required
def sample_orders_bulk_restore():
    """Restore soft-deleted sample orders."""
    ids = request.form.getlist('ids[]')
    if not ids:
        return jsonify(success=False, message='No orders selected.'), 400
    count = 0
    for sid in ids:
        so = SampleOrder.query.get(int(sid))
        if so and so.is_deleted:
            so.is_deleted = False
            so.deleted_at = None
            so.deleted_by = None
            count += 1
    db.session.commit()
    return jsonify(success=True, message=f'{count} order(s) restored.')


@crm.route('/sample-orders/bulk-permanent-delete', methods=['POST'])
@login_required
def sample_orders_bulk_permanent_delete():
    """Permanently delete sample orders."""
    ids = request.form.getlist('ids[]')
    if not ids:
        return jsonify(success=False, message='No orders selected.'), 400
    count = 0
    for sid in ids:
        so = SampleOrder.query.get(int(sid))
        if so and so.is_deleted:
            db.session.delete(so)
            count += 1
    db.session.commit()
    return jsonify(success=True, message=f'{count} order(s) permanently deleted.')


INVOICE_UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'invoices')
INVOICE_ALLOWED_EXT  = {'pdf', 'jpg', 'jpeg', 'png'}

def allowed_invoice(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in INVOICE_ALLOWED_EXT


@crm.route('/sample-orders/<int:id>/invoice-upload', methods=['POST'])
@login_required
def sample_order_invoice_upload(id):
    """Upload or replace invoice file for a sample order"""
    so = SampleOrder.query.get_or_404(id)
    f  = request.files.get('invoice_file')
    if not f or not f.filename:
        flash('Koi file select nahi ki.', 'danger')
        return redirect(url_for('crm.sample_orders_list'))

    if not allowed_invoice(f.filename):
        flash('Sirf PDF, JPG, JPEG ya PNG allowed hai.', 'danger')
        return redirect(url_for('crm.sample_orders_list'))

    os.makedirs(INVOICE_UPLOAD_FOLDER, exist_ok=True)

    # Delete old file if exists
    if so.invoice_file:
        old_path = os.path.join(INVOICE_UPLOAD_FOLDER, so.invoice_file)
        if os.path.exists(old_path):
            os.remove(old_path)

    ext   = f.filename.rsplit('.', 1)[1].lower()
    fname = secure_filename(f'{so.order_number}_invoice.{ext}')
    f.save(os.path.join(INVOICE_UPLOAD_FOLDER, fname))

    so.invoice_file = fname
    db.session.commit()
    flash('Invoice successfully upload ho gayi!', 'success')
    return redirect(url_for('crm.sample_orders_list'))


@crm.route('/sample-orders/<int:id>/invoice-download')
@login_required
def sample_order_invoice_download(id):
    """Download the uploaded invoice file"""
    so = SampleOrder.query.get_or_404(id)
    if not so.invoice_file:
        flash('Koi invoice upload nahi hui hai.', 'warning')
        return redirect(url_for('crm.sample_orders_list'))
    fpath = os.path.join(INVOICE_UPLOAD_FOLDER, so.invoice_file)
    if not os.path.exists(fpath):
        flash('File server pe nahi mili.', 'danger')
        return redirect(url_for('crm.sample_orders_list'))
    return send_file(fpath, as_attachment=True, download_name=so.invoice_file)


@crm.route('/sample-orders/<int:id>/invoice-delete', methods=['POST'])
@login_required
def sample_order_invoice_delete(id):
    """Remove uploaded invoice file"""
    so = SampleOrder.query.get_or_404(id)
    if so.invoice_file:
        fpath = os.path.join(INVOICE_UPLOAD_FOLDER, so.invoice_file)
        if os.path.exists(fpath):
            os.remove(fpath)
        so.invoice_file = None
        db.session.commit()
        flash('Invoice delete ho gayi.', 'success')
    return redirect(url_for('crm.sample_orders_list'))


def _build_sample_order_pdf(so, lead):
    """Build and return BytesIO PDF for a SampleOrder object."""
    import io, json as _json
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=12*mm, bottomMargin=15*mm)
    W = A4[0] - 30*mm
    styles = getSampleStyleSheet()
    def S(name, **kw): return ParagraphStyle(name, parent=styles['Normal'], **kw)
    normal_style = S('N', fontSize=9, leading=13)
    small_style  = S('Sm', fontSize=8, textColor=colors.HexColor('#6b7280'), leading=11)
    right_style  = S('R', fontSize=9, alignment=TA_RIGHT)
    center_style = S('C', fontSize=9, alignment=TA_CENTER)
    story = []
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'icons', 'hcp-logo.png')
    company_info = ('<b>HCP Wellness Pvt. Ltd.</b><br/>403, Maruti Vertex Elanza,<br/>'
                    'Opp. Global Hospital, Sindhu Bhavan Road, Bodakdev,<br/>'
                    'Ahmedabad-380054, Gujarat, India.<br/><b>GST :</b> 24AAFCH7246H1ZK')
    hdr_left = Image(logo_path, width=40*mm, height=18*mm, kind='proportional') if os.path.exists(logo_path) else Paragraph('<b>HCP</b>', normal_style)
    hdr_tbl = Table([[hdr_left, Paragraph(company_info, S('CI', fontSize=8.5, alignment=TA_RIGHT, leading=13))]],
                    colWidths=[W*0.35, W*0.65])
    hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT'),
                                  ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1e3a5f'), spaceAfter=5))
    bill_lines = []
    if so.bill_company: bill_lines.append(f'<b>{so.bill_company}</b>')
    if so.bill_address: bill_lines.append(so.bill_address.replace('\n','<br/>'))
    if so.bill_phone:   bill_lines.append(so.bill_phone)
    if so.bill_email:   bill_lines.append(so.bill_email)
    if so.bill_gst:     bill_lines.append(f'GST: {so.bill_gst}')
    bill_txt  = '<font size="7.5" color="#6b7280"><b>BILLING ADDRESS</b></font><br/>' + '<br/>'.join(bill_lines)
    order_txt = (f'<font size="7.5" color="#6b7280">Date</font><br/><b>{so.order_date.strftime("%d-%m-%Y")}</b><br/><br/>'
                 f'<font size="7.5" color="#6b7280">Order ID</font><br/><b>{so.order_number}</b>')
    addr_tbl = Table([[Paragraph(bill_txt, S('BT',fontSize=9,leading=14)),
                       Paragraph(order_txt, S('OI',fontSize=9,leading=13,alignment=TA_RIGHT))]],
                     colWidths=[W*0.55, W*0.45])
    addr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('TOPPADDING',(0,0),(-1,-1),8),
                                   ('BOTTOMPADDING',(0,0),(-1,-1),8),('LEFTPADDING',(0,0),(0,-1),0),
                                   ('RIGHTPADDING',(-1,0),(-1,-1),0),('LINEBELOW',(0,0),(-1,-1),0.5,colors.HexColor('#e5e7eb'))]))
    story.append(addr_tbl)
    story.append(Spacer(1, 5*mm))
    items = _json.loads(so.items_json or '[]')
    tbl_data = [[Paragraph('<b>Product Name</b>', S('TH',fontSize=9,fontName='Helvetica-Bold')),
                 Paragraph('<b>Rate (₹)</b>', S('THR',fontSize=9,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
                 Paragraph('<b>Quantity</b>', S('THC',fontSize=9,fontName='Helvetica-Bold',alignment=TA_CENTER)),
                 Paragraph('<b>Amount (₹)</b>', S('THA',fontSize=9,fontName='Helvetica-Bold',alignment=TA_RIGHT))]]
    for item in items:
        tbl_data.append([Paragraph(item.get('name',''), normal_style),
                         Paragraph(f"{float(item.get('rate',0)):,.2f}", right_style),
                         Paragraph(str(int(item.get('qty',0)) if float(item.get('qty',0))==int(float(item.get('qty',0))) else item.get('qty',0)), center_style),
                         Paragraph(f"{float(item.get('amount',0)):,.2f}", right_style)])
    col_w = [W*0.45, W*0.18, W*0.15, W*0.22]
    items_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#f1f5f9')),
                                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
                                    ('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7),
                                    ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
                                    ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8fafc')]),
                                    ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#e2e8f0')),
                                    ('BOX',(0,0),(-1,-1),1,colors.HexColor('#cbd5e1')),
                                    ('ALIGN',(1,0),(1,-1),'RIGHT'),('ALIGN',(2,0),(2,-1),'CENTER'),('ALIGN',(3,0),(3,-1),'RIGHT')]))
    story.append(items_tbl)
    story.append(Spacer(1, 4*mm))
    gst_pct = float(so.gst_pct or 0)
    totals_data = [
        [Paragraph('Sub Total', S('TL',fontSize=9,textColor=colors.HexColor('#6b7280'))), Paragraph(f'{float(so.sub_total):,.2f}', right_style)],
        [Paragraph(f'GST ({int(gst_pct) if gst_pct==int(gst_pct) else gst_pct}%)', S('TL2',fontSize=9,textColor=colors.HexColor('#6b7280'))), Paragraph(f'{float(so.gst_amount):,.2f}', right_style)],
        [Paragraph('<b>Total Amount</b>', S('TLB',fontSize=10,fontName='Helvetica-Bold',textColor=colors.HexColor('#1e3a5f'))),
         Paragraph(f'<b>{float(so.total_amount):,.2f}</b>', S('TRB',fontSize=10,fontName='Helvetica-Bold',alignment=TA_RIGHT,textColor=colors.HexColor('#1e3a5f')))],
    ]
    totals_tbl = Table(totals_data, colWidths=[W*0.6, W*0.4])
    totals_tbl.setStyle(TableStyle([('ALIGN',(1,0),(1,-1),'RIGHT'),('TOPPADDING',(0,0),(-1,-1),5),
                                     ('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),8),
                                     ('RIGHTPADDING',(0,0),(-1,-1),8),('LINEBELOW',(0,0),(-1,1),0.5,colors.HexColor('#e5e7eb')),
                                     ('BACKGROUND',(0,2),(-1,2),colors.HexColor('#e8f0fe')),
                                     ('BOX',(0,2),(-1,2),0.5,colors.HexColor('#c7d2fe'))]))
    story.append(totals_tbl)
    story.append(Spacer(1, 5*mm))
    terms = so.terms or ''
    story.append(Paragraph(f'<b>Terms & Conditions:</b><br/>{terms.replace(chr(10),"<br/>")}', small_style))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#d1d5db')))
    story.append(Paragraph(f'<i>Generated by ERP · {datetime.now().strftime("%d-%m-%Y %H:%M")}</i>',
                           S('Footer', fontSize=7.5, textColor=colors.HexColor('#9ca3af'), alignment=TA_RIGHT, spaceBefore=3)))
    doc.build(story)
    buf.seek(0)
    return buf


@crm.route('/sample-orders/<int:id>/reprint', methods=['POST'])
@login_required
def sample_order_reprint(id):
    """Re-generate PDF for an existing sample order"""
    import io, json as _json
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    so = SampleOrder.query.get_or_404(id)
    lead = so.lead

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=12*mm, bottomMargin=15*mm)
    W = A4[0] - 30*mm

    styles = getSampleStyleSheet()
    def S(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    normal_style = S('N', fontSize=9, leading=13)
    small_style  = S('Sm', fontSize=8, textColor=colors.HexColor('#6b7280'), leading=11)
    right_style  = S('R', fontSize=9, alignment=TA_RIGHT)
    center_style = S('C', fontSize=9, alignment=TA_CENTER)

    story = []

    # Header
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'icons', 'hcp-logo.png')
    company_info = ('<b>HCP Wellness Pvt. Ltd.</b><br/>403, Maruti Vertex Elanza,<br/>'
                    'Opp. Global Hospital, Sindhu Bhavan Road, Bodakdev,<br/>'
                    'Ahmedabad-380054, Gujarat, India.<br/><b>GST :</b> 24AAFCH7246H1ZK')
    hdr_left = Image(logo_path, width=40*mm, height=18*mm, kind='proportional') if os.path.exists(logo_path) else Paragraph('<b>HCP</b>', normal_style)
    hdr_tbl = Table([[hdr_left, Paragraph(company_info, S('CI', fontSize=8.5, alignment=TA_RIGHT, leading=13))]],
                    colWidths=[W*0.35, W*0.65])
    hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT'),
                                  ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6)]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1e3a5f'), spaceAfter=5))

    # Billing + Order info
    bill_lines = []
    if so.bill_company: bill_lines.append(f'<b>{so.bill_company}</b>')
    if so.bill_address: bill_lines.append(so.bill_address.replace('\n','<br/>'))
    if so.bill_phone:   bill_lines.append(so.bill_phone)
    if so.bill_email:   bill_lines.append(so.bill_email)
    if so.bill_gst:     bill_lines.append(f'GST: {so.bill_gst}')
    bill_txt  = '<font size="7.5" color="#6b7280"><b>BILLING ADDRESS</b></font><br/>' + '<br/>'.join(bill_lines)
    order_txt = (f'<font size="7.5" color="#6b7280">Date</font><br/><b>{so.order_date.strftime("%d-%m-%Y")}</b><br/><br/>'
                 f'<font size="7.5" color="#6b7280">Order ID</font><br/><b>{so.order_number}</b>')
    addr_tbl = Table([[Paragraph(bill_txt, S('BT',fontSize=9,leading=14)),
                       Paragraph(order_txt, S('OI',fontSize=9,leading=13,alignment=TA_RIGHT))]],
                     colWidths=[W*0.55, W*0.45])
    addr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('TOPPADDING',(0,0),(-1,-1),8),
                                   ('BOTTOMPADDING',(0,0),(-1,-1),8),('LEFTPADDING',(0,0),(0,-1),0),
                                   ('RIGHTPADDING',(-1,0),(-1,-1),0),('LINEBELOW',(0,0),(-1,-1),0.5,colors.HexColor('#e5e7eb'))]))
    story.append(addr_tbl)
    story.append(Spacer(1, 5*mm))

    # Items
    items = _json.loads(so.items_json or '[]')
    tbl_data = [[Paragraph('<b>Product Name</b>', S('TH',fontSize=9,fontName='Helvetica-Bold')),
                 Paragraph('<b>Rate (₹)</b>', S('THR',fontSize=9,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
                 Paragraph('<b>Quantity</b>', S('THC',fontSize=9,fontName='Helvetica-Bold',alignment=TA_CENTER)),
                 Paragraph('<b>Amount (₹)</b>', S('THA',fontSize=9,fontName='Helvetica-Bold',alignment=TA_RIGHT))]]
    for item in items:
        tbl_data.append([Paragraph(item.get('name',''), normal_style),
                         Paragraph(f"{float(item.get('rate',0)):,.2f}", right_style),
                         Paragraph(str(int(item.get('qty',0)) if float(item.get('qty',0))==int(float(item.get('qty',0))) else item.get('qty',0)), center_style),
                         Paragraph(f"{float(item.get('amount',0)):,.2f}", right_style)])
    col_w = [W*0.45, W*0.18, W*0.15, W*0.22]
    items_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#f1f5f9')),
                                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
                                    ('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7),
                                    ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
                                    ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8fafc')]),
                                    ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#e2e8f0')),
                                    ('BOX',(0,0),(-1,-1),1,colors.HexColor('#cbd5e1')),
                                    ('ALIGN',(1,0),(1,-1),'RIGHT'),('ALIGN',(2,0),(2,-1),'CENTER'),('ALIGN',(3,0),(3,-1),'RIGHT')]))
    story.append(items_tbl)
    story.append(Spacer(1, 4*mm))

    # Totals
    gst_pct = float(so.gst_pct or 0)
    totals_data = [
        [Paragraph('Sub Total', S('TL',fontSize=9,textColor=colors.HexColor('#6b7280'))), Paragraph(f'{float(so.sub_total):,.2f}', right_style)],
        [Paragraph(f'GST ({int(gst_pct) if gst_pct==int(gst_pct) else gst_pct}%)', S('TL2',fontSize=9,textColor=colors.HexColor('#6b7280'))), Paragraph(f'{float(so.gst_amount):,.2f}', right_style)],
        [Paragraph('<b>Total Amount</b>', S('TLB',fontSize=10,fontName='Helvetica-Bold',textColor=colors.HexColor('#1e3a5f'))),
         Paragraph(f'<b>{float(so.total_amount):,.2f}</b>', S('TRB',fontSize=10,fontName='Helvetica-Bold',alignment=TA_RIGHT,textColor=colors.HexColor('#1e3a5f')))],
    ]
    totals_tbl = Table(totals_data, colWidths=[W*0.6, W*0.4])
    totals_tbl.setStyle(TableStyle([('ALIGN',(1,0),(1,-1),'RIGHT'),('TOPPADDING',(0,0),(-1,-1),5),
                                     ('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),8),
                                     ('RIGHTPADDING',(0,0),(-1,-1),8),('LINEBELOW',(0,0),(-1,1),0.5,colors.HexColor('#e5e7eb')),
                                     ('BACKGROUND',(0,2),(-1,2),colors.HexColor('#e8f0fe')),
                                     ('BOX',(0,2),(-1,2),0.5,colors.HexColor('#c7d2fe'))]))
    story.append(totals_tbl)
    story.append(Spacer(1, 5*mm))

    terms = so.terms or ''
    story.append(Paragraph(f'<b>Terms & Conditions:</b><br/>{terms.replace(chr(10),"<br/>")}', small_style))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#d1d5db')))
    story.append(Paragraph(f'<i>Generated by ERP Demo · {datetime.now().strftime("%d-%m-%Y %H:%M")} · {current_user.full_name}</i>',
                           S('Footer', fontSize=7.5, textColor=colors.HexColor('#9ca3af'), alignment=TA_RIGHT, spaceBefore=3)))

    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=f'{so.order_number}.pdf')


# ══════════════════════════════════════
# CONTRIBUTION API
# ══════════════════════════════════════

@crm.route('/api/lead/<int:id>/contributions')
@login_required
def lead_contributions(id):
    """Get contribution scores for a lead."""
    from sqlalchemy import func
    rows = db.session.query(
        LeadContribution.user_id,
        func.sum(LeadContribution.points).label('total_pts'),
        func.count(LeadContribution.id).label('actions')
    ).filter_by(lead_id=id).group_by(LeadContribution.user_id).order_by(func.sum(LeadContribution.points).desc()).all()

    users = {u.id: u.full_name for u in User.query.all()}
    result = []
    for row in rows:
        result.append({
            'user_id':   row.user_id,
            'name':      users.get(row.user_id, 'Unknown'),
            'points':    int(row.total_pts),
            'actions':   int(row.actions),
        })
    return jsonify(contributions=result)


@crm.route('/api/emp-contribution-stats')
@login_required
def emp_contribution_stats():
    """Employee contribution leaderboard for performance dashboard."""
    from sqlalchemy import func
    from datetime import timedelta as _td3

    period = request.args.get('period', 'this_month')
    today  = datetime.now().date()

    if period == 'today':
        pf = datetime.combine(today, datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_week':
        pf = datetime.combine(today - _td3(days=today.weekday()), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_month':
        pf = datetime(today.year, today.month, 1)
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_month':
        first = today.replace(day=1); lp = first - _td3(days=1)
        pf = datetime(lp.year, lp.month, 1); pt = datetime.combine(lp, datetime.max.time())
    elif period == 'last_3_months':
        pf = datetime.combine(today - _td3(days=90), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_6_months':
        pf = datetime.combine(today - _td3(days=180), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'this_year':
        pf = datetime(today.year, 1, 1)
        pt = datetime.combine(today, datetime.max.time())
    elif period == 'last_year':
        pf = datetime(today.year - 1, 1, 1)
        pt = datetime(today.year - 1, 12, 31, 23, 59, 59)
    else:
        pf = datetime(today.year, today.month, 1)
        pt = datetime.combine(today, datetime.max.time())

    emp_id = request.args.get('emp_id', '', type=str).strip()

    q = db.session.query(
        LeadContribution.user_id,
        func.sum(LeadContribution.points).label('total_pts'),
        func.count(LeadContribution.id).label('actions'),
        func.sum(func.IF(LeadContribution.action_type == 'close', 1, 0)).label('closes'),
        func.sum(func.IF(LeadContribution.action_type == 'comment', 1, 0)).label('comments'),
    ).filter(
        LeadContribution.created_at >= pf,
        LeadContribution.created_at <= pt,
    )
    if emp_id:
        q = q.filter(LeadContribution.user_id == int(emp_id))
    rows = q.group_by(LeadContribution.user_id).order_by(func.sum(LeadContribution.points).desc()).all()

    users = {u.id: u.full_name for u in User.query.all()}
    result = []
    for i, row in enumerate(rows):
        result.append({
            'rank':     i + 1,
            'user_id':  row.user_id,
            'name':     users.get(row.user_id, 'Unknown'),
            'points':   int(row.total_pts or 0),
            'actions':  int(row.actions or 0),
            'closes':   int(row.closes or 0),
            'comments': int(row.comments or 0),
        })
    return jsonify(leaderboard=result)


# ══════════════════════════════════════
# CONTRIBUTION POINTS CONFIG
# ══════════════════════════════════════

DEFAULT_CONTRIB_CONFIG = [
    ('comment',       'Comment Added',         1,  'Discussion board mein comment karne pe'),
    ('edit',          'Lead Edited',            1,  'Lead record update karne pe'),
    ('status_change', 'Status Changed',         2,  'Lead status change karne pe (except close/cancel)'),
    ('close_slab1',   'Close: 1-7 days',       10,  'Lead 1-7 din mein close (active members mein divide)'),
    ('close_slab2',   'Close: 8-14 days',       8,  'Lead 8-14 din mein close (active members mein divide)'),
    ('close_slab3',   'Close: 15-21 days',      6,  'Lead 15-21 din mein close (active members mein divide)'),
    ('close_slab4',   'Close: 22-28 days',      4,  'Lead 22-28 din mein close (active members mein divide)'),
    ('close_slab5',   'Close: 29+ days',        0,  'Lead 29+ din baad close (default 0 points)'),
    ('cancel',        'Lead Cancelled',          0,  'Lead cancel karne pe'),
    ('follow_up',     'Follow Up Set',           1,  'Follow up date set karne pe'),
    ('reminder',      'Reminder Added',          1,  'Reminder add karne pe'),
]


@crm.route('/contribution-config', methods=['GET'])
@login_required
def contribution_config():
    """View contribution points configuration."""
    _seed_contrib_config()
    configs = ContributionConfig.query.order_by(ContributionConfig.id).all()
    return render_template('crm/contribution_config.html',
        configs=configs, active_page='contribution_config')


@crm.route('/contribution-config/save', methods=['POST'])
@login_required
def contribution_config_save():
    """Save contribution points configuration."""
    _seed_contrib_config()
    for action_type, label, default_pts, desc in DEFAULT_CONTRIB_CONFIG:
        pts_str = request.form.get(f'pts_{action_type}', str(default_pts))
        try:
            pts = max(0, int(pts_str))
        except ValueError:
            pts = default_pts
        cfg = ContributionConfig.query.filter_by(action_type=action_type).first()
        if cfg:
            cfg.points     = pts
            cfg.updated_by = current_user.id
            cfg.updated_at = datetime.utcnow()
    db.session.commit()
    flash('✅ Contribution points updated successfully!', 'success')
    return redirect(url_for('crm.contribution_config'))


def _seed_contrib_config():
    """Ensure all default configs exist in DB."""
    for action_type, label, default_pts, desc in DEFAULT_CONTRIB_CONFIG:
        if not ContributionConfig.query.filter_by(action_type=action_type).first():
            db.session.add(ContributionConfig(
                action_type=action_type, label=label,
                points=default_pts, description=desc
            ))
    db.session.commit()


# ══════════════════════════════════════
# EMPLOYEE RANK / LEADERBOARD PAGE
# ══════════════════════════════════════

@crm.route('/leaderboard')
@login_required
def leaderboard():
    all_users = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    return render_template('crm/leaderboard.html',
        all_users=all_users, active_page='leaderboard')


@crm.route('/api/leaderboard-data')
@login_required
def leaderboard_data():
    """Detailed leaderboard with slab breakdown, lead list per employee."""
    from sqlalchemy import func
    from datetime import timedelta as _tdl

    period   = request.args.get('period', 'this_month')
    emp_id   = request.args.get('emp_id', '')
    today    = datetime.now().date()

    # ── Period ──
    if period == 'today':
        pf = datetime.combine(today, datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
        plabel = 'Today'
    elif period == 'this_week':
        pf = datetime.combine(today - _tdl(days=today.weekday()), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
        plabel = 'This Week'
    elif period == 'this_month':
        pf = datetime(today.year, today.month, 1)
        pt = datetime.combine(today, datetime.max.time())
        plabel = today.strftime('%B %Y')
    elif period == 'last_month':
        first = today.replace(day=1); lp = first - _tdl(days=1)
        pf = datetime(lp.year, lp.month, 1); pt = datetime.combine(lp, datetime.max.time())
        plabel = lp.strftime('%B %Y')
    elif period == 'last_3_months':
        pf = datetime.combine(today - _tdl(days=90), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
        plabel = 'Last 3 Months'
    elif period == 'last_6_months':
        pf = datetime.combine(today - _tdl(days=180), datetime.min.time())
        pt = datetime.combine(today, datetime.max.time())
        plabel = 'Last 6 Months'
    elif period == 'this_year':
        pf = datetime(today.year, 1, 1)
        pt = datetime.combine(today, datetime.max.time())
        plabel = str(today.year)
    elif period == 'last_year':
        pf = datetime(today.year - 1, 1, 1)
        pt = datetime(today.year - 1, 12, 31, 23, 59, 59)
        plabel = str(today.year - 1)
    else:
        pf = datetime(today.year, today.month, 1)
        pt = datetime.combine(today, datetime.max.time())
        plabel = today.strftime('%B %Y')

    users_map = {u.id: u.full_name for u in User.query.filter_by(is_active=True).all()}

    # ── All contributions in period ──
    q = db.session.query(
        LeadContribution.user_id,
        LeadContribution.action_type,
        func.sum(LeadContribution.points).label('pts'),
        func.count(LeadContribution.id).label('cnt'),
    ).filter(
        LeadContribution.created_at >= pf,
        LeadContribution.created_at <= pt,
    )
    if emp_id:
        try: q = q.filter(LeadContribution.user_id == int(emp_id))
        except: pass

    rows = q.group_by(LeadContribution.user_id, LeadContribution.action_type).all()

    # Build per-user stats
    stats = {}
    for row in rows:
        uid = row.user_id
        if uid not in stats:
            stats[uid] = {
                'user_id': uid,
                'name': users_map.get(uid, 'Unknown'),
                'total_pts': 0, 'total_actions': 0,
                'comments': 0, 'edits': 0, 'status_changes': 0,
                'closes': 0, 'close_pts': 0,
                'slab1': 0, 'slab2': 0, 'slab3': 0, 'slab4': 0, 'slab5': 0,
                'reminders': 0, 'follow_ups': 0,
            }
        s = stats[uid]
        pts = int(row.pts or 0)
        cnt = int(row.cnt or 0)
        s['total_pts']     += pts
        s['total_actions'] += cnt
        at = row.action_type
        if at == 'comment':       s['comments']      += cnt
        elif at == 'edit':        s['edits']         += cnt
        elif at == 'status_change': s['status_changes'] += cnt
        elif at == 'reminder':    s['reminders']     += cnt
        elif at == 'follow_up':   s['follow_ups']    += cnt
        elif at.startswith('close_slab'):
            s['closes']    += cnt
            s['close_pts'] += pts
            slab_num = at[-1]
            s[f'slab{slab_num}'] = cnt

    # Sort by total points desc
    leaderboard = sorted(stats.values(), key=lambda x: x['total_pts'], reverse=True)
    for i, emp in enumerate(leaderboard):
        emp['rank'] = i + 1

    # ── Per-employee closed leads (for drill-down) ──
    if emp_id:
        try:
            uid = int(emp_id)
            closed_leads = Lead.query.filter(
                Lead.is_deleted == False,
                Lead.closed_at >= pf,
                Lead.closed_at <= pt,
                db.or_(Lead.assigned_to == uid, Lead.team_members.like(f'%{uid}%'))
            ).order_by(Lead.closed_at.desc()).all()

            leads_data = [{
                'id':      l.id,
                    'name':    l.contact_name or '',
                'company': l.company_name or '',
                'age':     l.lead_age,
                'closed':  l.closed_at.strftime('%d-%m-%Y') if l.closed_at else '',
                'status':  l.status,
            } for l in closed_leads]
        except:
            leads_data = []
    else:
        leads_data = []

    return jsonify(
        leaderboard=leaderboard,
        period_label=plabel,
        leads=leads_data,
    )


# ══════════════════════════════════════════════════════
# QUOTATION — CREATE / LIST / REPRINT / EMAIL SEND
# ══════════════════════════════════════════════════════

QUOT_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'quotations')


def _build_quotation_pdf(quot, lead):
    """Tally-style A4 Portrait Quotation PDF."""
    import io, json as _json
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=8*mm, bottomMargin=12*mm)
    W = A4[0] - 24*mm  # ~186mm

    def S(name, **kw):
        base = getSampleStyleSheet()['Normal']
        return ParagraphStyle(name, parent=base, **kw)

    BLK  = colors.black
    HBLU = colors.HexColor('#1e3a5f')
    DGRY = colors.HexColor('#d9d9d9')
    LGRY = colors.HexColor('#f2f2f2')
    GREY = colors.HexColor('#555555')

    n  = S('n',  fontSize=8, leading=11)
    nb = S('nb', fontSize=8, leading=11, fontName='Helvetica-Bold')
    sm = S('sm', fontSize=7, leading=10, textColor=GREY)
    c  = S('c',  fontSize=8, leading=11, alignment=TA_CENTER)
    r  = S('r',  fontSize=8, leading=11, alignment=TA_RIGHT)
    story = []

    date_str  = quot.quot_date.strftime('%d-%b-%Y')  if quot.quot_date  else '—'
    valid_str = quot.valid_until.strftime('%d-%b-%Y') if quot.valid_until else '—'

    # ══════════════════════════════════
    # TITLE
    # ══════════════════════════════════
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph('<b>QUOTATION</b>',
        S('qt', fontSize=16, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=5)))

    # ══════════════════════════════════
    # HEADER BOX
    # Layout exactly like Image 1:
    # ┌──────────────────┬────────────────────────────────────┐
    # │  Logo            │  Quot No.:      Dated.:            │
    # │                  │  QT-008/25-26   19-Mar-2026        │
    # ├──────────────────┼────────────────────────────────────┤
    # │  HCP Wellness    │  INVOICE TO                        │
    # │  address...      │  ABC Corp / address                │
    # └──────────────────┴────────────────────────────────────┘
    # ══════════════════════════════════
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'icons', 'hcp-logo.png')
    logo_img  = Image(logo_path, width=30*mm, height=14*mm, kind='proportional') if os.path.exists(logo_path) else Paragraph('<b>HCP</b>', nb)

    # Meta: Quot No + Date side by side
    meta_row = Table([
        [Paragraph('Quotation No. :', sm), Paragraph('Dated. :', sm)],
        [Paragraph(f'<b>{quot.quot_number}</b>',
                   S('qn', fontSize=11, fontName='Helvetica-Bold')),
         Paragraph(f'<b>{date_str}</b>',
                   S('dn', fontSize=11, fontName='Helvetica-Bold'))],
    ], colWidths=[W*0.25, W*0.25])
    meta_row.setStyle(TableStyle([
        ('TOPPADDING',    (0,0),(-1,-1), 2),
        ('BOTTOMPADDING', (0,0),(-1,-1), 2),
        ('LEFTPADDING',   (0,0),(-1,-1), 5),
    ]))

    # FROM address
    from_para = Paragraph(
        '<b>HCP Wellness Pvt. Ltd.</b><br/>'
        '403, Maruti Vertex Elanza, Opp. Global Hospital,<br/>'
        'Sindhu Bhavan Road, Bodakdev,<br/>'
        'Ahmedabad-380054, Gujarat.<br/>'
        'E-Mail : info@hcpwellness.in',
        S('fr', fontSize=7.5, leading=11)
    )

    # INVOICE TO
    bill_lines = []
    if quot.bill_company: bill_lines.append(f'<b>{quot.bill_company}</b>')
    if quot.bill_address:
        for ln in quot.bill_address.split('\n'):
            if ln.strip(): bill_lines.append(ln.strip())
    if quot.bill_phone:  bill_lines.append(quot.bill_phone)
    if quot.bill_email:  bill_lines.append(quot.bill_email)
    if quot.bill_gst:    bill_lines.append(f'GST: {quot.bill_gst}')

    to_para = Paragraph(
        '<font size="7" color="#555555">INVOICE TO</font><br/>'
        + '<br/>'.join(bill_lines),
        S('to', fontSize=8, leading=12)
    )

    # 2-col header: Left=logo/from, Right=meta/invoice_to
    hdr_tbl = Table([
        [logo_img,   meta_row],
        [from_para,  to_para ],
    ], colWidths=[W*0.40, W*0.60])

    hdr_tbl.setStyle(TableStyle([
        ('BOX',           (0,0),(-1,-1), 0.8, BLK),
        ('INNERGRID',     (0,0),(-1,-1), 0.5, colors.HexColor('#aaaaaa')),
        ('VALIGN',        (0,0),(-1,-1), 'TOP'),
        ('VALIGN',        (0,0),(1,0),   'MIDDLE'),
        ('TOPPADDING',    (0,0),(-1,-1), 7),
        ('BOTTOMPADDING', (0,0),(-1,-1), 7),
        ('LEFTPADDING',   (0,0),(-1,-1), 8),
        ('RIGHTPADDING',  (0,0),(-1,-1), 8),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 1*mm))

    # ══════════════════════════════════
    # ITEMS TABLE — no Due on column
    # ══════════════════════════════════
    items = _json.loads(quot.items_json or '[]')

    headers = [
        Paragraph('<b>Sl\nNo</b>',              S('h', fontSize=8, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph('<b>Description of Goods</b>', S('h', fontSize=8, fontName='Helvetica-Bold')),
        Paragraph('<b>Quantity</b>',             S('h', fontSize=8, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph('<b>Rate</b>',                 S('h', fontSize=8, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
        Paragraph('<b>per</b>',                  S('h', fontSize=8, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph('<b>Disc.%</b>',               S('h', fontSize=8, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph('<b>Amount</b>',               S('h', fontSize=8, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
    ]

    tbl_data     = [headers]
    total_amount = 0.0

    for idx, item in enumerate(items, 1):
        def sv(k): return str(item.get(k,'') or '').strip() or '—'
        uom     = sv('uom') if sv('uom') != '—' else 'nos'
        moq     = float(item.get('moq',    0) or 0)
        fc      = float(item.get('final_cost', 0) or 0)
        row_amt = fc * moq if moq else fc
        total_amount += row_amt

        desc_parts = [f'<b>{sv("name")}</b>']
        if sv('size')    != '—': desc_parts.append(f'Size: {sv("size")} {uom}')
        if sv('code')    != '—': desc_parts.append(f'Code: {sv("code")}')
        if sv('pm_spec') != '—': desc_parts.append(f'PM: {sv("pm_spec")}')
        if sv('category')!= '—': desc_parts.append(f'Category: {sv("category")}')

        tbl_data.append([
            Paragraph(str(idx), c),
            Paragraph('<br/>'.join(desc_parts), S('d', fontSize=8, leading=12)),
            Paragraph(f'{int(moq):,} nos' if moq else '—', c),
            Paragraph(f'{fc:,.2f}', r),
            Paragraph('nos', c),
            Paragraph('—', c),
            Paragraph(f'{row_amt:,.2f}', r),
        ])

    col_w = [W*0.05, W*0.38, W*0.14, W*0.12, W*0.07, W*0.07, W*0.17]
    items_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,0), DGRY),
        ('FONTSIZE',      (0,0),(-1,-1), 8),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
        ('LEFTPADDING',   (0,0),(-1,-1), 4),
        ('RIGHTPADDING',  (0,0),(-1,-1), 4),
        ('GRID',          (0,0),(-1,-1), 0.4, BLK),
        ('VALIGN',        (0,0),(-1,-1), 'TOP'),
        ('ALIGN',         (0,0),(0,-1), 'CENTER'),
        ('ALIGN',         (2,0),(2,-1), 'CENTER'),
        ('ALIGN',         (3,0),(3,-1), 'RIGHT'),
        ('ALIGN',         (4,0),(5,-1), 'CENTER'),
        ('ALIGN',         (6,0),(6,-1), 'RIGHT'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#fafafa')]),
    ]))
    story.append(items_tbl)

    # ── TOTAL ROW ──
    total_row = Table([[
        Paragraph(''), Paragraph(''),
        Paragraph('<b>Total</b>', S('tb', fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph(''), Paragraph(''), Paragraph(''),
        Paragraph(f'<b>Rs. {total_amount:,.2f}</b>',
                  S('ta', fontSize=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
    ]], colWidths=col_w)
    total_row.setStyle(TableStyle([
        ('BOX',           (0,0),(-1,-1), 0.8, BLK),
        ('LINEABOVE',     (0,0),(-1,0),  1,   BLK),
        ('TOPPADDING',    (0,0),(-1,-1), 5),
        ('BOTTOMPADDING', (0,0),(-1,-1), 5),
        ('LEFTPADDING',   (0,0),(-1,-1), 4),
        ('RIGHTPADDING',  (0,0),(-1,-1), 4),
        ('BACKGROUND',    (0,0),(-1,-1), LGRY),
    ]))
    story.append(total_row)

    # ── AMOUNT IN WORDS ──
    try:
        from num2words import num2words as _n2w
        rupees = int(total_amount); paise = round((total_amount - rupees)*100)
        words  = _n2w(rupees, lang='en_IN').title()
        if paise: words += ' and ' + _n2w(paise, lang='en_IN').title() + ' Paise'
        words += ' Only'
    except Exception:
        def _n(n):
            o=['','One','Two','Three','Four','Five','Six','Seven','Eight','Nine','Ten','Eleven',
               'Twelve','Thirteen','Fourteen','Fifteen','Sixteen','Seventeen','Eighteen','Nineteen']
            t=['','','Twenty','Thirty','Forty','Fifty','Sixty','Seventy','Eighty','Ninety']
            if n==0: return 'Zero'
            if n<20: return o[n]
            if n<100: return t[n//10]+(' '+o[n%10] if n%10 else '')
            if n<1000: return o[n//100]+' Hundred'+(' '+_n(n%100) if n%100 else '')
            if n<100000: return _n(n//1000)+' Thousand'+(' '+_n(n%1000) if n%1000 else '')
            if n<10000000: return _n(n//100000)+' Lakh'+(' '+_n(n%100000) if n%100000 else '')
            return _n(n//10000000)+' Crore'+(' '+_n(n%10000000) if n%10000000 else '')
        rupees = int(total_amount); paise = round((total_amount - rupees)*100)
        words  = _n(rupees) + (' and '+_n(paise)+' Paise' if paise else '') + ' Only'

    words_row = Table([[
        Paragraph(f'<b>Amount Chargeable (in words)</b><br/><i>INR {words}</i>',
                  S('aw', fontSize=8, leading=12)),
        Paragraph('<b>E. &amp; O.E</b>', S('eo', fontSize=8, alignment=TA_RIGHT)),
    ]], colWidths=[W*0.78, W*0.22])
    words_row.setStyle(TableStyle([
        ('BOX',           (0,0),(-1,-1), 0.8, BLK),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
        ('LEFTPADDING',   (0,0),(-1,-1), 5),
        ('RIGHTPADDING',  (0,0),(-1,-1), 5),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
    ]))
    story.append(words_row)
    story.append(Spacer(1, 4*mm))

    # ── KEY HIGHLIGHTS ──
    story.append(Paragraph('<b>Key Highlights of the Quotation:</b>',
        S('kh', fontSize=8, fontName='Helvetica-Bold', textColor=HBLU, spaceAfter=2)))
    for h in [
        '<b>Exclusions:</b> The quoted price does not include <b>GST and transportation charges.</b> These will be billed separately.',
        '<b>Validity:</b> This quotation remains valid for 30 days from the date of this quotation.',
        '<b>Note:</b> We have considered PM rates based upon specs shared. Any change in specs will lead to change in PM rates which eventually leads to change in final rates.',
    ]:
        story.append(Paragraph(f'&#8226; &nbsp;{h}',
            S('hl', fontSize=7.5, leading=12, leftIndent=8, spaceAfter=2)))

    # ── TERMS + SIGNATURE ──
    story.append(Spacer(1, 5*mm))
    terms_txt = (quot.terms or '').replace('\n', '<br/>')
    creator   = quot.creator.full_name if quot.creator else 'HCP Wellness'

    footer_tbl = Table([[
        Paragraph(
            f'<b>Company GST TIN</b> : 24AAFCH7246H1ZK<br/>'
            f'<b>Terms &amp; Conditions:</b><br/>{terms_txt}',
            S('tc', fontSize=7.5, leading=12)
        ),
        Paragraph(
            '<br/><br/><br/>'
            '___________________________<br/>'
            '<b>for HCP Wellness Pvt. Ltd.</b><br/><br/>'
            'Authorised Signatory',
            S('sg', fontSize=8, alignment=TA_CENTER)
        ),
    ]], colWidths=[W*0.60, W*0.40])
    footer_tbl.setStyle(TableStyle([
        ('VALIGN',      (0,0),(-1,-1), 'TOP'),
        ('ALIGN',       (1,0),(1,0),   'CENTER'),
        ('TOPPADDING',  (0,0),(-1,-1), 3),
        ('LEFTPADDING', (0,0),(-1,-1), 4),
    ]))
    story.append(footer_tbl)
    story.append(Spacer(1, 2*mm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#cccccc')))
    story.append(Paragraph(
        f'<i>Computer-generated &nbsp;·&nbsp; {datetime.now().strftime("%d-%m-%Y %H:%M")} &nbsp;·&nbsp; {creator}</i>',
        S('ft', fontSize=7, textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER, spaceBefore=2)
    ))

    doc.build(story)
    buf.seek(0)
    return buf



# ── Next Quotation Number ──
def _get_financial_year_qt():
    today = date.today()
    if today.month >= 4:
        return f'{str(today.year)[2:]}-{str(today.year+1)[2:]}'
    else:
        return f'{str(today.year-1)[2:]}-{str(today.year)[2:]}'

def _next_quot_number():
    """Generate next unique QT-001/25-26 number."""
    fy = _get_financial_year_qt()
    last = (Quotation.query
            .filter(Quotation.quot_number.like(f'QT-%/{fy}'))
            .order_by(Quotation.id.desc())
            .first())
    if last:
        try:
            num = int(last.quot_number.split('-')[1].split('/')[0]) + 1
        except Exception:
            num = Quotation.query.count() + 1
    else:
        num = 1
    candidate = f'QT-{num:03d}/{fy}'
    while Quotation.query.filter_by(quot_number=candidate).first():
        num += 1
        candidate = f'QT-{num:03d}/{fy}'
    return candidate

@crm.route('/quotations/next-number')
@login_required
def quotation_next_number():
    return jsonify(quot_number=_next_quot_number())


@crm.route('/leads/<int:id>/quotation', methods=['POST'])
@login_required
def lead_create_quotation(id):
    """Generate Quotation PDF, save to DB and return as download."""
    import json as _json
    lead = Lead.query.get_or_404(id)

    quot_number  = request.form.get('quot_number', '') or _next_quot_number()
    quot_date_s  = request.form.get('quot_date', date.today().strftime('%Y-%m-%d'))
    valid_until_s= request.form.get('valid_until', '')
    subject      = request.form.get('quot_subject', '')
    bill_company = request.form.get('bill_company', lead.company_name or '')
    bill_address = request.form.get('bill_address', '')
    bill_phone   = request.form.get('bill_phone', lead.phone or '')
    bill_email   = request.form.get('bill_email', lead.email or '')
    bill_gst     = request.form.get('bill_gst', '')
    terms        = request.form.get('terms', 'Payment: Advance\nDelivery: 7-10 working days\nQuotation valid as mentioned above.')
    notes        = request.form.get('notes', '')

    # New HCP-style fields per row
    item_names    = request.form.getlist('item_name[]')
    item_sizes    = request.form.getlist('item_size[]')
    item_codes    = request.form.getlist('item_code[]')
    item_uoms     = request.form.getlist('item_uom[]')
    item_costs    = request.form.getlist('item_cost[]')
    item_moqs     = request.form.getlist('item_moq[]')
    item_pm_specs = request.form.getlist('item_pm_spec[]')
    item_pm_costs = request.form.getlist('item_pm_cost[]')
    item_fg_costs = request.form.getlist('item_fg_cost[]')
    item_cats     = request.form.getlist('item_category[]')
    item_finals   = request.form.getlist('item_final_cost[]')

    items_list = []
    sub_total  = 0.0
    for i, name in enumerate(item_names):
        if not name.strip(): continue
        def _f(lst, idx): 
            try: return float(lst[idx]) if idx < len(lst) and lst[idx].strip() else 0.0
            except: return 0.0
        def _s(lst, idx): return lst[idx] if idx < len(lst) else ''
        final_cost = _f(item_finals, i)
        moq        = _f(item_moqs, i)
        sub_total += final_cost * moq
        items_list.append({
            'name':     name,
            'size':     _s(item_sizes, i),
            'code':     _s(item_codes, i),
            'uom':      _s(item_uoms, i),
            'cost':     _f(item_costs, i),
            'moq':      moq,
            'pm_spec':  _s(item_pm_specs, i),
            'pm_cost':  _f(item_pm_costs, i),
            'fg_cost':  _f(item_fg_costs, i),
            'category': _s(item_cats, i),
            'final_cost': final_cost,
        })

    # GST from form
    gst_pct      = float(request.form.get('quot_gst_pct', 0) or 0)
    gst_amount   = sub_total * gst_pct / 100
    total_amount = sub_total + gst_amount

    try:    quot_date_obj   = datetime.strptime(quot_date_s, '%Y-%m-%d').date()
    except: quot_date_obj   = date.today()
    try:    valid_until_obj = datetime.strptime(valid_until_s, '%Y-%m-%d').date() if valid_until_s else None
    except: valid_until_obj = None

    quot = Quotation(
        quot_number  = quot_number,
        lead_id      = id,
        quot_date    = quot_date_obj,
        valid_until  = valid_until_obj,
        subject      = subject,
        bill_company = bill_company,
        bill_address = bill_address,
        bill_phone   = bill_phone,
        bill_email   = bill_email,
        bill_gst     = bill_gst,
        gst_pct      = gst_pct,
        sub_total    = sub_total,
        gst_amount   = gst_amount,
        total_amount = total_amount,
        items_json   = _json.dumps(items_list),
        terms        = terms,
        notes        = notes,
        status       = 'draft',
        created_by   = current_user.id,
    )
    db.session.add(quot)
    log_activity(id, f'Quotation generated: {quot_number}')
    db.session.commit()  # commit first to get quot.id
    audit('quotations','INSERT', quot.id, quot_number, f'Quotation {quot_number} generated by {current_user.username} for {lead.contact_name}')

    flash(f'✅ Quotation {quot_number} created! You can download PDF from Quotations list.', 'success')
    return redirect(url_for('crm.quotations_list'))


@crm.route('/quotations')
@login_required
def quotations_list():
    search   = request.args.get('search', '').strip()
    tab      = request.args.get('tab', 'active')   # 'active' | 'deleted'
    status_f = request.args.get('status', '')       # draft|sent|accepted|rejected|''
    page     = request.args.get('page', 1, type=int)
    per_page = 20

    q = Quotation.query.join(Lead, Quotation.lead_id == Lead.id)

    if tab == 'deleted':
        q = q.filter(Quotation.is_deleted == True)
    else:
        q = q.filter((Quotation.is_deleted == False) | (Quotation.is_deleted == None))

    if search:
        q = q.filter(
            Quotation.quot_number.ilike(f'%{search}%') |
            Quotation.bill_company.ilike(f'%{search}%') |
            Lead.contact_name.ilike(f'%{search}%') |
            Lead.company_name.ilike(f'%{search}%') |
            Quotation.subject.ilike(f'%{search}%')
        )
    if status_f:
        q = q.filter(Quotation.status == status_f)

    q = q.order_by(Quotation.created_at.desc())
    pagination = q.paginate(page=page, per_page=per_page, error_out=False)

    # Count for tab badges
    active_count  = Quotation.query.filter((Quotation.is_deleted == False) | (Quotation.is_deleted == None)).count()
    deleted_count = Quotation.query.filter(Quotation.is_deleted == True).count()

    # AJAX request — return JSON rows only
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        import json as _json
        rows = []
        for qt in pagination.items:
            rows.append({
                'id':           qt.id,
                'quot_number':  qt.quot_number,
                'quot_date':    qt.quot_date.strftime('%d-%m-%Y') if qt.quot_date else '—',
                'valid_until':  qt.valid_until.strftime('%d-%m-%Y') if qt.valid_until else '—',
                'contact_name': qt.lead.contact_name if qt.lead else '—',
                'lead_id':      qt.lead_id,
                'bill_company': qt.bill_company or '',
                'subject':      qt.subject or '—',
                'sub_total':    float(qt.sub_total or 0),
                'gst_pct':      int(qt.gst_pct or 0),
                'gst_amount':   float(qt.gst_amount or 0),
                'total_amount': float(qt.total_amount or 0),
                'status':       qt.status or 'draft',
                'email_sent_at': qt.email_sent_at.strftime('%d-%m-%Y') if qt.email_sent_at else '',
                'email_sent_to': qt.email_sent_to or '',
                'bill_email':   qt.bill_email or '',
                'creator':      qt.creator.full_name if qt.creator else '—',
                'created_at':   qt.created_at.strftime('%d-%m-%Y %H:%M') if qt.created_at else '—',
                'deleted_at':   qt.deleted_at.strftime('%d-%m-%Y %H:%M') if qt.deleted_at else '',
                'is_deleted':   bool(qt.is_deleted),
            })
        return jsonify(
            rows=rows,
            total=pagination.total,
            page=pagination.page,
            pages=pagination.pages,
            active_count=active_count,
            deleted_count=deleted_count,
        )

    return render_template('crm/quotations/list.html',
        quotations=pagination.items, pagination=pagination,
        search=search, tab=tab,
        active_count=active_count, deleted_count=deleted_count,
        active_page='quotations')





@crm.route('/quotations/products/ajax')
@login_required
def quotation_products_ajax():
    import json as _json
    search    = request.args.get('search', '').strip()
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    # Fetch all quotations — filter at item level (items_json not searchable in DB)
    quotations = Quotation.query.join(Lead, Quotation.lead_id == Lead.id)        .order_by(Quotation.created_at.desc()).limit(1000).all()

    rows = []
    search_l = search.lower()
    for quot in quotations:
        try:
            items = _json.loads(quot.items_json or '[]')
        except: items = []
        for item in items:
            name     = item.get('name',     '').strip()
            category = item.get('category', '').strip()
            if not name: continue
            # Search: product name, category, company, contact name
            if search_l:
                haystack = (
                    name.lower() + ' ' +
                    category.lower() + ' ' +
                    (quot.bill_company or '').lower() + ' ' +
                    (quot.lead.contact_name if quot.lead else '').lower()
                )
                if search_l not in haystack:
                    continue
            size = str(item.get('size','') or '').strip()
            uom  = str(item.get('uom', '') or '').strip()
            rows.append({
                'quot_number': quot.quot_number,
                'quot_date':   quot.quot_date.strftime('%d-%m-%Y') if quot.quot_date else '—',
                'company':     quot.bill_company or (quot.lead.contact_name if quot.lead else '—'),
                'name':        name,
                'size_uom':    (size + ' ' + uom).strip() or '—',
                'cost':        item.get('cost',       0),
                'moq':         item.get('moq',        0),
                'pm_spec':     item.get('pm_spec',    '') or '—',
                'pm_cost':     item.get('pm_cost',    0),
                'fg_cost':     item.get('fg_cost',    0),
                'category':    category or '—',
                'final_cost':  item.get('final_cost', 0),
            })

    return jsonify(rows=rows, total=len(rows))

@crm.route('/quotations/products')
@login_required
def quotation_products_list():
    import json as _json
    search    = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    page      = request.args.get('page', 1, type=int)
    per_page  = 50

    q = Quotation.query.join(Lead, Quotation.lead_id == Lead.id)
    if search:
        q = q.filter(
            Quotation.quot_number.ilike(f'%{search}%') |
            Quotation.bill_company.ilike(f'%{search}%') |
            Lead.contact_name.ilike(f'%{search}%')
        )
    if date_from:
        try:
            from datetime import datetime as _dt
            q = q.filter(Quotation.quot_date >= _dt.strptime(date_from, '%Y-%m-%d').date())
        except: pass
    if date_to:
        try:
            from datetime import datetime as _dt
            q = q.filter(Quotation.quot_date <= _dt.strptime(date_to, '%Y-%m-%d').date())
        except: pass

    q = q.order_by(Quotation.created_at.desc())
    pagination = q.paginate(page=page, per_page=per_page, error_out=False)

    # Flatten all items from all quotations
    rows = []
    sr = (page - 1) * per_page
    for quot in pagination.items:
        try:
            items = _json.loads(quot.items_json or '[]')
        except: items = []
        for item in items:
            sr += 1
            rows.append({
                'sr':           sr,
                'quot_number':  quot.quot_number,
                'quot_id':      quot.id,
                'quot_date':    quot.quot_date.strftime('%d-%m-%Y') if quot.quot_date else '—',
                'company':      quot.bill_company or (quot.lead.contact_name if quot.lead else '—'),
                'lead_id':      quot.lead_id,
                'name':         item.get('name', '—'),
                'size':         item.get('size', '—'),
                'uom':          item.get('uom', ''),
                'cost':         item.get('cost', 0),
                'moq':          item.get('moq', '—'),
                'pm_spec':      item.get('pm_spec', '—'),
                'pm_cost':      item.get('pm_cost', 0),
                'fg_cost':      item.get('fg_cost', 0),
                'category':     item.get('category', '—'),
                'final_cost':   item.get('final_cost', 0),
                'status':       quot.status,
            })

    return render_template('crm/quotations/products.html',
        rows=rows, pagination=pagination,
        search=search, date_from=date_from, date_to=date_to,
        active_page='quotation_products')

@crm.route('/quotations/<int:id>/reprint', methods=['POST'])
@login_required
def quotation_reprint(id):
    """Re-download PDF for an existing quotation."""
    quot = Quotation.query.get_or_404(id)
    buf  = _build_quotation_pdf(quot, quot.lead)
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=f'Quotation_{quot.quot_number}.pdf')


@crm.route('/quotations/<int:id>/send-email', methods=['POST'])
@login_required
def quotation_send_email(id):
    """Send quotation PDF as email attachment to client."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.mime.base      import MIMEBase
    from email                import encoders

    quot     = Quotation.query.get_or_404(id)
    to_email = request.form.get('to_email', quot.bill_email or '').strip()

    if not to_email:
        flash('❌ Email address required!', 'danger')
        return redirect(url_for('crm.quotations_list'))

    # Always use Mail Master template — renders subject, body, items table
    from mail_routes import _get_or_create_quotation_template, _render_quot_template_vars
    creator_name = quot.creator.full_name if quot.creator else 'Administrator'
    t        = _get_or_create_quotation_template()
    subject  = request.form.get('subject', '').strip() or _render_quot_template_vars(t.subject, quot, creator_name)
    body_txt = request.form.get('body', '').strip()    or _render_quot_template_vars(t.body,    quot, creator_name)

    try:
        cfg = current_app.config
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From']    = f'HCP Wellness Pvt. Ltd. <{cfg.get("MAIL_USERNAME","info@hcpwellness.in")}>'
        msg['To']      = to_email
        msg['Reply-To']= cfg.get('MAIL_USERNAME', 'info@hcpwellness.in')

        # HTML body
        html_body = f'''<html><body style="font-family:Arial,sans-serif;font-size:14px;color:#333;">
        {body_txt}
        </body></html>'''
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        # PDF attachment
        pdf_buf  = _build_quotation_pdf(quot, quot.lead)
        att      = MIMEBase('application', 'octet-stream')
        att.set_payload(pdf_buf.read())
        encoders.encode_base64(att)
        att.add_header('Content-Disposition', 'attachment',
                       filename=f'Quotation_{quot.quot_number}.pdf')
        msg.attach(att)

        server = smtplib.SMTP(cfg['MAIL_SERVER'], cfg['MAIL_PORT'], timeout=15)
        server.ehlo()
        if cfg.get('MAIL_USE_TLS'): server.starttls()
        if cfg.get('MAIL_USERNAME') and cfg.get('MAIL_PASSWORD'):
            server.login(cfg['MAIL_USERNAME'], cfg['MAIL_PASSWORD'])
        server.sendmail(msg['From'], [to_email], msg.as_string())
        server.quit()

        # Update status & log
        quot.status        = 'sent'
        quot.email_sent_at = datetime.utcnow()
        quot.email_sent_to = to_email
        db.session.commit()
        log_activity(quot.lead_id, f'Quotation {quot.quot_number} emailed to {to_email}')
        flash(f'✅ Quotation successfully sent to {to_email}!', 'success')

    except Exception as e:
        flash(f'❌ Email failed: {e}', 'danger')

    return redirect(url_for('crm.quotations_list'))


@crm.route('/quotations/<int:id>/status', methods=['POST'])
@login_required
def quotation_status_update(id):
    """Update quotation status (accepted / rejected / draft)."""
    quot = Quotation.query.get_or_404(id)
    quot.status = request.form.get('status', quot.status)
    db.session.commit()
    flash('Quotation status updated.', 'success')
    return redirect(url_for('crm.quotations_list'))


# ── Bulk Actions ─────────────────────────────────────────────────────────────

@crm.route('/quotations/bulk-status', methods=['POST'])
@login_required
def quotations_bulk_status():
    """Bulk update status for selected quotations."""
    ids    = request.form.getlist('ids[]')
    status = request.form.get('status', '').strip()
    valid  = {'draft', 'sent', 'accepted', 'rejected'}
    if not ids:
        return jsonify(success=False, message='No quotations selected.'), 400
    if status not in valid:
        return jsonify(success=False, message='Invalid status.'), 400
    updated = 0
    for qid in ids:
        quot = Quotation.query.get(int(qid))
        if quot:
            quot.status = status
            updated += 1
    db.session.commit()
    return jsonify(success=True, message=f'{updated} quotation(s) updated to "{status}".')


@crm.route('/quotations/bulk-delete', methods=['POST'])
@login_required
def quotations_bulk_delete():
    """Soft-delete selected quotations (moved to Deleted tab)."""
    ids = request.form.getlist('ids[]')
    if not ids:
        return jsonify(success=False, message='No quotations selected.'), 400
    deleted = 0
    for qid in ids:
        quot = Quotation.query.get(int(qid))
        if quot and not quot.is_deleted:
            quot.is_deleted  = True
            quot.deleted_at  = datetime.utcnow()
            quot.deleted_by  = current_user.id
            log_activity(quot.lead_id, f'Quotation {quot.quot_number} moved to trash by {current_user.username}')
            deleted += 1
    db.session.commit()
    return jsonify(success=True, message=f'{deleted} quotation(s) moved to Deleted tab.')


@crm.route('/quotations/bulk-restore', methods=['POST'])
@login_required
def quotations_bulk_restore():
    """Restore soft-deleted quotations back to active."""
    ids = request.form.getlist('ids[]')
    if not ids:
        return jsonify(success=False, message='No quotations selected.'), 400
    restored = 0
    for qid in ids:
        quot = Quotation.query.get(int(qid))
        if quot and quot.is_deleted:
            quot.is_deleted = False
            quot.deleted_at = None
            quot.deleted_by = None
            log_activity(quot.lead_id, f'Quotation {quot.quot_number} restored by {current_user.username}')
            restored += 1
    db.session.commit()
    return jsonify(success=True, message=f'{restored} quotation(s) restored successfully.')


@crm.route('/quotations/bulk-permanent-delete', methods=['POST'])
@login_required
def quotations_bulk_permanent_delete():
    """Permanently delete quotations from Deleted tab."""
    ids = request.form.getlist('ids[]')
    if not ids:
        return jsonify(success=False, message='No quotations selected.'), 400
    count = 0
    for qid in ids:
        quot = Quotation.query.get(int(qid))
        if quot and quot.is_deleted:
            log_activity(quot.lead_id, f'Quotation {quot.quot_number} permanently deleted by {current_user.username}')
            db.session.delete(quot)
            count += 1
    db.session.commit()
    return jsonify(success=True, message=f'{count} quotation(s) permanently deleted.')


@crm.route('/quotations/bulk-email', methods=['POST'])
@login_required
def quotations_bulk_email():
    """Bulk send email using the shared _send_smtp helper from mail_routes."""
    from mail_routes import _get_or_create_quotation_template, _render_quot_template_vars, _send_smtp

    ids = request.form.getlist('ids[]')
    if not ids:
        return jsonify(success=False, message='No quotations selected.'), 400

    cfg        = current_app.config
    t          = _get_or_create_quotation_template()
    from_email = cfg.get('MAIL_USERNAME', 'info@hcpwellness.in')
    from_name  = 'HCP Wellness Pvt. Ltd.'
    sent = 0; skipped = 0; errors = []

    for qid in ids:
        try:
            quot = Quotation.query.get(int(qid))
        except Exception:
            continue
        if not quot:
            continue
        to_email = (quot.bill_email or '').strip()
        if not to_email:
            skipped += 1
            continue
        creator_name = quot.creator.full_name if quot.creator else 'Administrator'
        subject   = _render_quot_template_vars(t.subject, quot, creator_name)
        body_txt  = _render_quot_template_vars(t.body,    quot, creator_name)
        html_body = f'<html><body style="font-family:Arial,sans-serif;font-size:14px;color:#333;">{body_txt}</body></html>'
        try:
            pdf_bytes = _build_quotation_pdf(quot, quot.lead).read()
        except Exception as e:
            errors.append(f'{quot.quot_number}: PDF build failed – {e}')
            continue
        success, err = _send_smtp(
            to_email, subject, html_body, from_email, from_name,
            attachment_bytes=pdf_bytes,
            attachment_name=f'Quotation_{quot.quot_number}.pdf'
        )
        if success:
            quot.status        = 'sent'
            quot.email_sent_at = datetime.utcnow()
            quot.email_sent_to = to_email
            log_activity(quot.lead_id, f'Quotation {quot.quot_number} bulk-emailed to {to_email}')
            sent += 1
        else:
            errors.append(f'{quot.quot_number}: {err}')

    db.session.commit()
    parts = [f'{sent} email(s) sent']
    if skipped: parts.append(f'{skipped} skipped (no email on record)')
    if errors:  parts.append(f'{len(errors)} failed')
    return jsonify(success=True, message='; '.join(parts), errors=errors)

